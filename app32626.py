import streamlit as st
import pandas as pd
import geopandas as gpd
import numpy as np
import plotly.graph_objects as go
from shapely.geometry import Point, Polygon, MultiPolygon, box, shape
from shapely.ops import unary_union
import os, itertools, glob, math, simplekml, heapq, re, random, json, io, datetime, base64, smtplib, uuid
from concurrent.futures import ThreadPoolExecutor
import pulp
import urllib.request
import urllib.parse
import zipfile
import streamlit.components.v1 as components
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import gspread
from google.oauth2.service_account import Credentials
import pyproj

# --- PAGE CONFIG & INITIALIZE SESSION STATE ---
st.set_page_config(page_title="BRINC COS Drone Optimizer", layout="wide", initial_sidebar_state="expanded")

# This MUST run before any st.session_state checks to prevent KeyError
defaults = {
    'csvs_ready': False, 'df_calls': None, 'df_calls_full': None, 'df_stations': None,
    'active_city': "Victoria", 'active_state': "TX", 'estimated_pop': 65000,
    'k_resp': 2, 'k_guard': 0, 'r_resp': 2.0, 'r_guard': 8.0,
    'dfr_rate': 25, 'deflect_rate': 30, 'total_original_calls': 0, 'total_modeled_calls': 0,
    'onboarding_done': False, 'trigger_sim': False, 'city_count': 1,
    'brinc_user': 'steven.beltran',
    'pd_chief_name': '', 'pd_dept_name': '', 'pd_dept_email': '', 'pd_dept_phone': '',
    'session_start': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'session_id': str(uuid.uuid4())[:8],
    'data_source': 'unknown',   # 'cad_upload' | 'simulation' | 'demo' | 'brinc_file'
    'map_build_logged': False,  # prevent duplicate map-build rows per session
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

if 'target_cities' not in st.session_state:
    st.session_state['target_cities'] = [{"city": st.session_state.get('active_city', 'Victoria'), "state": st.session_state.get('active_state', 'TX')}]


GUARDIAN_FLIGHT_HOURS_PER_DAY = 23.5

SIMULATOR_DISCLAIMER_SHORT = (
    "Simulation output only. Coverage, station placement, response time, and ROI figures are model estimates based on uploaded data and configuration settings. "
    "They are not guarantees of real-world performance, legal compliance, FAA approval, procurement outcome, or financial results."
)

def _build_details_html(details):
    """Shared HTML block for deployment details used in email notifications."""
    if not details: return ""
    drone_list = "".join([
        f"<li><b>{d['name']}</b> ({d['type']}) @ {d['lat']:.4f}, {d['lon']:.4f}</li>"
        for d in details.get('active_drones', [])
    ])
    pop   = details.get('population', 0)
    calls = details.get('total_calls', 0)
    daily = details.get('daily_calls', 0)
    area  = details.get('area_sq_mi', 0)
    be    = details.get('break_even', 'N/A')
    src   = details.get('data_source', '—')
    sid   = details.get('session_id', '—')
    stime = details.get('session_start', '—')
    dur   = details.get('session_duration_min', '—')
    pd_c  = details.get('pd_chief', '—')
    pd_d  = details.get('pd_dept', '—')
    avg_t = details.get('avg_response_min', 0)
    time_saved = details.get('avg_time_saved_min', 0)
    area_cov = details.get('area_covered_pct', 0)
    return f"""
    <div style="margin-top:20px; padding-top:20px; border-top:1px solid #f0f0f0;">
        <h4 style="color:#555; margin-bottom:10px;">Session Info</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Session ID</td><td style="padding:4px;">{sid}</td></tr>
            <tr><td style="padding:4px; color:#888;">Session Start</td><td style="padding:4px;">{stime}</td></tr>
            <tr><td style="padding:4px; color:#888;">Session Duration</td><td style="padding:4px;">{dur} min</td></tr>
            <tr><td style="padding:4px; color:#888;">Data Source</td><td style="padding:4px;">{src}</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Jurisdiction</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Population</td><td style="padding:4px;">{pop:,}</td></tr>
            <tr><td style="padding:4px; color:#888;">Total Annual Calls</td><td style="padding:4px;">{calls:,}</td></tr>
            <tr><td style="padding:4px; color:#888;">Daily Calls</td><td style="padding:4px;">{daily:,}</td></tr>
            <tr><td style="padding:4px; color:#888;">Coverage Area</td><td style="padding:4px;">{area:,} sq mi</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Deployment Settings</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Strategy</td><td style="padding:4px;">{details.get('opt_strategy', '')}</td></tr>
            <tr><td style="padding:4px; color:#888;">Incremental Build</td><td style="padding:4px;">{details.get('incremental_build', False)}</td></tr>
            <tr><td style="padding:4px; color:#888;">Allow Overlap</td><td style="padding:4px;">{details.get('allow_redundancy', False)}</td></tr>
            <tr><td style="padding:4px; color:#888;">DFR Dispatch Rate</td><td style="padding:4px;">{details.get('dfr_rate', 0)}%</td></tr>
            <tr><td style="padding:4px; color:#888;">Deflection Rate</td><td style="padding:4px;">{details.get('deflect_rate', 0)}%</td></tr>
            <tr><td style="padding:4px; color:#888;">Total CapEx</td><td style="padding:4px;">${details.get('fleet_capex', 0):,.0f}</td></tr>
            <tr><td style="padding:4px; color:#888;">Annual Savings</td><td style="padding:4px;">${details.get('annual_savings', 0):,.0f}</td></tr>
            <tr><td style="padding:4px; color:#888;">Break-Even</td><td style="padding:4px;">{be}</td></tr>
            <tr><td style="padding:4px; color:#888;">Avg Response Time</td><td style="padding:4px;">{avg_t:.1f} min</td></tr>
            <tr><td style="padding:4px; color:#888;">Time Saved vs Patrol</td><td style="padding:4px;">{time_saved:.1f} min</td></tr>
            <tr><td style="padding:4px; color:#888;">Geographic Coverage</td><td style="padding:4px;">{area_cov:.1f}%</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Police Dept Contact</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Signatory</td><td style="padding:4px;">{pd_c}</td></tr>
            <tr><td style="padding:4px; color:#888;">Department</td><td style="padding:4px;">{pd_d}</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Active Drones Placed</h4>
        <ul style="font-size:12px; color:#444; padding-left:20px;">{drone_list}</ul>
    </div>
    """

def _build_sheets_row(city, state, event_type, k_resp, k_guard, coverage, name, email, details=None):
    """Build the flat list of values for a Google Sheets row — single source of truth."""
    d = details or {}
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session_start = d.get('session_start', now)
    dur = d.get('session_duration_min', '')
    try:
        if dur == '':
            start_dt = datetime.datetime.strptime(session_start, "%Y-%m-%d %H:%M:%S")
            dur = round((datetime.datetime.now() - start_dt).total_seconds() / 60, 1)
    except Exception:
        dur = ''
    return [
        # ── Identity ────────────────────────────────────────────────────────
        now,                                    # A: Timestamp
        d.get('session_id', ''),               # B: Session ID
        session_start,                          # C: Session Start
        dur,                                    # D: Session Duration (min)
        d.get('data_source', ''),              # E: Data Source
        # ── Who ─────────────────────────────────────────────────────────────
        name,                                   # F: BRINC Rep Name
        email,                                  # G: BRINC Rep Email
        d.get('pd_chief', ''),                 # H: PD Chief / Signatory
        d.get('pd_dept', ''),                  # I: Department Name
        d.get('pd_dept_email', ''),            # J: Department Email
        d.get('pd_dept_phone', ''),            # K: Department Phone
        # ── Where ───────────────────────────────────────────────────────────
        city,                                   # L: City
        state,                                  # M: State
        d.get('population', ''),               # N: Population
        d.get('area_sq_mi', ''),               # O: Area (sq mi)
        # ── Calls ───────────────────────────────────────────────────────────
        d.get('total_calls', ''),              # P: Total Annual Calls
        d.get('daily_calls', ''),              # Q: Daily Calls
        # ── Fleet ───────────────────────────────────────────────────────────
        event_type,                             # R: Event Type
        k_resp,                                 # S: Responders
        k_guard,                                # T: Guardians
        round(coverage, 1) if coverage else '', # U: Call Coverage %
        d.get('area_covered_pct', ''),         # V: Area Coverage %
        d.get('avg_response_min', ''),         # W: Avg Response (min)
        d.get('avg_time_saved_min', ''),       # X: Time Saved vs Patrol (min)
        # ── Financials ──────────────────────────────────────────────────────
        d.get('fleet_capex', ''),              # Y: Fleet CapEx
        d.get('annual_savings', ''),           # Z: Annual Savings
        d.get('break_even', ''),               # AA: Break-Even
        # ── Settings ────────────────────────────────────────────────────────
        d.get('opt_strategy', ''),             # AB: Opt Strategy
        d.get('dfr_rate', ''),                 # AC: DFR Rate %
        d.get('deflect_rate', ''),             # AD: Deflection Rate %
        d.get('incremental_build', ''),        # AE: Incremental Build
        d.get('allow_redundancy', ''),         # AF: Allow Overlap
        # ── Drones detail (JSON) ─────────────────────────────────────────────
        json.dumps([{"name": dr.get("name"), "type": dr.get("type"),
                     "lat": dr.get("lat"), "lon": dr.get("lon"),
                     "avg_time_min": dr.get("avg_time_min"),
                     "faa_ceiling": dr.get("faa_ceiling"),
                     "annual_savings": dr.get("annual_savings")}
                    for dr in d.get('active_drones', [])]),  # AG: Drone JSON
    ]

def _notify_email(city, state, file_type, k_resp, k_guard, coverage, name, email, details=None):
    try:
        gmail_address  = st.secrets.get("GMAIL_ADDRESS", "")
        app_password   = st.secrets.get("GMAIL_APP_PASSWORD", "")
        notify_address = st.secrets.get("NOTIFY_EMAIL", gmail_address)
        if not gmail_address or not app_password: return
        emoji = {"HTML": "📄", "KML": "🌏", "BRINC": "💾", "MAP_BUILD": "🗺️"}.get(file_type, "📥")
        subject = f"{emoji} BRINC {file_type.replace('_',' ').title()} — {city}, {state}"
        details_html = _build_details_html(details)
        d = details or {}
        pop  = d.get('population', 0)
        body = f"""
        <html><body style="font-family:Arial,sans-serif;color:#333;padding:20px;">
        <div style="max-width:560px;margin:0 auto;border:1px solid #ddd;border-radius:8px;overflow:hidden;">
            <div style="background:#000;padding:16px 20px;border-bottom:3px solid #00D2FF;">
                <span style="color:#00D2FF;font-size:18px;font-weight:900;letter-spacing:2px;">BRINC</span>
                <span style="color:#888;font-size:12px;margin-left:8px;">{file_type.replace('_',' ').title()} Notification</span>
            </div>
            <div style="padding:20px;">
                <table style="width:100%;border-collapse:collapse;font-size:14px;">
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;width:40%;">Event</td><td style="padding:8px 4px;font-weight:bold;">{emoji} {file_type.replace('_',' ').title()}</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Jurisdiction</td><td style="padding:8px 4px;font-weight:bold;">{city}, {state}</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Population</td><td style="padding:8px 4px;">{pop:,}</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Fleet</td><td style="padding:8px 4px;">{k_resp} Responder · {k_guard} Guardian</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Call Coverage</td><td style="padding:8px 4px;">{coverage:.1f}%</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">BRINC Rep</td><td style="padding:8px 4px;">{name if name else '—'}</td></tr>
                    <tr><td style="padding:8px 4px;color:#888;">Rep Email</td><td style="padding:8px 4px;">{f'<a href="mailto:{email}">{email}</a>' if email else '—'}</td></tr>
                </table>
                {details_html}
                <div style="margin-top:16px;font-size:11px;color:#bbb;">{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")} UTC</div>
            </div>
        </div>
        </body></html>
        """
        msg = MIMEMultipart("alternative")
        msg["Subject"], msg["From"], msg["To"] = subject, gmail_address, notify_address
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=8) as server:
            server.login(gmail_address, app_password)
            server.sendmail(gmail_address, notify_address, msg.as_string())
    except: pass

def _log_to_sheets(city, state, file_type, k_resp, k_guard, coverage, name, email, details=None):
    try:
        sheet_id = st.secrets.get("GOOGLE_SHEET_ID", "")
        creds_dict = st.secrets.get("gcp_service_account", {})
        if not sheet_id or not creds_dict: return
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(dict(creds_dict), scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(sheet_id).sheet1
        row = _build_sheets_row(city, state, file_type, k_resp, k_guard, coverage, name, email, details)
        sheet.append_row(row)
    except: pass

# --- GLOBAL CONFIGURATION ---
CONFIG = {
    "RESPONDER_COST": 80000, "GUARDIAN_COST": 160000, "RESPONDER_RANGE_MI": 2.0,
    "OFFICER_COST_PER_CALL": 82, "DRONE_COST_PER_CALL": 6,
    "DEFAULT_TRAFFIC_SPEED": 35.0, "RESPONDER_SPEED": 42.0, "GUARDIAN_SPEED": 60.0,
    # Guardian duty cycle: 60 min flight + 3 min charge = 63 min cycle
    # Daily airtime = (24*60) / 63 * 60 = 1371.4 min = 22.86 hrs
    "GUARDIAN_FLIGHT_MIN":  60,   # flight minutes per cycle
    "GUARDIAN_CHARGE_MIN":   3,   # charge minutes per cycle
    # Responder duty cycle: 30-min max flight per sortie, 11.6hr shift
    "RESPONDER_FLIGHT_MIN":   30,    # max flight minutes per sortie
    "RESPONDER_PATROL_HOURS": 11.6,
}
# Derived: compute Guardian daily airtime from duty cycle
CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"] = (
    (24 * 60) / (CONFIG["GUARDIAN_FLIGHT_MIN"] + CONFIG["GUARDIAN_CHARGE_MIN"])
) * CONFIG["GUARDIAN_FLIGHT_MIN"]
CONFIG["GUARDIAN_PATROL_HOURS"] = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"] / 60
STATE_FIPS = {"AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08", "CT": "09", "DE": "10", "FL": "12", "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33", "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55", "WY": "56"}
US_STATES_ABBR = {"Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR", "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY"}
KNOWN_POPULATIONS = {"Victoria": 65534, "New York": 8336817, "Los Angeles": 3822238, "Chicago": 2665039, "Houston": 1304379, "Phoenix": 1644409, "Philadelphia": 1567258, "San Antonio": 2302878, "San Diego": 1472530, "Dallas": 1299544, "San Jose": 1381162, "Austin": 974447, "Jacksonville": 971319, "Fort Worth": 956709, "Columbus": 907971, "Indianapolis": 880621, "Charlotte": 897720, "San Francisco": 971233, "Seattle": 749256, "Denver": 713252, "Washington": 678972, "Nashville": 683622, "Oklahoma City": 694800, "El Paso": 694553, "Boston": 650706, "Portland": 635067, "Las Vegas": 656274, "Detroit": 620376, "Memphis": 633104, "Louisville": 628594, "Baltimore": 620961, "Milwaukee": 620251, "Albuquerque": 677122, "Tucson": 564559, "Fresno": 677102, "Sacramento": 808418, "Kansas City": 697738, "Mesa": 504258, "Atlanta": 499127, "Omaha": 508901, "Colorado Springs": 483956, "Raleigh": 476587, "Miami": 449514, "Virginia Beach": 455369, "Oakland": 530763, "Minneapolis": 563332, "Tulsa": 547239, "Arlington": 398654, "New Orleans": 562503, "Wichita": 402263, "Cleveland": 900000, "Tampa": 449514, "Orlando": 316081}
DEMO_CITIES = [("Las Vegas", "NV"), ("Austin", "TX"), ("Seattle", "WA"), ("Denver", "CO"), ("Nashville", "TN"), ("Columbus", "OH"), ("Detroit", "MI"), ("San Diego", "CA"), ("Charlotte", "NC"), ("Portland", "OR"), ("Memphis", "TN"), ("Louisville", "KY"), ("Baltimore", "MD"), ("Milwaukee", "WI"), ("Albuquerque", "NM"), ("Tucson", "AZ"), ("Fresno", "CA"), ("Sacramento", "CA"), ("Kansas City", "MO"), ("Mesa", "AZ"), ("Atlanta", "GA"), ("Omaha", "NE"), ("Colorado Springs", "CO"), ("Raleigh", "NC"), ("Miami", "FL"), ("Minneapolis", "MN"), ("Tulsa", "OK"), ("Arlington", "TX"), ("Tampa", "FL"), ("New Orleans", "LA"), ("Wichita", "KS"), ("Cleveland", "OH"), ("Virginia Beach", "VA"), ("Oakland", "CA"), ("Indianapolis", "IN"), ("Jacksonville", "FL"), ("Fort Worth", "TX"), ("Boston", "MA"), ("El Paso", "TX"), ("Oklahoma City", "OK"), ("Boise", "ID"), ("Richmond", "VA"), ("Spokane", "WA"), ("Tacoma", "WA"), ("Aurora", "CO"), ("Anaheim", "CA"), ("Bakersfield", "CA"), ("Riverside", "CA"), ("Stockton", "CA"), ("Corpus Christi", "TX"), ("Lexington", "KY"), ("Henderson", "NV"), ("Saint Paul", "MN"), ("Anchorage", "AK"), ("Plano", "TX"), ("Lincoln", "NE"), ("Buffalo", "NY"), ("Fort Wayne", "IN"), ("Jersey City", "NJ"), ("Chula Vista", "CA"), ("Orlando", "FL"), ("St. Louis", "MO"), ("Madison", "WI"), ("Durham", "NC"), ("Lubbock", "TX"), ("Winston-Salem", "NC"), ("Garland", "TX"), ("Glendale", "AZ"), ("Hialeah", "FL"), ("Scottsdale", "AZ"), ("Irving", "TX"), ("Fremont", "CA"), ("Baton Rouge", "LA"), ("Birmingham", "AL"), ("Rochester", "NY"), ("Des Moines", "IA"), ("Montgomery", "AL"), ("Modesto", "CA"), ("Fayetteville", "NC"), ("Shreveport", "LA"), ("Akron", "OH"), ("Grand Rapids", "MI"), ("Huntington Beach", "CA"), ("Little Rock", "AR")]
FAST_DEMO_CITIES = [("Henderson", "NV"), ("Lincoln", "NE"), ("Boise", "ID"), ("Des Moines", "IA"), ("Madison", "WI"), ("Colorado Springs", "CO"), ("Richmond", "VA"), ("Raleigh", "NC"), ("Durham", "NC"), ("Fort Wayne", "IN"), ("Omaha", "NE"), ("Wichita", "KS"), ("Tulsa", "OK"), ("Spokane", "WA"), ("Tacoma", "WA"), ("Aurora", "CO"), ("Las Vegas", "NV"), ("Nashville", "TN"), ("Columbus", "OH"), ("Charlotte", "NC"), ("Louisville", "KY"), ("Indianapolis", "IN"), ("Memphis", "TN"), ("Detroit", "MI"), ("Milwaukee", "WI"), ("Minneapolis", "MN"), ("Seattle", "WA"), ("Denver", "CO"), ("Portland", "OR"), ("Austin", "TX")]
FAA_CEILING_COLORS = {0: {"line": "rgba(255,  20,  20, 0.95)", "fill": "rgba(255,  20,  20, 0.20)"}, 50: {"line": "rgba(255, 120,   0, 0.95)", "fill": "rgba(255, 120,   0, 0.18)"}, 100: {"line": "rgba(255, 210,   0, 0.95)", "fill": "rgba(255, 210,   0, 0.18)"}, 200: {"line": "rgba(180, 230,   0, 0.95)", "fill": "rgba(180, 230,   0, 0.16)"}, 300: {"line": "rgba( 80, 200,  50, 0.95)", "fill": "rgba( 80, 200,  50, 0.16)"}, 400: {"line": "rgba(  0, 180, 100, 0.95)", "fill": "rgba(  0, 180, 100, 0.15)"}}
FAA_DEFAULT_COLOR = {"line": "rgba(150,150,150,0.8)", "fill": "rgba(150,150,150,0.10)"}
STATION_COLORS = ["#00D2FF", "#39FF14", "#FFD700", "#FF007F", "#FF4500", "#00FFCC", "#FF3333", "#7FFF00", "#00FFFF", "#FF9900"]

# --- THEME VARIABLES ---
bg_main = "#000000"
bg_sidebar = "#111111"
text_main = "#ffffff"
text_muted = "#aaaaaa"
accent_color = "#00D2FF"
card_bg = "#111111"
card_border = "#333333"
card_text = "#eeeeee"
card_title = "#ffffff"
budget_box_bg = "#0a0a0a"
budget_box_border = "#00D2FF"
budget_box_shadow = "rgba(0, 210, 255, 0.15)"
map_style = "carto-darkmatter"
map_boundary_color = "#ffffff"
map_incident_color = "#00D2FF"
legend_bg = "rgba(0, 0, 0, 0.7)"
legend_text = "#ffffff"

HERO_MESSAGES = ["🚔 Building safer communities, one drone at a time…", "🛡️ Loading data because your officers deserve better tools…", "🫡 Honoring the men and women who answer the call every day…", "💙 Officers run toward danger so the rest of us don't have to…", "🚁 Optimizing so your team gets there first — every time…", "🌟 Every second we save is a life better protected…", "🤝 Technology in service of the community's greatest heroes…", "💪 Your officers deserve every advantage we can give them…", "🙏 Dedicated to the families who wait at home while heroes serve…", "🏅 Processing data worthy of those who wear the badge with pride…", "🌃 Mapping the city your officers protect through every shift…", "🔵 Building a network as reliable as the officers who depend on it…", "❤️ Because faster response means more lives saved…", "🌅 Creating tools that let officers come home safely every night…", "🦅 Guardian drones — always watching, always ready to assist…", "🏘️ Modeling coverage for the neighborhoods they protect and serve…", "📡 Connecting technology to the courage already on the streets…", "🧠 Smart systems for smarter, safer law enforcement…", "🌟 Every data point represents a community worth protecting…", "🚨 Fewer false alarms. More real backup. Better outcomes for all…"]
FAA_MESSAGES = ["✈️ Checking FAA airspace — keeping your drones and your pilots safe…", "🛫 Loading LAANC data — because safe skies mean more missions completed…", "🗺️ Mapping controlled airspace — so every flight is a legal, safe one…", "✈️ FAA compliance check in progress — protecting officers on the ground and drones in the air…", "🛡️ Pulling airspace boundaries — safe operations start before takeoff…", "🌐 Verifying flight corridors — your pilots deserve a clear path forward…", "📡 Syncing with FAA LAANC — because your department deserves zero surprises in the sky…", "🛩️ Loading aviation data — the same skies your officers look up to every night…"]
AIRFIELD_MESSAGES = ["🏗️ Locating nearby airfields — coordinating with the aviation community that shares your skies…", "📍 Mapping airports near each station — great neighbors make great operators…", "🛬 Finding local airfields — because your team coordinates with everyone keeping the community safe…", "✈️ Scanning for nearby aviation assets — your drones respect every aircraft they share the sky with…", "🗺️ Identifying airport proximity — so your officers always know what's overhead…", "🤝 Locating nearby airfields — collaboration between aviation and law enforcement saves lives…", "📡 Querying aviation infrastructure — the sky belongs to everyone who protects this community…"]
JURISDICTION_MESSAGES = ["🗺️ Identifying jurisdictions — every boundary represents a community counting on you…", "📐 Loading geographic boundaries — the lines officers cross every shift to keep people safe…", "🏙️ Mapping your jurisdiction — the streets your officers know better than anyone…", "🌆 Matching data to boundaries — every block is someone's home, someone's neighborhood…", "📍 Finding your coverage area — the community that trusts you with their safety…", "🗺️ Resolving jurisdictions — where every call for help deserves an answer…"]
SPATIAL_MESSAGES = ["⚡ Crunching coverage geometry — because your officers deserve precision, not guesswork…", "🧮 Computing spatial matrices — doing the math so your team can focus on what matters…", "📊 Building coverage model — every calculation brings faster response one step closer…", "🔬 Analyzing incident patterns — understanding the city so your officers can better protect it…", "💡 Optimizing station geometry — smart placement means no neighborhood is left behind…", "🧠 Modeling response zones — technology standing behind the officers who stand for us…"]

def get_hero_message(): return random.choice(HERO_MESSAGES)
def get_faa_message(): return random.choice(FAA_MESSAGES)
def get_airfield_message(): return random.choice(AIRFIELD_MESSAGES)
def get_jurisdiction_message(): return random.choice(JURISDICTION_MESSAGES)
def get_spatial_message(): return random.choice(SPATIAL_MESSAGES)

def get_base64_of_bin_file(bin_file):
    try:
        with open(bin_file, 'rb') as f: return base64.b64encode(f.read()).decode()
    except Exception: return None

# ============================================================
# COMMAND CENTER ANALYTICS GENERATOR
# ============================================================

def _detect_datetime_series_for_labels(df):
    """Return a best-effort parsed datetime series from common CAD field patterns."""
    if df is None or len(df) == 0:
        return None
    try:
        if 'date' in df.columns and 'time' in df.columns:
            s = pd.to_datetime(df['date'].astype(str).fillna('') + ' ' + df['time'].astype(str).fillna(''), errors='coerce')
            if s.notna().sum() > 0:
                return s
        if 'date' in df.columns:
            s = pd.to_datetime(df['date'], errors='coerce')
            if s.notna().sum() > 0:
                return s
        candidates = [
            'createdtime_central', 'created time', 'createdtime', 'call datetime', 'calldatetime',
            'timestamp', 'datetime', 'incident datetime', 'received time', 'time received',
            'dispatch datetime', 'event time', 'event datetime'
        ]
        col_map = {str(c).strip().lower(): c for c in df.columns}
        for cand in candidates:
            col = cand if cand in df.columns else col_map.get(cand)
            if col is not None:
                s = pd.to_datetime(df[col], errors='coerce')
                if s.notna().sum() > 0:
                    return s
    except Exception:
        return None
    return None


def generate_command_center_html(df, total_orig_calls, export_mode=False):
    """Generates the full Command Center visual suite with interactive Javascript filtering."""
    if df is None or df.empty:
        return "<div style='color:gray; padding:20px;'>Analytics unavailable. No incident records loaded.</div>"

    import calendar as _cal
    import json

    df_ana = df.copy()

    dt_obj = None
    if 'date' in df_ana.columns:
        _date_str = df_ana['date'].astype(str).fillna('')
        _time_str = df_ana['time'].astype(str).fillna('') if 'time' in df_ana.columns else ''
        if isinstance(_time_str, str):
            _combined = _date_str
        else:
            _combined = _date_str + ' ' + _time_str
        # Try ISO format first (what our parser stores), then fall back
        for _fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d']:
            try:
                _trial = pd.to_datetime(_combined, format=_fmt, errors='coerce')
                if _trial.notna().sum() > len(df_ana) * 0.5:
                    dt_obj = _trial
                    break
            except Exception:
                continue
        if dt_obj is None or dt_obj.dropna().empty:
            dt_obj = pd.to_datetime(_combined, errors='coerce')

    if dt_obj is None or dt_obj.dropna().empty:
        dt_candidates = [
            'createdtime_central', 'created time', 'createdtime', 'call datetime', 'calldatetime',
            'timestamp', 'datetime', 'incident datetime', 'received time', 'time received',
            'dispatch datetime', 'event time', 'event datetime'
        ]
        for cand in dt_candidates:
            if cand in df_ana.columns:
                trial = pd.to_datetime(df_ana[cand], errors='coerce')
                if trial.dropna().shape[0] > 0:
                    dt_obj = trial
                    break

    if dt_obj is None or dt_obj.dropna().empty:
        col_map = {str(c).strip().lower(): c for c in df_ana.columns}
        for cand in [
            'createdtime_central', 'created time', 'createdtime', 'call datetime', 'calldatetime',
            'timestamp', 'datetime', 'incident datetime', 'received time', 'time received',
            'dispatch datetime', 'event time', 'event datetime'
        ]:
            if cand in col_map:
                _col_real = col_map[cand]
                _fmt_try = ['%m/%d/%Y %I:%M %p','%m/%d/%Y %H:%M:%S','%Y-%m-%d %H:%M:%S',
                            '%Y-%m-%dT%H:%M:%S','%m/%d/%Y %H:%M','%Y-%m-%d %H:%M']
                for _fmt in _fmt_try:
                    try:
                        trial = pd.to_datetime(df_ana[_col_real], format=_fmt, errors='coerce')
                        if trial.notna().sum() > 0:
                            dt_obj = trial
                            break
                    except Exception:
                        continue
                if dt_obj is not None and dt_obj.dropna().shape[0] > 0:
                    break
                # Last resort: dateutil inference
                if dt_obj is None or dt_obj.dropna().empty:
                    trial = pd.to_datetime(df_ana[_col_real], errors='coerce')
                    if trial.dropna().shape[0] > 0:
                        dt_obj = trial
                        break

    # Universal scan: try every object column as a datetime source
    if dt_obj is None or dt_obj.dropna().empty:
        for _col in df_ana.select_dtypes(include='object').columns:
            try:
                _samp = df_ana[_col].dropna().head(20)
                _trial = pd.to_datetime(_samp, errors='coerce')
                if _trial.notna().sum() >= 10 and _trial.dt.year.between(2000, 2035).mean() > 0.8:
                    dt_obj = pd.to_datetime(df_ana[_col], errors='coerce')
                    break
            except Exception:
                continue

    if dt_obj is None or dt_obj.dropna().empty:
        return "<div style='color:gray; padding:20px;'>Analytics unavailable. Missing date/time fields.</div>"

    df_ana['dt_obj'] = dt_obj
    df_ana = df_ana.dropna(subset=['dt_obj'])
    if df_ana.empty: return "<div>No valid dates found in data.</div>"

    # 1. Parse dates and build a lightweight records array for JavaScript
    records = []
    for _, r in df_ana.iterrows():
        dt = r['dt_obj']
        p_val = str(r['priority']).upper().strip() if 'priority' in r else 'UNKNOWN'
        if p_val == 'NAN' or not p_val: p_val = 'UNKNOWN'
        records.append({
            'd': dt.strftime('%Y-%m-%d'),
            'h': dt.hour,
            'dow': dt.dayofweek, # Mon=0, Sun=6
            'p': p_val
        })
        
    # 2. Dynamically identify all priority types
    unique_pris = sorted(list(set(r['p'] for r in records)))
    options_html = '<option value="ALL">ALL PRIORITIES</option>'
    for p in unique_pris:
        # Clean up the display name slightly
        display_p = f"PRIORITY {p}" if len(p) <= 2 else p
        options_html += f'<option value="{p}">{display_p}</option>'

    # 3. Build the initial HTML shell (JS will populate the numbers)
    month_keys = sorted(list(set(r['d'][:7] for r in records)))
    cal_html = "<div style='display:grid; grid-template-columns:repeat(auto-fill, minmax(250px, 1fr)); gap:15px; margin-top:20px;'>"
    
    for mk in month_keys[:12]:
        yr, mo = int(mk.split('-')[0]), int(mk.split('-')[1])
        cal_html += f"<div style='background:#0c0c12; border:1px solid #1a1a26; border-radius:6px; padding:12px;'>"
        cal_html += f"<div style='display:flex; justify-content:space-between; align-items:baseline; border-bottom:1px solid #252535; padding-bottom:6px; margin-bottom:8px;'><span style='color:#00D2FF; font-weight:800; font-size:12px; text-transform:uppercase; letter-spacing:1px;'>{_cal.month_name[mo]} {yr}</span><span id='month-total-{mk}' style='color:#7777a0; font-size:10px; font-family:monospace;'>0 calls</span></div>"
        
        cal_html += "<div style='display:grid; grid-template-columns:repeat(7, 1fr); gap:2px; margin-bottom:4px;'>"
        for i, dname in enumerate(['Su','Mo','Tu','We','Th','Fr','Sa']):
            c = ['#FF6B6B','#4ECDC4','#45B7D1','#F0B429','#96CEB4','#DDA0DD','#FF9A8B'][i]
            cal_html += f"<div style='font-size:9px; text-align:center; color:{c}; font-weight:600;'>{dname}</div>"
        cal_html += "</div>"
        
        cal_html += "<div style='display:grid; grid-template-columns:repeat(7, 1fr); gap:2px;'>"
        first_dow_sun = (_cal.weekday(yr, mo, 1) + 1) % 7
        last_day = _cal.monthrange(yr, mo)[1]
        
        for _ in range(first_dow_sun): cal_html += "<div></div>"
            
        for d in range(1, last_day + 1):
            dk = f"{yr}-{mo:02d}-{d:02d}"
            dow_idx = (_cal.weekday(yr, mo, d) + 1) % 7 
            cal_html += f"<div class='day-cell' data-date='{dk}' data-mkey='{mk}' data-month='{_cal.month_name[mo]}' data-d='{d}' data-y='{yr}' data-dow='{dow_idx}' style='aspect-ratio:1; border-radius:2px; display:flex; flex-direction:column; align-items:center; justify-content:center; position:relative; font-family:monospace; cursor:default; border:1px solid transparent; transition:transform 0.1s;' onmouseover='showTooltip(this, event)' onmouseout='hideTooltip()'></div>"
            
        cal_html += "</div></div>"
    cal_html += "</div>"

    controls_html = f"""
    <div style="display:flex; gap:20px; align-items:center; background:#0c0c12; border:1px solid #1a1a26; padding:12px 18px; border-radius:6px; margin-bottom:20px;">
        <div style="display:flex; align-items:center; gap:10px;">
            <span style="font-size:10px; color:#7777a0; font-weight:bold; letter-spacing:1px; text-transform:uppercase;">Priority Filter:</span>
            <select id="pri-select" onchange="currentPri=this.value; updateDashboard();" style="background:#1a1a26; color:#00D2FF; border:1px solid #252535; padding:6px 12px; border-radius:4px; font-weight:bold; cursor:pointer;">
                {options_html}
            </select>
        </div>
        <div style="display:flex; align-items:center; gap:10px;">
            <span style="font-size:10px; color:#7777a0; font-weight:bold; letter-spacing:1px; text-transform:uppercase;">Shift Length:</span>
            <select id="shift-select" onchange="currentShift=parseInt(this.value); updateDashboard();" style="background:#1a1a26; color:#00D2FF; border:1px solid #252535; padding:6px 12px; border-radius:4px; font-weight:bold; cursor:pointer;">
                <option value="8" selected>8 HOURS</option>
                <option value="10">10 HOURS</option>
                <option value="12">12 HOURS</option>
            </select>
        </div>
    </div>
    """

    full_html = f"""
    <div style="background:#000; color:#e8e8f2; font-family: 'Barlow', sans-serif; padding:15px; border-radius:8px;">
        <style>
            .day-cell:hover {{ transform: scale(1.15); z-index: 10; box-shadow: 0 4px 12px rgba(0,0,0,0.5); }}
            .day-peak {{ border-color: #cc0000 !important; font-weight: 700; }}
            #dfr-tooltip {{ position: fixed; z-index: 9999; background: #09090f; border: 1px solid #252535; border-radius: 6px; padding: 12px 16px; font-family: monospace; font-size: 11px; color: #e8e8f2; pointer-events: none; box-shadow: 0 6px 24px rgba(0,0,0,0.8); display: none; min-width: 220px; }}
        </style>
        
        <div id="dfr-tooltip"></div>
        <div style="color:#00D2FF; font-weight:900; letter-spacing:3px; font-size:14px; text-transform:uppercase; margin-bottom:20px; border-bottom:1px solid #1a1a26; padding-bottom:10px;">Data Ingestion Analytics</div>
        
        {controls_html}
        
        <div style="display:grid; grid-template-columns: 1fr 1fr; gap:15px; margin-bottom:20px;">
            <div style="background:#0c0c12; border-left:4px solid #00D2FF; padding:15px; border-radius:4px; border-top:1px solid #1a1a26; border-right:1px solid #1a1a26; border-bottom:1px solid #1a1a26;">
                <div style="color:#00D2FF; font-size:26px; font-weight:900; font-family:monospace;" id="kpi-total-val">0</div>
                <div style="color:#7777a0; font-size:10px; text-transform:uppercase; letter-spacing:1px; margin-top:4px;">Displayed Incidents</div>
            </div>
            <div style="background:#0c0c12; border-left:4px solid #F0B429; padding:15px; border-radius:4px; border-top:1px solid #1a1a26; border-right:1px solid #1a1a26; border-bottom:1px solid #1a1a26;">
                <div style="color:#F0B429; font-size:26px; font-weight:900; font-family:monospace;" id="kpi-peak-val">0:00</div>
                <div style="color:#7777a0; font-size:10px; text-transform:uppercase; letter-spacing:1px; margin-top:4px;">Peak Activity Hour</div>
            </div>
        </div>
        
        <div style="display:grid; grid-template-columns: 3fr 2fr; gap:15px; margin-bottom:25px;">
            <div style="background:#06060a; border:1px solid #1a1a26; border-radius:6px; padding:15px;">
                <div style="margin-bottom:12px; font-size:10px; color:#7777a0; text-transform:uppercase; letter-spacing:1px; font-weight:bold;">Optimized DFR Shift Windows</div>
                <div id="shift-container"></div>
            </div>
            <div style="background:#06060a; border:1px solid #1a1a26; border-radius:6px; padding:15px; display:flex; flex-direction:column;">
                <div style="margin-bottom:12px; font-size:10px; color:#7777a0; text-transform:uppercase; letter-spacing:1px; font-weight:bold;">Call Volume by Day of Week</div>
                <div style="display:flex; justify-content:space-between; align-items:flex-end; flex-grow:1; padding:10px 5px 0;" id="dow-container"></div>
            </div>
        </div>
        
        <div style="display:flex; align-items:center; justify-content:space-between; margin-bottom:5px; padding-top:15px; border-top:1px solid #1a1a26;">
            <div style="font-size:14px; font-weight:800; color:#fff; letter-spacing:1px; text-transform:uppercase;">DFR Deployment Calendar</div>
            <div style="display:flex; gap:12px; font-family:monospace; font-size:9px; color:#8d93b8; flex-wrap:wrap;">
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#59B7FF; border:1px solid #1E5D91; border-radius:2px; box-shadow:0 0 6px rgba(89,183,255,0.18);"></div>VERY LOW</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#45E28A; border:1px solid #1D7D49; border-radius:2px; box-shadow:0 0 6px rgba(69,226,138,0.18);"></div>LOW</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#FFD84D; border:1px solid #8A6F00; border-radius:2px; box-shadow:0 0 6px rgba(255,216,77,0.18);"></div>MEDIUM</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#FF9F43; border:1px solid #9A4F00; border-radius:2px; box-shadow:0 0 6px rgba(255,159,67,0.18);"></div>HIGH</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#FF5B6E; border:1px solid #A51F2D; border-radius:2px; box-shadow:0 0 8px rgba(255,91,110,0.24);"></div>PEAK</div>
            </div>
        </div>
        
        {cal_html}
        
        <script>
            const rawData = {json.dumps(records)};
            const totalOrigCalls = {int(total_orig_calls) if total_orig_calls else len(records)};
            let currentShift = 8;
            let currentPri = 'ALL';
            window.dateHourly = {{}};
            const dowNames = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'];
            const dowColors = ['#4ECDC4','#45B7D1','#F0B429','#96CEB4','#DDA0DD','#FF9A8B','#FF6B6B'];
            const stripeColors = ['#FF6B6B','#4ECDC4','#45B7D1','#F0B429','#96CEB4','#DDA0DD','#FF9A8B'];

            function updateDashboard() {{
                let filtered = rawData;
                if(currentPri !== 'ALL') {{
                    filtered = rawData.filter(d => String(d.p) === currentPri);
                }}

                let total = filtered.length;
                if(currentPri === 'ALL') {{ total = totalOrigCalls; }}
                document.getElementById('kpi-total-val').innerHTML = total.toLocaleString();

                let hourly = new Array(24).fill(0);
                let daily = {{}};
                let dowCounts = new Array(7).fill(0);
                window.dateHourly = {{}};
                let mTotal = {{}};
                let mMax = {{}};

                filtered.forEach(r => {{
                    hourly[r.h]++;
                    daily[r.d] = (daily[r.d] || 0) + 1;
                    dowCounts[r.dow]++;
                    
                    if(!window.dateHourly[r.d]) window.dateHourly[r.d] = new Array(24).fill(0);
                    window.dateHourly[r.d][r.h]++;
                    
                    let m = r.d.substring(0,7);
                    mTotal[m] = (mTotal[m] || 0) + 1;
                    if(!mMax[m] || daily[r.d] > mMax[m]) mMax[m] = daily[r.d];
                }});

                // Peak Hour
                let peakHr = hourly.indexOf(Math.max(...hourly));
                if(total === 0) peakHr = 0;
                document.getElementById('kpi-peak-val').innerText = peakHr + ':00';

                // Shifts
                let shiftHtml = "";
                [8, 10, 12].forEach(win => {{
                    let bestV = 0, bestS = 0;
                    for(let s=0; s<24; s++) {{
                        let v = 0;
                        for(let h=0; h<win; h++) v += hourly[(s+h)%24];
                        if(v > bestV) {{ bestV = v; bestS = s; }}
                    }}
                    let pct = total > 0 ? (bestV/total)*100 : 0;
                    let isActive = (win === currentShift);
                    let bgColor = isActive ? "rgba(0,210,255,0.08)" : "#0c0c12";
                    let brColor = isActive ? "#00D2FF" : "#252535";
                    let badge = isActive ? "<div style='font-size:8px; color:#00D2FF; margin-left:10px; border:1px solid #00D2FF; padding:1px 4px; border-radius:2px;'>SELECTED</div>" : "";
                    let eHr = (bestS + win) % 24;
                    let pS = String(bestS).padStart(2,'0');
                    let pE = String(eHr).padStart(2,'0');
                    
                    shiftHtml += `<div style="display:flex; align-items:center; background:${{bgColor}}; border:1px solid ${{brColor}}; padding:8px; margin-bottom:5px; border-radius:4px; transition:all 0.2s;">
                        <div style="width:50px; font-weight:800; color:#fff; font-size:13px;">${{win}}hr</div>
                        <div style="width:110px; font-family:monospace; color:#00D2FF; font-size:12px;">${{pS}}:00 - ${{pE}}:00</div>
                        <div style="flex-grow:1; background:#1a1a26; height:8px; border-radius:4px; margin:0 15px; position:relative;">
                            <div style="position:absolute; left:${{(bestS/24)*100}}%; width:${{(win/24)*100}}%; background:#00D2FF; height:100%; border-radius:4px; opacity:0.6;"></div>
                        </div>
                        <div style="width:50px; text-align:right; font-family:monospace; color:#00D2FF; font-size:13px;">${{pct.toFixed(1)}}%</div>
                        ${{badge}}
                    </div>`;
                }});
                document.getElementById('shift-container').innerHTML = shiftHtml;

                // DOW
                let maxDow = Math.max(...dowCounts, 1);
                let dowHtml = "";
                for(let i=0; i<7; i++) {{
                    let hPct = (dowCounts[i]/maxDow)*100;
                    dowHtml += `<div style="flex:1; display:flex; flex-direction:column; align-items:center;">
                        <div style="background:#1a1a26; width:22px; height:80px; position:relative; border-radius:2px;">
                            <div style="position:absolute; bottom:0; width:100%; height:${{hPct}}%; background:${{dowColors[i]}}; border-radius:2px; transition:height 0.3s;"></div>
                        </div>
                        <span style="font-size:10px; color:#7777a0; margin-top:6px; font-family:monospace;">${{dowNames[i]}}</span>
                    </div>`;
                }}
                document.getElementById('dow-container').innerHTML = dowHtml;

                // Month Headers
                document.querySelectorAll('[id^="month-total-"]').forEach(el => {{
                    let m = el.id.replace('month-total-','');
                    let cnt = mTotal[m] || 0;
                    el.innerText = cnt.toLocaleString() + ' calls';
                }});

                // Calendar Cells
                document.querySelectorAll('.day-cell').forEach(cell => {{
                    if(!cell.hasAttribute('data-date')) return;
                    let d = cell.getAttribute('data-date');
                    let mkey = cell.getAttribute('data-mkey');
                    let cnt = daily[d] || 0;
                    let max = mMax[mkey] || 1;
                    let ratio = max > 0 ? cnt / max : 0;
                    
                    let bg, fc, cls;
                    if (cnt === 0) {{ bg='#08080f'; fc='#333'; cls='day-zero'; }}
                    else if (ratio >= 0.85) {{ bg='#3d0a0a'; fc='#ff4444'; cls='day-peak'; }}
                    else if (ratio >= 0.55) {{ bg='#3d1a00'; fc='#ff8c00'; cls='day-high'; }}
                    else if (ratio >= 0.25) {{ bg='#2d2d00'; fc='#d4c000'; cls='day-med'; }}
                    else {{ bg='#0d3320'; fc='#2ecc71'; cls='day-low'; }}

                    cell.className = 'day-cell ' + cls;
                    cell.style.background = bg;
                    cell.style.color = fc;
                    cell.setAttribute('data-count', cnt);
                    cell.setAttribute('data-ratio', ratio);
                    
                    let domD = cell.getAttribute('data-d');
                    let html = `<span style='font-size:11px; z-index:1; font-weight:bold;'>${{domD}}</span>`;
                    if(cnt > 0) {{
                        html += `<span style='font-size:8px; opacity:0.7; margin-top:1px;'>${{cnt}}</span>`;
                        let dowIdx = parseInt(cell.getAttribute('data-dow'));
                        html += `<div style='position:absolute; bottom:0; left:0; right:0; height:2px; background:${{stripeColors[dowIdx]}}; opacity:0.7; border-radius:0 0 2px 2px;'></div>`;
                    }}
                    cell.innerHTML = html;
                }});
            }}

            function showTooltip(el, ev) {{
                const cnt = parseInt(el.getAttribute('data-count'));
                if (cnt === 0) return;
                
                const dk = el.getAttribute('data-date');
                const ratio = parseFloat(el.getAttribute('data-ratio'));
                const mName = el.getAttribute('data-month');
                const d = el.getAttribute('data-d');
                const y = el.getAttribute('data-y');
                const dow = parseInt(el.getAttribute('data-dow'));
                
                let loadText = '';
                if (ratio >= 0.85) loadText = '<span style="color:#FF5B6E">■ PEAK</span> — Full crew';
                else if (ratio >= 0.65) loadText = '<span style="color:#FF9F43">■ HIGH</span> — Priority deploy';
                else if (ratio >= 0.45) loadText = '<span style="color:#FFD84D">■ MEDIUM</span> — Standard ops';
                else if (ratio >= 0.25) loadText = '<span style="color:#45E28A">■ LOW</span> — Light staffing';
                else loadText = '<span style="color:#59B7FF">■ VERY LOW</span> — Opportunistic coverage';
                
                const hrArr = window.dateHourly[dk] || Array(24).fill(0);
                let bestV = 0, bestS = 0;
                for (let s=0; s<24; s++) {{
                    let v = 0;
                    for (let h=0; h<currentShift; h++) v += hrArr[(s+h)%24];
                    if (v > bestV) {{ bestV = v; bestS = s; }}
                }}
                const dayPct = cnt > 0 ? Math.round((bestV / cnt) * 100) : 0;
                const eHr = (bestS + currentShift) % 24;
                const fmt = (h) => (h%12 || 12) + (h<12 ? 'AM' : 'PM');
                
                const tt = document.getElementById('dfr-tooltip');
                tt.innerHTML = `
                    <div style="color:#00D2FF; margin-bottom:6px; font-size:12px; font-weight:bold; border-bottom:1px solid #252535; padding-bottom:4px;">${{mName}} ${{d}}, ${{y}} · ${{dowNames[dow]}}</div>
                    <div style="margin-bottom:8px; font-size:13px;">Calls: <span style="color:#fff; font-weight:bold;">${{cnt}}</span>  ·  ${{loadText}}</div>
                    <div style="background:#1a1a26; padding:8px; border-radius:4px;">
                        <div style="color:#7777a0; font-size:9px; letter-spacing:1px; text-transform:uppercase; margin-bottom:4px;">Best ${{currentShift}}hr Shift</div>
                        <div style="color:#00D2FF; font-size:14px; font-weight:bold; margin-bottom:2px;">${{fmt(bestS)}} – ${{fmt(eHr)}}</div>
                        <div style="color:#aaa; font-size:10px;">Covers <span style="color:#fff;">${{dayPct}}%</span> of daily volume</div>
                    </div>
                `;
                
                tt.style.display = 'block';
                
                let left = ev.clientX + 15;
                let top = ev.clientY - 20;
                if (left + 220 > window.innerWidth) left = ev.clientX - 235;
                if (top + 100 > window.innerHeight) top = ev.clientY - 110;
                
                tt.style.left = left + 'px';
                tt.style.top = top + 'px';
            }}
            
            function hideTooltip() {{
                document.getElementById('dfr-tooltip').style.display = 'none';
            }}

            // Run once to populate the dashboard on load
            updateDashboard();
        </script>
    </div>
    """
    return full_html

# ============================================================
# AGGRESSIVE DATA PARSER
# ============================================================
def aggressive_parse_calls(uploaded_files):
    all_calls_list = []
    CV = {
        'date': ['received date','incident date','call date','call creation date','calldatetime','call datetime','calltime','timestamp','date','datetime','dispatch date','time received','incdate','date_rept','date_occu','createdtime','created_time','receivedtime','received_time','eventtime','event_time','incidenttime','incident_time','reportedtime','reported_time','entrytime','entry_time','time_central','time_stamp','created'],
        'time': ['call creation time','call time','dispatch time','received time','time', 'hour', 'hour_rept','hour_occu'],
        'priority': ['call priority', 'priority level', 'priority', 'pri', 'urgency'],
        'lat': ['latitude','lat','y coord','ycoord','ycoor','addressy','geoy','y_coord','map_y',
                'point_y','gps_lat','gps_latitude','ylat','coord_y','northing','y_wgs','lat_wgs',
                'incident_lat','inc_lat','event_lat','y_coordinate','address_y','ylocation'],
        'lon': ['longitude','lon','long','x coord','xcoord','xcoor','addressx','geox','x_coord',
                'map_x','point_x','gps_lon','gps_long','gps_longitude','xlon','coord_x','easting',
                'x_wgs','lon_wgs','incident_lon','inc_lon','event_lon','x_coordinate','address_x','xlocation']
    }

    def _infer_city_from_location_text(raw_df):
        text_cols = [c for c in raw_df.columns if c in ['location', 'address', 'incident_location', 'addr', 'street']]
        if not text_cols:
            return None

        s = raw_df[text_cols[0]].dropna().astype(str).str.upper().str.strip()
        if s.empty:
            return None

        s = s.str.replace(r':.*$', '', regex=True)
        s = s.str.replace(r'CNTY', 'COUNTY', regex=True)
        s = s.str.replace(r'[^A-Z0-9 /-]', ' ', regex=True)
        s = s.str.replace(r'\s+', ' ', regex=True).str.strip()

        candidates = []
        for val in s:
            padded = f' {val} '
            if ' MOBILE ' in padded:
                candidates.append('Mobile')
                continue

            m = re.search(r'([A-Z]{3,}(?:\s+[A-Z]{3,}){0,2})$', val)
            if m:
                city = m.group(1).title()
                if city not in {'County', 'City'}:
                    candidates.append(city)

        if not candidates:
            return None

        vc = pd.Series(candidates).value_counts()
        return vc.index[0] if not vc.empty else None

    def _infer_state_from_text(raw_df, inferred_city=None):
        for col in ['state', 'state_name']:
            if col in raw_df.columns:
                top = raw_df[col].dropna().astype(str).str.strip().value_counts()
                if not top.empty:
                    state_val = top.index[0]
                    state_up = str(state_val).upper()
                    if state_up in STATE_FIPS:
                        return state_up
                    state_title = str(state_val).title()
                    if state_title in US_STATES_ABBR:
                        return US_STATES_ABBR[state_title]

        if inferred_city == 'Mobile':
            return 'AL'
        return None


    def _choose_priority_column(raw_df):
        exact_names = ['priority', 'call priority', 'priority level', 'pri']
        exact = [c for c in raw_df.columns if c.strip().lower() in exact_names]
        if exact:
            exact.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return exact[0]

        loose_names = ['priority', 'call priority', 'priority level', 'pri', 'urgency']
        loose = [c for c in raw_df.columns if any(k in c for k in loose_names)]
        if loose:
            loose.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return loose[0]
        return None

    def parse_priority(raw):
        s = str(raw).strip().upper()
        if not s or s == 'NAN': return None
        # Smart inference for PD offenses if priority column is missing
        if any(w in s for w in ['ROBBERY','BURGLARY','ASSAULT','SHOOTING','STABBING','CRITICAL','EMERG']): return 1
        if any(w in s for w in ['ACCIDENT','DISTURBANCE','THEFT','MED','ALARM']): return 2
        if any(w in s for w in ['NON REPORTABLE','FOUND PROPERTY','INFO','ROUTINE','MISC']): return 4
        
        m = re.search(r'^(\d+)', s)
        if m: return int(m.group(1))
        return 3

    for cfile in uploaded_files:
        try:
            fname = cfile.name.lower()
            excel_exts = ('.xlsx', '.xls', '.xlsb', '.xlsm')

            if fname.endswith(excel_exts):
                # ── Excel path ────────────────────────────────────────────────
                raw_bytes = cfile.getvalue()
                engine = 'openpyxl'
                if fname.endswith('.xls'):
                    engine = 'xlrd'
                elif fname.endswith('.xlsb'):
                    engine = 'pyxlsb'

                def _sheet_score(ws):
                    score = 0
                    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
                    if not rows:
                        return -1
                    header = rows[0] or []
                    header_norm = [str(h).strip().lower() for h in header if h is not None]
                    if not header_norm:
                        return -1
                    hints = ['latitude', 'longitude', 'lat', 'lon', 'priority', 'location', 'date', 'time']
                    score += sum(10 for h in header_norm if any(k == h or k in h for k in hints))
                    score += sum(1 for h in header_norm if h and not re.match(r'^column\d+$', h))
                    if len(rows) > 1 and rows[1] and any(v is not None and str(v).strip() != '' for v in rows[1]):
                        score += 25
                    # Penalize external-data placeholder sheets
                    if len(header_norm) == 1 and header_norm[0].startswith('externaldata_'):
                        score -= 100
                    return score

                try:
                    import openpyxl as _oxl
                    _wb = _oxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                    _sheet_name = max(_wb.sheetnames, key=lambda sn: _sheet_score(_wb[sn]))
                    _ws = _wb[_sheet_name]
                    _row_iter = _ws.iter_rows(values_only=True)
                    _headers_raw = next(_row_iter)
                    if _headers_raw is None:
                        raise ValueError("Selected Excel sheet has no header row.")
                    _real_idx = [
                        i for i, h in enumerate(_headers_raw)
                        if h is not None and not (str(h).startswith('Column') and str(h)[6:].isdigit())
                    ]
                    if not _real_idx:
                        _real_idx = [i for i, h in enumerate(_headers_raw) if h is not None]
                    _real_headers = [str(_headers_raw[i]).lower().strip() for i in _real_idx]
                    _rows_data = []
                    for _row in _row_iter:
                        if _row is None:
                            continue
                        _trimmed = [_row[i] if i < len(_row) else None for i in _real_idx]
                        if any(v is not None and str(v).strip() != '' for v in _trimmed):
                            _rows_data.append(_trimmed)
                    _wb.close()
                    raw_df = pd.DataFrame(_rows_data, columns=_real_headers)
                    raw_df = raw_df.dropna(how='all')
                    raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
                except Exception as _xe:
                    raw_df = None
                    # Try all sheets with pandas and pick the one that looks most like CAD data
                    try:
                        _all = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, sheet_name=None)
                        best_score = -10**9
                        best_df = None
                        for _sn, _df in _all.items():
                            _df.columns = [str(c).lower().strip() for c in _df.columns]
                            _score = 0
                            for _c in _df.columns:
                                if _c in ('latitude', 'longitude', 'priority', 'location'):
                                    _score += 20
                                elif any(k in _c for k in ['lat', 'lon', 'priority', 'location', 'date', 'time']):
                                    _score += 5
                            _score += min(len(_df), 100)
                            if len(_df.columns) == 1 and str(_df.columns[0]).startswith('externaldata_'):
                                _score -= 100
                            if _score > best_score:
                                best_score = _score
                                best_df = _df
                        if best_df is not None:
                            raw_df = best_df
                    except Exception:
                        pass
                    if raw_df is None:
                        raw_df = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, dtype=str)
                        raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
            else:
                # ── CSV / TXT path ────────────────────────────────────────────
                content = cfile.getvalue().decode('utf-8', errors='ignore')
                first_line = content.split('\n')[0]
                delim = ',' if first_line.count(',') > first_line.count('\t') else '\t'
                raw_df = pd.read_csv(io.StringIO(content), sep=delim, dtype=str)
                raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
            
            res = pd.DataFrame()
            exact_coord_names = {
                'lat': ['latitude', 'lat', 'gps_lat', 'gps_latitude'],
                'lon': ['longitude', 'lon', 'long', 'gps_lon', 'gps_longitude']
            }
            for field in ['lat', 'lon']:
                found_exact = [c for c in raw_df.columns if c.strip().lower() in exact_coord_names[field]]
                found_loose = [c for c in raw_df.columns if any(s in c for s in CV[field])]
                found = found_exact or found_loose
                if found:
                    res[field] = pd.to_numeric(raw_df[found[0]], errors='coerce')

            # ── Fallback: no column name matched — scan numeric columns by value range ──
            # Lat: -90 to 90, Lon: -180 to 180. Pick best candidate for each.
            if 'lat' not in res.columns or 'lon' not in res.columns:
                numeric_cols = []
                for c in raw_df.columns:
                    series = pd.to_numeric(raw_df[c], errors='coerce').dropna()
                    if len(series) > 10:
                        numeric_cols.append((c, series))

                lat_candidates, lon_candidates = [], []
                for c, series in numeric_cols:
                    mn, mx = series.min(), series.max()
                    # Reject if already assigned
                    if c in (res.get('_lat_col',''), res.get('_lon_col','')):
                        continue
                    # Large integer coords (State Plane) — treat as potential coord pair
                    if mx > 1000:
                        lat_candidates.append((c, series))
                        lon_candidates.append((c, series))
                        continue
                    if -90 <= mn and mx <= 90 and mn < -1:
                        lat_candidates.append((c, series))
                    if -180 <= mn and mx <= 180 and (mn < -90 or mx > 90):
                        lon_candidates.append((c, series))

                # Prefer candidate whose name hints at lat/lon
                def _score(name, hints):
                    return sum(1 for h in hints if h in name)

                if 'lat' not in res.columns and lat_candidates:
                    lat_candidates.sort(key=lambda x: -_score(x[0], ['lat','y','north']))
                    best_lat_col = lat_candidates[0][0]
                    res['lat'] = pd.to_numeric(raw_df[best_lat_col], errors='coerce')

                if 'lon' not in res.columns and lon_candidates:
                    # Don't reuse the lat column
                    used = res.get('lat', pd.Series()).name if 'lat' in res.columns else None
                    lon_candidates = [(c, s) for c, s in lon_candidates if c != used]
                    if lon_candidates:
                        lon_candidates.sort(key=lambda x: -_score(x[0], ['lon','long','x','east']))
                        best_lon_col = lon_candidates[0][0]
                        res['lon'] = pd.to_numeric(raw_df[best_lon_col], errors='coerce')
            
            _p_col = _choose_priority_column(raw_df)
            p_found = [_p_col] if _p_col else []
            if _p_col:
                parsed_priority = raw_df[_p_col].apply(parse_priority)
                parsed_priority = pd.to_numeric(parsed_priority, errors='coerce')
                parsed_priority = parsed_priority.where(parsed_priority.isin([1, 2, 3, 4, 5, 6, 7, 8, 9]))
                if parsed_priority.dropna().empty:
                    res['priority'] = 3
                else:
                    res['priority'] = parsed_priority.fillna(3).astype(int)
            else:
                # No trustworthy priority field — keep the app usable with a neutral default
                res['priority'] = 3
            
            # ── Event type description — carried through for CAD analytics charts ──
            _desc_hints = ['desc','type','nature','offense','calltype','call_type','event_type',
                           'eventtype','calldesc','incident_type','agencyeventtype']
            _desc_found = [c for c in raw_df.columns
                           if any(h in c for h in _desc_hints)
                           and c not in (p_found[:1] if p_found else [])]
            if _desc_found:
                # Pick the column with the most unique text values (most descriptive)
                _best_desc = max(_desc_found, key=lambda c: raw_df[c].dropna().nunique())
                if raw_df[_best_desc].dropna().nunique() > 2:
                    res['call_type_desc'] = raw_df[_best_desc].astype(str).str.strip()

            d_found = [c for c in raw_df.columns if any(s in c for s in CV['date'])]
            t_found = [c for c in raw_df.columns if any(s in c for s in CV['time'])]

            # Fallback: if no date column found by name hint, scan all string columns
            # for any that successfully parse as datetime (catches columns like
            # 'createdtime_central', 'call_ts', 'event_dttm', etc.)
            if not d_found:
                for _col in raw_df.columns:
                    if _col in (t_found or []):
                        continue
                    try:
                        _test = pd.to_datetime(raw_df[_col].dropna().head(50), errors='coerce')
                        _valid = _test.dropna()
                        if len(_valid) >= 10 and _valid.dt.year.between(2000, 2035).mean() > 0.8:
                            d_found = [_col]
                            break
                    except Exception:
                        continue

            if d_found:
                # Build the raw string series to parse — combine date+time cols if separate
                if t_found and d_found[0] != t_found[0]:
                    _raw_dt_str = raw_df[d_found[0]].fillna('') + ' ' + raw_df[t_found[0]].fillna('')
                else:
                    _raw_dt_str = raw_df[d_found[0]]

                # Try explicit common formats first (orders of magnitude faster than
                # dateutil fallback on large files, and avoids NaT on ghost rows).
                # Format detection: sample the first non-null value.
                _sample_vals = _raw_dt_str.dropna().str.strip()
                _sample_vals = _sample_vals[_sample_vals != ''].head(5)
                _fmt_candidates = [
                    '%m/%d/%Y %I:%M %p',   # 2/14/2025 6:03 PM  (Mobile AL)
                    '%m/%d/%Y %H:%M:%S',   # 2/14/2025 18:03:00
                    '%m/%d/%Y %H:%M',      # 2/14/2025 18:03
                    '%Y-%m-%d %H:%M:%S',   # 2025-02-14 18:03:00
                    '%Y-%m-%dT%H:%M:%S',   # ISO 8601
                    '%Y-%m-%d %H:%M',      # 2025-02-14 18:03
                    '%Y/%m/%d %H:%M:%S',
                    '%d/%m/%Y %H:%M:%S',
                    '%m-%d-%Y %H:%M:%S',
                ]
                dt_series = None
                if not _sample_vals.empty:
                    for _fmt in _fmt_candidates:
                        try:
                            _trial = pd.to_datetime(_sample_vals.iloc[0], format=_fmt, errors='raise')
                            # Format matched — apply to full series
                            dt_series = pd.to_datetime(_raw_dt_str, format=_fmt, errors='coerce')
                            break
                        except Exception:
                            continue
                if dt_series is None:
                    # Final fallback: let pandas infer (slow but handles edge cases)
                    dt_series = pd.to_datetime(_raw_dt_str, errors='coerce')

                res['date'] = dt_series.dt.strftime('%Y-%m-%d')
                res['time'] = dt_series.dt.strftime('%H:%M:%S')

            # --- COORDINATE CONVERSION (STATE PLANE / LARGE-INTEGER DETECTOR) ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns:
                res = res[(res['lat'] != 0) & (res['lon'] != 0)].dropna(subset=['lat', 'lon'])
                if not res.empty:
                    max_val = max(res['lat'].abs().max(), res['lon'].abs().max())
                    if max_val > 1000:
                        converted = False
                        # Strategy 1: Try common State Plane CRS at /100 and /1 scales
                        candidate_crs = [
                            "EPSG:2278",  # TX South Central (ftUS)
                            "EPSG:2277",  # TX Central (ftUS)
                            "EPSG:2276",  # TX North Central (ftUS)
                            "EPSG:2279",  # TX South (ftUS)
                            "EPSG:32140", # TX South Central (m)
                        ]
                        for scale in [100.0, 1.0]:
                            for crs in candidate_crs:
                                try:
                                    transformer = pyproj.Transformer.from_crs(crs, "EPSG:4326", always_xy=True)
                                    test_lons, test_lats = transformer.transform(
                                        res['lon'].values[:20] / scale,
                                        res['lat'].values[:20] / scale
                                    )
                                    if (24 < float(test_lats.mean()) < 50 and
                                            -130 < float(test_lons.mean()) < -60 and
                                            float(test_lats.std()) < 5 and
                                            float(test_lons.std()) < 5):
                                        lons, lats = transformer.transform(
                                            res['lon'].values / scale, res['lat'].values / scale
                                        )
                                        res['lon'], res['lat'] = lons, lats
                                        converted = True
                                        break
                                except Exception:
                                    continue
                            if converted:
                                break

                        # Strategy 2: If CRS conversion failed, anchor to city column geocode
                        if not converted:
                            try:
                                city_name = None
                                for col in ['city', 'city_name', 'municipality', 'jurisdiction']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            city_name = top.index[0]
                                            break
                                state_name = None
                                for col in ['state', 'state_name']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            state_name = top.index[0]
                                            break
                                if city_name:
                                    query_str = f"{city_name}, {state_name}" if state_name else city_name
                                    geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(query_str)}&limit=1"
                                    req = urllib.request.Request(geo_url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"})
                                    with urllib.request.urlopen(req, timeout=10) as resp:
                                        geo_data = json.loads(resp.read().decode("utf-8"))
                                    if geo_data:
                                        anchor_lat = float(geo_data[0]["lat"])
                                        anchor_lon = float(geo_data[0]["lon"])
                                        raw_cx = res["lon"].median()
                                        raw_cy = res["lat"].median()
                                        city_radius_deg = 0.35
                                        raw_spread = max(res["lon"].std(), res["lat"].std(), 1)
                                        deg_per_unit = city_radius_deg / raw_spread
                                        res["lon"] = anchor_lon + (res["lon"] - raw_cx) * deg_per_unit
                                        res["lat"] = anchor_lat + (res["lat"] - raw_cy) * deg_per_unit
                                        converted = True
                            except Exception:
                                pass

                        if converted:
                            res = res[
                                (res["lat"] > 18) & (res["lat"] < 72) &
                                (res["lon"] > -170) & (res["lon"] < -60)
                            ]

            # City/state detection: store top values on rows for location detection
            top_city_name = None
            for col in ["city", "city_name", "municipality", "jurisdiction"]:
                if col in raw_df.columns:
                    top_city = raw_df[col].dropna().astype(str).str.strip().value_counts()
                    if not top_city.empty:
                        top_city_name = str(top_city.index[0]).title()
                        res["_csv_city"] = top_city_name
                        break

            if "_csv_city" not in res.columns:
                inferred_city = _infer_city_from_location_text(raw_df)
                if inferred_city:
                    top_city_name = inferred_city
                    res["_csv_city"] = inferred_city

            inferred_state = _infer_state_from_text(raw_df, top_city_name)
            if inferred_state:
                res["_csv_state"] = inferred_state

            all_calls_list.append(res)
        except: continue
        
    if not all_calls_list: return pd.DataFrame()
    # Only keep frames that actually have lat/lon columns — Excel sheets
    # without coordinate data should not crash the concat
    valid = [df for df in all_calls_list if 'lat' in df.columns and 'lon' in df.columns]
    if not valid: return pd.DataFrame()
    combined = pd.concat(valid, ignore_index=True)
    # Safe dropna — columns guaranteed to exist now
    combined = combined.dropna(subset=['lat', 'lon'])
    # IMPORTANT: keep the full parsed CAD dataset here.
    #
    # The optimizer is sampled later (after upload) for performance, but the
    # parsed dataframe itself must preserve every incident so:
    #   1) Total Incidents shows the true uploaded count
    #   2) the stations map can render a much denser full-history call cloud
    #   3) export/reporting math stays tied to the source file, not a k-means
    #      surrogate created during parsing
    return combined

def _build_cad_charts_html(df_calls):
    """Generate a self-contained HTML block with Chart.js charts for the PDF/HTML export.
    Returns an empty string if no real CAD data is available."""
    if df_calls is None or df_calls.empty:
        return ""
    try:
        # Priority counts
        pri_labels, pri_vals = [], []
        if 'priority' in df_calls.columns:
            pc = df_calls['priority'].dropna().astype(str).value_counts().sort_index()
            pri_labels = [f"Priority {p}" for p in pc.index]
            pri_vals   = pc.values.tolist()

        # Top event types
        type_labels, type_vals = [], []
        for _c in ['call_type_desc','agencyeventtypecodedesc','eventdesc','calldesc','description','nature','event_desc']:
            if _c in df_calls.columns and df_calls[_c].dropna().nunique() > 2:
                tc = df_calls[_c].dropna().str.strip().value_counts().head(10)
                type_labels = tc.index.tolist()
                type_vals   = tc.values.tolist()
                break

        # Concentration curve
        conc_x, conc_y, n_cells, pct10 = [], [], 0, 0
        try:
            LAT_MI, LON_MI, BIN = 69.0, 55.0, 0.5
            _df = df_calls[['lat','lon']].dropna().copy()
            _df['_bl'] = (_df['lat'] / (BIN/LAT_MI)).round().astype(int)
            _df['_bn'] = (_df['lon'] / (BIN/LON_MI)).round().astype(int)
            bins = _df.groupby(['_bl','_bn']).size().sort_values(ascending=False)
            total = int(bins.sum())
            cum = bins.cumsum()
            n_cells = len(bins)
            top10 = max(1, int(n_cells * 0.1))
            pct10  = round(float(cum.iloc[top10-1]) / total * 100, 1)
            # Downsample to max 200 points for the export chart
            step = max(1, n_cells // 200)
            conc_x = list(range(1, n_cells+1, step))
            conc_y = [round(float(cum.iloc[min(i-1, n_cells-1)]) / total * 100, 1) for i in conc_x]
        except Exception:
            pass

        total_calls = len(df_calls)

        import json
        pri_labels_js  = json.dumps(pri_labels)
        pri_vals_js    = json.dumps(pri_vals)
        type_labels_js = json.dumps(type_labels)
        type_vals_js   = json.dumps(type_vals)
        conc_x_js      = json.dumps(conc_x)
        conc_y_js      = json.dumps(conc_y)

        has_pri   = "true" if pri_vals   else "false"
        has_types = "true" if type_vals  else "false"
        has_conc  = "true" if conc_x     else "false"

        bar_height = max(260, len(type_labels) * 28 + 60) if type_labels else 260

        return f"""
<h2 style="color:#111; font-size:22px; font-weight:800; margin-top:40px; margin-bottom:20px;
           padding-bottom:10px; border-bottom:2px solid #eee;">Incident Data Analysis</h2>
<p style="font-size:13px; color:#666; margin-bottom:20px;">
  Summary of <strong>{total_calls:,}</strong> calls for service used to optimise drone placement.
</p>
<div style="display:grid; grid-template-columns:1fr 1fr; gap:24px; margin-bottom:24px;">
  <div>
    <p style="font-size:12px; font-weight:700; color:#555; text-transform:uppercase;
              letter-spacing:0.5px; margin:0 0 8px;">Call Priority Breakdown</p>
    <div style="position:relative; height:220px;"><canvas id="expPriChart"></canvas></div>
  </div>
  <div>
    <p style="font-size:12px; font-weight:700; color:#555; text-transform:uppercase;
              letter-spacing:0.5px; margin:0 0 8px;">Call Density Concentration</p>
    <div style="position:relative; height:220px;"><canvas id="expConcChart"></canvas></div>
  </div>
</div>
<p style="font-size:12px; font-weight:700; color:#555; text-transform:uppercase;
          letter-spacing:0.5px; margin:0 0 8px;">Top Call Types</p>
<div style="position:relative; height:{bar_height}px; margin-bottom:24px;">
  <canvas id="expTypeChart"></canvas>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
(function(){{
  var priL={pri_labels_js}, priV={pri_vals_js};
  var typL={type_labels_js}, typV={type_vals_js};
  var cX={conc_x_js}, cY={conc_y_js};
  var hasPri={has_pri}, hasTypes={has_types}, hasConc={has_conc};

  if(hasPri && priL.length) {{
    new Chart(document.getElementById('expPriChart'), {{
      type:'doughnut',
      data:{{
        labels:priL,
        datasets:[{{
          data:priV,
          backgroundColor:['#E24B4A','#3B8BD4','#9FE1CB','#888780','#EF9F27'],
          borderWidth:0
        }}]
      }},
      options:{{responsive:true,maintainAspectRatio:false,cutout:'55%',
        plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}},padding:8}}}}}}
      }}
    }});
  }}

  if(hasConc && cX.length) {{
    new Chart(document.getElementById('expConcChart'), {{
      type:'line',
      data:{{
        labels:cX,
        datasets:[{{
          data:cY, borderColor:'#00D2FF', backgroundColor:'rgba(0,210,255,0.12)',
          fill:true, tension:0.3, pointRadius:0, borderWidth:2
        }}]
      }},
      options:{{responsive:true,maintainAspectRatio:false,
        plugins:{{legend:{{display:false}},
          annotation:{{annotations:{{}}}}
        }},
        scales:{{
          x:{{title:{{display:true,text:'Cells ranked by density',font:{{size:10}}}},
             ticks:{{maxTicksLimit:6}}}},
          y:{{title:{{display:true,text:'% of calls',font:{{size:10}}}},min:0,max:100,
             ticks:{{callback:function(v){{return v+'%'}}}}}}
        }}
      }}
    }});
  }}

  if(hasTypes && typL.length) {{
    new Chart(document.getElementById('expTypeChart'), {{
      type:'bar',
      data:{{
        labels:typL,
        datasets:[{{data:typV,backgroundColor:'#00D2FF',borderRadius:3,borderSkipped:false}}]
      }},
      options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',
        plugins:{{legend:{{display:false}}}},
        scales:{{
          x:{{ticks:{{callback:function(v){{return v>=1000?Math.round(v/1000)+'k':v}}}}}},
          y:{{ticks:{{font:{{size:11}}}}}}
        }}
      }}
    }});
  }}
}})();
</script>
"""
    except Exception:
        return ""


def _build_cad_charts(df_calls, text_main, text_muted, card_bg, card_border, accent_color):
    """Render three Plotly charts summarising the uploaded CAD dataset."""
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    if df_calls is None or df_calls.empty:
        return

    layout_base = dict(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color=text_muted, size=11),
        margin=dict(l=10, r=10, t=32, b=10),
        hoverlabel=dict(bgcolor=card_bg, font_size=12, font_color=text_main, bordercolor=accent_color),
    )
    grid_color = card_border

    # ── Chart 1: Priority breakdown (donut) ──────────────────────────────────
    if 'priority' in df_calls.columns:
        pri_counts = df_calls['priority'].dropna().astype(str).value_counts().sort_index()
        if not pri_counts.empty:
            colors = ['#E24B4A','#3B8BD4','#9FE1CB','#888780','#EF9F27','#5DCAA5']
            labels = [f"Priority {p}" for p in pri_counts.index]
            fig_pri = go.Figure(go.Pie(
                labels=labels, values=pri_counts.values,
                hole=0.55, marker_colors=colors[:len(pri_counts)],
                textinfo='percent', hovertemplate='<b>%{label}</b><br>%{value:,} calls (%{percent})<extra></extra>'
            ))
            fig_pri.update_layout(**layout_base, height=240,
                title=dict(text='Call Priority Distribution', font=dict(size=13, color=text_main), x=0),
                showlegend=True,
                legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5,
                            font=dict(size=10, color=text_muted))
            )
            st.plotly_chart(fig_pri, use_container_width=True, config={'displayModeBar': False})

    # ── Chart 2: Top event types (horizontal bar) ─────────────────────────────
    desc_col = None
    for _c in ['call_type_desc','agencyeventtypecodedesc','eventdesc','calldesc','description','nature','event_desc']:
        if _c in df_calls.columns and df_calls[_c].dropna().nunique() > 2:
            desc_col = _c
            break

    if desc_col:
        top_types = df_calls[desc_col].dropna().str.strip().value_counts().head(12)
        if not top_types.empty:
            fig_types = go.Figure(go.Bar(
                x=top_types.values, y=top_types.index,
                orientation='h',
                marker_color=accent_color,
                text=[f'{v:,}' for v in top_types.values],
                textposition='outside',
                hovertemplate='<b>%{y}</b><br>%{x:,} calls<extra></extra>',
            ))
            fig_types.update_layout(**layout_base,
                height=max(280, len(top_types) * 30 + 60),
                title=dict(text='Top Call Types', font=dict(size=13, color=text_main), x=0),
                xaxis=dict(showgrid=True, gridcolor=grid_color, title='Calls'),
                yaxis=dict(showgrid=False, autorange='reversed'),
                showlegend=False,
            )
            st.plotly_chart(fig_types, use_container_width=True, config={'displayModeBar': False})

    # ── Chart 3: Call density concentration (Lorenz curve) ───────────────────
    try:
        LAT_MI, LON_MI = 69.0, 55.0
        BIN = 0.5
        bin_lat = BIN / LAT_MI
        bin_lon = BIN / LON_MI
        _df = df_calls[['lat','lon']].dropna().copy()
        _df['_bl'] = (_df['lat'] / bin_lat).round().astype(int)
        _df['_bn'] = (_df['lon'] / bin_lon).round().astype(int)
        bins = _df.groupby(['_bl','_bn']).size().sort_values(ascending=False)
        total = bins.sum()
        cum_pct = (bins.cumsum() / total * 100).values
        n_cells = len(bins)
        top10 = int(n_cells * 0.1)
        pct_from_top10 = round(float(cum_pct[min(top10-1, len(cum_pct)-1)]), 1)
        x_vals = list(range(1, n_cells + 1))

        fig_conc = go.Figure()
        fig_conc.add_trace(go.Scatter(
            x=x_vals, y=list(cum_pct),
            fill='tozeroy', line=dict(color=accent_color, width=2),
            fillcolor=f'rgba(0,210,255,0.12)',
            hovertemplate='Top %{x} cells → %{y:.1f}% of calls<extra></extra>',
            name='Cumulative calls'
        ))
        fig_conc.add_annotation(
            x=top10, y=pct_from_top10,
            text=f'Top 10% of cells<br>= {pct_from_top10}% of calls',
            showarrow=True, arrowhead=2, arrowcolor=accent_color,
            font=dict(size=10, color=text_main),
            bgcolor=card_bg, bordercolor=accent_color, borderwidth=1
        )
        fig_conc.update_layout(**layout_base,
            height=220,
            title=dict(text=f'Call Density Concentration — {n_cells} active 0.5-mi cells', font=dict(size=13, color=text_main), x=0),
            xaxis=dict(title='Cells ranked by call density', showgrid=True, gridcolor=grid_color),
            yaxis=dict(title='% of total calls', showgrid=True, gridcolor=grid_color, range=[0,105]),
            showlegend=False,
        )
        st.plotly_chart(fig_conc, use_container_width=True, config={'displayModeBar': False})
    except Exception:
        pass



def _safe_df_to_records(df):
    """Safely serialize a DataFrame to a JSON-safe list of records.
    Returns an empty list if df is None, empty, or serialization fails."""
    if df is None:
        return []
    try:
        if hasattr(df, 'empty') and df.empty:
            return []
        return json.loads(df.replace({float('nan'): None}).to_json(orient='records'))
    except Exception:
        try:
            return json.loads(df.fillna('').to_json(orient='records'))
        except Exception:
            return []


def _make_random_stations(df_calls, n=40, boundary_geom=None, epsg_code=None):
    """Fallback station generator based on call-density hotspots.

    If a city boundary is supplied, only incidents inside that boundary are used and
    final station coordinates are snapped to the nearest in-boundary incident so every
    suggested site remains inside the geographic area.
    """
    if df_calls is None or df_calls.empty:
        return pd.DataFrame()

    work = df_calls.copy()
    work['lat'] = pd.to_numeric(work['lat'], errors='coerce')
    work['lon'] = pd.to_numeric(work['lon'], errors='coerce')
    work = work.dropna(subset=['lat', 'lon']).reset_index(drop=True)
    if work.empty:
        return pd.DataFrame()

    if boundary_geom is not None and epsg_code is not None:
        try:
            work_gdf = gpd.GeoDataFrame(work, geometry=gpd.points_from_xy(work.lon, work.lat), crs="EPSG:4326").to_crs(epsg=int(epsg_code))
            inside_mask = work_gdf.within(boundary_geom)
            if inside_mask.any():
                work = work.loc[inside_mask.values].reset_index(drop=True)
        except Exception:
            pass
        if work.empty:
            return pd.DataFrame()

    lats = work['lat'].dropna().values
    lons = work['lon'].dropna().values
    if len(lats) == 0:
        return pd.DataFrame()

    q1_la, q3_la = np.percentile(lats, 5), np.percentile(lats, 95)
    q1_lo, q3_lo = np.percentile(lons, 5), np.percentile(lons, 95)
    iqr_la, iqr_lo = q3_la - q1_la, q3_lo - q1_lo
    buf_la, buf_lo = max(iqr_la * 0.5, 0.01), max(iqr_lo * 0.5, 0.01)
    mask = ((lats >= q1_la - buf_la) & (lats <= q3_la + buf_la) &
            (lons >= q1_lo - buf_lo) & (lons <= q3_lo + buf_lo))
    clean_lats, clean_lons = lats[mask], lons[mask]
    if len(clean_lats) == 0:
        clean_lats, clean_lons = lats, lons

    base_coords = np.column_stack([clean_lats, clean_lons])
    if len(base_coords) == 0:
        return pd.DataFrame()

    try:
        from sklearn.cluster import MiniBatchKMeans as _KM
        k = min(n, len(base_coords))
        km = _KM(n_clusters=k, random_state=42, batch_size=1024, n_init=3)
        km.fit(base_coords)
        centroids = km.cluster_centers_
    except Exception:
        np.random.seed(42)
        idx = np.random.choice(len(base_coords), min(n, len(base_coords)), replace=False)
        centroids = base_coords[idx]

    # Snap every centroid to the nearest actual in-boundary call to guarantee the
    # proposed station remains inside the jurisdiction geometry.
    snapped = []
    for cen_lat, cen_lon in centroids:
        d2 = (base_coords[:, 0] - cen_lat) ** 2 + (base_coords[:, 1] - cen_lon) ** 2
        nearest = base_coords[int(np.argmin(d2))]
        snapped.append((float(nearest[0]), float(nearest[1])))

    if not snapped:
        return pd.DataFrame()

    deduped = list(dict.fromkeys((round(lat, 6), round(lon, 6)) for lat, lon in snapped))
    station_lats = np.array([lat for lat, _ in deduped])
    station_lons = np.array([lon for _, lon in deduped])

    k_actual = len(station_lats)
    types = (['Police'] * max(1, math.ceil(k_actual * 0.5)) +
             ['Fire']   * max(1, math.ceil(k_actual * 0.3)) +
             ['School'] * max(1, math.ceil(k_actual * 0.2)))[:k_actual]
    return pd.DataFrame({
        'name': [f"{types[i]} Station {i+1}" for i in range(k_actual)],
        'lat':  station_lats,
        'lon':  station_lons,
        'type': types,
    })

def generate_stations_from_calls(df_calls, max_stations=100):
    """Query OpenStreetMap for real stations; fall back gracefully if unavailable."""
    lats = df_calls['lat'].dropna().values
    lons = df_calls['lon'].dropna().values
    if len(lats) == 0: return None, "No coordinates available to generate stations."

    q1_la, q3_la = np.percentile(lats, 25), np.percentile(lats, 75)
    q1_lo, q3_lo = np.percentile(lons, 25), np.percentile(lons, 75)
    iqr_la = q3_la - q1_la
    iqr_lo = q3_lo - q1_lo
    mask = (lats >= q1_la - 2.5 * iqr_la) & (lats <= q3_la + 2.5 * iqr_la) & (lons >= q1_lo - 2.5 * iqr_lo) & (lons <= q3_lo + 2.5 * iqr_lo)
    cen_lat, cen_lon = lats[mask].mean(), lons[mask].mean()

    # Try a slightly larger radius first, then fall back to tighter if too many results
    for R in [0.25, 0.45]:
        bbox = f"{cen_lat - R},{cen_lon - R},{cen_lat + R},{cen_lon + R}"
        query = (
            f'[out:json][timeout:30];'
            f'(node["amenity"="fire_station"]({bbox});'
            f'node["amenity"="police"]({bbox});'
            f'node["amenity"="school"]({bbox});'
            f'way["amenity"="fire_station"]({bbox});'
            f'way["amenity"="police"]({bbox});'
            f'way["amenity"="school"]({bbox});'
            f');out center;'
        )

        data = None
        osm_urls = [
            'https://overpass-api.de/api/interpreter',
            'https://overpass.kumi.systems/api/interpreter',
            'https://overpass.openstreetmap.ru/api/interpreter',  # third mirror
        ]
        for osm_url in osm_urls:
            try:
                req = urllib.request.Request(
                    f"{osm_url}?data={urllib.parse.quote(query)}",
                    headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'}
                )
                with urllib.request.urlopen(req, timeout=25) as resp:
                    data = json.loads(resp.read().decode('utf-8'))
                break
            except Exception:
                continue

        if data is None:
            continue  # try next radius / give up

        elements = data.get('elements', [])
        rows = []
        for el in elements:
            tags = el.get('tags', {})
            lat = el.get('lat') or (el.get('center') or {}).get('lat')
            lon = el.get('lon') or (el.get('center') or {}).get('lon')
            if lat is None or lon is None: continue
            amenity = tags.get('amenity', '')
            type_label = 'Fire' if amenity == 'fire_station' else 'Police' if amenity == 'police' else 'School'
            fac_name = tags.get('name', f"{type_label} Station")
            rows.append({'name': fac_name, 'lat': round(lat, 6), 'lon': round(lon, 6), 'type': type_label})

        if rows:
            df_s = pd.DataFrame(rows).drop_duplicates(subset=['lat', 'lon']).reset_index(drop=True)
            # Enforce unique names
            counts = {}
            new_names = []
            for n in df_s['name']:
                if n in counts:
                    counts[n] += 1
                    new_names.append(f"{n} ({counts[n]})")
                else:
                    counts[n] = 0
                    new_names.append(n)
            df_s['name'] = new_names
            if len(df_s) > max_stations:
                priority_order = {'Police': 0, 'Fire': 1, 'School': 2}
                df_s['_pri'] = df_s['type'].map(priority_order).fillna(3)
                df_s = df_s.sort_values('_pri').head(max_stations).drop(columns='_pri').reset_index(drop=True)
            return df_s, f"Auto-generated {len(df_s)} stations from OpenStreetMap."

    # ── All OSM attempts failed — use random stations from call locations ──
    df_fallback = _make_random_stations(df_calls, n=40)
    if not df_fallback.empty:
        return df_fallback, "⚠️ OpenStreetMap unavailable — using estimated station locations from call data. Upload a stations CSV for accuracy."
    return None, "Could not generate stations — no valid call coordinates."

    return df_s, f"Auto-generated {len(df_s)} stations from OpenStreetMap."

# ============================================================
# CACHED DATA FUNCTIONS
# ============================================================
@st.cache_data
def get_address_from_latlon(lat, lon):
    url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=18&addressdetails=1"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_DFR_Optimizer_App/2.0'})
        with urllib.request.urlopen(req, timeout=2) as response:
            data = json.loads(response.read().decode('utf-8'))
            if 'address' in data:
                addr = data['address']
                road = addr.get('road', '')
                house_number = addr.get('house_number', '')
                city = addr.get('city', addr.get('town', addr.get('village', '')))
                if road:
                    return f"{house_number} {road}, {city}".strip(', ')
    except Exception:
        pass
    # Fallback to coordinates if an exact street address isn't found
    return f"{lat:.5f}, {lon:.5f}"

@st.cache_data
def forward_geocode(address_str):
    import urllib.parse
    url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(address_str)}&limit=1"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data:
                return float(data[0]['lat']), float(data[0]['lon'])
    except Exception: pass
    return None, None

def lookup_county_for_city(city_name, state_abbr):
    """Use Nominatim reverse-geocode to find the county name for a city that
    doesn't directly match a county name in the local parquet."""
    try:
        lat, lon = forward_geocode(f"{city_name}, {state_abbr}, USA")
        if lat is None: return None
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=8&addressdetails=1"
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        county_raw = data.get('address', {}).get('county', '')
        # Nominatim returns "Winnebago County" — strip the suffix
        county_name = county_raw.replace(' County', '').replace(' Parish', '').replace(' Borough', '').strip()
        return county_name if county_name else None
    except Exception:
        return None

@st.cache_data
def fetch_county_boundary_local(state_abbr, county_name_input):
    # 1. Clean the input
    search_name = county_name_input.lower().strip()
    if search_name.endswith(" county"):
        search_name = search_name.replace(" county", "").strip()
        
    state_fips = STATE_FIPS.get(state_abbr)
    if not state_fips: return False, None
    
    # 2. Look for our new ultra-compressed parquet file
    local_file = "counties_lite.parquet"
    if not os.path.exists(local_file):
        st.error(f"Missing {local_file}! Please ensure it is uploaded to your repository.")
        return False, None
                
    # 3. Read directly from the Parquet file instantly
    try:
        # Geopandas reads Parquet files in milliseconds!
        gdf = gpd.read_parquet(local_file)
        
        # Filter for the exact State FIPS code and County Name
        match = gdf[(gdf['STATEFP'] == state_fips) & (gdf['NAME'].str.lower() == search_name)]
        
        if not match.empty:
            # Put the word "County" back on for the UI displays
            match = match.copy()
            match['NAME'] = match['NAME'] + " County"
            return True, match[['NAME', 'geometry']]
    except Exception as e:
        st.error(f"Error reading local database: {e}")
        pass
        
    return False, None

@st.cache_data
def fetch_place_boundary_local(state_abbr, place_name_input):
    """Look up a city/town/CDP boundary from the local places_lite.parquet.
    Returns (True, GeoDataFrame) on success, (False, None) if not found or
    the file doesn't exist yet (falls back to county lookup in caller)."""
    local_file = "places_lite.parquet"
    if not os.path.exists(local_file):
        return False, None   # file not yet added — caller falls back to county

    state_fips = STATE_FIPS.get(state_abbr)
    if not state_fips: return False, None

    search_name = place_name_input.lower().strip()
    # Strip common suffixes the user might have typed
    for suffix in [" city", " town", " village", " borough", " township", " cdp"]:
        if search_name.endswith(suffix):
            search_name = search_name[:-len(suffix)].strip()
            break

    try:
        gdf = gpd.read_parquet(local_file)
        state_rows = gdf[gdf["STATEFP"] == state_fips]

        # Exact match first
        match = state_rows[state_rows["NAME"].str.lower() == search_name]

        # Partial match fallback (e.g. "Fort Worth" matching "Fort Worth city")
        if match.empty:
            match = state_rows[state_rows["NAME"].str.lower().str.startswith(search_name)]
            if not match.empty:
                match = match.copy()
                match["_diff"] = match["NAME"].str.len() - len(search_name)
                match = match.sort_values("_diff").head(1)

        if match.empty:
            return False, None

        result = match.copy()
        # Use NAMELSAD for display if available (e.g. "Rockford city"), else NAME
        name_col = "NAMELSAD" if "NAMELSAD" in result.columns else "NAME"
        result["NAME"] = result[name_col].astype(str)
        return True, result[["NAME", "geometry"]]

    except Exception:
        return False, None

@st.cache_data
def reverse_geocode_state(lat, lon):
    url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=10&addressdetails=1"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode('utf-8'))
            address = data.get('address', {})
            state = address.get('state', '')
            city = address.get('city', address.get('town', address.get('village', address.get('county', 'Unknown City'))))
            return state, city
    except Exception: return None, None

@st.cache_data
def fetch_census_population(state_fips, place_name, is_county=False):
    if is_county:
        url = f"https://api.census.gov/data/2020/dec/pl?get=P1_001N,NAME&for=county:*&in=state:{state_fips}"
    else:
        url = f"https://api.census.gov/data/2020/dec/pl?get=P1_001N,NAME&for=place:*&in=state:{state_fips}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            search_name = place_name.lower().strip()
            for row in data[1:]:
                place_full = row[1].lower().split(',')[0].strip()
                if place_full == search_name or place_full.startswith(search_name + " "): return int(row[0])
    except Exception: pass
    return None

SHAPEFILE_DIR = "jurisdiction_data"
if not os.path.exists(SHAPEFILE_DIR): os.makedirs(SHAPEFILE_DIR)

@st.cache_data
def fetch_tiger_city_shapefile(state_fips, city_name, output_dir):
    # Check if we already downloaded and cached this state's places file
    temp_dir = os.path.join(output_dir, f"temp_tiger_{state_fips}")
    cached_shp = os.path.join(temp_dir, f"tl_2023_{state_fips}_place.shp")
    gdf = None

    if os.path.exists(cached_shp):
        try:
            gdf = gpd.read_file(cached_shp)
        except Exception:
            gdf = None

    if gdf is None:
        # Download from Census TIGER — try 2023 then 2022 as fallback
        for year in ["2023", "2022"]:
            url = f"https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_{state_fips}_place.zip"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"})
                with urllib.request.urlopen(req, timeout=45) as resp:
                    zip_data = resp.read()
                zip_file = zipfile.ZipFile(io.BytesIO(zip_data))
                os.makedirs(temp_dir, exist_ok=True)
                zip_file.extractall(temp_dir)
                shp_files = glob.glob(os.path.join(temp_dir, "*.shp"))
                if shp_files:
                    gdf = gpd.read_file(shp_files[0])
                    break
            except Exception:
                continue

    if gdf is None:
        return False, None

    try:
        search_name = city_name.lower().strip()
        exact_mask = gdf['NAME'].str.lower().str.strip() == search_name
        if exact_mask.any():
            city_gdf = gdf[exact_mask].copy()
        else:
            # Partial match — prefer the longest name match to avoid tiny place with same substring
            partial = gdf[gdf['NAME'].str.lower().str.contains(search_name, case=False, na=False)].copy()
            if partial.empty:
                return False, None
            # Pick the row whose NAME most closely matches (shortest extra chars)
            partial['_diff'] = partial['NAME'].str.len() - len(search_name)
            city_gdf = partial.sort_values('_diff').head(1)

        if city_gdf.empty:
            return False, None

        city_gdf = city_gdf.dissolve(by='NAME').reset_index()
        if city_gdf.crs is None:
            city_gdf = city_gdf.set_crs(epsg=4269)
        city_gdf = city_gdf.to_crs(epsg=4326)
        save_path = os.path.join(output_dir, f"{city_name.replace(' ', '_')}_{state_fips}.shp")
        city_gdf.to_file(save_path)
        return True, city_gdf
    except Exception:
        return False, None

def generate_mock_faa_grid(minx, miny, maxx, maxy):
    features = []
    x_steps = np.linspace(minx, maxx, 20)
    y_steps = np.linspace(miny, maxy, 20)
    mock_airports = [{"lon": minx + 0.3 * (maxx - minx), "lat": miny + 0.3 * (maxy - miny), "radius": 0.15, "name": "Mock Intl (MCK)"}]
    for i in range(len(x_steps) - 1):
        for j in range(len(y_steps) - 1):
            cell_poly = [[x_steps[i], y_steps[j]], [x_steps[i+1], y_steps[j]], [x_steps[i+1], y_steps[j+1]], [x_steps[i], y_steps[j+1]], [x_steps[i], y_steps[j]]]
            cell_center = Point((x_steps[i] + x_steps[i+1]) / 2, (y_steps[j] + y_steps[j+1]) / 2)
            ceiling, arpt_name = None, ""
            for ap in mock_airports:
                dist_ratio = cell_center.distance(Point(ap["lon"], ap["lat"])) / ap["radius"]
                if dist_ratio < 1.0:
                    if   dist_ratio < 0.15: ceiling, arpt_name = 0,   ap["name"]
                    elif dist_ratio < 0.35: ceiling, arpt_name = 50,  ap["name"]
                    elif dist_ratio < 0.55: ceiling, arpt_name = 100, ap["name"]
                    else:                   ceiling, arpt_name = 200, ap["name"]
                    break
            if ceiling is not None:
                features.append({"type": "Feature", "geometry": {"type": "Polygon", "coordinates": [cell_poly]}, "properties": {"CEILING": ceiling, "ARPT_Name": arpt_name}})
    return {"type": "FeatureCollection", "features": features}

@st.cache_data
def load_faa_parquet(minx, miny, maxx, maxy):
    if not os.path.exists("faa_uasfm.parquet"): return generate_mock_faa_grid(minx, miny, maxx, maxy)
    try:
        gdf = gpd.read_parquet("faa_uasfm.parquet")
        pad = 0.05
        filtered = gdf.cx[minx-pad:maxx+pad, miny-pad:maxy+pad]
        if filtered.empty: return {"type": "FeatureCollection", "features": []}
        return json.loads(filtered.to_json())
    except Exception as e: return generate_mock_faa_grid(minx, miny, maxx, maxy)

def add_faa_laanc_layer_to_plotly(fig, faa_geojson, is_dark=True):
    if not faa_geojson or not faa_geojson.get("features"): return
    text_lons, text_lats, text_strings, text_hovers = [], [], [], []
    for feature in faa_geojson.get("features", []):
        geom = feature.get("geometry")
        props = feature.get("properties", {})
        ceiling = props.get("CEILING")
        arpt = props.get("ARPT_Name") or props.get("ARPT_NAME") or "Unknown Airport"
        if ceiling is None or geom is None or geom.get("type") != "Polygon": continue
        snapped = min(FAA_CEILING_COLORS.keys(), key=lambda v: abs(v - ceiling))
        colors = FAA_CEILING_COLORS.get(snapped, FAA_DEFAULT_COLOR)
        coords = geom["coordinates"][0]
        bx, by = zip(*coords)
        fig.add_trace(go.Scattermapbox(mode="lines", lon=list(bx), lat=list(by), fill="toself", fillcolor=colors["fill"], line=dict(color=colors["line"], width=1.5), hoverinfo="text", text=f"<b>{ceiling} ft AGL</b><br>{arpt}", name=f"LAANC {ceiling}ft", showlegend=False))
        try:
            centroid = shape(geom).centroid
            text_lons.append(centroid.x); text_lats.append(centroid.y); text_strings.append(str(ceiling)); text_hovers.append(f"{ceiling} ft — {arpt}")
        except Exception: pass
    if text_lons:
        fig.add_trace(go.Scattermapbox(mode="text", lon=text_lons, lat=text_lats, text=text_strings, hovertext=text_hovers, hoverinfo="text", textfont=dict(size=10, color="#ffffff" if is_dark else "#000000"), showlegend=False, name="LAANC Labels"))

def get_station_faa_ceiling(lat, lon, faa_geojson):
    if not faa_geojson or 'features' not in faa_geojson: return "400 ft (Class G)"
    pt = Point(lon, lat)
    for feature in faa_geojson['features']:
        if 'geometry' in feature and feature['geometry']:
            try:
                s = shape(feature['geometry'])
                if s.contains(pt):
                    val = feature['properties'].get('CEILING')
                    if val is not None: return f"{val} ft (Controlled)"
            except Exception: pass
    return "400 ft (Class G)"

@st.cache_data
def fetch_airfields(minx, miny, maxx, maxy):
    pad = 0.2
    query = f"""[out:json];(node["aeroway"~"aerodrome|heliport"]({miny-pad},{minx-pad},{maxy+pad},{maxx+pad});way["aeroway"~"aerodrome|heliport"]({miny-pad},{minx-pad},{maxy+pad},{maxx+pad}););out center;"""
    try:
        req = urllib.request.Request("https://overpass-api.de/api/interpreter", data=query.encode('utf-8'), headers={'User-Agent': 'BRINC_Optimizer'})
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            airfields = []
            for el in data.get('elements', []):
                lat = el.get('lat') or el.get('center', {}).get('lat')
                lon = el.get('lon') or el.get('center', {}).get('lon')
                name = el.get('tags', {}).get('name', 'Unknown Airfield')
                if lat and lon: airfields.append({'name': name, 'lat': lat, 'lon': lon})
            return airfields
    except Exception: return []

def get_nearest_airfield(lat, lon, airfields):
    if not airfields: return "No data"
    min_dist = float('inf')
    best = None
    for af in airfields:
        lat1, lon1, lat2, lon2 = map(math.radians, [lat, lon, af['lat'], af['lon']])
        a = math.sin((lat2-lat1)/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin((lon2-lon1)/2)**2
        dist = 3958.8 * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
        if dist < min_dist:
            y = math.sin(lon2-lon1)*math.cos(lat2)
            x = math.cos(lat1)*math.sin(lat2) - math.sin(lat1)*math.cos(lat2)*math.cos(lon2-lon1)
            bearing = (math.degrees(math.atan2(y, x)) + 360) % 360
            dirs = ['N','NNE','NE','ENE','E','ESE','SE','SSE','S','SSW','SW','WSW','W','WNW','NW','NNW']
            min_dist = dist
            best = (af['name'], dist, dirs[int((bearing+11.25)/22.5) % 16])
    if best:
        n = best[0][:18] + ("..." if len(best[0]) > 18 else "")
        return f"{best[1]:.1f}mi {best[2]} ({n})"
    return "No data"

def generate_random_points_in_polygon(polygon, num_points):
    points = []
    minx, miny, maxx, maxy = polygon.bounds
    while len(points) < num_points:
        x_coords = np.random.uniform(minx, maxx, 1000)
        y_coords = np.random.uniform(miny, maxy, 1000)
        for x, y in zip(x_coords, y_coords):
            if len(points) >= num_points: break
            if polygon.contains(Point(x, y)): points.append((y, x))
    return points

def generate_clustered_calls(polygon, num_points):
    points = []
    minx, miny, maxx, maxy = polygon.bounds
    hotspots = []
    while len(hotspots) < random.randint(5, 15):
        hx, hy = random.uniform(minx, maxx), random.uniform(miny, maxy)
        if polygon.contains(Point(hx, hy)): hotspots.append((hx, hy))
    target_clustered = int(num_points * 0.75)
    while len(points) < target_clustered:
        hx, hy = random.choice(hotspots)
        px, py = np.random.normal(hx, 0.02), np.random.normal(hy, 0.02)
        if polygon.contains(Point(px, py)): points.append((py, px))
    while len(points) < num_points:
        px, py = random.uniform(minx, maxx), random.uniform(miny, maxy)
        if polygon.contains(Point(px, py)): points.append((py, px))
    np.random.shuffle(points)
    return points

def estimate_grants(population):
    if population > 1000000: return "$1.5M - $3.0M+"
    elif population > 500000: return "$500k - $1.5M"
    elif population > 250000: return "$250k - $500k"
    elif population > 100000: return "$100k - $250k"
    else: return "$25k - $100k"

def get_circle_coords(lat, lon, r_mi=2.0):
    angles = np.linspace(0, 2*np.pi, 100)
    c_lats = lat + (r_mi/69.172) * np.sin(angles)
    c_lons = lon + (r_mi/(69.172 * np.cos(np.radians(lat)))) * np.cos(angles)
    return c_lats, c_lons

def format_3_lines(name_str):
    match = re.search(r'\s(\d{1,5}\s+[A-Za-z])', name_str)
    if match:
        idx = match.start()
        line1 = name_str[:idx].strip()
        rest = name_str[idx:].strip()
        if ',' in rest:
            parts = rest.split(',', 1)
            return f"{line1}<br>{parts[0].strip()},<br>{parts[1].strip()}"
        return f"{line1}<br>{rest}"
    if ',' in name_str:
        parts = name_str.split(',')
        if len(parts) >= 3:
            return f"{parts[0].strip()},<br>{parts[1].strip()},<br>{','.join(parts[2:]).strip()}"
    return name_str

def _build_unit_cards_html(active_drones, text_main, text_muted, card_bg, card_border, card_title, accent_color, columns_per_row=2):
    if not active_drones:
        return ""
    # Per-type daily airtime budgets derived from CONFIG duty cycles:
    #   Guardian: 60 min flight + 3 min charge → (24*60/63)*60 = 1371.4 min = 22.86 hr
    #   Responder: patrol-unit model, 11.6 hr shift equivalent
    _GUARDIAN_DAILY_MINS  = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"]   # ~1371.4
    _GUARDIAN_DAILY_HOURS = CONFIG["GUARDIAN_PATROL_HOURS"]        # ~22.86
    _RESPONDER_DAILY_MINS  = CONFIG["RESPONDER_PATROL_HOURS"] * 60 # 11.6 * 60 = 696
    _RESPONDER_DAILY_HOURS = CONFIG["RESPONDER_PATROL_HOURS"]      # 11.6
    columns_per_row = max(1, int(columns_per_row))

    cards_html = []
    for d in active_drones:
        short_name  = format_3_lines(d["name"])
        d_color     = d["color"]
        d_type      = d["type"]
        d_step      = d["deploy_step"]
        d_savings   = d["annual_savings"]
        d_flights   = d["marginal_flights"]
        d_shared    = d["shared_flights"]
        d_deflected = d["marginal_deflected"]
        d_time      = d["avg_time_min"]
        d_faa       = d["faa_ceiling"]
        d_airport   = d["nearest_airport"]
        d_cost      = d["cost"]
        d_be        = d["be_text"]
        d_lat       = d["lat"]
        d_lon       = d["lon"]
        d_address   = get_address_from_latlon(d_lat, d_lon)
        gmaps_url   = f"https://www.google.com/maps/search/?api=1&query={d_lat},{d_lon}"

        # Pick duty-cycle values for this drone type
        is_guardian = (d_type == "GUARDIAN")
        max_patrol_mins  = _GUARDIAN_DAILY_MINS  if is_guardian else _RESPONDER_DAILY_MINS
        max_patrol_hours = _GUARDIAN_DAILY_HOURS if is_guardian else _RESPONDER_DAILY_HOURS

        # Uptime tooltip: show the duty-cycle breakdown for Guardians
        if is_guardian:
            _g_fl  = CONFIG["GUARDIAN_FLIGHT_MIN"]
            _g_ch  = CONFIG["GUARDIAN_CHARGE_MIN"]
            _g_cyc = _g_fl + _g_ch
            _cycles_per_day = (24 * 60) / _g_cyc
            uptime_tooltip = (
                f"{_g_fl}min flight + {_g_ch}min charge = {_g_cyc}min cycle · "
                f"{_cycles_per_day:.1f} cycles/day · "
                f"{max_patrol_hours:.2f}hr airtime"
            )
        else:
            uptime_tooltip = f"{max_patrol_hours}hr patrol shift"

        total_daily_flights = d_flights + d_shared
        patrol_time_line = ""
        if total_daily_flights > 0:
            # Raw calculation: patrol budget / flights = available min per flight
            raw_mins_per_flight = max_patrol_mins / total_daily_flights
            # Cap at drone's physical max flight time
            max_single_flight = CONFIG["GUARDIAN_FLIGHT_MIN"] if is_guardian else CONFIG["RESPONDER_FLIGHT_MIN"]
            mins_per_flight = min(raw_mins_per_flight, max_single_flight)
            capped = raw_mins_per_flight > max_single_flight
            # Always show the line so low-volume Responders display correctly
            patrol_color = "#F0B429" if mins_per_flight < 15 else "#2ecc71" if mins_per_flight >= max_single_flight * 0.9 else "#00D2FF"
            cap_note = f" (max {max_single_flight}min)" if capped else ""
            patrol_time_line = (
                f'<div style="border-top:1px dashed rgba(255,255,255,0.15); margin-top:5px; '
                f'padding-top:5px; font-size:0.58rem; color:{text_muted};" '
                f'title="{uptime_tooltip}">'
                f'{total_daily_flights:.1f} flights ÷ {max_patrol_hours:.2f}hr max = '
                f'<span style="font-weight:800; color:{patrol_color};">{mins_per_flight:.1f} min/flight{cap_note}</span></div>'
            )

        # Concurrency / value breakdown
        d_util         = d.get('utilization', 0)
        d_blocked      = d.get('blocked_per_day', 0)
        d_base_annual  = d.get('base_annual', d_savings)
        d_conc_annual  = d.get('concurrent_annual', 0)
        d_best         = d.get('best_case_annual', d_savings)
        d_best_be      = d.get('best_be_text', d_be)
        util_pct       = f"{d_util*100:.1f}%"
        util_color     = "#dc3545" if d_util > 0.75 else "#F0B429" if d_util > 0.4 else "#2ecc71"
        has_concurrent = d_shared > 0.1 and d_conc_annual > 0
        # Breakdown label for the value box
        if has_concurrent:
            _excl_str = f"${d_base_annual:,.0f} exclusive"
            _conc_str = f"+ ${d_conc_annual:,.0f} concurrent"
        else:
            _excl_str = "exclusive zone coverage"
            _conc_str = ""

        cards_html.append(f'''
<div class="unit-card" style="background:{card_bg}; border-top:3px solid {d_color}; border:1px solid {card_border}; border-top:3px solid {d_color}; border-radius:6px; padding:10px; display:flex; flex-direction:column; box-sizing:border-box;">
  <!-- Header: name + type badge -->
  <div style="margin-bottom:6px;">
    <div style="font-weight:700; font-size:0.72rem; color:{card_title}; line-height:1.3; margin-bottom:1px;">{short_name}</div>
    <div style="font-size:0.55rem; color:#777; text-transform:uppercase; letter-spacing:0.5px;">{"🔒 " if d.get("pinned") else ""}{d_type} · Phase #{d_step}</div>
    <div style="font-size:0.6rem; margin-top:3px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">
      <a href="{gmaps_url}" target="_blank" style="color:{accent_color}; text-decoration:none; font-weight:600;">📍 {d_address} ↗</a>
    </div>
  </div>

  <!-- Annual value box -->
  <div style="background:rgba(0,210,255,0.07); border:1px solid rgba(0,210,255,0.15); border-radius:4px; padding:6px 8px; text-align:center; margin-bottom:4px;">
    <div style="font-size:0.55rem; color:{text_muted}; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:1px;">Annual Value</div>
    <div style="font-size:1.1rem; font-weight:900; color:{accent_color}; line-height:1.1;">${d_best:,.0f}</div>
    {patrol_time_line}
  </div>

  <!-- Value breakdown box -->
  <div style="border:1px solid rgba(57,255,20,0.18); border-radius:4px; padding:4px 6px; text-align:center; margin-bottom:6px; background:rgba(57,255,20,0.04);"
       title="Exclusive: calls only this drone covers. Concurrent: calls handled while partner is airborne.">
    <div style="display:flex; justify-content:space-around; font-size:0.57rem; margin-bottom:1px;">
      <div style="text-align:center;">
        <div style="color:{accent_color}; font-weight:700;">${d_base_annual:,.0f}</div>
        <div style="color:{text_muted}; font-size:0.48rem;">exclusive</div>
      </div>
      <div style="color:{text_muted}; font-size:0.65rem; opacity:0.5; padding-top:2px;">+</div>
      <div style="text-align:center;">
        <div style="color:#39FF14; font-weight:700;">${d_conc_annual:,.0f}</div>
        <div style="color:{text_muted}; font-size:0.48rem;">concurrent</div>
      </div>
    </div>
    <div style="font-size:0.48rem; color:{text_muted}; opacity:0.65;">{util_pct} util · ROI {d_best_be}</div>
  </div>

  <!-- Stats grid -->
  <div style="display:grid; grid-template-columns:1fr 1fr; gap:2px 6px; font-size:0.59rem; flex:1; margin-bottom:6px;">
    <div style="color:{text_muted};">Zone Flights/day</div>
    <div style="text-align:right; font-weight:700; color:{accent_color};">{d.get("zone_flights",d_flights):.1f}</div>
    <div style="color:{text_muted};">Shared Flights</div>
    <div style="text-align:right; font-weight:700; color:{card_title};">{d_shared:.1f}</div>
    <div style="color:{text_muted};">Utilization</div>
    <div style="text-align:right; font-weight:700; color:{util_color};">{util_pct}</div>
    <div style="color:{text_muted};">Resolved/day</div>
    <div style="text-align:right; font-weight:700; color:{card_title};">{d_deflected:.1f}</div>
    <div style="color:{text_muted};">Avg Response</div>
    <div style="text-align:right; font-weight:700; color:{card_title};">{d_time:.1f} min</div>
    <div style="color:{text_muted};">FAA Ceiling</div>
    <div style="text-align:right; font-weight:700; color:{card_title};">{d_faa}</div>
    <div style="color:{text_muted}; font-size:0.54rem;">Airfield</div>
    <div style="text-align:right; font-weight:600; color:{card_title}; font-size:0.52rem; word-break:break-word;">{d_airport}</div>
  </div>

  <!-- CapEx + ROI footer -->
  <div style="border-top:1px solid {card_border}; padding-top:5px; display:grid; grid-template-columns:1fr 1fr; gap:2px 6px; font-size:0.59rem; margin-bottom:6px;">
    <div style="color:{text_muted};">CapEx</div>
    <div style="text-align:right; font-weight:700; color:{card_title};">${d_cost:,.0f}</div>
    <div style="color:{text_muted};">Base ROI</div>
    <div style="text-align:right; font-weight:800; color:{accent_color};">{d_be}</div>
  </div>

  <!-- Pin buttons — rendered via session_state keys set by JS postMessage -->
  <div style="display:grid; grid-template-columns:1fr 1fr; gap:4px;">
    {"'''<div style=\'background:rgba(255,215,0,0.15); border:1px solid rgba(255,215,0,0.4); border-radius:4px; padding:3px 6px; font-size:0.55rem; font-weight:700; color:#FFD700; text-align:center; cursor:pointer;\'>&nbsp;🔒 GUARDIAN LOCKED</div>'''" if d.get("pinned") and d_type=="GUARDIAN" else "'''<div style=\'border:1px dashed rgba(255,215,0,0.25); border-radius:4px; padding:3px 6px; font-size:0.55rem; color:rgba(255,215,0,0.5); text-align:center;\'><span style=\'opacity:0.6\'>🦅 lock as guard</span></div>'''" }
    {"'''<div style=\'background:rgba(0,210,255,0.15); border:1px solid rgba(0,210,255,0.4); border-radius:4px; padding:3px 6px; font-size:0.55rem; font-weight:700; color:#00D2FF; text-align:center; cursor:pointer;\'>&nbsp;🔒 RESPONDER LOCKED</div>'''" if d.get("pinned") and d_type=="RESPONDER" else "'''<div style=\'border:1px dashed rgba(0,210,255,0.25); border-radius:4px; padding:3px 6px; font-size:0.55rem; color:rgba(0,210,255,0.5); text-align:center;\'><span style=\'opacity:0.6\'>🚁 lock as resp</span></div>'''" }
  </div>
</div>''')

    grid = (
        '<div style="display:grid; grid-template-columns:repeat(' + str(columns_per_row) + ', minmax(0, 1fr));'
        ' gap:10px; align-items:start; margin-bottom:12px; width:100%; box-sizing:border-box;">'
        + "".join(cards_html)
        + '</div>'
    )
    # Wrap in a style-scoped div to prevent Streamlit container from collapsing width
    return (
        '<style>'
        '.unit-card-grid { display:grid; gap:10px; align-items:stretch; width:100%; box-sizing:border-box; }'
        '.unit-card-grid > .unit-card { min-width:0; height:100%; }'
        '</style>'
        '<div class="unit-card-grid" style="grid-template-columns:repeat(' + str(columns_per_row) + ', minmax(0,1fr));">'
        + "".join(cards_html)
        + '</div>'
    )

def to_kml_color(hex_str):
    h = hex_str.lstrip('#')
    return f"ff{h[4:6]}{h[2:4]}{h[0:2]}" if len(h) == 6 else "ff0000ff"

def calculate_zoom(min_lon, max_lon, min_lat, max_lat):
    lon_diff = max_lon - min_lon
    lat_diff = max_lat - min_lat
    if lon_diff <= 0 or lat_diff <= 0: return 12
    return min(max(min(np.log2(360/lon_diff), np.log2(180/lat_diff)) + 1.6, 5), 18)

def generate_kml(active_gdf, active_drones, calls_gdf):
    kml = simplekml.Kml()
    kml.document.name = "BRINC DFR Deployment Plan"
    kml.document.description = (
        "SIMULATOR DISCLAIMER: This file was generated by the BRINC Drones Coverage Optimization Simulator. "
        "All coverage zones, station locations, and incident data are model estimates based on user-provided inputs. "
        "Real-world results will vary. This file does not constitute a legal recommendation, binding proposal, "
        "contract, or guarantee of any product, service, or financial outcome. "
        "All deployments require FAA authorization, local ordinances review, and formal procurement."
    )
    fol_bounds = kml.newfolder(name="Jurisdictions")
    for _, row in active_gdf.iterrows():
        geoms = [row.geometry] if isinstance(row.geometry, Polygon) else row.geometry.geoms
        for geom in geoms:
            pol = fol_bounds.newpolygon(name=row.get('DISPLAY_NAME', 'Boundary'))
            pol.outerboundaryis = list(geom.exterior.coords)
            pol.style.linestyle.color = simplekml.Color.red
            pol.style.linestyle.width = 3
            pol.style.polystyle.color = simplekml.Color.changealphaint(30, simplekml.Color.red)
    fol_stations = kml.newfolder(name="Station Points")
    fol_rings = kml.newfolder(name="Coverage Rings")
    for d in active_drones:
        kml_c = to_kml_color(d['color'])
        pnt = fol_stations.newpoint(name=f"[{d['type'][:3]}] {d['name']}")
        pnt.coords = [(d['lon'], d['lat'])]
        pnt.style.iconstyle.icon.href = 'http://maps.google.com/mapfiles/kml/paddle/blu-blank.png'
        lats, lons = get_circle_coords(d['lat'], d['lon'], r_mi=d['radius_m']/1609.34)
        ring_coords = list(zip(lons, lats))
        ring_coords.append(ring_coords[0])
        pol = fol_rings.newpolygon(name=f"Range: {d['name']}")
        pol.outerboundaryis = ring_coords
        pol.style.linestyle.color = kml_c
        pol.style.linestyle.width = 2
        pol.style.polystyle.color = simplekml.Color.changealphaint(60, kml_c)
    fol_calls = kml.newfolder(name="Incident Data (Sample)")
    calls_export = calls_gdf.to_crs(epsg=4326)
    if len(calls_export) > 2000:
        calls_export = calls_export.sample(2000, random_state=42)
    for _, row in calls_export.iterrows():
        pnt = fol_calls.newpoint()
        pnt.coords = [(row.geometry.x, row.geometry.y)]
        pnt.style.iconstyle.scale = 0.5
        pnt.style.iconstyle.icon.href = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'
    return kml.kml()

@st.cache_data
def find_relevant_jurisdictions(calls_df, stations_df, shapefile_dir):
    points_list = []
    if calls_df is not None: points_list.append(calls_df[['lat', 'lon']])
    if stations_df is not None: points_list.append(stations_df[['lat', 'lon']])
    if not points_list: return None
    full_points = pd.concat(points_list)
    full_points = full_points[(full_points.lat.abs() > 1) & (full_points.lon.abs() > 1)]
    scan_points = full_points.sample(50000, random_state=42) if len(full_points) > 50000 else full_points
    points_gdf = gpd.GeoDataFrame(scan_points, geometry=gpd.points_from_xy(scan_points.lon, scan_points.lat), crs="EPSG:4326")
    total_bounds = points_gdf.total_bounds
    shp_files = glob.glob(os.path.join(shapefile_dir, "*.shp"))
    relevant_polys = []
    for shp_path in shp_files:
        try:
            gdf_chunk = gpd.read_file(shp_path, bbox=tuple(total_bounds))
            if not gdf_chunk.empty:
                if gdf_chunk.crs is None: gdf_chunk.set_crs(epsg=4269, inplace=True)
                gdf_chunk = gdf_chunk.to_crs(epsg=4326)
                hits = gpd.sjoin(gdf_chunk, points_gdf, how="inner", predicate="intersects")
                if not hits.empty:
                    subset = gdf_chunk.loc[hits.index.unique()].copy()
                    subset['data_count'] = hits.index.value_counts()
                    name_col = next((c for c in ['NAME','DISTRICT','NAMELSAD'] if c in subset.columns), subset.columns[0])
                    subset['DISPLAY_NAME'] = subset[name_col].astype(str)
                    relevant_polys.append(subset)
        except Exception: continue
    if not relevant_polys: return None
    master_gdf = pd.concat(relevant_polys, ignore_index=True).sort_values(by='data_count', ascending=False)
    master_gdf = master_gdf.dissolve(by='DISPLAY_NAME', aggfunc={'data_count': 'sum'}).reset_index()
    master_gdf = master_gdf.sort_values(by='data_count', ascending=False)
    if master_gdf['data_count'].sum() > 0:
        master_gdf['pct_share'] = master_gdf['data_count'] / master_gdf['data_count'].sum()
        master_gdf['cum_share'] = master_gdf['pct_share'].cumsum()
        mask = (master_gdf['cum_share'] <= 0.98) | (master_gdf['pct_share'] > 0.01)
        mask.iloc[0] = True
        return master_gdf[mask]
    return master_gdf

@st.cache_data(show_spinner=False)
def build_display_calls(df_calls_full, _city_m, epsg_code, max_points=300000, seed=42):
    if df_calls_full is None or len(df_calls_full) == 0:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    df = df_calls_full.copy()
    if 'lat' not in df.columns or 'lon' not in df.columns:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    df['lat'] = pd.to_numeric(df['lat'], errors='coerce')
    df['lon'] = pd.to_numeric(df['lon'], errors='coerce')
    df = df.dropna(subset=['lat', 'lon']).reset_index(drop=True)
    if df.empty:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.lon, df.lat), crs="EPSG:4326")
    try:
        gdf_m = gdf.to_crs(epsg=int(epsg_code))
        calls_in_city = gdf_m[gdf_m.within(_city_m)] if _city_m is not None else gdf_m
    except Exception:
        calls_in_city = gdf

    if calls_in_city.empty:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    if len(calls_in_city) <= max_points:
        return calls_in_city.to_crs(epsg=4326)

    sampled = calls_in_city.copy()
    minx, miny, maxx, maxy = sampled.total_bounds
    span_x = max(maxx - minx, 1.0)
    span_y = max(maxy - miny, 1.0)
    target_cells = max(25, int(np.sqrt(max_points) * 0.7))
    nx = max(25, min(120, target_cells))
    ny = max(25, min(120, int(target_cells * (span_y / span_x))))

    sampled['_gx'] = np.floor((sampled.geometry.x - minx) / span_x * nx).clip(0, nx - 1).astype(int)
    sampled['_gy'] = np.floor((sampled.geometry.y - miny) / span_y * ny).clip(0, ny - 1).astype(int)
    sampled['_cell'] = sampled['_gx'].astype(str) + '_' + sampled['_gy'].astype(str)

    counts = sampled['_cell'].value_counts()
    alloc = np.maximum(1, np.floor(counts / counts.sum() * max_points).astype(int))
    shortfall = int(max_points - alloc.sum())
    if shortfall > 0:
        remainders = (counts / counts.sum() * max_points) - np.floor(counts / counts.sum() * max_points)
        for cell in remainders.sort_values(ascending=False).index[:shortfall]:
            alloc.loc[cell] += 1

    parts = []
    for cell, group in sampled.groupby('_cell', sort=False):
        take = int(min(len(group), alloc.get(cell, 1)))
        if take >= len(group):
            parts.append(group)
        elif take > 0:
            parts.append(group.sample(take, random_state=seed))

    if not parts:
        display_calls = sampled.sample(max_points, random_state=seed)
    else:
        display_calls = pd.concat(parts, ignore_index=False)
        if len(display_calls) > max_points:
            display_calls = display_calls.sample(max_points, random_state=seed)

    display_calls = display_calls.drop(columns=['_gx', '_gy', '_cell'], errors='ignore')
    return display_calls.to_crs(epsg=4326)

@st.cache_resource
def precompute_spatial_data(df_calls, df_calls_full, df_stations_all, _city_m, epsg_code, resp_radius_mi, guard_radius_mi, center_lat, center_lon, bounds_hash):
    gdf_calls = gpd.GeoDataFrame(df_calls, geometry=gpd.points_from_xy(df_calls.lon, df_calls.lat), crs="EPSG:4326")
    gdf_calls_utm = gdf_calls.to_crs(epsg=int(epsg_code))
    try: calls_in_city = gdf_calls_utm[gdf_calls_utm.within(_city_m)]
    except: calls_in_city = gdf_calls_utm
    radius_resp_m = resp_radius_mi * 1609.34
    radius_guard_m = guard_radius_mi * 1609.34
    station_metadata = []
    total_calls = len(calls_in_city)
    n = len(df_stations_all)
    resp_matrix = np.zeros((n, total_calls), dtype=bool)
    guard_matrix = np.zeros((n, total_calls), dtype=bool)
    dist_matrix_r = np.zeros((n, total_calls))
    dist_matrix_g = np.zeros((n, total_calls))
    display_calls = build_display_calls(df_calls_full if df_calls_full is not None else df_calls, _city_m, epsg_code, max_points=300000)
    max_dist = max(((row['lon']-center_lon)**2 + (row['lat']-center_lat)**2)**0.5 for _, row in df_stations_all.iterrows()) or 1.0
    if not calls_in_city.empty:
        calls_array = np.array(list(zip(calls_in_city.geometry.x, calls_in_city.geometry.y)))
        for idx_pos, (i, row) in enumerate(df_stations_all.iterrows()):
            s_pt_m = gpd.GeoSeries([Point(row['lon'], row['lat'])], crs="EPSG:4326").to_crs(epsg=int(epsg_code)).iloc[0]
            dists = np.sqrt((calls_array[:,0]-s_pt_m.x)**2 + (calls_array[:,1]-s_pt_m.y)**2)
            dists_mi = dists / 1609.34
            mask_r = dists <= radius_resp_m
            mask_g = dists <= radius_guard_m
            resp_matrix[idx_pos, :] = mask_r
            guard_matrix[idx_pos, :] = mask_g
            dist_matrix_r[idx_pos, :] = dists_mi
            dist_matrix_g[idx_pos, :] = dists_mi
            full_buf_2m = s_pt_m.buffer(radius_resp_m)
            try: clipped_2m = full_buf_2m.intersection(_city_m)
            except: clipped_2m = full_buf_2m
            full_buf_guard = s_pt_m.buffer(radius_guard_m)
            try: clipped_guard = full_buf_guard.intersection(_city_m)
            except: clipped_guard = full_buf_guard
            dist_c = ((row['lon']-center_lon)**2 + (row['lat']-center_lat)**2)**0.5
            station_metadata.append({
                'name': row['name'], 'lat': row['lat'], 'lon': row['lon'],
                'clipped_2m': clipped_2m, 'clipped_guard': clipped_guard,
                'avg_dist_r': dists_mi[mask_r].mean() if mask_r.any() else resp_radius_mi*(2/3),
                'avg_dist_g': dists_mi[mask_g].mean() if mask_g.any() else guard_radius_mi*(2/3),
                'centrality': 1.0 - (dist_c / max_dist)
            })
    return calls_in_city, display_calls, resp_matrix, guard_matrix, dist_matrix_r, dist_matrix_g, station_metadata, total_calls

def solve_mclp(resp_matrix, guard_matrix, dist_r, dist_g, num_resp, num_guard, allow_redundancy, incremental=True, forced_r=None, forced_g=None):
    """MCLP optimizer. forced_r / forced_g are lists of station indices that must
    be included as Responders / Guardians regardless of coverage score."""
    forced_r = list(forced_r or [])
    forced_g = list(forced_g or [])
    n_stations, n_calls = resp_matrix.shape
    if n_calls == 0 or (num_resp == 0 and num_guard == 0): return [], [], [], []
    df_profiles = pd.DataFrame(resp_matrix.T).astype(int).astype(str)
    df_profiles['g'] = pd.DataFrame(guard_matrix.T).astype(int).astype(str).agg(''.join, axis=1)
    df_profiles['r'] = df_profiles.drop(columns='g').agg(''.join, axis=1)
    grouped = df_profiles.groupby(['r', 'g'], sort=False)
    weights = grouped.size().values
    unique_idx = grouped.head(1).index
    u_resp = resp_matrix[:, unique_idx]
    u_guard = guard_matrix[:, unique_idx]
    u_dist_r = dist_r[:, unique_idx]
    u_dist_g = dist_g[:, unique_idx]
    n_u = len(weights)

    def run_lp(target_r, target_g, locked_r, locked_g):
        model = pulp.LpProblem("DroneCoverage", pulp.LpMaximize)
        x_r = pulp.LpVariable.dicts("r_st", range(n_stations), 0, 1, pulp.LpBinary)
        x_g = pulp.LpVariable.dicts("g_st", range(n_stations), 0, 1, pulp.LpBinary)
        model += pulp.lpSum(x_r[i] for i in range(n_stations)) == target_r
        model += pulp.lpSum(x_g[i] for i in range(n_stations)) == target_g
        for r in locked_r: model += x_r[r] == 1
        for g in locked_g: model += x_g[g] == 1
        if not allow_redundancy:
            for s in range(n_stations): model += x_r[s] + x_g[s] <= 1
        y = pulp.LpVariable.dicts("cl", range(n_u), 0, 1, pulp.LpBinary)
        penalty = 0.00001
        model += pulp.lpSum(y[i]*weights[i] for i in range(n_u)) - pulp.lpSum(
            x_r[s]*np.sum(u_dist_r[s,:])*penalty + x_g[s]*np.sum(u_dist_g[s,:])*penalty
            for s in range(n_stations))
        for i in range(n_u):
            cover = [x_r[s] for s in range(n_stations) if u_resp[s,i]] + [x_g[s] for s in range(n_stations) if u_guard[s,i]]
            if cover: model += y[i] <= pulp.lpSum(cover)
            else: model += y[i] == 0
        model.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=10, gapRel=0.0))
        return (
            [i for i in range(n_stations) if (pulp.value(x_r[i]) or 0) > 0.5],
            [i for i in range(n_stations) if (pulp.value(x_g[i]) or 0) > 0.5]
        )

    if not incremental:
        res_r, res_g = run_lp(num_resp, num_guard, forced_r, forced_g)
        return res_r, res_g, res_r, res_g
    # Start with forced pins already locked in
    curr_r, curr_g = list(forced_r), list(forced_g)
    chrono_r, chrono_g = list(forced_r), list(forced_g)
    # Add remaining Guardians one at a time (incremental)
    for tg in range(len(forced_g) + 1, num_guard + 1):
        next_r, next_g = run_lp(0, tg, curr_r, curr_g)
        chrono_g.extend([x for x in next_g if x not in curr_g])
        curr_r, curr_g = next_r, next_g
    # Add remaining Responders one at a time
    for tr in range(len(forced_r) + 1, num_resp + 1):
        next_r, next_g = run_lp(tr, num_guard, curr_r, curr_g)
        chrono_r.extend([x for x in next_r if x not in curr_r])
        curr_r, curr_g = next_r, next_g
    return curr_r, curr_g, chrono_r, chrono_g

@st.cache_resource
def compute_all_elbow_curves(n_calls, _resp_matrix, _guard_matrix, _geos_r, _geos_g, total_area, _bounds_hash, max_stations=100):
    n_st_calls = min(_resp_matrix.shape[0], max_stations)
    n_st_area_r = min(len(_geos_r), 25)
    n_st_area_g = min(len(_geos_g), 25)

    def greedy_calls(matrix):
        uncovered = np.ones(n_calls, dtype=bool)
        curve = [0.0]
        cov_count = 0
        import heapq as hq
        pq = [(-matrix[i].sum(), i) for i in range(n_st_calls)]
        hq.heapify(pq)
        for _ in range(n_st_calls):
            if not pq: break
            best_s, best_cov = -1, -1
            while pq:
                neg_gain, idx = hq.heappop(pq)
                actual_gain = (matrix[idx] & uncovered).sum()
                if not pq or actual_gain >= -pq[0][0]:
                    best_s, best_cov = idx, actual_gain
                    break
                else:
                    hq.heappush(pq, (-actual_gain, idx))
            if best_s != -1 and best_cov > 0:
                uncovered = uncovered & ~matrix[best_s]
                cov_count += best_cov
                curve.append((cov_count / max(1, n_calls)) * 100)
                if cov_count == n_calls: break
            else:
                break
        return curve

    def greedy_area(geos, limit):
        if total_area <= 0 or limit <= 0: return [0.0]
        current_union = Polygon()
        curve = [0.0]
        import heapq as hq
        geos_sub = geos[:limit]
        
        pq = [(-geos_sub[i].area, i) for i in range(len(geos_sub))]
        hq.heapify(pq)
        
        for _ in range(len(geos_sub)):
            if not pq: break
            best_s, best_gain = -1, -1
            
            while pq:
                neg_gain, idx = hq.heappop(pq)
                try:
                    actual_gain = current_union.union(geos_sub[idx]).area - current_union.area
                except Exception:
                    actual_gain = 0
                    
                if not pq or actual_gain >= -pq[0][0]:
                    best_s, best_gain = idx, actual_gain
                    break
                else:
                    hq.heappush(pq, (-actual_gain, idx))
                    
            if best_s != -1 and best_gain > 0:
                try:
                    current_union = current_union.union(geos_sub[best_s])
                    curve.append((current_union.area / total_area) * 100)
                except Exception:
                    pass
            else:
                break
        return curve

    with ThreadPoolExecutor() as executor:
        f_cr = executor.submit(greedy_calls, _resp_matrix[:n_st_calls])
        f_cg = executor.submit(greedy_calls, _guard_matrix[:n_st_calls])
        f_ar = executor.submit(greedy_area, _geos_r, n_st_area_r)
        f_ag = executor.submit(greedy_area, _geos_g, n_st_area_g)
        c_r, c_g, a_r, a_g = f_cr.result(), f_cg.result(), f_ar.result(), f_ag.result()

    max_len = max(len(c_r), len(c_g), len(a_r), len(a_g))
    def pad(c):
        r = list(c)
        while len(r) < max_len: r.append(np.nan)
        return r
    return pd.DataFrame({
        'Drones': range(max_len),
        'Responder (Calls)': pad(c_r),
        'Responder (Area)':  pad(a_r),
        'Guardian (Calls)':  pad(c_g),
        'Guardian (Area)':   pad(a_g)
    })

# ============================================================
# APP FLOW 
# ============================================================

if not st.session_state['csvs_ready']:

    # GRAB THE LOGO FOR THE UPLOAD PAGE
    logo_b64 = get_base64_of_bin_file("logo.png")
    hero_logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:48px; margin-bottom:15px; filter: brightness(0) invert(1);">' if logo_b64 else f'<div style="font-size:2.5rem; font-weight:900; letter-spacing:4px; color:#ffffff; margin-bottom:15px;">BRINC</div>'

    st.markdown(f"""
    <style>
    @keyframes pulseGlow {{
        0%, 100% {{ opacity: 0.55; }}
        50%       {{ opacity: 1.0; }}
    }}
    @keyframes fadeUp {{
        from {{ opacity:0; transform:translateY(14px); }}
        to   {{ opacity:1; transform:translateY(0); }}
    }}
    .brinc-hero {{
        position: relative;
        text-align: center;
        padding: 52px 24px 40px;
        margin-bottom: 36px;
        border-radius: 12px;
        background: radial-gradient(ellipse at 50% 0%,
            rgba(0,210,255,0.13) 0%, rgba(0,0,0,0) 68%);
        border-bottom: 1px solid rgba(0,210,255,0.15);
        overflow: hidden;
        animation: fadeUp 0.5s ease both;
    }}
    .brinc-hero::before {{
        content: '';
        position: absolute; inset: 0;
        background:
            repeating-linear-gradient(0deg,
                transparent, transparent 39px,
                rgba(0,210,255,0.025) 39px,
                rgba(0,210,255,0.025) 40px),
            repeating-linear-gradient(90deg,
                transparent, transparent 79px,
                rgba(0,210,255,0.025) 79px,
                rgba(0,210,255,0.025) 80px);
        pointer-events: none;
    }}
    .brinc-eyebrow {{
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.62rem;
        font-weight: 700;
        letter-spacing: 4px;
        color: {accent_color};
        text-transform: uppercase;
        opacity: 0.7;
        margin-bottom: 12px;
    }}
    .brinc-h1 {{
        font-family: 'Manrope', sans-serif;
        font-size: clamp(2rem, 4vw, 3rem);
        font-weight: 900;
        color: #ffffff;
        letter-spacing: -0.5px;
        line-height: 1.08;
        margin-bottom: 12px;
    }}
    .brinc-h1 em {{
        font-style: normal;
        color: {accent_color};
    }}
    .brinc-tagline {{
        font-size: 0.88rem;
        color: #666;
        max-width: 500px;
        margin: 0 auto 22px;
        line-height: 1.65;
    }}
    .brinc-badges {{
        display: flex;
        flex-wrap: wrap;
        justify-content: center;
        gap: 8px;
        margin-top: 4px;
    }}
    .brinc-badge {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(0,210,255,0.07);
        border: 1px solid rgba(0,210,255,0.2);
        border-radius: 100px;
        padding: 4px 13px;
        font-size: 0.64rem;
        font-weight: 700;
        color: {accent_color};
        letter-spacing: 0.8px;
        text-transform: uppercase;
    }}
    .brinc-badge.pulse {{
        animation: pulseGlow 3s ease-in-out infinite;
    }}
    .path-card {{
        background: #080808;
        border: 1px solid #1c1c1c;
        border-radius: 10px;
        padding: 22px 18px 16px;
        position: relative;
        overflow: hidden;
        transition: border-color 0.2s ease, box-shadow 0.2s ease;
    }}
    .path-card::after {{
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: var(--accent);
        border-radius: 10px 10px 0 0;
    }}
    .path-card:hover {{
        border-color: rgba(255,255,255,0.12);
        box-shadow: 0 0 28px rgba(0,210,255,0.05);
    }}
    .pc-icon  {{ font-size: 1.5rem; display:block; margin-bottom:9px; }}
    .pc-tag   {{ font-size:0.55rem; font-weight:800; letter-spacing:2.5px;
                 text-transform:uppercase; color:var(--accent); margin-bottom:5px; }}
    .pc-title {{ font-size:1rem; font-weight:800; color:#fff;
                 line-height:1.25; margin-bottom:7px; }}
    .pc-desc  {{ font-size:0.7rem; color:#555; line-height:1.6; margin-bottom:0; }}
    .field-footnote {{
        font-size: 0.63rem; color: #3a3a3a; line-height: 1.75;
        margin-top: 10px; border-top: 1px solid #141414;
        padding-top: 10px;
    }}
    .demo-cities {{
        font-size: 0.65rem; color: #444; line-height: 1.9;
        margin-top: 10px;
    }}
    .demo-cities b {{ color: #555; }}
    .demo-check {{
        font-size: 0.63rem; color: #333; line-height: 1.8;
        margin-top: 12px; border-top: 1px solid #141414;
        padding-top: 10px;
    }}
    .demo-check span {{ color: {accent_color}; margin-right: 5px; }}
    </style>

    <div class="brinc-hero">
        {hero_logo_html}
        <div class="brinc-eyebrow">BRINC Drones · DFR Platform</div>
        <div class="brinc-h1">
            Coverage. Operations.<br><em>Savings.</em>
        </div>
        <div class="brinc-tagline">
            Optimize drone-as-first-responder deployments for any US jurisdiction.
            Model coverage, forecast ROI, and generate grant-ready proposals in minutes.
        </div>
        <div class="brinc-badges">
            <div class="brinc-badge pulse">🛰 3D Swarm Simulation</div>
            <div class="brinc-badge">🗺 Census Boundaries</div>
            <div class="brinc-badge">📄 Grant Narrative Export</div>
            <div class="brinc-badge">✈️ FAA LAANC Overlay</div>
            <div class="brinc-badge">⚡ MCLP Optimizer</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    path_sim_col, path_upload_col, path_demo_col = st.columns(3, gap="medium")

    with path_sim_col:
        st.markdown(f"""
        <div class="path-card" style="--accent:{accent_color};">
            <span class="pc-icon">🗺</span>
            <div class="pc-tag">Path 01</div>
            <div class="pc-title">Simulate Any<br>US Region</div>
            <div class="pc-desc">No data needed. Real Census boundaries + realistic 911 call distribution generated automatically. Stack multiple jurisdictions in one run.</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        for i in range(st.session_state.city_count):
            c1, c2 = st.columns([3, 1])
            c_val = st.session_state['target_cities'][i]['city'] if i < len(st.session_state['target_cities']) else ""
            s_val = st.session_state['target_cities'][i]['state'] if i < len(st.session_state['target_cities']) else "FL"
            c_name = c1.text_input(
                f"City or County {i+1}", value=c_val, key=f"c_{i}",
                placeholder="e.g. Orlando OR Orange County",
                help="Official municipality or county name."
            )
            state_idx = list(STATE_FIPS.keys()).index(s_val) if s_val in STATE_FIPS else 8
            s_name = c2.selectbox(
                f"State {i+1}", list(STATE_FIPS.keys()), index=state_idx,
                key=f"s_{i}",
                label_visibility="collapsed" if i > 0 else "visible"
            )
            if i < len(st.session_state['target_cities']):
                st.session_state['target_cities'][i] = {"city": c_name, "state": s_name}
            else:
                st.session_state['target_cities'].append({"city": c_name, "state": s_name})

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        st.file_uploader(
            "Optional: Custom stations (CSV or Excel)",
            type=['csv', 'xlsx', 'xls', 'xlsm', 'xlsb'],
            key="sim_station_uploader",
            help="Include 'lat'/'lon' OR 'address' columns. Max ~20 locations for auto-geocoding."
        )

        col_add, col_run = st.columns([1, 1])
        if st.session_state.city_count < 10:
            if col_add.button("＋ City", use_container_width=True, key="add_city_btn"):
                st.session_state.city_count += 1
                st.rerun()
        submit_demo = col_run.button("▶ Run", use_container_width=True, key="run_sim_btn",
                                     help="Fetch boundaries and launch the simulation.")

    with path_upload_col:
        st.markdown(f"""
        <div class="path-card" style="--accent:#39FF14;">
            <span class="pc-icon">📂</span>
            <div class="pc-tag">Path 02</div>
            <div class="pc-title">Upload CAD<br>or .brinc Save</div>
            <div class="pc-desc">
                Drop <b>any</b> CAD export CSV — no renaming needed.
                Or, drop a previously saved <b>.brinc</b> file to instantly restore your deployment.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        uploaded_files = st.file_uploader(
            "Drop your CAD export (+ optional stations CSV)",
            accept_multiple_files=True,
            type=['csv', 'xlsx', 'xls', 'xlsb', 'xlsm', 'brinc', 'json', 'txt'],
            label_visibility="collapsed",
            help="One file = raw CAD export. Two files = calls + stations. OR drop a .brinc file to restore a previous session."
        )

        st.markdown("""
        <div class="field-footnote">
            <b style='color:#555;'>1 file</b> — any CAD export (CSV or Excel); stations auto-built from OSM<br>
            <b style='color:#555;'>2+ files</b> — calls + stations (CSV or Excel, any column names)<br>
            <b style='color:#39FF14;'>.brinc file</b> — instantly restore a saved deployment<br>
            Max 25,000 calls · 100 stations
        </div>
        """, unsafe_allow_html=True)

        def _looks_like_stations(fname):
            n = fname.lower()
            return any(k in n for k in ['station','dept','agency','facility','police','fire','loc'])

        if uploaded_files and len(uploaded_files) >= 1:
            
            # --- 1. INTELLIGENTLY CHECK FOR .BRINC FILE ---
            # Browsers sometimes append .json to .brinc files on download
            brinc_file = None
            for f in uploaded_files:
                fname = f.name.lower()
                if '.brinc' in fname or fname.endswith('.json'):
                    try:
                        # Quick peek inside to see if it has our save data keys
                        f.seek(0)
                        peek = json.loads(f.getvalue().decode('utf-8'))
                        if 'k_resp' in peek and 'calls_data' in peek:
                            brinc_file = f
                            break
                    except:
                        pass

            if brinc_file:
                with st.spinner("💾 Restoring saved deployment..."):
                    try:
                        brinc_file.seek(0)
                        save_data = json.loads(brinc_file.getvalue().decode('utf-8'))
                        
                        st.session_state['active_city'] = save_data.get('city', 'Unknown')
                        st.session_state['active_state'] = save_data.get('state', 'US')
                        st.session_state['k_resp'] = save_data.get('k_resp', 2)
                        st.session_state['k_guard'] = save_data.get('k_guard', 0)
                        st.session_state['r_resp'] = save_data.get('r_resp', 2.0)
                        st.session_state['r_guard'] = save_data.get('r_guard', 8.0)
                        st.session_state['dfr_rate'] = save_data.get('dfr_rate', 25)
                        st.session_state['deflect_rate'] = save_data.get('deflect_rate', 30)
                        
                        if save_data.get('calls_data'):
                            df_c = pd.DataFrame(save_data['calls_data'])
                            # Safely cast to numeric so the map geometry doesn't crash
                            if 'lat' in df_c.columns: df_c['lat'] = pd.to_numeric(df_c['lat'], errors='coerce')
                            if 'lon' in df_c.columns: df_c['lon'] = pd.to_numeric(df_c['lon'], errors='coerce')
                            st.session_state['df_calls'] = df_c
                            st.session_state['df_calls_full'] = df_c.copy()
                            st.session_state['total_original_calls'] = len(df_c)
                            st.session_state['total_modeled_calls'] = len(df_c)
                        
                        if save_data.get('stations_data'):
                            df_s = pd.DataFrame(save_data['stations_data'])
                            if 'lat' in df_s.columns: df_s['lat'] = pd.to_numeric(df_s['lat'], errors='coerce')
                            if 'lon' in df_s.columns: df_s['lon'] = pd.to_numeric(df_s['lon'], errors='coerce')
                            st.session_state['df_stations'] = df_s
                            
                        st.session_state['data_source'] = 'brinc_file'
                        st.session_state['map_build_logged'] = False
                        st.session_state['csvs_ready'] = True
                        st.toast("✅ Deployment restored successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error loading .brinc file: {e}")
                        st.stop()

            else:
                # --- 2. OTHERWISE, PROCESS AS NORMAL CSV CAD DATA ---
                st.session_state['active_city'] = ""
                st.session_state['active_state'] = ""
                st.session_state['target_cities'] = []

                f_list = list(uploaded_files)
                call_files = []
                station_file = None

                for f in f_list:
                    if _looks_like_stations(f.name):
                        station_file = f
                    else:
                        call_files.append(f)
                
                if len(f_list) == 2 and not station_file:
                    f0, f1 = f_list
                    f0.seek(0); sz0 = len(f0.read()); f0.seek(0)
                    f1.seek(0); sz1 = len(f1.read()); f1.seek(0)
                    if sz0 >= sz1:
                        call_files = [f0]
                        station_file = f1
                    else:
                        call_files = [f1]
                        station_file = f0

                if call_files:
                    with st.spinner("🔍 Detecting column types in CAD export…"):
                        df_c = aggressive_parse_calls(call_files)

                    if df_c is None or df_c.empty:
                        st.error("❌ Calls file error: Could not parse valid coordinates.")
                        st.stop()

                    df_c_full = df_c.reset_index(drop=True).copy()
                    st.session_state['total_original_calls'] = len(df_c_full)

                    if len(df_c_full) > 25000:
                        df_c = df_c_full.sample(25000, random_state=42).reset_index(drop=True)
                        st.session_state['total_modeled_calls'] = len(df_c)
                        st.toast(f"⚠️ Optimization modeled with {len(df_c):,} representative calls out of {len(df_c_full):,} total incidents.")
                    else:
                        df_c = df_c_full.copy()
                        st.session_state['total_modeled_calls'] = len(df_c)

                    if station_file is not None:
                        with st.spinner("🔍 Reading stations file…"):
                            try:
                                sfname = station_file.name.lower()
                                if sfname.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
                                    engine = 'xlrd' if sfname.endswith('.xls') else 'pyxlsb' if sfname.endswith('.xlsb') else 'openpyxl'
                                    df_s = pd.read_excel(io.BytesIO(station_file.getvalue()), engine=engine)
                                else:
                                    df_s = pd.read_csv(station_file)
                                df_s.columns = [str(c).lower().strip() for c in df_s.columns]
                                if 'latitude' in df_s.columns: df_s = df_s.rename(columns={'latitude':'lat'})
                                if 'longitude' in df_s.columns: df_s = df_s.rename(columns={'longitude':'lon'})
                                if 'station_name' in df_s.columns: df_s = df_s.rename(columns={'station_name':'name'})
                                if 'station_type' in df_s.columns: df_s = df_s.rename(columns={'station_type':'type'})
                                
                                if 'lat' in df_s.columns and 'lon' in df_s.columns:
                                    df_s['lat'] = pd.to_numeric(df_s['lat'], errors='coerce')
                                    df_s['lon'] = pd.to_numeric(df_s['lon'], errors='coerce')
                                else:
                                    raise ValueError("Could not find lat/lon columns.")

                                if 'name' not in df_s.columns: 
                                    df_s['name'] = [f"Site {i+1}" for i in range(len(df_s))]
                                else:
                                    df_s['name'] = df_s['name'].fillna('').astype(str).str.strip()
                                    df_s['name'] = df_s['name'].replace(r'(?i)^(null|<null>|nan|none)$', '', regex=True)
                                    df_s['name'] = [n if n else f"Site {i+1}" for i, n in enumerate(df_s['name'])]

                                counts = {}
                                new_names = []
                                for n in df_s['name']:
                                    if n in counts:
                                        counts[n] += 1
                                        new_names.append(f"{n} ({counts[n]})")
                                    else:
                                        counts[n] = 0
                                        new_names.append(n)
                                df_s['name'] = new_names

                                if 'type' not in df_s.columns: df_s['type'] = 'Police'
                                df_s = df_s.dropna(subset=['lat', 'lon']).reset_index(drop=True)
                                osm_note = "Loaded stations from file."
                            except Exception as e:
                                df_s, osm_note = None, f"Failed: {e}"
                        if df_s is None or df_s.empty:
                            st.error(f"❌ Stations file error: {osm_note}")
                            st.stop()
                    else:
                        with st.spinner("🌐 No stations file detected — querying OpenStreetMap for police, fire & schools…"):
                            df_s, osm_note = generate_stations_from_calls(df_c)
                        if df_s is None or df_s.empty:
                            # Final safety net: scatter stations across call bounding box
                            df_s = _make_random_stations(df_c, n=40)
                            osm_note = "⚠️ Could not reach any map source — using estimated station positions from call data."
                            st.warning(osm_note)
                        else:
                            st.toast(f"✅ {osm_note}")

                    if len(df_s) > 100:
                        df_s = df_s.sample(100, random_state=42).reset_index(drop=True)

                    detected_city = None
                    detected_state = None

                    with st.spinner(get_jurisdiction_message()):
                        # Priority 1: city/state extracted directly from the CAD export
                        if '_csv_city' in df_c.columns:
                            city_val = str(df_c['_csv_city'].iloc[0]).strip().title()
                            if city_val and city_val.lower() not in ('nan', 'none', ''):
                                detected_city = city_val

                        if '_csv_state' in df_c.columns:
                            state_val = str(df_c['_csv_state'].iloc[0]).strip().upper()
                            if state_val in STATE_FIPS:
                                detected_state = state_val
                            elif state_val.title() in US_STATES_ABBR:
                                detected_state = US_STATES_ABBR[state_val.title()]

                        # If the export gives us a city but not a state, forward-geocode the city name.
                        if detected_city and not detected_state:
                            try:
                                geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(detected_city)}&limit=1&countrycodes=us"
                                req_geo = urllib.request.Request(geo_url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
                                with urllib.request.urlopen(req_geo, timeout=8) as resp_geo:
                                    geo_result = json.loads(resp_geo.read().decode('utf-8'))
                                if geo_result:
                                    display_name = geo_result[0].get('display_name', '')
                                    parts = [p.strip() for p in display_name.split(',')]
                                    state_full = parts[2] if len(parts) >= 3 else ''
                                    if state_full in US_STATES_ABBR:
                                        detected_state = US_STATES_ABBR[state_full]
                            except Exception:
                                pass

                        # Priority 2: reverse-geocode the centroid of the calls, not the first row.
                        if not detected_city or not detected_state:
                            try:
                                cen_lat = float(df_c['lat'].median())
                                cen_lon = float(df_c['lon'].median())
                                detected_state_full, detected_city_rg = reverse_geocode_state(cen_lat, cen_lon)
                                if detected_state_full and detected_state_full in US_STATES_ABBR:
                                    if not detected_state:
                                        detected_state = US_STATES_ABBR[detected_state_full]
                                    if not detected_city and detected_city_rg and detected_city_rg != 'Unknown City':
                                        detected_city = detected_city_rg
                            except Exception:
                                pass

                        if detected_city and detected_state:
                            st.session_state['active_city'] = detected_city
                            st.session_state['active_state'] = detected_state
                            st.session_state['target_cities'] = [{"city": detected_city, "state": detected_state}]
                            st.toast(f"📍 Detected: {detected_city}, {detected_state}")

                    # Only clip calls to the station bbox once we know the correct jurisdiction.
                    if detected_city and detected_state:
                        lat_min, lat_max = df_s['lat'].min(), df_s['lat'].max()
                        lon_min, lon_max = df_s['lon'].min(), df_s['lon'].max()
                        clip_mask_modeled = (
                            (df_c['lat'] >= lat_min - 0.5) & (df_c['lat'] <= lat_max + 0.5) &
                            (df_c['lon'] >= lon_min - 0.5) & (df_c['lon'] <= lon_max + 0.5)
                        )
                        df_c = df_c[clip_mask_modeled].reset_index(drop=True)
                        clip_mask_full = (
                            (df_c_full['lat'] >= lat_min - 0.5) & (df_c_full['lat'] <= lat_max + 0.5) &
                            (df_c_full['lon'] >= lon_min - 0.5) & (df_c_full['lon'] <= lon_max + 0.5)
                        )
                        df_c_full = df_c_full[clip_mask_full].reset_index(drop=True)

                    st.session_state['df_calls']             = df_c
                    st.session_state['df_calls_full']        = df_c_full
                    st.session_state['df_stations']          = df_s
                    st.session_state['total_original_calls'] = len(df_c_full)
                    st.session_state['total_modeled_calls']  = len(df_c)

                    with st.spinner(get_jurisdiction_message()):
                        # ── BOUNDARY LOOKUP: fetch & save shapefile NOW so
                        # find_relevant_jurisdictions() can use it on the map page ──
                        detected_city_for_boundary = st.session_state.get('active_city', '')
                        detected_state_for_boundary = st.session_state.get('active_state', '')
                        if detected_city_for_boundary and detected_state_for_boundary and detected_state_for_boundary in STATE_FIPS:
                            # Boundary lookup priority:
                            # 1. places_lite.parquet  — exact city shape (best)
                            # 2. counties_lite.parquet — same-name county
                            # 3. Geocode → find county → counties_lite.parquet
                            b_success, b_gdf = fetch_place_boundary_local(
                                detected_state_for_boundary, detected_city_for_boundary)
                            if not b_success:
                                for name_try in [detected_city_for_boundary,
                                                 detected_city_for_boundary + " County"]:
                                    b_success, b_gdf = fetch_county_boundary_local(
                                        detected_state_for_boundary, name_try)
                                    if b_success and b_gdf is not None:
                                        break
                            if not b_success:
                                county_name = lookup_county_for_city(
                                    detected_city_for_boundary, detected_state_for_boundary)
                                if county_name:
                                    b_success, b_gdf = fetch_county_boundary_local(
                                        detected_state_for_boundary, county_name)
                                    if b_success and b_gdf is not None:
                                        b_gdf = b_gdf.copy()
                                        b_gdf['NAME'] = detected_city_for_boundary + " (" + county_name + " County)"
                            if b_success and b_gdf is not None:
                                try:
                                    safe_n = detected_city_for_boundary.replace(" ", "_").replace("/", "_")
                                    b_gdf.to_file(os.path.join(
                                        SHAPEFILE_DIR,
                                        f"{safe_n}_{detected_state_for_boundary}.shp"))
                                except Exception:
                                    pass

                    st.session_state['data_source'] = 'cad_upload'
                    st.session_state['map_build_logged'] = False
                    st.session_state['csvs_ready'] = True
                    st.rerun()

    with path_demo_col:
        st.markdown(f"""
        <div class="path-card" style="--accent:#FFD700;">
            <span class="pc-icon">⚡</span>
            <div class="pc-tag">Path 03</div>
            <div class="pc-title">1-Click Demo<br>Large US City</div>
            <div class="pc-desc">Instantly spin up a fully pre-configured scenario for a major US city. Ideal for live stakeholder presentations and platform walkthroughs.</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        if st.button("⚡ Launch Random Demo City", use_container_width=True, key="demo_btn"):
            random.seed(datetime.datetime.now().microsecond + os.getpid())
            already_used = st.session_state.get('_last_demo_city', '')
            candidates = [c for c in FAST_DEMO_CITIES if c[0] != already_used]
            rcity, rstate = random.choice(candidates)
            st.session_state['_last_demo_city'] = rcity
            st.session_state['target_cities'] = [{"city": rcity, "state": rstate}]
            st.session_state.city_count = 1
            for i in range(10):
                st.session_state.pop(f"c_{i}", None)
                st.session_state.pop(f"s_{i}", None)
            st.session_state['trigger_sim'] = True
            st.rerun()

        city_chips = "  ·  ".join([f"{c}" for c, _ in DEMO_CITIES[:12]]) + "  · and more…"
        st.markdown(f"""
        <div class="demo-cities">
            <b>Available Cities</b><br>
            {city_chips}
        </div>
        <div class="demo-check">
            <span>✓</span>Real Census boundaries<br>
            <span>✓</span>Clustered 911 simulation<br>
            <span>✓</span>100 station candidates<br>
            <span>✓</span>Full optimization & export
        </div>
        """, unsafe_allow_html=True)

    st.markdown(f"""
    <div style="text-align:center; margin-top:8px; font-size:0.63rem; color:#2a2a2a;">
        BRINC Drones, Inc. · <a href="https://brincdrones.com" target="_blank"
        style="color:#333; text-decoration:none;">brincdrones.com</a>
        · All coverage estimates are for planning purposes only.
    </div>
    """, unsafe_allow_html=True)

    if submit_demo or st.session_state.get('trigger_sim', False):
        if st.session_state.get('trigger_sim', False):
            st.session_state['trigger_sim'] = False

        active_targets = [loc for loc in st.session_state['target_cities'] if loc['city'].strip()]
        if not active_targets:
            st.error("Please enter at least one valid city name.")
            st.stop()

        if len(active_targets) == 1:
            st.session_state['active_city']  = active_targets[0]['city']
            st.session_state['active_state'] = active_targets[0]['state']
        else:
            st.session_state['active_city']  = f"{active_targets[0]['city']} & {len(active_targets)-1} others"
            st.session_state['active_state'] = active_targets[0]['state']

        prog = st.progress(0, text="🫡 Preparing tools worthy of those who serve…")
        all_gdfs = []
        total_estimated_pop = 0

        for i, loc in enumerate(active_targets):
            c_name = loc['city'].strip()
            s_name = loc['state']
            is_county = c_name.lower().endswith(" county")
            
            prog.progress(10 + int((i / len(active_targets)) * 20),
                          text=f"🗺️ Mapping {c_name}, {s_name} — because every block they patrol matters…")
            
            # Boundary lookup priority:
            # 1. places_lite.parquet  — exact city/town shape (best)
            # 2. counties_lite.parquet — county that shares the city name
            # 3. Geocode city → find its county → counties_lite.parquet
            success, temp_gdf = fetch_place_boundary_local(s_name, c_name)
            if not success:
                # Direct county name match (e.g. "Gilmer County" or "Gilmer")
                success, temp_gdf = fetch_county_boundary_local(s_name, c_name)
            if not success:
                success, temp_gdf = fetch_county_boundary_local(s_name, c_name + " County")
            if not success:
                # City doesn't share its county's name — geocode to find the county
                county_name = lookup_county_for_city(c_name, s_name)
                if county_name:
                    success, temp_gdf = fetch_county_boundary_local(s_name, county_name)
                    if success and temp_gdf is not None:
                        temp_gdf = temp_gdf.copy()
                        temp_gdf['NAME'] = c_name + " (" + county_name + " County)"
            if success:
                is_county = True

            # County boundaries come from the parquet, not from TIGER, so they are
            # never written to SHAPEFILE_DIR. find_relevant_jurisdictions() only scans
            # that directory, so without this save it always falls back to
            # "Auto-Generated Boundary". Save any successfully loaded county GDF now.
            if success and is_county and temp_gdf is not None:
                try:
                    safe_name = c_name.replace(" ", "_").replace("/", "_")
                    county_shp_path = os.path.join(SHAPEFILE_DIR, f"{safe_name}_{s_name}.shp")
                    temp_gdf.to_file(county_shp_path)
                except Exception:
                    pass  # If save fails, fall back gracefully

            if success:
                all_gdfs.append(temp_gdf)
                pop = fetch_census_population(STATE_FIPS[s_name], c_name, is_county=is_county)
                if pop:
                    total_estimated_pop += pop
                    st.toast(f"✅ {c_name} population verified: {pop:,}")
                else:
                    gdf_proj   = temp_gdf.to_crs(epsg=3857)
                    area_sq_mi = gdf_proj.geometry.area.sum() / 2589988.11
                    est = KNOWN_POPULATIONS.get(c_name, int(area_sq_mi * 3500))
                    total_estimated_pop += est
                    st.toast(f"⚠️ {c_name} population estimated: {est:,}")
            else:
                st.warning(f"⚠️ Could not find a boundary for {c_name}, {s_name}. Try another city.")
                if st.session_state.get('_last_demo_city') == c_name:
                    random.seed(datetime.datetime.now().microsecond + os.getpid())
                    candidates = [c for c in DEMO_CITIES if c[0] != c_name]
                    rcity, rstate = random.choice(candidates)
                    st.session_state['_last_demo_city'] = rcity
                    st.session_state['target_cities'] = [{"city": rcity, "state": rstate}]
                    for j in range(10):
                        st.session_state.pop(f"c_{j}", None)
                        st.session_state.pop(f"s_{j}", None)
                    st.rerun()

        if not all_gdfs:
            prog.empty()
            st.error("❌ Could not find Census boundaries for any of the entered locations. Check spelling.")
            st.stop()

        prog.progress(35, text="💙 Boundaries loaded — honoring the officers who know every street…")
        active_city_gdf = pd.concat(all_gdfs, ignore_index=True)
        city_poly = active_city_gdf.geometry.union_all()
        st.session_state['estimated_pop'] = total_estimated_pop

        annual_cfs = int(total_estimated_pop * 0.6)
        st.session_state['total_original_calls'] = annual_cfs
        simulated_points_count = min(int(annual_cfs / 12), 25000)

        prog.progress(55, text="🚔 Modeling 911 calls — every one represents someone who needed help…")
        np.random.seed(42)
        call_points = generate_clustered_calls(city_poly, simulated_points_count)
        
        base_date = datetime.datetime.now() - datetime.timedelta(days=30)
        fake_dts = [(base_date + datetime.timedelta(days=random.randint(0, 30), hours=random.randint(0, 23), minutes=random.randint(0, 59))) for _ in range(simulated_points_count)]
        
        df_demo = pd.DataFrame({
            'lat':      [p[0] for p in call_points],
            'lon':      [p[1] for p in call_points],
            'priority': np.random.choice([1, 2, 3], simulated_points_count, p=[0.15, 0.35, 0.50]),
            'date':     [d.strftime('%Y-%m-%d') for d in fake_dts],
            'time':     [d.strftime('%H:%M:%S') for d in fake_dts]
        })
        st.session_state['df_calls'] = df_demo
        st.session_state['df_calls_full'] = df_demo.copy()
        st.session_state['total_modeled_calls'] = len(df_demo)

        # --- PROCESS OPTIONAL CUSTOM STATIONS ---
        custom_stations_used = False
        sim_uploader = st.session_state.get('sim_station_uploader')
        
        if sim_uploader is not None:
            prog.progress(80, text="🏅 Geocoding custom stations from CSV…")
            import time
            try:
                sim_uploader.seek(0)
                s_df = pd.read_csv(sim_uploader)
                s_df.columns = [str(c).lower().strip() for c in s_df.columns]
                
                # Detect what columns the user provided
                lat_col = next((c for c in s_df.columns if c in ['lat', 'latitude', 'y']), None)
                lon_col = next((c for c in s_df.columns if c in ['lon', 'long', 'longitude', 'x']), None)
                addr_col = next((c for c in s_df.columns if any(a in c for a in ['address', 'street', 'location'])), None)
                name_col = next((c for c in s_df.columns if any(n in c for n in ['name', 'station', 'facility', 'dept'])), None)
                type_col = next((c for c in s_df.columns if any(t in c for t in ['type', 'category'])), None)
                
                parsed_stations = []
                for idx, row in s_df.iterrows():
                    s_name = str(row[name_col]) if name_col and pd.notna(row[name_col]) else f"Custom Station {idx+1}"
                    s_type = str(row[type_col]) if type_col and pd.notna(row[type_col]) else 'Custom'
                    s_lat, s_lon = None, None
                    
                    if lat_col and lon_col and pd.notna(row[lat_col]) and pd.notna(row[lon_col]):
                        s_lat, s_lon = float(row[lat_col]), float(row[lon_col])
                    elif addr_col and pd.notna(row[addr_col]):
                        addr_str = str(row[addr_col])
                        # Attempt geocoding
                        s_lat, s_lon = forward_geocode(addr_str)
                        if s_lat is None:
                            # Fallback: Try appending the city and state to the address string
                            s_lat, s_lon = forward_geocode(f"{addr_str}, {active_targets[0]['city']}, {active_targets[0]['state']}")
                        if s_lat is None:
                            st.toast(f"⚠️ Could not geocode: {addr_str}")
                        time.sleep(1) # Slow down requests slightly to prevent API blocking
                        
                    if s_lat and s_lon:
                        parsed_stations.append({
                            'name': s_name,
                            'lat': s_lat,
                            'lon': s_lon,
                            'type': s_type
                        })
                        
                if parsed_stations:
                    st.session_state['df_stations'] = pd.DataFrame(parsed_stations)
                    custom_stations_used = True
                else:
                    st.warning("⚠️ Could not geocode or parse your custom stations. Falling back to 100 random stations.")
            except Exception as e:
                st.warning(f"⚠️ Error reading custom stations: {e}. Falling back to random stations.")

        # --- FALLBACK: PULL REAL STATIONS FROM OPENSTREETMAP ---
        if not custom_stations_used:
            prog.progress(80, text="🌐 Querying OpenStreetMap for real police, fire & schools…")
            
            # Use the simulated calls we just made to find the bounding box, and pull real OSM data!
            df_s, osm_note = generate_stations_from_calls(st.session_state['df_calls'])
            
            if df_s is not None and not df_s.empty:
                st.session_state['df_stations'] = df_s
                st.toast(f"✅ {osm_note}")
            else:
                # Absolute worst-case scenario (no internet or OSM API is down): fall back to random
                st.warning("⚠️ Could not reach OpenStreetMap for real stations. Falling back to random placements.")
                station_points = generate_random_points_in_polygon(city_poly, 100)
                types = ['Police', 'Fire', 'EMS'] * 34
                st.session_state['df_stations'] = pd.DataFrame({
                    'name': [f'Station {i+1}' for i in range(len(station_points))],
                    'lat':  [p[0] for p in station_points],
                    'lon':  [p[1] for p in station_points],
                    'type': types[:len(station_points)]
                })

        prog.progress(100, text="✅ Ready — built for the communities they protect and serve.")
        st.session_state['inferred_daily_calls_override'] = int(annual_cfs / 365)
        st.session_state['data_source'] = 'simulation'
        st.session_state['map_build_logged'] = False
        st.session_state['csvs_ready'] = True
        st.rerun()

# ============================================================
# MAIN MAP INTERFACE
# ============================================================
if st.session_state['csvs_ready']:
    components.html("<script>window._brincHasData = true;</script>", height=0)

    df_calls = st.session_state['df_calls'].copy()
    df_calls_full = st.session_state.get('df_calls_full')
    if df_calls_full is None:
        df_calls_full = df_calls.copy()
    else:
        df_calls_full = df_calls_full.copy()
    df_stations_all = st.session_state['df_stations'].copy()
    full_total_calls = int(st.session_state.get('total_original_calls', len(df_calls_full) if df_calls_full is not None else len(df_calls)) or 0)
    full_daily_calls = max(1, int(full_total_calls / 365)) if full_total_calls else 1

    # ── MAP BUILD EVENT: log to sheets once per session ──────────────────────
    if not st.session_state.get('map_build_logged', False):
        try:
            _map_city  = st.session_state.get('active_city', '')
            _map_state = st.session_state.get('active_state', '')
            _brinc_raw = st.session_state.get('brinc_user', 'steven.beltran').strip()
            if not _brinc_raw: _brinc_raw = 'steven.beltran'
            _map_name  = " ".join([w.capitalize() for w in _brinc_raw.split('.')])
            _map_email = f"{_brinc_raw}@brincdrones.com"
            _map_pop   = st.session_state.get('estimated_pop', 0)
            _map_calls = st.session_state.get('total_original_calls', 0)
            _map_daily = max(1, int(_map_calls / 365))
            _session_start = st.session_state.get('session_start', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            try:
                _start_dt = datetime.datetime.strptime(_session_start, '%Y-%m-%d %H:%M:%S')
                _dur_min  = round((datetime.datetime.now() - _start_dt).total_seconds() / 60, 1)
            except Exception:
                _dur_min = ''
            _map_details = {
                'session_id':       st.session_state.get('session_id', ''),
                'session_start':    _session_start,
                'session_duration_min': _dur_min,
                'data_source':      st.session_state.get('data_source', 'unknown'),
                'population':       _map_pop,
                'total_calls':      _map_calls,
                'daily_calls':      _map_daily,
                'area_sq_mi':       0,
                'fleet_capex':      0,
                'annual_savings':   0,
                'break_even':       'N/A',
                'opt_strategy':     '',
                'dfr_rate':         st.session_state.get('dfr_rate', 0),
                'deflect_rate':     st.session_state.get('deflect_rate', 0),
                'incremental_build': False,
                'allow_redundancy': False,
                'avg_response_min': 0,
                'avg_time_saved_min': 0,
                'area_covered_pct': 0,
                'pd_chief':         st.session_state.get('pd_chief_name', ''),
                'pd_dept':          st.session_state.get('pd_dept_name', ''),
                'pd_dept_email':    st.session_state.get('pd_dept_email', ''),
                'pd_dept_phone':    st.session_state.get('pd_dept_phone', ''),
                'active_drones':    [],
            }
            _log_to_sheets(_map_city, _map_state, 'MAP_BUILD', 0, 0, 0.0,
                           _map_name, _map_email, _map_details)
            st.session_state['map_build_logged'] = True
        except Exception:
            pass

    with st.spinner(get_jurisdiction_message()):
        master_gdf = find_relevant_jurisdictions(df_calls, df_stations_all, SHAPEFILE_DIR)

    if master_gdf is None or master_gdf.empty:
        # ── Fallback 1: load any saved shapefile directly (spatial join may have
        #    failed if coordinate conversion was imperfect, but the shapefile exists) ──
        shp_files = glob.glob(os.path.join(SHAPEFILE_DIR, "*.shp"))
        if shp_files:
            try:
                # Pick the shapefile whose name best matches active_city
                active_city_key = st.session_state.get('active_city', '').replace(' ', '_').lower()
                best = None
                for sf in shp_files:
                    if active_city_key and active_city_key in os.path.basename(sf).lower():
                        best = sf
                        break
                if best is None:
                    best = shp_files[0]  # just use the first one
                fallback_gdf = gpd.read_file(best)
                if fallback_gdf.crs is None:
                    fallback_gdf = fallback_gdf.set_crs(epsg=4269)
                fallback_gdf = fallback_gdf.to_crs(epsg=4326)
                name_col = next((c for c in ['NAME', 'DISTRICT', 'NAMELSAD'] if c in fallback_gdf.columns), fallback_gdf.columns[0])
                fallback_gdf['DISPLAY_NAME'] = fallback_gdf[name_col].astype(str)
                fallback_gdf['data_count'] = len(df_calls)
                master_gdf = fallback_gdf[['DISPLAY_NAME', 'data_count', 'geometry']]
            except Exception:
                master_gdf = None

    if master_gdf is None or master_gdf.empty:
        # ── Fallback 2: bounding box around call points ──
        min_lon, min_lat = df_calls['lon'].min(), df_calls['lat'].min()
        max_lon, max_lat = df_calls['lon'].max(), df_calls['lat'].max()
        lon_pad = (max_lon - min_lon) * 0.1
        lat_pad = (max_lat - min_lat) * 0.1
        poly = box(min_lon-lon_pad, min_lat-lat_pad, max_lon+lon_pad, max_lat+lat_pad)
        master_gdf = gpd.GeoDataFrame({'DISPLAY_NAME':['Auto-Generated Boundary'],'data_count':[len(df_calls)]}, geometry=[poly], crs="EPSG:4326")

    # --- DRAW SIDEBAR LOGO FIRST SO IT IS AT THE ABSOLUTE TOP ---
    logo_b64 = get_base64_of_bin_file("logo.png")
    if logo_b64:
        st.sidebar.markdown(f"""
        <div style="background-color: transparent; padding: 40px 20px 10px 20px; margin: -60px -20px 20px -20px; text-align: center; pointer-events: none;">
            <img src="data:image/png;base64,{logo_b64}" style="height: 36px; filter: brightness(0) invert(1);">
        </div>
        """, unsafe_allow_html=True)
    else:
        st.sidebar.markdown(f"""
        <div style="background-color: transparent; padding: 40px 20px 10px 20px; margin: -60px -20px 20px -20px; text-align: center; pointer-events: none;">
            <div style="font-size:26px; font-weight:900; letter-spacing:3px; color:#ffffff;">BRINC</div>
        </div>
        """, unsafe_allow_html=True)

    st.sidebar.markdown('<div class="sidebar-section-header">① Configure</div>', unsafe_allow_html=True)

    total_pts = master_gdf['data_count'].sum()
    master_gdf['LABEL'] = master_gdf['DISPLAY_NAME'] + " (" + (master_gdf['data_count']/total_pts*100).round(1).astype(str) + "%)"
    options_map = dict(zip(master_gdf['LABEL'], master_gdf['DISPLAY_NAME']))
    all_options = master_gdf['LABEL'].tolist()
    
    default_selection = [all_options[0]] if all_options else []
    selected_labels = st.sidebar.multiselect("Jurisdictions", options=all_options, default=default_selection,
                                             help="Select which geographic areas to include in coverage analysis.")

    if not selected_labels:
        st.warning("Please select at least one jurisdiction from the sidebar.")
        st.stop()
        
    selected_names = [options_map[l] for l in selected_labels]
    active_gdf = master_gdf[master_gdf['DISPLAY_NAME'].isin(selected_names)]
    if selected_names and st.session_state.get('active_city') == "Orlando":
        st.session_state['active_city'] = selected_names[0]

    filter_expander = st.sidebar.expander("⚙️ Data Filters", expanded=False)
    with filter_expander:
        if 'type' in df_stations_all.columns:
            all_types = sorted(df_stations_all['type'].dropna().astype(str).unique().tolist())
            if all_types:
                selected_types = st.multiselect("Facility Type", options=all_types, default=all_types,
                                                help="Filter which station types are eligible for drone deployment.")
                if not selected_types:
                    st.warning("Select at least one facility type.")
                    st.stop()
                df_stations_all = df_stations_all[df_stations_all['type'].astype(str).isin(selected_types)].copy().reset_index(drop=True)
                df_stations_all['name'] = "[" + df_stations_all['type'].astype(str) + "] " + df_stations_all['name'].astype(str)
        priority_source = df_calls_full if (df_calls_full is not None and 'priority' in df_calls_full.columns) else df_calls
        if 'priority' in priority_source.columns:
            all_priorities = sorted(pd.Series(priority_source['priority']).dropna().astype(int).unique().tolist())
            if all_priorities:
                selected_priorities = st.multiselect("Incident Priority", options=all_priorities, default=all_priorities,
                                                     help="Filter which call priorities to include in coverage scoring.")
                if not selected_priorities:
                    st.warning("Select at least one priority level.")
                    st.stop()
                df_calls = df_calls[df_calls['priority'].isin(selected_priorities)].copy().reset_index(drop=True)
                if df_calls_full is not None and 'priority' in df_calls_full.columns:
                    df_calls_full = df_calls_full[df_calls_full['priority'].isin(selected_priorities)].copy().reset_index(drop=True)

    if len(df_stations_all) == 0:
        st.error("No stations match the selected filters."); st.stop()
    if len(df_calls) == 0:
        st.error("No calls match the selected filters."); st.stop()

    disp_expander = st.sidebar.expander("👁️ Display Options", expanded=False)
    with disp_expander:
        show_boundaries = st.toggle("Jurisdiction Boundaries", value=True)
        show_heatmap    = st.toggle("911 Call Heatmap", value=False)
        show_health     = st.toggle("Health Score", value=False)
        show_satellite  = st.toggle("Satellite Imagery", value=False)
        show_cards      = True
        show_faa        = st.toggle("FAA LAANC Airspace", value=False)
        simulate_traffic = st.toggle("Simulate Ground Traffic", value=False)
        traffic_level   = st.slider("Traffic Congestion", 0, 100, 40) if simulate_traffic else 40

    strat_expander = st.sidebar.expander("⚙️ Deployment Strategy", expanded=False)
    with strat_expander:
        incremental_build = st.toggle("Phased Rollout", value=True,
            help="Place drones one at a time in priority order. Disable to find the global optimum in a single pass.")

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:8px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Deployment Mode</div>", unsafe_allow_html=True)
        deployment_mode = st.radio(
            "Deployment Mode", 
            ("Complement — push apart", "Independent — each maximises own area", "Shared — allow full overlap"),
            index=st.session_state.get('deployment_mode_idx', 1),
            label_visibility="collapsed",
            help=(
                "Complement: Responders fill gaps left by Guardians — no wasted overlap. "
                "Independent: each fleet optimises on its own objective; overlap allowed but not forced. "
                "Shared: both fleets optimise together against the same call set — hotspot stacking."
            )
        )
        _mode_map = {"Complement — push apart": 0, "Independent — each maximises own area": 1, "Shared — allow full overlap": 2}
        st.session_state['deployment_mode_idx'] = _mode_map.get(deployment_mode, 1)

        # Derived flags used by the optimizer
        allow_redundancy  = (deployment_mode != "Complement — push apart")
        complement_mode   = (deployment_mode == "Complement — push apart")
        shared_mode       = (deployment_mode == "Shared — allow full overlap")

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:10px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Guardian Objective</div>", unsafe_allow_html=True)
        guard_strategy_raw = st.radio(
            "Guardian Objective",
            ("Call Coverage", "Land Coverage"),
            index=st.session_state.get('guard_strat_idx', 1),
            horizontal=True,
            label_visibility="collapsed",
            help="What the Guardian optimizer maximises. Land Coverage = wide area patrol. Call Coverage = respond to highest-volume locations."
        )
        st.session_state['guard_strat_idx'] = 0 if guard_strategy_raw == "Call Coverage" else 1
        guard_strategy = "Maximize Call Coverage" if guard_strategy_raw == "Call Coverage" else "Maximize Land Coverage"

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:10px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Responder Objective</div>", unsafe_allow_html=True)
        resp_strategy_raw = st.radio(
            "Responder Objective",
            ("Call Coverage", "Land Coverage"),
            index=st.session_state.get('resp_strat_idx', 0),
            horizontal=True,
            label_visibility="collapsed",
            help="What the Responder optimizer maximises. Call Coverage = densest incident areas. Land Coverage = broadest geographic reach."
        )
        st.session_state['resp_strat_idx'] = 0 if resp_strategy_raw == "Call Coverage" else 1
        resp_strategy = "Maximize Call Coverage" if resp_strategy_raw == "Call Coverage" else "Maximize Land Coverage"

    # Keep opt_strategy for any code that still references it (used in export/logs)
    opt_strategy = guard_strategy  # primary strategy label for reporting

    st.sidebar.markdown('<div class="sidebar-section-header">② Optimize Fleet</div>', unsafe_allow_html=True)

    minx, miny, maxx, maxy = active_gdf.to_crs(epsg=4326).total_bounds
    center_lon = (minx + maxx) / 2
    center_lat = (miny + maxy) / 2
    dynamic_zoom = calculate_zoom(minx, maxx, miny, maxy)
    utm_zone = int((center_lon + 180) / 6) + 1
    epsg_code = int(f"326{utm_zone}") if center_lat > 0 else int(f"327{utm_zone}")

    city_m = None
    city_boundary_geom = None
    try:
        active_utm = active_gdf.to_crs(epsg=epsg_code)
        raw_union = (active_utm.geometry.union_all() if hasattr(active_utm.geometry, 'union_all')
                     else active_utm.geometry.unary_union)
        # buffer(0.1).buffer(-0.1) cleans self-intersections but can collapse thin geometries.
        # Use a larger initial buffer and validate before shrinking.
        clean_geom = raw_union.buffer(1.0).buffer(-1.0)
        if clean_geom.is_empty or not clean_geom.is_valid:
            clean_geom = raw_union.buffer(0)  # zero-buffer repair only
        if clean_geom.is_empty:
            clean_geom = raw_union          # use as-is if still empty
        city_m = clean_geom
        city_boundary_geom = gpd.GeoSeries([clean_geom], crs=epsg_code).to_crs(epsg=4326).iloc[0]
    except Exception as e:
        st.error(f"Geometry Error: {e}"); st.stop()

    # --- GEOGRAPHIC FILTERING FOR STATIONS ---
    # Keep stations inside city boundary + generous buffer.
    # If OSM found nothing inside the boundary (e.g. small cities with few public
    # buildings tagged), fall back to call-density-derived synthetic stations so
    # the tool never dead-ends on legitimate data.
    if not df_stations_all.empty and city_m is not None:
        st_gdf = gpd.GeoDataFrame(df_stations_all,
                                   geometry=gpd.points_from_xy(df_stations_all.lon, df_stations_all.lat),
                                   crs="EPSG:4326")
        st_gdf_utm = st_gdf.to_crs(epsg=epsg_code)

        # Keep candidate sites strictly inside the jurisdiction whenever possible.
        mask = st_gdf_utm.within(city_m)
        df_inside = df_stations_all[mask].reset_index(drop=True)

        if df_inside.empty:
            st.info(
                "ℹ️ No OSM public buildings were found inside the jurisdiction boundary. "
                "Using call-density station placement — stations are snapped to incident "
                "locations that fall inside the city limits."
            )
            try:
                df_stations_all = _make_random_stations(df_calls, n=60, boundary_geom=city_m, epsg_code=epsg_code)
            except Exception:
                df_stations_all = pd.DataFrame()

            # Absolute last resort: build a simple grid from call quantiles
            if df_stations_all.empty:
                try:
                    _lats = df_calls['lat'].dropna()
                    _lons = df_calls['lon'].dropna()
                    _grid_lats = np.linspace(_lats.quantile(0.1), _lats.quantile(0.9), 8)
                    _grid_lons = np.linspace(_lons.quantile(0.1), _lons.quantile(0.9), 8)
                    _glat, _glon = np.meshgrid(_grid_lats, _grid_lons)
                    df_stations_all = pd.DataFrame({
                        'name':  [f'Station {i+1}' for i in range(len(_glat.ravel()))],
                        'lat':   _glat.ravel(),
                        'lon':   _glon.ravel(),
                        'type':  (['Police', 'Fire', 'School'] * 30)[:len(_glat.ravel())],
                    })
                except Exception:
                    df_stations_all = pd.DataFrame()
        else:
            df_stations_all = df_inside

        if not df_stations_all.empty:
            try:
                _final_st_gdf = gpd.GeoDataFrame(df_stations_all, geometry=gpd.points_from_xy(df_stations_all.lon, df_stations_all.lat), crs="EPSG:4326").to_crs(epsg=epsg_code)
                _final_mask = _final_st_gdf.within(city_m)
                if _final_mask.any():
                    df_stations_all = df_stations_all[_final_mask].reset_index(drop=True)
            except Exception:
                pass

        if df_stations_all.empty:
            st.error(
                "⚠️ No station candidates could be generated. Please upload a CAD file "
                "with valid coordinates, or switch to Simulation mode."
            )
            st.stop()

    # ── Inject custom stations (bypass boundary clip & type filter) ──────────
    _custom_st = st.session_state.get('custom_stations', pd.DataFrame())
    if not _custom_st.empty:
        # Apply the same type-prefix rename the filter block uses, so pin lookups match
        _cst_renamed = _custom_st.copy()
        _cst_renamed['name'] = "[" + _cst_renamed['type'].astype(str) + "] " + _cst_renamed['name'].astype(str)
        # Drop columns that might not exist in df_stations_all to avoid concat issues
        _keep_cols = [c for c in _cst_renamed.columns if c in list(df_stations_all.columns) + ['name','lat','lon','type','custom']]
        _cst_renamed = _cst_renamed[_keep_cols]
        df_stations_all = pd.concat([df_stations_all, _cst_renamed], ignore_index=True)

    n = len(df_stations_all)

    # Dynamic Sliders based on Area Size
    area_sq_mi = city_m.area / 2589988.11 if city_m and not city_m.is_empty else 100.0
    r_resp_est = st.session_state.get('r_resp', 2.0)
    r_guard_est = st.session_state.get('r_guard', 8.0)
    
    max_resp_calc = min(n, int(math.ceil(area_sq_mi / (math.pi * (r_resp_est**2)))) + 5)
    max_guard_calc = min(n, int(math.ceil(area_sq_mi / (math.pi * (r_guard_est**2)))) + 5)

    # Safely pull the default values without exceeding the allowed maximums
    val_r = min(st.session_state.get('k_resp', 2), max_resp_calc)
    val_g = min(st.session_state.get('k_guard', 0), max_guard_calc)

    k_responder = st.sidebar.slider("🚁 Responder Count", 0, max(1, max_resp_calc), val_r, help="Short-range tactical drones (2-3mi radius).")
    k_guardian  = st.sidebar.slider("🦅 Guardian Count", 0, max(1, max_guard_calc), val_g, help="Long-range overwatch drones (5-8mi radius).")
    
    resp_radius_mi  = st.sidebar.slider("🚁 Responder Range (mi)", 2.0, 3.0, float(st.session_state.get('r_resp', 2.0)), step=0.5)
    guard_radius_mi = st.sidebar.slider("🦅 Guardian Range (mi) [⚡ 5mi Rapid]", 1, 8, int(st.session_state.get('r_guard', 8)), help="The 5-mile rapid response focus zone will automatically be highlighted inside the maximum perimeter.")

    st.session_state.update({'k_resp': k_responder, 'k_guard': k_guardian, 'r_resp': resp_radius_mi, 'r_guard': guard_radius_mi})

    # ── MANUAL STATION PINS ───────────────────────────────────────────────────
    # ── Sync pinned station lists (expander removed — pinning via card buttons) ──
    _station_names = df_stations_all['name'].tolist() if not df_stations_all.empty else []
    _saved_g = [s for s in st.session_state.get('pinned_guard_names', []) if s in _station_names]
    _saved_r = [s for s in st.session_state.get('pinned_resp_names',  []) if s in _station_names]
    st.session_state['pinned_guard_names'] = _saved_g
    st.session_state['pinned_resp_names']  = _saved_r
    pinned_guard_names = _saved_g
    pinned_resp_names  = _saved_r

    # Warn in sidebar if pin count exceeds slider
    if len(pinned_guard_names) > k_guardian:
        st.sidebar.warning(f"⚠️ Raise Guardian Count ≥ {len(pinned_guard_names)} to use all Guardian pins.")
    if len(pinned_resp_names) > k_responder:
        st.sidebar.warning(f"⚠️ Raise Responder Count ≥ {len(pinned_resp_names)} to use all Responder pins.")

    # ── ADD CUSTOM STATION BY ADDRESS ─────────────────────────────────────────
    add_expander = st.sidebar.expander("➕ Add Custom Station", expanded=False)
    with add_expander:
        st.markdown(
            f"<div style='font-size:0.7rem; color:{text_muted}; margin-bottom:8px;'>"
            "Enter an address to add a custom deployment location. "
            "The Census geocoder resolves it to lat/lon. Custom stations persist for this session.</div>",
            unsafe_allow_html=True
        )
        # Use value= (not key=) so we can clear via session_state buffer without
        # triggering the "cannot modify after instantiation" error.
        if 'cs_addr_buf'  not in st.session_state: st.session_state['cs_addr_buf']  = ""
        if 'cs_label_buf' not in st.session_state: st.session_state['cs_label_buf'] = ""
        if 'cs_type_buf'  not in st.session_state: st.session_state['cs_type_buf']  = "Police"

        _custom_addr  = st.text_input("Address", value=st.session_state['cs_addr_buf'],
                                       placeholder="123 Main St, Mobile, AL",
                                       key="custom_station_addr")
        _custom_label = st.text_input("Label (optional)", value=st.session_state['cs_label_buf'],
                                       placeholder="Fire Station 7",
                                       key="custom_station_label")
        _type_opts    = ["Police", "Fire", "School", "Government", "Hospital", "Library", "Other"]
        _type_idx     = _type_opts.index(st.session_state['cs_type_buf']) if st.session_state['cs_type_buf'] in _type_opts else 0
        _custom_type  = st.selectbox("Type", _type_opts, index=_type_idx,
                                      key="custom_station_type")

        # Explicit drone role selector — user decides, no auto-guessing
        if 'cs_role_buf' not in st.session_state: st.session_state['cs_role_buf'] = "🦅 Lock as Guardian"
        _role_opts = ["🦅 Lock as Guardian", "🚁 Lock as Responder"]
        _role_idx  = _role_opts.index(st.session_state['cs_role_buf']) if st.session_state['cs_role_buf'] in _role_opts else 0
        _custom_role = st.radio("Pin as", _role_opts, index=_role_idx,
                                horizontal=True, key="custom_station_role",
                                help="Choose which fleet this station will be locked into.")

        # Sync buffer keys from live widget values each run
        st.session_state['cs_addr_buf']  = _custom_addr
        st.session_state['cs_label_buf'] = _custom_label
        st.session_state['cs_type_buf']  = _custom_type
        st.session_state['cs_role_buf']  = _custom_role

        if st.button("📍 Geocode & Add Station", use_container_width=True, key="geocode_btn"):
            _addr_to_geocode = _custom_addr.strip()
            if _addr_to_geocode:
                try:
                    import urllib.request, urllib.parse, json as _json
                    _params = urllib.parse.urlencode({
                        "address": _addr_to_geocode,
                        "benchmark": "2020",
                        "format": "json"
                    })
                    _url = f"https://geocoding.geo.census.gov/geocoder/locations/onelineaddress?{_params}"
                    with urllib.request.urlopen(_url, timeout=8) as _resp:
                        _data = _json.loads(_resp.read().decode())
                    _matches = _data.get("result", {}).get("addressMatches", [])
                    if _matches:
                        _coords = _matches[0]["coordinates"]
                        _geo_lat = float(_coords["y"])
                        _geo_lon = float(_coords["x"])
                        _matched_addr = _matches[0].get("matchedAddress", _addr_to_geocode)
                        _label = _custom_label.strip() or _matched_addr
                        # The type-prefix rename later produces "[Type] label"
                        # Store both original and prefixed name so pin lookup works
                        _prefixed_label = f"[{_custom_type}] {_label}"
                        _new_row = pd.DataFrame([{
                            "name":   _label,          # original, pre-prefix
                            "lat":    _geo_lat,
                            "lon":    _geo_lon,
                            "type":   _custom_type,
                            "custom": True,
                        }])
                        # Store in dedicated 'custom_stations' key so boundary clip
                        # and type filter can't drop them
                        _cst = st.session_state.get('custom_stations', pd.DataFrame())
                        st.session_state['custom_stations'] = pd.concat(
                            [_cst, _new_row], ignore_index=True
                        ) if not _cst.empty else _new_row

                        # Pin using the user's explicit role choice
                        if _custom_role == "🦅 Lock as Guardian":
                            _pg = list(st.session_state.get('pinned_guard_names', []))
                            if _prefixed_label not in _pg:
                                _pg.append(_prefixed_label)
                            # Remove from responder list if it was there
                            st.session_state['pinned_resp_names'] = [
                                x for x in st.session_state.get('pinned_resp_names', [])
                                if x != _prefixed_label]
                            st.session_state['pinned_guard_names'] = _pg
                            if st.session_state.get('k_guard', 0) < len(_pg):
                                st.session_state['k_guard'] = len(_pg)
                            _pin_note = f"🦅 Pinned as Guardian."
                        else:
                            _pr = list(st.session_state.get('pinned_resp_names', []))
                            if _prefixed_label not in _pr:
                                _pr.append(_prefixed_label)
                            st.session_state['pinned_guard_names'] = [
                                x for x in st.session_state.get('pinned_guard_names', [])
                                if x != _prefixed_label]
                            st.session_state['pinned_resp_names'] = _pr
                            if st.session_state.get('k_resp', 0) < len(_pr):
                                st.session_state['k_resp'] = len(_pr)
                            _pin_note = f"🚁 Pinned as Responder."

                        st.success(f"✅ Added & locked: **{_label}** ({_geo_lat:.4f}, {_geo_lon:.4f})\n{_pin_note}")
                        st.caption(f"Matched address: {_matched_addr}")
                        # Clear buffers (role intentionally kept so user can add more of same type)
                        st.session_state['cs_addr_buf']  = ""
                        st.session_state['cs_label_buf'] = ""
                        # Clear ALL optimizer caches so the new station
                        # enters both the spatial precompute AND the LP solver
                        for _ck in ['_opt_cache_key', '_opt_best_combo',
                                    '_opt_chrono_r', '_opt_chrono_g']:
                            st.session_state.pop(_ck, None)
                        st.rerun()
                    else:
                        st.warning("⚠️ Address not found. Try including city and state — e.g. '123 Main St, Mobile AL'.")
                except Exception as _ge:
                    st.error(f"Geocoding failed: {_ge}")
            else:
                st.warning("Enter an address first.")

        # Show custom stations added this session
        _cst_display = st.session_state.get('custom_stations', pd.DataFrame())
        if not _cst_display.empty:
            _custom_added = _cst_display['name'].tolist()
        else:
            _custom_added = []
        if _custom_added:
            st.markdown(f"<div style='font-size:0.65rem; color:{text_muted}; margin-top:6px;'>"
                        f"Pinned this session:</div>", unsafe_allow_html=True)
            _pg_set = set(st.session_state.get('pinned_guard_names', []))
            _pr_set = set(st.session_state.get('pinned_resp_names', []))
            _cst_disp = st.session_state.get('custom_stations', pd.DataFrame())
            for _cn in _custom_added[:8]:
                # Check both original and prefixed name
                _cst_row = _cst_disp[_cst_disp['name'] == _cn].iloc[0] if not _cst_disp.empty and (_cst_disp['name'] == _cn).any() else None
                _pfx = f"[{_cst_row['type']}] {_cn}" if _cst_row is not None else _cn
                _is_g = _pfx in _pg_set or _cn in _pg_set
                _badge = "🦅" if _is_g else "🚁"
                _color = "#FFD700" if _is_g else "#00D2FF"
                st.markdown(f"<div style='font-size:0.65rem; color:{_color}; padding:1px 0;'>{_badge} {_pfx}</div>",
                            unsafe_allow_html=True)
            if st.button("🗑 Remove all custom stations", key="remove_custom",
                         use_container_width=True):
                _cst_to_rm = st.session_state.get('custom_stations', pd.DataFrame())
                # Build set of both original and prefixed names to un-pin
                _rm_names = set()
                if not _cst_to_rm.empty:
                    for _, _row in _cst_to_rm.iterrows():
                        _rm_names.add(str(_row['name']))
                        _rm_names.add(f"[{_row['type']}] {_row['name']}")
                st.session_state['custom_stations'] = pd.DataFrame()
                st.session_state['pinned_guard_names'] = [
                    x for x in st.session_state.get('pinned_guard_names', []) if x not in _rm_names]
                st.session_state['pinned_resp_names']  = [
                    x for x in st.session_state.get('pinned_resp_names',  []) if x not in _rm_names]
                if '_opt_cache_key' in st.session_state:
                    del st.session_state['_opt_cache_key']
                st.rerun()

    # Convert pin names → station indices for the optimizer
    _name_to_idx = {row['name']: i for i, row in df_stations_all.iterrows()}
    locked_g_pins = [_name_to_idx[n] for n in pinned_guard_names if n in _name_to_idx]
    locked_r_pins = [_name_to_idx[n] for n in pinned_resp_names  if n in _name_to_idx]

    bounds_hash = f"{minx}_{miny}_{maxx}_{maxy}_{n}_{resp_radius_mi}_{guard_radius_mi}"

    prog2 = st.sidebar.empty()
    prog2.caption(get_spatial_message())
    calls_in_city, display_calls, resp_matrix, guard_matrix, dist_matrix_r, dist_matrix_g, station_metadata, total_calls = precompute_spatial_data(
        df_calls, df_calls_full, df_stations_all, city_m, epsg_code, resp_radius_mi, guard_radius_mi, center_lat, center_lon, bounds_hash
    )
    df_curve = compute_all_elbow_curves(
        total_calls, resp_matrix, guard_matrix,
        [s['clipped_2m'] for s in station_metadata],
        [s['clipped_guard'] for s in station_metadata],
        city_m.area if city_m else 1.0, bounds_hash,
        max_stations=100
    )
    prog2.empty()

    # (Scored station table removed — station scores shown in Add Custom Station expander)

    def get_max_drones(col_name):
        series = df_curve[col_name].dropna()
        if len(series) == 0: return 1
        idx_99 = series[series >= 99.0].first_valid_index()
        fallback = series.index[-1]
        return int(df_curve.loc[idx_99 if idx_99 is not None else fallback, 'Drones'])

    with st.spinner(get_faa_message()):
        faa_geojson = load_faa_parquet(minx, miny, maxx, maxy)
    with st.spinner(get_airfield_message()):
        airfields = fetch_airfields(minx, miny, maxx, maxy)

    st.sidebar.markdown('<div class="sidebar-section-header">③ Budget & Export</div>', unsafe_allow_html=True)

    # We use the strat_expander we defined earlier in the sidebar to inject the sliders
    with strat_expander:
        st.markdown("---")
        inferred_daily = st.session_state.get('inferred_daily_calls_override', full_daily_calls)
        inferred_daily = max(1, int(inferred_daily))
        calls_per_day = st.slider("Total Daily Calls (citywide)", 1, max(100, inferred_daily*3), inferred_daily)
        st.caption(f"Derived from the full uploaded CAD total ({full_total_calls:,} incidents), not the optimization sample.")

        st.markdown(f"<div style='font-size:0.72rem; color:{text_muted}; margin-top:8px; margin-bottom:2px;'>DFR Dispatch Rate (%)</div>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:0.65rem; color:#666; margin-bottom:4px;'>What % of in-range calls will the drone be sent to?</div>", unsafe_allow_html=True)
        dfr_dispatch_rate = st.slider("DFR Dispatch Rate", 1, 100, st.session_state.get('dfr_rate',25), label_visibility="collapsed") / 100.0

        st.markdown(f"<div style='font-size:0.72rem; color:{text_muted}; margin-top:8px; margin-bottom:2px;'>Calls Resolved Without Officer Dispatch (%)</div>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:0.65rem; color:#666; margin-bottom:4px;'>Of drone-attended calls, what % close without a patrol car?</div>", unsafe_allow_html=True)
        deflection_rate = st.slider("Resolution Rate", 0, 100, st.session_state.get('deflect_rate',30), label_visibility="collapsed") / 100.0

        st.session_state['dfr_rate']    = int(dfr_dispatch_rate * 100)
        st.session_state['deflect_rate'] = int(deflection_rate * 100)

    # ── OPTIMIZATION ──────────────────────────────────────────────────
    active_resp_names, active_guard_names = [], []
    active_resp_idx, active_guard_idx = [], []  
    chrono_r, chrono_g = [], []
    best_combo = None

    _pins_key = f"{sorted(locked_g_pins)}_{sorted(locked_r_pins)}"
    opt_cache_key = f"{k_responder}_{k_guardian}_{resp_radius_mi}_{guard_radius_mi}_{guard_strategy}_{resp_strategy}_{deployment_mode}_{incremental_build}_{bounds_hash}_{_pins_key}"

    if k_responder + k_guardian > n:
        st.error("⚠️ Over-Deployment: Total drones exceed available stations.")
        active_resp_names, active_guard_names = [], []
        chrono_r, chrono_g = [], []
        best_combo = None
    elif k_responder == 0 and k_guardian == 0:
        active_resp_names, active_guard_names = [], []
        chrono_r, chrono_g = [], []
        best_combo = None
    else:
        if st.session_state.get('_opt_cache_key') != opt_cache_key:
            stage_bar = st.empty()

            # ── HELPER: greedy area-coverage for one fleet ───────────────────
            def _greedy_area(matrix, geo_list, k, forced, exclude_set):
                """Greedily pick k stations maximising unary_union area,
                starting from forced pins and skipping exclude_set."""
                chosen = list(forced)
                chrono  = list(forced)
                current_union = unary_union([geo_list[i] for i in chosen]) if chosen else None
                for _ in range(k - len(forced)):
                    best_s, best_gain = -1, -1.0
                    for s in range(len(geo_list)):
                        if s in chosen or s in exclude_set:
                            continue
                        g = geo_list[s]
                        new_area = current_union.union(g).area if current_union else g.area
                        gain = new_area - (current_union.area if current_union else 0)
                        if gain > best_gain:
                            best_gain, best_s = gain, s
                    if best_s != -1:
                        chosen.append(best_s)
                        chrono.append(best_s)
                        g = geo_list[best_s]
                        current_union = current_union.union(g) if current_union else g
                return chosen, chrono

            # ── PASS 1: Optimise Guardians independently ─────────────────────
            stage_bar.info("🦅 Optimising Guardian fleet…")
            if k_guardian > 0:
                if guard_strategy == "Maximize Call Coverage":
                    # solve_mclp returns (r_best, g_best, chrono_r, chrono_g)
                    # Pass 1 runs Guardians only (num_resp=0) so r_best=[] and g_best has the result
                    _, g_best, _, chrono_g = solve_mclp(
                        resp_matrix, guard_matrix, dist_matrix_r, dist_matrix_g,
                        0, k_guardian, True, incremental=incremental_build,
                        forced_r=[], forced_g=locked_g_pins
                    )
                else:
                    g_best, chrono_g = _greedy_area(
                        guard_matrix,
                        [station_metadata[i]['clipped_guard'] for i in range(n)],
                        k_guardian, locked_g_pins, set()
                    )
                g_best = list(g_best)
            else:
                g_best, chrono_g = [], []

            # ── PASS 2: Optimise Responders around Guardian result ────────────
            stage_bar.info("🚁 Optimising Responder fleet…")
            if k_responder > 0:
                # In complement mode, mask out calls already covered by Guardians
                # so Responders fill the gaps rather than stacking on the same calls.
                if complement_mode and g_best and total_calls > 0:
                    guard_covered = guard_matrix[g_best].any(axis=0)
                    # Build a reduced matrix: zero out already-covered calls for Responders
                    resp_matrix_eff = resp_matrix.copy()
                    resp_matrix_eff[:, guard_covered] = False
                    dist_matrix_r_eff = dist_matrix_r.copy()
                else:
                    resp_matrix_eff    = resp_matrix
                    dist_matrix_r_eff  = dist_matrix_r

                # In complement mode, Responders also can't reuse Guardian stations
                _excl = set(g_best) if not allow_redundancy else set()

                if resp_strategy == "Maximize Call Coverage":
                    r_best, _, chrono_r, _ = solve_mclp(
                        resp_matrix_eff, guard_matrix, dist_matrix_r_eff, dist_matrix_g,
                        k_responder, 0, allow_redundancy, incremental=incremental_build,
                        forced_r=locked_r_pins, forced_g=[]
                    )
                    # Filter out Guardian stations if complement mode
                    if complement_mode:
                        r_best = [s for s in r_best if s not in set(g_best)]
                        # Pad back to k_responder if exclusion removed some
                        if len(r_best) < k_responder:
                            remaining = [s for s in range(n)
                                         if s not in r_best and s not in set(g_best)]
                            r_best += remaining[:k_responder - len(r_best)]
                else:
                    _excl_resp = set(g_best) if complement_mode else set()
                    r_best, chrono_r = _greedy_area(
                        resp_matrix_eff,
                        [station_metadata[i]['clipped_2m'] for i in range(n)],
                        k_responder, locked_r_pins, _excl_resp
                    )
            else:
                r_best, chrono_r = [], []

            best_combo = (tuple(r_best), tuple(g_best))
            stage_bar.empty()
            st.toast("✅ Independent optimisation complete!", icon="✅")

            st.session_state['_opt_cache_key']  = opt_cache_key
            st.session_state['_opt_best_combo'] = best_combo
            st.session_state['_opt_chrono_r']   = chrono_r
            st.session_state['_opt_chrono_g']   = chrono_g
        else:
            best_combo = st.session_state.get('_opt_best_combo')
            chrono_r   = st.session_state.get('_opt_chrono_r', [])
            chrono_g   = st.session_state.get('_opt_chrono_g', [])

        if best_combo is not None:
            r_best, g_best = best_combo
            active_resp_names  = [station_metadata[i]['name'] for i in r_best]
            active_guard_names = [station_metadata[i]['name'] for i in g_best]
            active_resp_idx  = list(r_best)
            active_guard_idx = list(g_best)
        else:
            active_resp_names, active_guard_names = [], []
            active_resp_idx, active_guard_idx = [], []

    # ── METRICS ───────────────────────────────────────────────────────
    # ── SPLIT METRICS: Guardian and Responder computed independently ─────────
    area_covered_perc = overlap_perc = calls_covered_perc = 0.0
    guard_calls_perc  = guard_area_perc  = 0.0
    resp_calls_perc   = resp_area_perc   = 0.0
    cov_r = np.zeros(total_calls, bool) if total_calls > 0 else np.zeros(0, bool)
    cov_g = np.zeros(total_calls, bool) if total_calls > 0 else np.zeros(0, bool)

    ordered_deployments_raw = []
    for idx in chrono_g:
        if idx in active_guard_idx: ordered_deployments_raw.append((idx,'GUARDIAN'))
    for idx in chrono_r:
        if idx in active_resp_idx: ordered_deployments_raw.append((idx,'RESPONDER'))
    for idx in active_resp_idx:
        if idx not in chrono_r: ordered_deployments_raw.append((idx,'RESPONDER'))
    for idx in active_guard_idx:
        if idx not in chrono_g: ordered_deployments_raw.append((idx,'GUARDIAN'))

    active_color_map = {}
    c_idx = 0
    for idx, d_type in ordered_deployments_raw:
        key = f"{idx}_{d_type}"
        if key not in active_color_map:
            active_color_map[key] = STATION_COLORS[c_idx % len(STATION_COLORS)]
            c_idx += 1

    guard_geos = [station_metadata[i]['clipped_guard'] for i in active_guard_idx]
    resp_geos  = [station_metadata[i]['clipped_2m']    for i in active_resp_idx]
    active_geos = resp_geos + guard_geos

    city_area = city_m.area if (city_m and not city_m.is_empty) else 1.0

    # Guardian-only metrics
    if guard_geos:
        guard_area_perc = (unary_union(guard_geos).area / city_area) * 100
    if active_guard_idx and total_calls > 0:
        cov_g = guard_matrix[active_guard_idx].any(axis=0)
        guard_calls_perc = cov_g.sum() / total_calls * 100

    # Responder-only metrics
    if resp_geos:
        resp_area_perc = (unary_union(resp_geos).area / city_area) * 100
    if active_resp_idx and total_calls > 0:
        cov_r = resp_matrix[active_resp_idx].any(axis=0)
        resp_calls_perc = cov_r.sum() / total_calls * 100

    # Combined metrics
    if active_geos:
        area_covered_perc = (unary_union(active_geos).area / city_area) * 100
    if total_calls > 0:
        calls_covered_perc = (np.logical_or(cov_r, cov_g).sum() / total_calls) * 100
    if len(active_geos) >= 2:
        inters = [active_geos[i].intersection(active_geos[j])
                  for i in range(len(active_geos))
                  for j in range(i+1, len(active_geos))
                  if not active_geos[i].is_empty and not active_geos[j].is_empty
                  and active_geos[i].intersects(active_geos[j])]
        if inters:
            overlap_perc = (unary_union(inters).area / city_area) * 100

    # ── BUDGET CALCULATIONS ───────────────────────────────────────────
    actual_k_responder = len(active_resp_names)
    actual_k_guardian  = len(active_guard_names)
    capex_resp  = actual_k_responder * CONFIG["RESPONDER_COST"]
    capex_guard = actual_k_guardian  * CONFIG["GUARDIAN_COST"]
    fleet_capex = capex_resp + capex_guard

    annual_savings = 0
    break_even_text = "N/A"
    daily_drone_only_calls = 0
    covered_daily_calls = 0
    daily_dfr_responses = 0

    if fleet_capex > 0:
        covered_daily_calls    = calls_per_day * (calls_covered_perc / 100.0)
        daily_dfr_responses    = covered_daily_calls * dfr_dispatch_rate
        daily_drone_only_calls = daily_dfr_responses * deflection_rate
        if daily_drone_only_calls > 0:
            monthly_savings = (CONFIG["OFFICER_COST_PER_CALL"] - CONFIG["DRONE_COST_PER_CALL"]) * daily_drone_only_calls * 30.4
            annual_savings  = monthly_savings * 12
            break_even_text = f"{fleet_capex / monthly_savings:.1f} MONTHS"

    if fleet_capex > 0:
        st.sidebar.markdown(f"""
        <div style="background:{budget_box_bg}; border:1px solid {budget_box_border}; padding:12px; border-radius:4px;
             text-align:center; margin:8px 0 12px 0; box-shadow:0 2px 5px {budget_box_shadow};">
            <div style="font-size:0.7rem; color:{text_muted}; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Annual Capacity Value</div>
            <div style="font-size:1.8rem; font-weight:900; color:{budget_box_border}; font-family:monospace;">${annual_savings:,.0f}</div>
            <div style="border-top:1px solid {card_border}; margin:8px 0;"></div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:3px;">
                <span style="color:{text_muted};">Calls in range:</span>
                <span style="color:{text_main}; font-weight:700;">{covered_daily_calls:.1f}/day</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:3px;">
                <span style="color:{text_muted};">DFR flights ({int(dfr_dispatch_rate*100)}%):</span>
                <span style="color:{text_main}; font-weight:700;">{daily_dfr_responses:.1f}/day</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:8px;">
                <span style="color:{text_muted};">Resolved no dispatch:</span>
                <span style="color:{text_main}; font-weight:700;">{daily_drone_only_calls:.1f}/day</span>
            </div>
            <div style="border-top:1px dashed {card_border}; margin:6px 0;"></div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:3px;">
                <span style="color:{text_muted};">Fleet CapEx:</span>
                <span style="color:{text_main}; font-weight:700;">${fleet_capex:,.0f}</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem;">
                <span style="color:{text_muted};">Break-even:</span>
                <span style="color:{budget_box_border}; font-weight:700;">{break_even_text}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.sidebar.info("👈 Set Responder/Guardian counts above to calculate budget impact.")

    # ── BUILD DRONE OBJECTS ───────────────────────────────────────────
    active_drones = []
    cumulative_mask = np.zeros(total_calls, dtype=bool) if total_calls > 0 else None
    step = 1
    for idx, d_type in ordered_deployments_raw:
        if d_type == 'RESPONDER':
            cov_array = resp_matrix[idx]; cost = CONFIG["RESPONDER_COST"]
            speed_mph = CONFIG["RESPONDER_SPEED"]; avg_dist = station_metadata[idx]['avg_dist_r']
            radius_m  = resp_radius_mi * 1609.34
        else:
            cov_array = guard_matrix[idx]; cost = CONFIG["GUARDIAN_COST"]
            speed_mph = CONFIG["GUARDIAN_SPEED"]; avg_dist = station_metadata[idx]['avg_dist_g']
            radius_m  = guard_radius_mi * 1609.34
        map_color    = active_color_map[f"{idx}_{d_type}"]
        avg_time_min = (avg_dist / speed_mph) * 60
        d_lat = station_metadata[idx]['lat']; d_lon = station_metadata[idx]['lon']

        _is_pinned = (d_type == 'GUARDIAN' and idx in locked_g_pins) or (d_type == 'RESPONDER' and idx in locked_r_pins)
        d = {
            'idx': idx, 'name': station_metadata[idx]['name'],
            'lat': d_lat, 'lon': d_lon, 'type': d_type, 'cost': cost,
            'cov_array': cov_array, 'color': map_color,
            'pinned': _is_pinned,
            'deploy_step': step if (idx in chrono_r or idx in chrono_g) else "MANUAL",
            'avg_time_min': avg_time_min, 'speed_mph': speed_mph, 'radius_m': radius_m,
            'faa_ceiling': get_station_faa_ceiling(d_lat, d_lon, faa_geojson),
            'nearest_airport': get_nearest_airfield(d_lat, d_lon, airfields)
        }

        if total_calls > 0 and cumulative_mask is not None:
            # ── DEDUPLICATION: track unique calls added for combined KPI totals ──
            marginal_mask     = cov_array & ~cumulative_mask
            marginal_historic = np.sum(marginal_mask)
            d['assigned_indices'] = np.where(marginal_mask)[0]
            cumulative_mask   = cumulative_mask | cov_array

            # ── RAW ZONE COVERAGE: how many calls fall in this drone's zone ───
            # Used for per-unit economics. Independent of iteration order so
            # Responders are never penalised for a Guardian claiming the same calls.
            _raw_zone_calls = int(np.sum(cov_array))  # all calls inside this drone's radius
            _raw_zone_perc  = _raw_zone_calls / total_calls

            # Shared zone: calls covered by at least one OTHER active drone
            all_cov = np.vstack([resp_matrix[i] for i in active_resp_idx] + [guard_matrix[i] for i in active_guard_idx]) if (active_resp_idx or active_guard_idx) else np.zeros((1, total_calls), dtype=bool)
            shared_mask   = d['cov_array'] & (all_cov.sum(axis=0) > 1)
            _shared_calls = int(np.sum(shared_mask))
            _excl_calls   = _raw_zone_calls - _shared_calls  # calls ONLY this drone covers

            # ── UTILIZATION: based on full zone call load, not marginal residual ─
            # Responders should reflect how busy they truly are in their patrol zone.
            # daily calls dispatched to this drone = zone calls × dispatch rate
            _is_guard    = (d_type == 'GUARDIAN')
            _budget_min  = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"] if _is_guard else (CONFIG["RESPONDER_PATROL_HOURS"] * 60)
            _zone_flights = _raw_zone_perc * calls_per_day * dfr_dispatch_rate
            _util = min(0.99, (_zone_flights * avg_time_min) / max(1.0, _budget_min))

            # ── BASE VALUE: calls uniquely covered (non-shared zone) ──────────
            # These calls have no other drone to fall back on — pure incremental value.
            _excl_daily        = (_excl_calls / total_calls) * calls_per_day
            _excl_flights      = _excl_daily * dfr_dispatch_rate
            _excl_deflected    = _excl_flights * deflection_rate
            _cost_delta        = CONFIG["OFFICER_COST_PER_CALL"] - CONFIG["DRONE_COST_PER_CALL"]
            _base_monthly      = _cost_delta * _excl_deflected * 30.4
            _base_annual       = _base_monthly * 12

            # ── CONCURRENT VALUE: shared-zone calls captured while partner is busy ─
            # Guardian is airborne util% of the time → Responder handles that fraction.
            # Responder is airborne util% of the time → Guardian handles that fraction.
            # Net concurrent gain = shared_calls × partner_util × deflection × cost_delta
            # For a Responder sharing zone with a Guardian: partner = Guardian utilization
            # We approximate partner utilization as _util (symmetric; use actual if available)
            _shared_daily      = (_shared_calls / total_calls) * calls_per_day
            _shared_dfr        = _shared_daily * dfr_dispatch_rate
            # Calls this drone handles while its partner is busy (partner util ≈ _util)
            _concurrent_daily  = _shared_dfr * _util
            _concurrent_month  = _cost_delta * (_concurrent_daily * deflection_rate) * 30.4
            _concurrent_annual = _concurrent_month * 12

            # ── BEST CASE: base + full concurrent (partner always available) ──
            _best_monthly  = _base_monthly + _concurrent_month
            _best_annual   = _base_annual  + _concurrent_annual

            # ── STORE — use best_case as primary display value ─────────────────
            # Base (excl-only) is the conservative floor.
            # Best case is the headline figure shown in the card.
            d['marginal_perc']     = marginal_historic / total_calls  # for KPI dedup only
            d['marginal_flights']  = _excl_flights          # exclusive zone flights
            d['marginal_deflected']= _excl_deflected
            d['shared_flights']    = _shared_dfr             # shared zone DFR flights
            d['zone_flights']      = _zone_flights           # total zone flights (for util)
            d['utilization']       = _util
            d['blocked_per_day']   = _concurrent_daily
            d['monthly_savings']   = _best_monthly           # headline = best case
            d['annual_savings']    = _best_annual
            d['base_annual']       = _base_annual            # conservative floor
            d['concurrent_annual'] = _concurrent_annual
            d['best_case_annual']  = _best_annual            # same as headline now
            d['concurrent_monthly']= _concurrent_month
            d['be_text']   = f"{d['cost']/_best_monthly:.1f} MO" if _best_monthly > 0 else "N/A"
            d['best_be_text'] = d['be_text']
        else:
            d.update({'assigned_indices':[],'annual_savings':0,'marginal_flights':0,
                      'marginal_deflected':0,'shared_flights':0,'be_text':"N/A",
                      'utilization':0,'concurrent_monthly':0,'best_case_annual':0,
                      'blocked_per_day':0,'best_be_text':"N/A",'base_annual':0,
                      'concurrent_annual':0,'zone_flights':0})
        active_drones.append(d)
        step += 1

    pop_metric = st.session_state.get('estimated_pop', 250000)
    grant_bracket = estimate_grants(pop_metric)
    st.sidebar.markdown(f"""
    <div style="margin-top:12px; background:{card_bg}; border:1px solid {budget_box_border}; padding:10px; border-radius:4px; margin-bottom:10px;">
        <div style="font-size:0.68rem; color:{text_muted}; font-weight:bold; text-transform:uppercase;">Est. Grant Eligibility</div>
        <div style="font-size:1.1rem; color:{budget_box_border}; font-weight:bold; font-family:monospace;">{grant_bracket}</div>
    </div>
    <div style="font-size:0.73rem; color:{text_muted}; line-height:1.5; margin-bottom:10px;">
        <a href="https://bja.ojp.gov/program/jag/overview" target="_blank" style="color:{accent_color}; font-weight:bold;">DOJ Byrne JAG</a> — UAS procurement eligible<br>
        <a href="https://www.fema.gov/grants/preparedness/homeland-security" target="_blank" style="color:{accent_color}; font-weight:bold;">FEMA HSGP</a> — CapEx offset for tactical deployments
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    if show_health:
        norm_redundancy = min(overlap_perc/35.0, 1.0)*100
        health_score = (calls_covered_perc*0.50) + (area_covered_perc*0.35) + (norm_redundancy*0.15)
        h_color, h_label = (accent_color,"OPTIMAL") if health_score>=80 else ("#94c11f","GOOD") if health_score>=70 else ("#ffc107","MARGINAL") if health_score>=55 else ("#dc3545","ESSENTIAL")
        st.markdown(f"""<div style="background:{card_bg}; border-left:5px solid {h_color}; border:1px solid {card_border};
            padding:10px; border-radius:4px; color:{text_main}; margin-bottom:10px;
            display:flex; align-items:center; justify-content:space-between;">
            <span style="font-size:1.4em; font-weight:bold; color:{h_color};">Department Health Score: {health_score:.1f}%</span>
            <span style="font-size:1.2em; background:rgba(128,128,128,0.15); padding:2px 10px; border-radius:4px;">{h_label}</span>
            </div>""", unsafe_allow_html=True)

    # Safely compute traffic impacts with strict float casting to prevent any TypeErrors
    if simulate_traffic:
        avg_ground_speed = float(CONFIG["DEFAULT_TRAFFIC_SPEED"]) * (1 - float(traffic_level) / 100.0)
        eval_dist  = float(guard_radius_mi if active_guard_names else resp_radius_mi)
        eval_speed = float(CONFIG["GUARDIAN_SPEED"] if active_guard_names else CONFIG["RESPONDER_SPEED"])
        
        if (active_resp_names or active_guard_names) and avg_ground_speed > 0:
            time_saved = ((eval_dist * 1.4 / avg_ground_speed) - (eval_dist / eval_speed)) * 60
            gain_val = f"{time_saved:.1f} min"
        else:
            gain_val = "N/A"
    else:
        gain_val = None

    orig_calls = int(st.session_state.get('total_original_calls', full_total_calls or (len(df_calls_full) if df_calls_full is not None else total_calls)) or total_calls)
    modeled_calls = int(st.session_state.get('total_modeled_calls', total_calls) or total_calls)
    displayed_points = len(display_calls) if display_calls is not None else 0
    call_str = f"{orig_calls:,}"

    # Calculate Date Range of CAD data (if available)
    date_range_str = "Simulated / Unknown"
    _date_src_df = df_calls_full if df_calls_full is not None else df_calls
    _label_dt = _detect_datetime_series_for_labels(_date_src_df)
    if _label_dt is not None:
        try:
            _label_dt = pd.to_datetime(_label_dt, errors='coerce').dropna()
            if not _label_dt.empty:
                min_date = _label_dt.min().strftime('%b %Y')
                max_date = _label_dt.max().strftime('%b %Y')
                date_range_str = f"{min_date} – {max_date}" if min_date != max_date else min_date
        except Exception:
            pass

    avg_resp_time = sum(d['avg_time_min'] for d in active_drones) / len(active_drones) if active_drones else 0.0

    # 1. THE SINGLE-LINE EXECUTIVE HEADER
    logo_b64 = get_base64_of_bin_file("logo.png")
    main_logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:24px; vertical-align:middle; margin-right:15px; filter: brightness(0) invert(1);">' if logo_b64 else f'<span style="font-size:1.5rem; font-weight:900; letter-spacing:2px; color:#ffffff; margin-right:15px;">BRINC</span>'

    header_html = f"""
    <div style="margin-top: 5px; margin-bottom: 15px; padding-bottom: 12px; border-bottom: 1px solid {card_border}; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px;">
        <div style="display: flex; align-items: center; flex-wrap: wrap; font-size: 0.9rem;">
            {main_logo_html}
            <span style="color: {accent_color}; font-family: 'IBM Plex Mono', monospace; font-size: 0.8rem; letter-spacing: 1px; text-transform: uppercase; margin-right: 12px;">Strategic Deployment Plan</span>
            <span style="font-weight: 800; color: {text_main}; font-size: 1.1rem; margin-right: 12px;">{st.session_state.get('active_city', 'Unknown City')}, {st.session_state.get('active_state', 'US')}</span>
            <span style="color: {text_muted}; margin-right: 12px;">• Serving {st.session_state.get('estimated_pop', 0):,} residents across ~{int(area_sq_mi):,} sq miles</span>
        </div>
        <div style="display: flex; align-items: center; font-size: 0.85rem; color: {text_muted}; gap: 15px;">
            <span>Data Period: <span style="color:#fff;">{date_range_str}</span></span>
            <span style="color:{card_border};">|</span>
            <span style="font-weight: 800; color: {text_main}; font-size: 0.95rem;">{actual_k_responder} <span style="color:#888; font-weight:normal;">Resp</span> · {actual_k_guardian} <span style="color:#888; font-weight:normal;">Guard</span></span>
        </div>
    </div>
    """
    st.markdown(header_html, unsafe_allow_html=True)

    # Cleanly evaluate dynamic CSS to avoid f-string syntax errors
    border_css = 'border-right: 1px solid #222; padding-right: 10px;' if gain_val is not None else ''

    # If traffic simulation is on, nest the time saved right inside the Avg Response box!
    if gain_val is not None:
        resp_content = (
            f'<div style="font-size: 2.2rem; font-weight: 800; color: {accent_color}; font-family: \'IBM Plex Mono\', monospace; line-height: 1.1;">{avg_resp_time:.1f}m</div>'
            f'<div style="font-size: 0.7rem; color: #39FF14; font-weight: 800; text-transform: uppercase; margin-top: 4px;">▼ Saves {gain_val}</div>'
        )
    else:
        resp_content = f'<div style="font-size: 2.2rem; font-weight: 800; color: {accent_color}; font-family: \'IBM Plex Mono\', monospace;">{avg_resp_time:.1f}m</div>'

    # 2. SPLIT KPI BAR — Guardian row + Responder row + combined summary
    def _kpi_cell(label, value, color=accent_color, border=True):
        br = f"border-right: 1px solid #222; padding-right: 10px;" if border else ""
        return (
            f'<div style="{br} text-align: center;">'
            f'<div style="font-size: 0.68rem; color: {text_muted}; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom:2px;">{label}</div>'
            f'<div style="font-size: 1.9rem; font-weight: 800; color: {color}; font-family: \'IBM Plex Mono\', monospace;">{value}</div>'
            f'</div>'
        )

    _GUARD_COL = "#FFD700"   # gold for Guardian
    _RESP_COL  = "#00D2FF"   # cyan for Responder
    _COMB_COL  = "#39FF14"   # green for combined

    kpi_html = (
        # ── Row 1: summary totals ──────────────────────────────────────────
        f'<div style="background:{card_bg}; border:1px solid {card_border}; border-radius:8px; padding:16px 20px; margin-bottom:8px;">'
        f'<div style="font-size:0.65rem; color:{text_muted}; text-transform:uppercase; letter-spacing:1px; margin-bottom:10px;">Fleet Summary</div>'
        f'<div style="display:grid; grid-template-columns:repeat(5,1fr); gap:8px;">'
        + _kpi_cell("Total Incidents", call_str)
        + _kpi_cell("Combined Coverage", f"{calls_covered_perc:.1f}%", _COMB_COL)
        + _kpi_cell("Land Covered", f"{area_covered_perc:.1f}%", _COMB_COL)
        + _kpi_cell("Zone Overlap", f"{overlap_perc:.1f}%", text_muted)
        + _kpi_cell("Avg Response", f"{avg_resp_time:.1f}m", accent_color, border=False)
        + f'</div></div>'

        # ── Row 2: Guardian-specific metrics ──────────────────────────────
        + f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:8px;">'

        + f'<div style="background:{card_bg}; border:1px solid #3a3000; border-top:3px solid {_GUARD_COL}; border-radius:8px; padding:14px 16px;">'
        + f'<div style="font-size:0.65rem; color:{_GUARD_COL}; text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; font-weight:700;">🦅 Guardian Fleet — {actual_k_guardian} unit{"s" if actual_k_guardian!=1 else ""} · {guard_strategy_raw}</div>'
        + f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:6px;">'
        + _kpi_cell("Call Coverage", f"{guard_calls_perc:.1f}%", _GUARD_COL)
        + _kpi_cell("Area Coverage", f"{guard_area_perc:.1f}%", _GUARD_COL, border=False)
        + f'</div></div>'

        # ── Row 3: Responder-specific metrics ─────────────────────────────
        + f'<div style="background:{card_bg}; border:1px solid #003a3a; border-top:3px solid {_RESP_COL}; border-radius:8px; padding:14px 16px;">'
        + f'<div style="font-size:0.65rem; color:{_RESP_COL}; text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; font-weight:700;">🚁 Responder Fleet — {actual_k_responder} unit{"s" if actual_k_responder!=1 else ""} · {resp_strategy_raw}</div>'
        + f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:6px;">'
        + _kpi_cell("Call Coverage", f"{resp_calls_perc:.1f}%", _RESP_COL)
        + _kpi_cell("Area Coverage", f"{resp_area_perc:.1f}%", _RESP_COL, border=False)
        + f'</div></div>'

        + f'</div>'
    )
    
    st.markdown(kpi_html, unsafe_allow_html=True)
    if orig_calls != modeled_calls:
        model_note = f"Optimization modeled via {modeled_calls:,} representative CAD samples from {orig_calls:,} total incidents."
    else:
        model_note = f"Optimization modeled via all {modeled_calls:,} available incidents."
    if displayed_points and displayed_points < orig_calls:
        map_note = f"Map renders {displayed_points:,} incident points using Plotly's WebGL-backed map layer for dense full-call visualization."
    elif displayed_points:
        map_note = f"Map renders all {displayed_points:,} incident points."
    else:
        map_note = ""
    full_daily_note = f"Citywide daily-call estimates are based on the full uploaded CAD total of {orig_calls:,} incidents." if orig_calls else ""
    note_bits = [model_note]
    if map_note:
        note_bits.append(map_note)
    if full_daily_note:
        note_bits.append(full_daily_note)
    st.markdown(f"<div style='font-size:0.65rem;color:gray;margin-top:-10px;margin-bottom:12px;text-align:right;'>{' '.join(note_bits)}</div>", unsafe_allow_html=True)

    cards_below_map = bool(show_cards)
    map_col = st.container()

    with map_col:
        fig = go.Figure()

        if show_boundaries and city_boundary_geom is not None and not city_boundary_geom.is_empty:
            geoms_to_draw = [city_boundary_geom] if isinstance(city_boundary_geom, Polygon) else list(city_boundary_geom.geoms)
            for gi, geom in enumerate(geoms_to_draw):
                bx, by = geom.exterior.coords.xy
                fig.add_trace(go.Scattermapbox(mode="lines", lon=list(bx), lat=list(by),
                    line=dict(color=map_boundary_color, width=2), name="Jurisdiction Boundary",
                    hoverinfo='skip', showlegend=(gi==0)))

        if show_heatmap and not display_calls.empty:
            fig.add_trace(go.Densitymapbox(lat=display_calls.geometry.y, lon=display_calls.geometry.x,
                z=np.ones(len(display_calls)), radius=12, colorscale='Inferno', opacity=0.6,
                showscale=False, name="Heatmap", hoverinfo='skip'))

        if not display_calls.empty:
            point_size = 1 if len(display_calls) > 150000 else 2 if len(display_calls) > 50000 else 3 if len(display_calls) > 20000 else 4
            point_opacity = 0.06 if len(display_calls) > 150000 else 0.10 if len(display_calls) > 50000 else 0.18 if len(display_calls) > 20000 else 0.28 if len(display_calls) > 10000 else 0.4
            fig.add_trace(go.Scattermapbox(lat=display_calls.geometry.y, lon=display_calls.geometry.x,
                mode='markers', marker=dict(size=point_size, color=map_incident_color, opacity=point_opacity),
                name="Incident Data", hoverinfo='skip'))

        if show_faa and faa_geojson:
            add_faa_laanc_layer_to_plotly(fig, faa_geojson, is_dark=not show_satellite)

        for d in active_drones:
            clats, clons = get_circle_coords(d['lat'], d['lon'], r_mi=d['radius_m']/1609.34)
            lbl = f"{d['name'].split(',')[0]} ({'Resp' if d['type']=='RESPONDER' else 'Guard'})"
            
            # Determine if this is an extended Guardian (so we can relax the outer ring)
            is_extended_guardian = (d['type'] == 'GUARDIAN' and d['radius_m']/1609.34 > 5.0)
            
            # The outer ring becomes relaxed (thinner, more transparent) if > 5 miles
            outer_width = 1.5 if is_extended_guardian else 4.5
            outer_opac = 0.4 if is_extended_guardian else 1.0
            
            fig.add_trace(go.Scattermapbox(
                lat=list(clats)+[None,d['lat']], lon=list(clons)+[None,d['lon']],
                mode='lines+markers',
                opacity=outer_opac,
                marker=dict(size=[0]*len(clats)+[0,20], color=d['color']),
                line=dict(color=d['color'], width=outer_width),
                fill='toself', fillcolor='rgba(0,0,0,0)', name=lbl, hoverinfo='name'))

            # The 5-mile Rapid Response ring gets the "Important" styling (thick, solid, heavier fill)
            if is_extended_guardian:
                f_lats, f_lons = get_circle_coords(d['lat'], d['lon'], r_mi=5.0)
                fig.add_trace(go.Scattermapbox(
                    lat=list(f_lats), lon=list(f_lons),
                    mode='lines',
                    line=dict(color=d['color'], width=4.5),
                    opacity=1.0,
                    fill='toself',
                    fillcolor=f"rgba({int(d['color'][1:3],16)},{int(d['color'][3:5],16)},{int(d['color'][5:7],16)},0.12)",
                    name=f"Rapid Response 5mi · {d['name'].split(',')[0]}",
                    hoverinfo='text',
                    text=f"⚡ Rapid Response Focus Zone — 5mi<br>{d['name'].split(',')[0]}",
                    showlegend=False
                ))

            # Star marker for manually pinned stations
            if d.get('pinned'):
                fig.add_trace(go.Scattermapbox(
                    lat=[d['lat']], lon=[d['lon']], mode='markers',
                    marker=dict(size=18, color=d['color'], symbol='star'),
                    name=f"📍 {d['name'].split(',')[0]} (Pinned)",
                    hovertemplate=f"<b>🔒 PINNED</b><br>{d['name']}<br>{d['type']}<extra></extra>",
                    showlegend=False
                ))

            if simulate_traffic:
                t_color = "#28a745" if traffic_level<35 else "#ffc107" if traffic_level<75 else "#dc3545"
                t_fill  = f"rgba({'40,167,69' if traffic_level<35 else '255,193,7' if traffic_level<75 else '220,53,69'}, 0.15)"
                t_label = "Light" if traffic_level<35 else "Moderate" if traffic_level<75 else "Heavy"
                gs = CONFIG["DEFAULT_TRAFFIC_SPEED"]*(1-traffic_level/100)
                if gs > 0:
                    gr_mi = (gs/60) * (d['radius_m']/1609.34/d['speed_mph'])*60
                    ga = np.linspace(0,2*np.pi,9)
                    fig.add_trace(go.Scattermapbox(
                        lat=list(d['lat']+(gr_mi/69.172)*np.sin(ga)),
                        lon=list(d['lon']+(gr_mi/(69.172*np.cos(np.radians(d['lat']))))*np.cos(ga)),
                        mode='lines', line=dict(color=t_color, width=2.5),
                        fill='toself', fillcolor=t_fill,
                        name=f"Ground ({t_label})", hoverinfo='skip'))

        mapbox_cfg = dict(center=dict(lat=center_lat, lon=center_lon), zoom=dynamic_zoom, style=map_style)
        if show_satellite:
            mapbox_cfg["style"] = "carto-positron"
            mapbox_cfg["layers"] = [{"below":"traces","sourcetype":"raster",
                "sourceattribution":"Esri, Maxar, Earthstar Geographics",
                "source":["https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}"]}]

        fig.update_layout(uirevision="LOCKED_MAP", mapbox=mapbox_cfg,
            margin=dict(l=0,r=0,t=0,b=0), height=800, font=dict(size=18),
            showlegend=True,
            legend=dict(yanchor="top", y=0.98, xanchor="left", x=0.02,
                        bgcolor=legend_bg, bordercolor=accent_color, borderwidth=1,
                        font=dict(size=12, color=legend_text), itemclick="toggle"))

        st.plotly_chart(fig, use_container_width=True, config={"scrollZoom": True})

    # ── UNIT ECONOMICS CARDS (directly below map, no toggle) ─────────────────
    st.markdown("---")
    st.markdown(f"<h4 style='margin-top:2px; border-bottom:1px solid {card_border}; padding-bottom:8px; color:{text_main};'>Unit Economics</h4>", unsafe_allow_html=True)
    st.markdown(
        f"<div style='font-size:0.6rem; color:#666; background:rgba(240,180,41,0.07); border-left:3px solid #F0B429; padding:5px 8px; border-radius:0 3px 3px 0; margin-bottom:10px;'>{SIMULATOR_DISCLAIMER_SHORT}</div>",
        unsafe_allow_html=True
    )
    st.markdown(
        """<style>
        .unit-card { transition: transform 0.2s ease-out, box-shadow 0.2s ease-out; }
        .unit-card:hover { transform: translateY(-2px); box-shadow: 0 8px 20px rgba(0,210,255,0.12); }
        </style>""",
        unsafe_allow_html=True
    )
    if active_drones:
        _n_cols = 4 if len(active_drones) >= 4 else (2 if len(active_drones) > 1 else 1)
        _saved_gnames = list(st.session_state.get('pinned_guard_names', []))
        _saved_rnames = list(st.session_state.get('pinned_resp_names',  []))

        # Render card HTML + lock buttons together inside st.columns so buttons
        # are always directly below their card in the same Streamlit column.
        import math as _math
        _n_rows = _math.ceil(len(active_drones) / _n_cols)
        for _row_idx in range(_n_rows):
            _row_drones = active_drones[_row_idx * _n_cols : (_row_idx + 1) * _n_cols]
            _row_cols   = st.columns(_n_cols)
            for _slot, _d in enumerate(_row_drones):
                _ci    = _row_idx * _n_cols + _slot
                _dname = _d['name']
                _dtype = _d['type']
                _is_pg = _dname in _saved_gnames
                _is_pr = _dname in _saved_rnames
                with _row_cols[_slot]:
                    # ── Card HTML ─────────────────────────────────────────────
                    st.markdown(
                        _build_unit_cards_html(
                            [_d], text_main, text_muted, card_bg, card_border,
                            card_title, accent_color, columns_per_row=1
                        ),
                        unsafe_allow_html=True
                    )
                    # ── Lock / Switch / Unpin buttons ─────────────────────────
                    if _is_pg or _is_pr:
                        _switch_label = "🚁 Switch to Resp" if _is_pg else "🦅 Switch to Guard"
                        _bc1, _bc2 = st.columns([3, 2])
                        with _bc1:
                            if st.button(_switch_label, key=f"switch_{_ci}",
                                         use_container_width=True):
                                if _is_pg:
                                    st.session_state['pinned_guard_names'] = [x for x in _saved_gnames if x != _dname]
                                    _pr = list(st.session_state.get('pinned_resp_names', []))
                                    if _dname not in _pr: _pr.append(_dname)
                                    st.session_state['pinned_resp_names'] = _pr
                                    if st.session_state.get('k_resp', 0) < len(_pr):
                                        st.session_state['k_resp'] = len(_pr)
                                else:
                                    st.session_state['pinned_resp_names'] = [x for x in _saved_rnames if x != _dname]
                                    _pg = list(st.session_state.get('pinned_guard_names', []))
                                    if _dname not in _pg: _pg.append(_dname)
                                    st.session_state['pinned_guard_names'] = _pg
                                    if st.session_state.get('k_guard', 0) < len(_pg):
                                        st.session_state['k_guard'] = len(_pg)
                                st.rerun()
                        with _bc2:
                            if st.button("✕ Unpin", key=f"unpin_{_ci}",
                                         use_container_width=True, type="primary"):
                                st.session_state['pinned_guard_names'] = [x for x in _saved_gnames if x != _dname]
                                st.session_state['pinned_resp_names']  = [x for x in _saved_rnames if x != _dname]
                                st.rerun()
                    else:
                        _ba, _bb = st.columns(2)
                        with _ba:
                            if st.button("🦅 lock as guard", key=f"pin_g_{_ci}",
                                         use_container_width=True):
                                _pg = list(st.session_state.get('pinned_guard_names', []))
                                if _dname not in _pg: _pg.append(_dname)
                                st.session_state['pinned_guard_names'] = _pg
                                st.session_state['pinned_resp_names']  = [x for x in _saved_rnames if x != _dname]
                                if st.session_state.get('k_guard', 0) < len(_pg):
                                    st.session_state['k_guard'] = len(_pg)
                                st.rerun()
                        with _bb:
                            if st.button("🚁 lock as resp", key=f"pin_r_{_ci}",
                                         use_container_width=True):
                                _pr = list(st.session_state.get('pinned_resp_names', []))
                                if _dname not in _pr: _pr.append(_dname)
                                st.session_state['pinned_resp_names']  = _pr
                                st.session_state['pinned_guard_names'] = [x for x in _saved_gnames if x != _dname]
                                if st.session_state.get('k_resp', 0) < len(_pr):
                                    st.session_state['k_resp'] = len(_pr)
                                st.rerun()
    else:
        st.markdown(
            f"""
            <div style="background:{card_bg}; border:1px dashed {card_border}; border-radius:6px; padding:22px; text-align:center; margin-top:8px;">
                <div style="font-size:2rem; margin-bottom:8px;">🚁</div>
                <div style="font-weight:700; color:{text_main}; margin-bottom:6px;">No drones deployed yet</div>
                <div style="font-size:0.8rem; color:{text_muted};">
                    Use the <b>Responder / Guardian Count</b> sliders in the sidebar to deploy drones and see per-unit economics here.
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    # ── COVERAGE CURVE + STATION RING CHART (side by side, directly below cards) ──
    st.markdown("---")
    st.markdown(f"<h4 style='border-bottom:1px solid {card_border}; padding-bottom:8px; color:{text_main};'>Coverage Curve</h4>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:0.8rem; color:{text_muted}; margin-bottom:8px;'>How added drones improve coverage — and where returns flatten.</div>", unsafe_allow_html=True)

    _curve_col, _ring_col = st.columns([3, 2], gap="medium")

    with _curve_col:
        if not df_curve.empty:
            fig_curve = go.Figure()
            for col, color, dash in [('Responder (Calls)',accent_color,'solid'),('Guardian (Calls)','#FFD700','solid'),
                                      ('Responder (Area)',accent_color,'dash'),('Guardian (Area)','#FFD700','dash')]:
                y_data = df_curve[col].dropna()
                x_data = df_curve.loc[y_data.index,'Drones']
                if not y_data.empty:
                    fig_curve.add_trace(go.Scatter(x=x_data, y=y_data, mode='lines+markers', name=col,
                        line=dict(color=color,width=2,dash=dash), marker=dict(size=4),
                        hovertemplate=f"<b>{col}</b><br>Drones: %{{x}}<br>Coverage: %{{y:.1f}}%<extra></extra>"))
                    if 'Calls' in col:
                        idx_90 = y_data[y_data >= 90.0].first_valid_index()
                        if idx_90 is not None:
                            fig_curve.add_trace(go.Scatter(x=[int(x_data.loc[idx_90])], y=[y_data.loc[idx_90]],
                                mode='markers', marker=dict(color=color,size=12,symbol='star',line=dict(color='white',width=1)),
                                showlegend=False, hoverinfo='skip'))
            fig_curve.update_layout(
                xaxis_title="Drones", yaxis_title="Coverage %",
                xaxis=dict(showgrid=True, gridcolor=card_border, tickfont=dict(color=text_muted)),
                yaxis=dict(showgrid=True, gridcolor=card_border, tickfont=dict(color=text_muted),
                           tickvals=[0,20,40,60,80,90,100], range=[0,105]),
                legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1,
                            font=dict(size=9,color=text_muted)),
                margin=dict(l=10,r=10,t=20,b=10), height=320,
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                hoverlabel=dict(bgcolor=card_bg, font_size=13, font_color=text_main, bordercolor=accent_color)
            )
            st.plotly_chart(fig_curve, use_container_width=True, config={'displayModeBar':False})
        else:
            st.info("Run optimization to generate coverage curve.")

    with _ring_col:
        # Split ring: outer ring = Guardians (gold), inner ring = Responders (cyan)
        if active_drones and total_calls > 0:
            _g_drones = [d for d in active_drones if d['type'] == 'GUARDIAN']
            _r_drones = [d for d in active_drones if d['type'] == 'RESPONDER']

            def _build_ring_data(drones, fleet_cov_mask):
                """Build labels/values/colors for one fleet's ring slices."""
                labels, values, colors = [], [], []
                remaining = int(fleet_cov_mask.sum()) if fleet_cov_mask is not None else 0
                for d in drones:
                    _m = int(d.get('marginal_perc', 0) * total_calls)
                    if _m > 0:
                        labels.append(d['name'].split(',')[0][:18])
                        values.append(_m)
                        colors.append(d['color'])
                        remaining = max(0, remaining - _m)
                return labels, values, colors

            _g_labels, _g_vals, _g_cols = _build_ring_data(_g_drones, cov_g)
            _r_labels, _r_vals, _r_cols = _build_ring_data(_r_drones, cov_r)

            # Uncovered slice for combined view
            _combined_covered = int(np.logical_or(cov_r, cov_g).sum()) if total_calls > 0 else 0
            _uncovered = max(0, total_calls - _combined_covered)

            # Build a single donut: Guardian slices (gold ring) + Responder slices (cyan ring)
            # separated by a small "uncovered" gap
            all_labels = _g_labels + _r_labels + (["Uncovered"] if _uncovered > 0 else [])
            all_values = _g_vals   + _r_vals   + ([_uncovered] if _uncovered > 0 else [])
            all_colors = _g_cols   + _r_cols   + (["#1a1a1a"] if _uncovered > 0 else [])

            if all_values:
                fig_ring = go.Figure(go.Pie(
                    labels=all_labels,
                    values=all_values,
                    hole=0.58,
                    marker=dict(colors=all_colors, line=dict(color='#000', width=1.5)),
                    textinfo='none',
                    hovertemplate='<b>%{label}</b><br>%{value:,} calls (%{percent})<extra></extra>',
                    sort=False,
                ))
                _cov_pct = round(_combined_covered / total_calls * 100, 1)
                _mode_short = "▶◀" if complement_mode else "↔" if shared_mode else "⊕"
                fig_ring.update_layout(
                    annotations=[dict(
                        text=f"<b>{_cov_pct}%</b><br><span style='font-size:9px'>{_mode_short} combined</span>",
                        x=0.5, y=0.5, font_size=15, showarrow=False,
                        font=dict(color=text_main)
                    )],
                    showlegend=True,
                    legend=dict(
                        orientation='v', x=1.02, y=0.5,
                        font=dict(size=9, color=text_muted),
                        bgcolor='rgba(0,0,0,0)',
                        groupclick='toggleitem',
                    ),
                    margin=dict(l=0, r=0, t=10, b=10),
                    height=320,
                    paper_bgcolor='rgba(0,0,0,0)',
                    hoverlabel=dict(bgcolor=card_bg, font_size=12, font_color=text_main),
                )
                st.plotly_chart(fig_ring, use_container_width=True, config={'displayModeBar':False})

                # Mode legend below the ring
                _mode_label = {
                    "Complement — push apart": "▶◀ Complement — Responders fill Guardian gaps",
                    "Independent — each maximises own area": "⊕ Independent — each fleet optimised separately",
                    "Shared — allow full overlap": "↔ Shared — both fleets maximise same call set",
                }.get(deployment_mode, "")
                st.markdown(
                    f"<div style='font-size:0.65rem; color:{text_muted}; text-align:center; margin-top:-8px;'>{_mode_label}</div>",
                    unsafe_allow_html=True
                )
        else:
            st.markdown(
                f"<div style='color:{text_muted}; font-size:0.8rem; padding:40px 0; text-align:center;'>Deploy drones to see call distribution ring.</div>",
                unsafe_allow_html=True
            )


    # Resolve real incident datetime coverage for labels on the stations page
    _label_dt_series = _detect_datetime_series_for_labels(df_calls_full if df_calls_full is not None else df_calls)
    _label_has_real_dates = _label_dt_series is not None and getattr(_label_dt_series, "notna", lambda: pd.Series([], dtype=bool))().sum() > 0

    # ── CAD DATA CHARTS (moved into CAD Ingestion Analytics below) ───────────
    _cad_src = st.session_state.get('data_source', '')
    _has_real_calls = _cad_src in ('cad_upload', 'brinc_file') or (
        'df_calls' in st.session_state and st.session_state['df_calls'] is not None
        and len(st.session_state['df_calls']) > 100
    )

    # ── 3D SWARM SIMULATION ───────────────────────────────────────────
    if fleet_capex > 0:
        st.markdown("---")
        st.markdown(f"<h3 style='color:{text_main};'>🚁 3D Swarm Simulation</h3>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:0.82rem; color:{text_muted}; margin-bottom:10px;'>Animated deck.gl simulation of all DFR flights over a compressed 24-hour day. Use the speed slider to accelerate or slow the simulation. Great for council presentations.</div>", unsafe_allow_html=True)

        show_sim = st.toggle("🎬 Enable 3D Simulation", value=False)
        if show_sim:
            calls_lonlat = calls_in_city.to_crs(epsg=4326)
            calls_coords = np.column_stack((calls_lonlat.geometry.x, calls_lonlat.geometry.y))

            sim_assignments = {i:[] for i in range(len(active_drones))}
            for c_idx, cc in enumerate(calls_coords):
                best_d, best_dist = -1, float('inf')
                for d_idx, d in enumerate(active_drones):
                    if d['cov_array'][c_idx] if c_idx < len(d['cov_array']) else False:
                        dist = (cc[0]-d['lon'])**2 + (cc[1]-d['lat'])**2
                        if dist < best_dist:
                            best_dist, best_d = dist, d_idx
                if best_d != -1:
                    sim_assignments[best_d].append(c_idx)

            stations_json, flights_json, legend_html_sim = [], [], ""
            total_sim_flights = 0
            for d_idx, d in enumerate(active_drones):
                hex_c = d['color'].lstrip('#')
                rgb = [int(hex_c[j:j+2],16) for j in (0,2,4)]
                stations_json.append({"name":d['name'].split(',')[0][:30],"lon":d['lon'],"lat":d['lat'],"color":rgb,"radius":d['radius_m']})
                legend_html_sim += f'<div style="margin-bottom:3px;"><span style="display:inline-block;width:9px;height:9px;background:{d["color"]};margin-right:7px;border-radius:50%;"></span>{d["name"].split(",")[0][:28]} ({d["type"][:3]})</div>'
                frac = len(sim_assignments[d_idx])/len(calls_coords) if calls_coords.shape[0]>0 else 0
                monthly_for_drone = int(frac * calls_per_day * 30 * dfr_dispatch_rate)
                pool = sim_assignments[d_idx]

                if not pool: sim_calls = []
                elif monthly_for_drone > len(pool): sim_calls = random.choices(pool, k=monthly_for_drone)
                else: sim_calls = random.sample(pool, monthly_for_drone)

                total_sim_flights += len(sim_calls)
                for ci in sim_calls:
                    lon1,lat1 = calls_coords[ci]
                    lon0,lat0 = d['lon'],d['lat']
                    dist_mi = math.sqrt((lon1-lon0)**2+(lat1-lat0)**2)*69.172
                    vis_time = max((dist_mi/d['speed_mph'])*3600*8, 240)
                    launch = random.randint(0, 2592000)
                    arc_h = min(max(dist_mi*90, 80), 400)
                    t0 = launch
                    t1 = launch + vis_time * 0.15
                    t2 = launch + vis_time * 0.40
                    t3 = launch + vis_time * 0.75
                    t4 = launch + vis_time * 0.90
                    t5 = launch + vis_time
                    mx1 = lon0 + 0.15*(lon1-lon0);  my1 = lat0 + 0.15*(lat1-lat0)
                    mx2 = lon0 + 0.35*(lon1-lon0);  my2 = lat0 + 0.35*(lat1-lat0)
                    mx3 = lon0 + 0.65*(lon1-lon0);  my3 = lat0 + 0.65*(lat1-lat0)
                    mx4 = lon0 + 0.85*(lon1-lon0);  my4 = lat0 + 0.85*(lat1-lat0)
                    flights_json.append({
                        "path": [[lon0, lat0, 0], [mx1, my1, arc_h*0.75], [mx2, my2, arc_h], [mx3, my3, arc_h], [mx4, my4, arc_h*0.75], [lon1, lat1, 0]],
                        "timestamps": [t0, t1, t2, t3, t4, t5],
                        "color": rgb
                    })

            warn_html_sim = ""
            if len(flights_json) > 3000:
                flights_json = random.sample(flights_json, 3000)
                warn_html_sim = f'<div style="background:#440000;border:1px solid #ff4b4b;color:#ffbbbb;padding:5px;font-size:10px;border-radius:4px;margin-bottom:8px;">⚠️ Capped at 3,000 flights for performance (actual: {total_sim_flights:,})</div>'

            drone_svg = "data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='white'%3E%3Cpath d='M18 6a2 2 0 100-4 2 2 0 000 4zm-12 0a2 2 0 100-4 2 2 0 000 4zm12 12a2 2 0 100-4 2 2 0 000 4zm-12 0a2 2 0 100-4 2 2 0 000 4z'/%3E%3Cpath stroke='white' stroke-width='2' stroke-linecap='round' d='M8.5 8.5l7 7m0-7l-7 7'/%3E%3Ccircle cx='12' cy='12' r='2' fill='white'/%3E%3C/svg%3E"

            sim_html = f"""<!DOCTYPE html><html><head>
            <script src="https://unpkg.com/deck.gl@8.9.35/dist.min.js"></script>
            <script src="https://unpkg.com/maplibre-gl@3.0.0/dist/maplibre-gl.js"></script>
            <link href="https://unpkg.com/maplibre-gl@3.0.0/dist/maplibre-gl.css" rel="stylesheet"/>
            <style>
              body{{margin:0;padding:0;overflow:hidden;background:#000;font-family:Manrope,sans-serif;}}
              #map{{width:100vw;height:100vh;position:absolute;}}
              #ui{{position:absolute;top:16px;left:16px;background:rgba(17,17,17,0.92);padding:16px;border-radius:8px;
                   color:white;border:1px solid #333;z-index:10;box-shadow:0 4px 10px rgba(0,0,0,0.5);width:260px;}}
              button{{background:#00D2FF;color:black;border:none;padding:10px;cursor:pointer;font-weight:bold;
                      border-radius:4px;width:100%;font-size:13px;text-transform:uppercase;margin-bottom:8px;}}
              button:disabled{{background:#444;color:#888;cursor:not-allowed;}}
              #timeDisplay{{font-family:monospace;font-size:16px;color:#00ffcc;font-weight:bold;text-align:center;margin-bottom:8px;}}
            </style></head><body>
            <div id="ui">
              <h3 style="margin:0 0 8px;color:#00D2FF;font-size:14px;">DFR SWARM SIMULATION</h3>
              {warn_html_sim}
              <div style="font-size:11px;color:#aaa;margin-bottom:10px;">
                {total_sim_flights:,} flights over 30 days at {int(dfr_dispatch_rate*100)}% dispatch rate
              </div>
              <div style="margin-bottom:10px;">
                <label style="font-size:11px;color:#ccc;">Speed: <span id="speedLabel">1</span>x</label>
                <input type="range" id="speedSlider" min="1" max="100" value="1" style="width:100%;margin-top:4px;">
              </div>
              <button id="runBtn">▶ LAUNCH SWARM</button>
              <div id="timeDisplay">00:00</div>
              <div style="margin-top:10px;border-top:1px solid #333;padding-top:8px;">
                <div style="font-size:10px;color:#888;text-transform:uppercase;margin-bottom:5px;">Stations</div>
                <div style="font-size:10px;color:#ddd;max-height:100px;overflow-y:auto;">{legend_html_sim}</div>
              </div>
            </div>
            <div id="map"></div>
            <script>
              const stations={json.dumps(stations_json)};
              const flights={json.dumps(flights_json)};
              const speedSlider=document.getElementById('speedSlider');
              const speedLabel=document.getElementById('speedLabel');
              speedSlider.oninput=()=>speedLabel.innerText=speedSlider.value;
              const map=new deck.DeckGL({{
                container:'map',
                mapStyle:'https://basemaps.cartocdn.com/gl/dark-matter-gl-style/style.json',
                initialViewState:{{longitude:{center_lon},latitude:{center_lat},zoom:{dynamic_zoom},pitch:50,bearing:0}},
                controller:true
              }});
              let time=0,timer=null,lastTime=0;
              function render(){{
                map.setProps({{layers:[
                  new deck.ScatterplotLayer({{id:'rings',data:stations,getPosition:d=>[d.lon,d.lat],
                    getFillColor:d=>[d.color[0],d.color[1],d.color[2],25],
                    getLineColor:d=>[d.color[0],d.color[1],d.color[2],220],
                    lineWidthMinPixels:2,stroked:true,filled:true,getRadius:d=>d.radius}}),
                  new deck.ScatterplotLayer({{id:'pads',data:stations,getPosition:d=>[d.lon,d.lat],
                    getFillColor:d=>[d.color[0],d.color[1],d.color[2],120],getRadius:180}}),
                  new deck.IconLayer({{id:'icons',data:stations,
                    getIcon:d=>({{url:"{drone_svg}",width:24,height:24,anchorY:12}}),
                    getPosition:d=>[d.lon,d.lat],getSize:36,sizeScale:1}}),
                  new deck.TripsLayer({{id:'flights',data:flights,getPath:d=>d.path,
                    getTimestamps:d=>d.timestamps,getColor:d=>d.color,
                    opacity:0.85,widthMinPixels:5,trailLength:13500,currentTime:time,rounded:true}}),
                  new deck.ScatterplotLayer({{id:'landed',data:flights,getPosition:d=>d.path[5],
                    getFillColor:d=>time>=d.timestamps[5]?[d.color[0],d.color[1],d.color[2],255]:[0,0,0,0],
                    getRadius:25,radiusMinPixels:3,updateTriggers:{{getFillColor:time}}}})
                ]}});
                let day=Math.floor(time/86400)+1;
                let h=Math.floor((time%86400)/3600).toString().padStart(2,'0');
                let m=Math.floor((time%3600)/60).toString().padStart(2,'0');
                document.getElementById('timeDisplay').innerText=`Day ${{day}} · ${{h}}:${{m}}`;
              }}
              const animate=()=>{{
                let now=performance.now();
                let dt=Math.min(now-lastTime,100);
                lastTime=now;
                time+=dt/1000*43200*parseFloat(speedSlider.value);
                render();
                if(time<2592000){{timer=requestAnimationFrame(animate);}}
                else{{
                  document.getElementById('runBtn').disabled=false;
                  document.getElementById('runBtn').innerText='↺ RESTART';
                  time=0;
                }}
              }};
              document.getElementById('runBtn').onclick=()=>{{
                document.getElementById('runBtn').disabled=true;
                document.getElementById('runBtn').innerText='SIMULATING…';
                time=0;lastTime=performance.now();
                if(timer)cancelAnimationFrame(timer);
                animate();
              }};
              render();
            </script></body></html>"""

            components.html(sim_html, height=700)

    # ── COMMAND CENTER ANALYTICS DASHBOARD ──
    st.markdown("---")
    st.markdown(f"<h3 style='color:{text_main};'>📊 CAD Ingestion Analytics</h3>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:0.82rem; color:{text_muted}; margin-bottom:10px;'>Temporal patterns derived from your uploaded CAD data — hourly volumes, day-of-week distribution, optimal DFR shift windows, and a higher-contrast 5-band call-volume calendar.</div>", unsafe_allow_html=True)

    _analytics_df = df_calls_full if (df_calls_full is not None and not df_calls_full.empty) else df_calls
    analytics_html_block = generate_command_center_html(
        _analytics_df,
        total_orig_calls=st.session_state.get('total_original_calls', full_total_calls or total_calls)
    )
    components.html(analytics_html_block, height=1700, scrolling=False)

    if _has_real_calls and _analytics_df is not None and not _analytics_df.empty:
        # Collapse gap between components.html block and the plotly charts below
        st.markdown("<div style='margin-top:-32px;'></div>", unsafe_allow_html=True)
        _build_cad_charts(_analytics_df, text_main, text_muted, card_bg, card_border, accent_color)

    # ── EXPORT BUTTONS — always visible in sidebar ──
    st.sidebar.markdown("---")

    brinc_user = st.sidebar.text_input("BRINC Email Prefix (first.last)", value=st.session_state.get('brinc_user', 'steven.beltran'), key='brinc_user', help="Enter 'first.last' to auto-generate your name and @brincdrones.com email address.")
    st.sidebar.caption("*(Press **Enter** after typing to apply changes)*")

    user_clean = brinc_user.strip()
    if not user_clean: user_clean = "steven.beltran"

    # Police dept fields removed from sidebar — read from session state defaults
    pd_chief_name = st.session_state.get('pd_chief_name', '')
    pd_dept_name  = st.session_state.get('pd_dept_name', '')
    pd_dept_email = st.session_state.get('pd_dept_email', '')
    pd_dept_phone = st.session_state.get('pd_dept_phone', '')
    prop_email = f"{user_clean}@brincdrones.com"
    prop_name = " ".join([word.capitalize() for word in user_clean.split('.')])

    # Always define these so download buttons work regardless of fleet_capex
    prop_city  = st.session_state.get('active_city', 'City')
    prop_state = st.session_state.get('active_state', 'FL')
    _safe_city_base = prop_city.replace(" ", "_").replace("/", "_")

    if fleet_capex > 0:

        prop_city  = st.session_state.get('active_city', 'City')
        prop_state = st.session_state.get('active_state', 'FL')

        # Police dept signatory — falls back gracefully if not filled in
        pd_chief  = pd_chief_name.strip()  if pd_chief_name.strip()  else f"Chief of Police, {prop_city}"
        pd_dept   = pd_dept_name.strip()   if pd_dept_name.strip()   else f"{prop_city} Police Department"
        pd_email  = pd_dept_email.strip()  if pd_dept_email.strip()  else ""
        pd_phone  = pd_dept_phone.strip()  if pd_dept_phone.strip()  else ""
        pd_email_html = f'📧 <a href="mailto:{pd_email}">{pd_email}</a><br>' if pd_email else ""
        pd_phone_html = f'📞 {pd_phone}<br>' if pd_phone else ""
        pop_metric = st.session_state.get('estimated_pop', 250000)
        current_time_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        safe_city = prop_city.replace(" ","_").replace("/","_")

        _session_start = st.session_state.get('session_start', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        try:
            _start_dt = datetime.datetime.strptime(_session_start, '%Y-%m-%d %H:%M:%S')
            _dur_min  = round((datetime.datetime.now() - _start_dt).total_seconds() / 60, 1)
        except Exception:
            _dur_min = ''

        avg_resp_time    = sum(d['avg_time_min'] for d in active_drones) / len(active_drones) if active_drones else 0.0
        avg_ground_speed = CONFIG["DEFAULT_TRAFFIC_SPEED"] * (1 - traffic_level / 100)
        avg_time_saved   = ((sum((d['radius_m']/1609.34*1.4/avg_ground_speed)*60 for d in active_drones) / len(active_drones)) - avg_resp_time) if active_drones and avg_ground_speed > 0 else 0.0
        _calls_lons = (df_calls_full if df_calls_full is not None else df_calls)['lon'].dropna()
        _calls_lats = (df_calls_full if df_calls_full is not None else df_calls)['lat'].dropna()
        minx = float(_calls_lons.min()) if len(_calls_lons) else 0
        maxx = float(_calls_lons.max()) if len(_calls_lons) else 0
        miny = float(_calls_lats.min()) if len(_calls_lats) else 0
        maxy = float(_calls_lats.max()) if len(_calls_lats) else 0
        area_sq_mi_est   = max(1, int((maxx - minx) * (maxy - miny) * 3280))
        export_details = {
            # Session
            "session_id":            st.session_state.get('session_id', ''),
            "session_start":         _session_start,
            "session_duration_min":  _dur_min,
            "data_source":           st.session_state.get('data_source', 'unknown'),
            # Jurisdiction
            "population":            pop_metric,
            "total_calls":           st.session_state.get('total_original_calls', 0),
            "daily_calls":           max(1, int(st.session_state.get('total_original_calls', 0) / 365)),
            "area_sq_mi":            area_sq_mi_est,
            # Financials
            "opt_strategy":          opt_strategy,
            "incremental_build":     incremental_build,
            "allow_redundancy":      allow_redundancy,
            "dfr_rate":              int(dfr_dispatch_rate*100),
            "deflect_rate":          int(deflection_rate*100),
            "fleet_capex":           fleet_capex,
            "annual_savings":        annual_savings,
            "break_even":            break_even_text,
            # Operational
            "avg_response_min":      round(avg_resp_time, 2),
            "avg_time_saved_min":    round(avg_time_saved, 2),
            "area_covered_pct":      round(area_covered_perc, 1),
            # PD Contact
            "pd_chief":              st.session_state.get('pd_chief_name', ''),
            "pd_dept":               st.session_state.get('pd_dept_name', ''),
            "pd_dept_email":         st.session_state.get('pd_dept_email', ''),
            "pd_dept_phone":         st.session_state.get('pd_dept_phone', ''),
            # Drones
            "active_drones": [{
                "name": d['name'], "type": d['type'],
                "lat": d['lat'],   "lon": d['lon'],
                "avg_time_min":    d.get('avg_time_min', 0),
                "faa_ceiling":     d.get('faa_ceiling', ''),
                "annual_savings":  d.get('annual_savings', 0),
            } for d in active_drones],
        }

        export_dict = {
            "city": prop_city, "state": prop_state,
            "_disclaimer": (
                "SIMULATION TOOL: All figures in this file are model estimates based on user-provided inputs. "
                "Real-world results will vary. This is not a legal recommendation, binding proposal, contract, "
                "or guarantee of any product, service, or financial outcome."
            ),
            "k_resp": k_responder, "k_guard": k_guardian,
            "r_resp": resp_radius_mi, "r_guard": guard_radius_mi,
            "dfr_rate": int(dfr_dispatch_rate*100), "deflect_rate": int(deflection_rate*100),
            "calls_data": _safe_df_to_records(
                st.session_state.get('df_calls_full') if st.session_state.get('df_calls_full') is not None
                else st.session_state.get('df_calls')
            ),
            "stations_data": _safe_df_to_records(st.session_state.get('df_stations')),
            "faa_geojson": faa_geojson
        }

        fig_for_export = go.Figure()
        for d in active_drones:
            clats, clons = get_circle_coords(d['lat'], d['lon'], r_mi=d['radius_m']/1609.34)
            fig_for_export.add_trace(go.Scattermapbox(
                lat=list(clats)+[None,d['lat']], lon=list(clons)+[None,d['lon']],
                mode='lines+markers', line=dict(color=d['color'], width=3),
                marker=dict(size=[0]*len(clats)+[0,16], color=d['color']),
                fill='toself', fillcolor='rgba(0,0,0,0)', name=d['name'][:30]
            ))
        fig_for_export.update_layout(
            mapbox=dict(center=dict(lat=center_lat, lon=center_lon), zoom=dynamic_zoom, style="carto-darkmatter"),
            margin=dict(l=0,r=0,t=0,b=0), height=500, showlegend=True,
            legend=dict(bgcolor=legend_bg, font=dict(color=legend_text, size=11))
        )
        map_html_str = fig_for_export.to_html(full_html=False, include_plotlyjs='cdn', default_height='500px', default_width='100%')
        station_rows = "".join(f"<tr><td>{d['name']}</td><td>{d['type']}</td><td>{d['avg_time_min']:.1f} min</td><td>{d['faa_ceiling']}</td><td>${d['cost']:,}</td></tr>" for d in active_drones)

        all_bldgs_rows = ""
        _type_colors = {
            "Police": ("rgba(0,210,255,0.1)","#0066aa"),
            "Fire": ("rgba(220,53,69,0.1)","#aa0022"),
            "School": ("rgba(255,215,0,0.12)","#7a6000"),
            "Hospital": ("rgba(34,197,94,0.1)","#006622"),
            "Government": ("rgba(139,92,246,0.1)","#4b0082"),
            "Library": ("rgba(249,115,22,0.1)","#8a3300"),
        }
        for _, row in df_stations_all.iterrows():
            gmaps_link = f"https://www.google.com/maps/search/?api=1&query={row['lat']},{row['lon']}"
            _rtype = str(row.get('type', 'Facility'))
            _tc = _type_colors.get(_rtype, ("rgba(0,0,0,0.04)","#555"))
            _short_name = str(row['name'])[:45] + ("…" if len(str(row['name']))>45 else "")
            all_bldgs_rows += (
                f'''<div class="infra-item">
                  <span class="i-name" title="{row['name']}">{_short_name}</span>
                  <span class="i-type" style="background:{_tc[0]};color:{_tc[1]}">{_rtype}</span>
                  <a class="i-link" href="{gmaps_link}" target="_blank">↗</a>
                </div>'''
            )

        logo_b64 = get_base64_of_bin_file("logo.png")
        logo_html_str = f'<img src="data:image/png;base64,{logo_b64}" style="height:32px;">' if logo_b64 else '<div style="font-size:24px;font-weight:900;letter-spacing:3px;color:#fff;">BRINC</div>'

        jurisdiction_list = ", ".join(selected_names) if selected_names else prop_city
        all_station_types = df_stations_all['type'].dropna().unique().tolist() if 'type' in df_stations_all.columns else []
        police_dept_names = [d['name'] for d in active_drones if '[Police]' in d['name']]
        fire_dept_names   = [d['name'] for d in active_drones if '[Fire]' in d['name']]
        ems_dept_names    = [d['name'] for d in active_drones if '[EMS]' in d['name']]

        police_stations = [d['name'] for d in active_drones if 'Police' in d.get('name','') or (
            'type' in df_stations_all.columns and
            'Police' in str(df_stations_all[df_stations_all['name'].str.contains(
                d['name'].split(']')[-1].strip(), na=False, regex=False
            )]['type'].values[:1])
        )]

        dept_summary_parts = []
        if police_dept_names: dept_summary_parts.append(f"{len(police_dept_names)} Police station{'s' if len(police_dept_names)>1 else ''}")
        if fire_dept_names:   dept_summary_parts.append(f"{len(fire_dept_names)} Fire station{'s' if len(fire_dept_names)>1 else ''}")
        if ems_dept_names:    dept_summary_parts.append(f"{len(ems_dept_names)} EMS station{'s' if len(ems_dept_names)>1 else ''}")
        dept_summary = ", ".join(dept_summary_parts) if dept_summary_parts else f"{len(active_drones)} municipal stations"
        police_names_str = (", ".join([n.replace('[Police] ','') for n in police_dept_names[:6]]) + ("..." if len(police_dept_names)>6 else "")) if police_dept_names else "municipal facilities"
        total_fleet = actual_k_responder + actual_k_guardian
        analytics_html_export = generate_command_center_html(df_calls_full if df_calls_full is not None else df_calls, total_orig_calls=st.session_state.get('total_original_calls', full_total_calls or total_calls), export_mode=True)
        cad_charts_html_export = _build_cad_charts_html(df_calls_full if df_calls_full is not None else df_calls)

        prepared_for_city = st.session_state.get('active_city', prop_city) or prop_city
        prepared_by_name = prop_name
        # ── Export personalization: extract call-type and priority stats ────────
        _exp_df = df_calls_full if (df_calls_full is not None and not df_calls_full.empty) else df_calls
        _exp_top_calls, _exp_pri_str, _exp_date_range = [], "mixed priorities", "recent period"
        try:
            for _cc in ['call_type_desc','agencyeventtypecodedesc','calldesc','description']:
                if _cc in _exp_df.columns:
                    _tc = _exp_df[_cc].dropna().str.strip().value_counts().head(5)
                    if not _tc.empty:
                        _exp_top_calls = list(zip(_tc.index.tolist(), _tc.values.tolist()))
                    break
            if 'priority' in _exp_df.columns:
                _pc = _exp_df['priority'].dropna().value_counts().sort_index()
                _pri_parts = [f"Priority {k}: {v:,} ({v/max(len(_exp_df),1)*100:.0f}%)" for k,v in _pc.items() if str(k).strip().isdigit()]
                _exp_pri_str = " · ".join(_pri_parts[:4]) if _pri_parts else "mixed priorities"
            if 'date' in _exp_df.columns:
                _dd = pd.to_datetime(_exp_df['date'], errors='coerce').dropna()
                if not _dd.empty:
                    _exp_date_range = f"{_dd.min().strftime('%b %Y')} – {_dd.max().strftime('%b %Y')}"
        except Exception:
            pass
        _top5_html = ""
        for _cname, _ccnt in (_exp_top_calls or []):
            _cpct = f"{_ccnt/max(len(_exp_df),1)*100:.1f}%"
            _is_prop = any(w in str(_cname).upper() for w in ['THEFT','BURGLAR','VANDAL','ROBBERY','TRESPASS','LARCEN','AUTO','VEHICLE','BREAK','SHOPLI'])
            _cflag = ' <span style="color:#f59e0b;font-size:10px">property-related</span>' if _is_prop else ''
            _top5_html += f'<tr><td><strong>{_cname}</strong>{_cflag}</td><td style="font-family:monospace;color:#374151">{_ccnt:,}</td><td style="color:#6b7280">{_cpct}</td></tr>'
        _guardian_img_b64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCARlB9ADASIAAhEBAxEB/8QAHQABAAEFAQEBAAAAAAAAAAAAAAECAwQFBgcICf/EAF8QAAIBAwIDBAYECQcJAwkFCQABAgMEEQUhBjFBElFhcQcTIoGRoRQyscEIIzNCUmJy0fAVJIKSssLhFiU0Q1Njc6LxNbPSCRcmRWR0g5OjNjdEhJTD4mV1pLTTVIX/xAAaAQEBAAMBAQAAAAAAAAAAAAAAAQIDBAUG/8QAMhEBAQACAQMCBAUEAQQDAAAAAAECEQMEEjEFISIyM0ETYXGBwSNCUbHwBiRykRQ08f/aAAwDAQACEQMRAD8A+MgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABk6fp9/qNV0tPsbm7qLdxoUpTa9yRVqOmalpsox1HT7uzlLkq9GVNv4pAYgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlJt4Syza6XwzxHqjxpmgare/8C0nP7EBqQdpbei3jmrj1uifQ133lzSt8e6ckzLufRJxXRtfpCraHU3x2I6rRUs/0pJfMDgAb7WODeKtIt5XOoaDf0raO7uI0nOiv/iRzH5mhAAAAAAAAAAAAAAAAAAAAAABesbW4vr2hZWlGVa4uKkaVKnFZc5yeEl4ttIsnr34LHCX8vcfrWrml2rLRYqvusqVd5VNe7Ep574LvA+q/QtwjbcI8EWGhQhTVanTUrqpBb1Kz3m89Vl4WeiR0/F3DOm69w5faRf28K9C6oyptVIqXZbTSktua5p9GXNClHDbe77zdTlTlDGQPy61Syr6bqd1p9zHs17WtOjUXdKMmn80Yx9A/hP8AopraVxLqvF9ncRhY30/pDVRKMI1HjtR7Wdm3lpPnnCPn4AAAAAAAAAAAAAAAG20XhniPW4Opo+gapqMFzla2k6qXvimBqQbbV+GeI9Hj2tW0DVbCPfc2lSmvjJI1IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAN5wzwjxPxNNrQtDvr+Cl2Z1adJ+qg/1pv2Y+9oDRg9R070O3UIRnr/Eul6e3vKha5vKq/qYp5/pmxocDejuzmoXGo6zqdVfmRq0qGf6MY1H8wPHQe+WGg8I0Mfyb6P53Ul+fdRuKjf8AXlGHyN9YQ1q3X+aeGdL0rucKFvQl8YRcvmB886Vw7xBqyzpWhanfrvtrSdRf8qZ0Nt6LOOq0VKposbRPn9Mu6NvJecak0/ke4VrPi2//ANN1ijFd0p1K2Pi0iinwvWlvcazcvvVKnCC+xv5geT23of1RqLvOJNAtv0oRqVqsl/UpuL/rGyh6L+FbRxnqPF99VS+tClp9Okn5TnVf9k9Pp8KaUvy30q4/4tzNp+7ODLttC0a2eaOmWkH3+qTfxYHmdrwr6NLeSlRsNW1Ka/Nq3/rYv3UaUX/zG6stK0Oi1LS/R1ZqXSVS1nVXwuakl8j0GFOEFiEIxXgsFQHMWa4opNfybpWm6Wv0qCpW0l/8mn95Z1yHFaoOrfavbVY8+zVlVqfNyR2NJZy+4430jaireynFSxhAef33G91YXcqFawt7hRf1qdVw+1MztJ470e9rRoVlVtKr2XrEnFvzR5tqFZ1a06jeW2yxolFXOqQ7X1YvtMD3ZKjXoT7OHTrU3Cbjt2otYafg10PLOPODZ2NSd9ptNyoPeVOK5eX7vh3HZ6JqHqYxpy3p/YdDJUq9FxklOElvnqB84g9C494MlRlPUNMg5JvM6aXPy8ft8+fnr2eGAAAAAAAAAAAAAAAAAAAA90/B99KPDXBfC1fSdXp3MK1e+lVdalSUoqLhBLtbp/mvkn8zwsz9M0rUb+lVrWtrUnQpRbq1eUILxk9kB9x8K+lbgLUH2qXFOmUo/wDtFb1D+FTB0tx6TvR3aUnOtxpocsLLjSvadSXwi2z87KeMtS7zb0bSyryp1LKv6q4i0/VV2uzJ+EuXuYH2lxj6V/Rfq2h3+k3dSprFvVtpyrUI2dVxlTSy2nKMY5XNNPOcY3wfCcsdp9nPZztk7/T9evretCjWn6mcPrZgn7t9uX2mn1bhh1Yu60WXr485W/58f2f0l4c/BgcuCWmm0001zTIAAAAAAAAAAADp/RVw6uKvSDo2h1IOdCvcJ3CTx+Kj7U9+nspr3n6G6BZW9lZUbW1oU6FClFRp06cFGMIrkklskfJv4GHD30riPVuIqtPMLWjG2oyf6c32pNeKUUv6R9gWUMRQGZGEZRcZRUk1hprmecekT0IejvjChUncaJS0y+kvZvNOiqM0+eWkuzL+km/FHpceRbrvEQPzt9Mvow1r0aa5TtL+rC8sLntOzvacezGqlzUo79mSysrL57NnBn1p+HDqlCHCmi6RJZr3F+7iG3KNODjL51InyWAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADs+A/RrxNxdS+m21Cnp+kxf4zUr6XqrePPaL5ze2MRT354A4w9E4N9EnEOtWtLVNXqUeHdIqbxub5P1lVf7qkvan8l4npXD+gcFcDQVbSreOs6pT3erajSXYpvfejReVHo05ZZreIeLbm+uZ1nXqXNeXOvWefgu75AZOn6B6PuE6fbs9Ghq1xFNO91nFVN4/MoL2F4drtMazxhquq9mk7ibow2pqptTguS7FNYUdvI46vc1q9X1lWpKcu9v+MF+3ny3A31rRpXU072dS6/VqP2P6q2Ox0JUKMFGjSp013Qil9hwthWxJbnV6NX3SyB1q5At0akXTTckiireWtJfjLinHzkBfBqLriTRbZP1t/RXlI0956Q+HbfOLn1jX6IHXkHm976VtOhn6PbVJvxNHe+li+llW1pCC72B7JsUynCK9qcV5s8DvfSNxDcZUa6pJ/oo1FxxHrV4/wAbqFw0+5gfS9rUp14SjQqQnLlhSR5p6U9I16ssU7Oc4yeF2WcBo9bXnGUtOq385t/6pNvPuOxvOKPSDY6BGje2FadJNdmtXs5dqP8AS2+YHlmtWF9YZjd2tWi/1o4LGl1lb4S+tUe/kdPxJxDqGs0ezqFOnVwt8eyzkKaXrfWQz2U+T5oD0LTK3boxeToNK1B0ZKnUeYP5HH6JWzSjubunLZAdm+xWpYaU4SW+eTPNfSDwfKE56lp0M53nBLn/AI/b58+t0nUHSkqdR5g/kb1qnWpNSSnCS38QPnEHoHpD4RdCc9S0+DcXvOC6+Pn9vmefgAAAAAAAAAAAAAAAAD0DQuMaWm8IUtHnp6urKpScK/Zk6dSNTtyeYzw+jjzTPPyYylHPZk1nnhksl8rjlcbuOpp6fos6U7uyub2lGSaUK1NNrZ/nRlv06LmYVtaaPTuXG7vbydJPeNGgsteDlLbr0KJ0K8VTt1e9lwoKc4LpJ79l+Kzv5FmyUoWlzW9RK4nTcJObbUYR9rKfm+z8Br813+TZaX2rm8r2lrmraxzKlC6koyx3Ka2UvPCZl05XNnWkqXrVKm/bo1F2alPzX3mht9bvqNJ0oyg4N5w1kv0eIblzjG9pxuaEdoRz2Z0v2J84+W68CsW/uoaXr0M3i9Td4wrmC9p/tL877fE5jWtFvdKmnXgqlCT9ivT3hL39H4M3SnbXUHcW1btpLMpqPZnDf/WQX9qOUZdnqVa2i6NxCNWhUWGpJShNfYwOIB1uo8OW17F3GiTUKnN2tSWz/Yk/sfx6HK16VWhVlRrU50qkHiUJxw0/FAUAAAAAABsuFtJq67xLpujUW1O9uadDtJZ7KlJJy9yy/cB9mfgtcPfyJ6LdNnOCjX1DN7Uff2/qf8ige120cJGh4cs6VpY0LahBQpUoKEIpYSilhI6OgtgLq5GLdyxFmVLkavVq0aVvOcmkkm2+4D4o/DB13+VPSlHTKdWUqWl2kKbh0jUn7cmvOLh8Dxc3vpA1mXEPG+s605ucbu8qVKbf6HaagvdFJGiAAAAAAAAAAAAAAAAAAAAAAABMYylJRinKTeEkt2BB9keh70E8FR4S0281/RPp+r1qEa1w7qrPs05SSfYUE1HbON02eI+iT0K8WcR65YXuraRV0/RKdaNSvK7i4TrQi03CMH7XtLbLSWMvL5P7g0i3lRpJOOAPM+MvQDwBrWiXNCz0S00y/dKStrm17VP1c8ey2k8SWeaa+B8Pa7pd7oms3ekalRlQvLOtKjWg+kovD93ifqBRdJT/AB0ZSjj81nyr+Gf6Oavb/wDOFpkVOmpRo39OMMShHZQqN5339l7dY9ETfvrSb99afLgAKoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbLhvQdZ4j1OGmaHptxqF3PlTowzhd7fKK8XhAa06LgngriTjG7lQ0LTp1qdPevczfYoUF3zqPZeXN9Ez1Lhr0UcOcPKNzxpeLWtQW/wDJVhVcaFN91Wst2/CHVczfcQ8XpWkNOoRt7Wyo/kdPsaapUKXX6q2z4vL3A1/DfAHBfCKjcai6XFerw39tONhRfhHnV9+I+Be4n4zuL6ovWV/pDprs04RxGjSXdGK2S8jlNR1S5vG1UniHSEdl/iYDeQMm+vri7n2q1Ry7l0XkjGznmQOm4EldOahvJpJdWarUdYtbROKl6yp+imczqOr3V22u12YdIrkB2V1xNaWbxSfrprpHkYFTjTWJyxbuNCPhzOM7dRPtM7rReGLejodPXuIbmVpZTx6uEY5nU8gLD4n1+rT7L1Gql3JmO7jVrx/lrus3+jl/YbOrxholjD1Gh8OUFjZ3F3ic34qPJPzb8jGjxZGrVVWtWvIT6erxFLyUWkvcgKbfhvX7x/i9Nu556yi19ptbT0ccS18OVtTop/pzMjT+NIxwv8oL6n4VJTf70bFa5bagsVOI+1no7pw+WUBYpei+tBdq/wBYtKC67l6PBnB1p/pvESqNc1Ta+4R0GzvX2o6nVq5/RrRkTLg2D+pdN/txb+8B6j0cWPKldXcl35Sf2B8QcK2+1jwzCfc6m5ZqcH30F+IqWsl45X3GJccMa7BezbwqL9SovvwBtYcfata0nDSNKtrWL6Qp/wCJz2o+kPjL6Wp3FRTSe0JUk4/aZFPStYtl+M0+4f7Me19hgapC4jNOpa1YY/Sg0BpdV4hratcupdW1vTnJYbpxx8jVrtW2JygpQk+Zf1KhKVf1kI7l+3oynaNqLqUX9eP50GBlaNdU4SilL8W3tn819x1FCeYrc4GSlbVMp5i+vRo6PQNTjUjGjUl4Rb+wDo4s2ulag6UlSqvMH8jTwZcTA66pCnXouEkpQmt0eUekHhSen1p39nDtUJZlOKXLx/f8fLutN1WFCao1pey+XgbqtToX1q4SxOnJc0B87A6/jThC7024nc2dGVS2k8tQWez5eH2HIAAAAAAAAAAAAAABbvCPTuD/AEb1KVktY4n7VtBx7VG0e1R90pLp5czO9G2gaNplnaa3VlT1DUakY1aT507dtbYXWa6t8ny5ZfUavdVrqlN1JN555YHkvFVOypar/MqKpQjHD3znx8zBoVZUuHdShj2KzpxznqpJr5ZNjxDbS+mTaXU12rYoaFRoY3q1nJ/0V/8AvAaMAAV0alSjUjUpTlCcd1KLw0ba01aM12LhQi5c32fYl5pcn4o0wA6qhOcJqVtN55+rcs58YvqjPnX0/WKKoapSzOK7MK0dqkPf1XgzkNPq3XroW9vTlXdSSUaKi5OUnywlvnyPo30VegfUNYpUNW439dp1s8Sp2UXi5mu6pL81f837IHjek+jXijXdWp2XDtjLVKdR49fT9mFNd9RvaHv59MntPDv4LFP6NCfEPFUvXte1SsaC7MfKc+f9VH0RoWkaZoem09N0iyo2VpTXs06UcLzb5t97e7M4D54vfwV9Emv5lxZqFF/723hU+xxOd1L8FjW6cmtP4rsK66evtZ0v7LkfVKJTaA+ML78HD0hUV2rSpo9/HPOjcyj/AG4xPUvQf6A3wvxBbcS67r9tcXttGXq7W1g+xCUouLcpyw3s3theZ77nqUuMWmmlh8/EDJsXQXs06qlg2NOqo+Jp6ijUlTnKFNSpx7MZRgk8eLxv5sqjlSz23juwgNy68JLHJ+J5v6fdf/kD0Ya9qEKnYqq1lSpPqpz9iPzkjs3OTj2fWSx3Gi4z4W0LjDSHpPENnK7s3OM3TVadPMly3g0wPzjB9m6t+DZ6OLuXatf5X0/9W3u1Jf8A1IyZ4X+ED6IoejarY3mn6lVvdNvZSpxVdJVaU0s4eNpJrLykuQHkwAAAAAAAAAAAmMZSeIxcn3JGTDTr+azGzr473BpfFgYoM6OlXWcTdCn+1Wj9ibZep6O3L2ruDX+6pzk/mkvmBqwb6joMXL6t5Vj3KnGn8239hm0OHcS7X0FyXdWr5X/KkBygO0paJCjLtONhR84es/ttlyNCztpOX8qxot8/UKNP+ykByUNNv5QjUVnWUJLMZSi0mvNnr34NOucOcH8SXk+KI2VKtdxpwtLqcoVPU4cu0m032O1mO+y9nc5PTqvCcK1aeoSqXMnhxlLMm3vn7jKra7wdClKlT0l1ItYa7GAPu/RLjTb+zp3djdULihUipQqUpqcZLvTWzNoqlGKwpI/N691i1o3Ebrh9X+lXMdvWUbiUW15p5z7yr/L/AI8oper404iiu5anW/8AEB+jjqRb+sjF1ixsdX0m60zUKUK9pdUpUq1OXKcJLDXwZ+fFn6V/SRaSTpcaa1LH+1upVF/zNm4oenj0q0oY/wAq6jXRStKEm/e4Acz6VeE6nBPHuqcNyreup21VOhU6zpSSlBvxw1nxTOXNrxXxBq3FGu3Gta3dO6va+O3PsqKwlhJJbJJI1QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAt3hAC9Z21zeXVK0s7ercXFWShTpUoOc5yfJJLds9H4K9D+s6pa0dX4muI8N6PP2oVLmDdxcL/AHVLm/2nhb53PTNOuuHeCLGVvwlYR01yj2Kuo3DU72uv2/zE/wBGGAOJ4U9DUbOnT1D0g38tNg0pQ0m0cZ3lRc1239Wkt098vnsmdtecRafoekvR9BsbfQ9OfO1tN6lbxq1H7U35vHgclq3EVxcTn6hyh2nmU5PM5eJopzlOTlJtt823uwNpqWt3NzmEH6qn3Re782aptsgAAWLu6oWtPt1qiiu7qzVVLy/1BuNlT9TR/wBpIDY3+o21nF+tmnLpFcznNR1e8u040k6VLuXX3m1ttGowfbqt1qnNykZT0+k/zUBxM1POZZKDtZ6RbzW8TB1DRbOjQlVlKUcLIHNRbksPdRMq8vby9nGV3c1q7iuzH1k2+yu5dy8DvvRXwppXEGg6z9O7UpUoxqQnCfZlT7O/Nprfdcn9hjytZ0O1C04P06tSXLFV1JP3yeX7gOCyTk6y91C1tJYv+C6NBv8AThOGfsMb+WuHZfW4Xof0a1Rf3gOdTGTo1qvCr+tw1jyuan/iKlqHCEueg1o+VxP94HORbTynhmXb6nqVDHqNQu6WP0K0l95ufpXBcuemX0P2a7+8ntcES509Vh5Vo/8AgAxrfi3iOgkoatXeP08T+1M2Vr6QuIqSxOpbV/8AiUUv7OCzGjwPL/X6vHzqQ/8AAVxs+CJctR1OPm4P+6BtbX0nahH/AEnTLWp+xKUPtybC29J1rLa60mtBd9Oqpfakc8tK4Nn9XXL2HnCL/cUT0HhqW9Lilx7lO1X2qYHXQ454TvJdm6sasc/nVraEl8my7Q1T0fVqrcJ2lKUtn+InTz57JHCz4csn+Q4m0uX/ABO3H7EyzLhu4zijqOlV307N0o/2sAegXGg8C6hlU761g5b4pXiz8G2Wv/NxpDxUstSu4Z5NuM18kjgZ8Na5F+zYTr/8CUav9lspp6Rq1s3UurW7saUMuVSpBw5dI5x2n4LcD06lwpXo0lBahGtJcnKl2fvZZu9Bv6FCdSLpVOym8Rby/LYjhXiW1ttPlYXN1VvpWuIwrwh2nVi84WM5bWPsNouK9Me8qV3Bd8qWPvA8n1/VmpOFOT7Xh0LGi8U63Y1EqNf1kesZ7pmz4w0izvtdrXmkTf0erH1tVSi4qk/zny5ddi1b6DK0uHSqzpznDHbUcvs55Z2A6vTONoV6SjqOmV4pr2p0o+sj71zKbzS+DOIJt0rihSuJfoT7E8+KfMx7CyUEuysPwNo7ahWgo3djQuV/vaeX8eYHMaj6NblZnp1/Tqxf1Y1Fh/FHN6vwprml0qla5tM0aazKpCacV956bT0vTI/kHqOny6fRrp9lf0ZZR5pxRqt7farcW1fUK91bW86kKPrMLKWUpNLbIGhAAAAAAAAAAHRcD67U0nVKVGrNuyrTUakW9oZ27S+89hvoqFF+R8+Hta1W3loVjVr3NKM6ltTk+1NJ5cUBzmrUFOrOWOpxfElVSvIUI5xRgk/N7/Zj4Ha6pe2sbarWVenNRTb7MkzzmvUlWrTqy+tOTk/eBQAXrK2uL28oWdrSlWuK9SNKlTjznKTwkvNsCyd56LfRTxX6QLiM9NtfommKWKuo3KcaUe9R6zltyj15tcz3L0S/g32Gm+p1TjuUNQu9pQ06m/xFN8/bl+e/Bez+0fQNtQo21vTtrejTo0acVGnTpxUYwS5JJbJAcF6LPRJwnwBQhWsrf6dqvZxU1G5inUzjfsLlBeW/e2egAAAAAAAAAASSUlSTeyAnIxs5SfZiubZg63q2n6Lp9e+v68YUqFOVWo+kYxWW35JHzv6WfTjqVW2ceE7iwhTjPeU5qc5R35JPbp3gey8bcc0tEoToaRZVNV1HlGlTnGKi++Um8JeWX4HyZ6WqfpO4r4ocuJNMvacYZlaUZRVOhTg3zi2+zvjm23sbSy9OPG1O1lTo3UaVaWezKlZw7v1k/Ey+G9Wv+JdLu9Q4p1C41C6rV0o1K9XHYhFPCSWEl7ctljmB53X4KuoWcHG4p/S2t6M9svui+XvbSOa+h3f0iVuras60HiVNQfaT7mj3OpX0C2WJ1LReDaZx/Fl1oNzqKqwvJyjKCThTk+zleHwA4KGl6hKXZlazpP8A32Ka+MsFdPSqzn2atxa0fF1O3/Y7R0DvdDpfUtJVH3sh67Qp7ULCkvMDUUdGjKWJ3FSS6Ojbyln+t2TKoaBFve2vaq6NyjT+6RkVOIrzD7FOlTXhExK2u30+d3heDQGfR4exzsaGO+rWlJ/JpGRDSqVD69axorwowbXveWc3V1KrP61zUl72Y8rrPPtMDr8afSi1PV6sl1jCeF8EY0rrQKbz6urWffLf7TmY3MM+1Tk/6RkUb+0g8ysIz/amwN49b0+l+Q06P9LBRPiSvj8VbUYe4x7XiShbtSp6VaRa5P1cW/i0bNce139e3g/OKYGpq8Q6hLP84jBdywjDratdVPr3dR+TZ0q45hL8pY0ZedGL+4f5YaZU/LaNYz/atKb+4DkJ3cpfWlOT8WW3Xf6PzOynxFwxXj2a/D1jh/oUVTfxjgxqlzwPWWHpdxbt9aVxL+82Byvrpfool15YwoQXjudGtO4TuX2be+1Ki3+lCNT7Eif8l9Ok/wAXxDTiv99buD+0C/wFpekX0Kl5rV5SoxVVUaMas+xTcsZ9qXT4osX/AAnrUrqrVoWUa9tLLhK1rRrwUd+tPPc+fc+5mDe+piriwoV+xbUqkeWMVJRi12+/fLfN8xps8UZU7fNZreWISeF7kY6u/aspZrVi/Q4R1qpJxnbQt995XFRUYrdrd1HFc4te5nZW1HgrSeD56XxFXsNW1irJRtlp8u3KhnC9qsl2Nn0TkcNeQq1ZpOCj0w85+wx6dvCSk5Tperiuay8MXG3zVmUn2ax7PBBuaGiVb6mqtlKlGMfZqOpPsrteC54w0Xo8LXf517Yw85yf2RMmDQA6B8K3P5uoWDfd2p/+EofCuqf6t2tTyrJfbgDRA3VThfXILP0JS/Yqwl9jLEuH9cX/AKpvJfs0nL7ANYDJr6ff0HitZXNL9ulJfajHlGUfrRa80BAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAN5wdwlxDxfqX0Dh7S697VW9SUVinSXfOb9mK2fNnsnDPo64P4R7NzrdWhxXrEN/UwbWn0JeL2dZ8u6IHmXAPo24l4vh9Mt6NPT9Ig/wAbqV6/V0I+EXznLwinvzwevcO6LwbwHTVbRbaOratDnrGo01inLvo0XtDdZUpZlzLHFPGde8nGFSsq3ql2aVGmlCjRXRRitkvI4u+v7i7n2q1RyXSK2ivcB0nEHFdxeXU6zrVLmvP61erJy+By9xcVq9Rzq1JTk+rZZbb3AAENpLLeF4muutVipOlaQ9dU5ZX1V7wNhVqU6UXOpJRiubbNVW1OtcydLT6Ta5OrLkiyrapcT9be1XUfSC2ijPpKMIqMUkl0QGNbabD1nrrubuKv63Je42cUksJJItKRWmBcRUihMqTArNJxZV7Fh2U95PBucnNcYVM+qp97yB6F6G6btfR9xFf8nVdOhF+by/sIgsbrYzeEqf0L0N2y5Svb+U/NRWPvMKIGbbX9zRXZU+1Hk4y3RbuLPh2/SV/oNnJ5y5Uoeqk/fHDZYRXEDX3fAfC932pWd7e2M3yjJqpBe54fzNReejHUo5lp2p2V5HHKbdOT926+Z1SZXCc4v2ZNeTA8z1DhDiWwTdfR7lxX51KPrF8Y5NJOE6c3CpCUJLmpLDR7hR1G7pfVqv3l2vf0byChqNhbXce6rSjL7QPCck5PX7vhvg2+WZ6dOzm/zreo4/J5XyNTd+jbT6ylLTNecX+bC4pZ+Ml+4DzfIydbf+jnia3bdCjb30Es9qhWX2Sw/kc3qGl6np+Pp+n3dqm8J1aMop+Ta3Ax1Jrk2veXIXFWD9mcl7zHyTkDNpX9xTl2oyXa72ss6Hhz13EFeNtdX8KUqVSHqlUk3u8+0vaWMY6d5yWTO4fhCtrdnSqVPVxnVjFy22y+e4Hb6no3GFG77FtThd0Kb2nLs1e37qjlg3tPV7+jHs1eDK8Goxy6MVhvOHjCfLmcxDhnW7n/ADfpl9SXq8tylKUXPzxnPyNZe6PxZpFCpSuK8qNOUsqp65pLyfTmTa2NzqFzrGo67ClY6Bd2VGSabr0F7csvftSjheRqq99daTr0vX16lO4pvE6K7MVy69lYkY+l65rWnxU6mvTqUmnmm6k5v57r4mWtNrcS8M3usUaebuyrNyfacpVoNZlu93jms9M+ANL0+Mrzl6+pHwzgtriu7k97mp/WOWhiUezJZXTwHqG/qMqO2tuK7uKx9KqY7u0yY6vY1JOVawsKrfN1LaEm/ijmrDhzX71RdpptzUjL6s+ziL/pPY2S4G4ux/2cl/8AmKf/AIgM+tV4euPy2jWP9Cn6v+zgtu24RksPRqaferir/wCIs0eAeK6jxKhRpLvlXi/syZsPRrxC/r3tjDw9ZN/3QNfU0fhOrLKje0fCnXWP+aLKJ8O8MTj+KvdQpvvnKEl/ZRu6Pox1J/ltXt4fswlL9xu9K9Dc76GY8V0I1MfUdu08931vsyBwX+SWly+pr8l4StV/4y1V4Ny/5trVlNf7yMofYmd5e+iatYVvU3WtVoTa7WFSi9vdItx9G9OP/ry591NfvA4KfBepqOaV3p9Z90azX9pIWvA/EFxPs+rtKa/Snd08fJtnoUfR5GOMa7dr/wCGv3la4DqQw46/c++iv3gcRW9G2v0rWdx66xrdhZ9XRqynOXgko4z7zK4d4D1nU5ud/GdjSguzH10X2nhbJLnj4HZQ4MvIfU4hqrzt1/4jIpcNatS+pxFL/wDT/wD74HOVPRVWrP8AG8QJR6RjaYS93bEfRHT669J+Vr/+8dZZVtS0etGlqtX19tUeI11nEX0znkdFkDzJ+iOl016f/wCl/wD3iaHoprW1xSubXiOVKtSmp05q1w4yTymn2+aZ6ZKUYpuUkkt228HH6txRezv68NJdGpbUWqfbjFy7c+u/LC5Ae/8ABnpIqfQqFrxL6upcxiozuaUOxGb7+zl4O8/lvR3a/Sp39vSpYy5VZqOPifKVTj+y0+1hUo6HXvLxflo1pdmnGPesZaeVs/f3FHCupS4y1qj/AJS63eW1K7l6u1p0VKahJyxCPlzzJ92WwPqeHEfD9ZL6Pq2nVm3jELmGftLtDVtJrwlOF1SxF9lty2yeJab6K7fUnKen8VXUKabS9bbuTyuf56LGr+izinTaUq9lrVreQjyj2JRm/cov7QPoKEaVSCnTlmMllNcmiI0oyWYzTPlq8vvSVo0Zqlc6rGFKPafqZ1HhLq4N5wv2TdeiP0qVrniOjp3FF66lvdTxC6jiCjUbyu3jbDe2dsdeuA+i/UvpJNkOjMuW1rSoU08Sk1vl7tvGDB1HXbOw7XrqNfC64QGQ6c10ZHZl3M0dfjaxg8U6MJN8vby/gjBrcbXkpyhb2tKm11cd0B09zVo2tF1rqap011fXwS6s0eocRdrNGxp4fLtS/wADR3d9falNO6qtrOcFy3oqPQCu803+VdC1S2ulKv8ASrSpRmm/rKUWseHPofO+v+jvVLiz7PCfD+p/SW8RxKaj8ZSwscz1P0o+lHT+CrSVna3ka+sTWY2lGKnNbbObeVBbrnv3Jnguoel30i6nVqSuOJby3pyeVStsUlFd2YpMDecJ+hz0i1L+N3qOkVpQgnhV7ql2ctY6z259zOc9JXBmp8FThG81jSqkrirLs2VpeOtVopLnNJJLuMO54mv72DlqN7e3k+nra8pb+bZdstc1GVD2L+6prk4xrSx9oHHyqzfOb+BTKSkubz3vc7aWo1ai/HU7Ss/0q1rTqP4yi2UdrTpxxX0XT6rf53ZnT/sSigOJcZdn2a2/jBJfE6j0dcL09curivqNWpKytorMITce3N8lnHLCeceBkRsdBnn1+m1455fRrpwx/XjM7zhOxtdM4apxtIVoRuJyrNVZqUt9lukukU+XUDVV+DuGIvEdLX/zqn/iOF4/0S00y8ofydbVKdOrTbcU3JZT33bb7j1Kq22c/wAbaPW1TSqUrf1frqFTOZzUV2Xs93st8AeSNNPDTRB1N5w3rVrGPajb1e1yjb3tGs/hCTZj3mga5a0fW3eiX1Gnz7dW0kl8WgOeBnOnDlKjHPvX2FPqaPWEl5Sx9oGGDK+j02/rTivJP9xS7dZxGosd8lj7MgY4L0reSeIzhPyePtInb1Yc4p5/Rkn9gFoFcqNWKzKlOK73FlAAvUbm5o/kq9Wn+zNosgDY0dWulF+subiUt8Ptt/x1FtfUoet9Z66frMdrLUs45c0a4Abu31Oyo1o1lSq9uDyvqrf3RMWhc0KVGSSjPPNTi038Ga4AbmhfLsulav1W+ew+vkyJXlyn7Umacy6F44xUK0PWR6PqgM2F/Wi/rMyaWrVI82zWyubV8o1F7l+8odah07fwA30NbmvzmXY8QVI8pyOadaj+v8B6+kvzZsDqY8TVo8qk/iXf8q7hx7M6s5R7m8o5D6TDpR+MiHcvpTgviB1f8u2c5Zq2NpUffO3g/tRTV1DRK/5XS7P+hSUP7ODlPpE/0YfAh16n6vwA6pPheccPTKcX3xrVM/2i2tO4YqSz6y9h4RrRx84nMevqd6+CHrqnevggOknoOi1JfiNVr0l3Tpxn9jRRU4XotfzbWKE3/vKTh9jZz3rqv6XyKlc11yqMDZ6hw7e2lvKuq1tcRisyVGbbS78NI0xsLDVrm2rxm5esj1i+4jV6Vr2oXVnOPqq2fxfWm1zXluBgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlJt4Sy2eo8E+h3U763o6vxhdPhzSJ+1CFSGbu5XdTpc0n+lLCWU8MDzfS9PvtUv6VhptnXvLqtLs06NGm5zk/BI9i4W9D2naOoXvpCvZOv9aOi2NROq/CtVW0F4Ry/FHYW+p6Fwfps7DhGwho1tOPZq3Updu9uvGdTms4z2Y4SeTi9V1+tXcoW+aUXzk95P9wHXa1xTb2WmR0fTra20vTKf5PT7GPZi/Gb5zfe5HD6lrFzeNx7Xq6f6EX9r6mtnOU2229+ZAEtt82QCxd3dC1h2q00u5dWBfMO91Ghbews1KvSEd2au61G5usxp5oUn/Wf7i1RhCC9lbvm+rAu1Z3N483E3Cn/s4v7WX6MYU4qMEorwLEWXIyAyYyK4sx4suRkBkRZXFliLK4sDITKkyyn4laYF3OEcjxPU9ZqMYdyOplLEWcfe5udc7C3bmooD2u+grPgThewSw/orryXjOX+Bpkb7jrFHULSxjytLKjS8mops0CAuIqRbRWmBcRUmW0ypMCsEIkCGN/InmQBdp3Fem8wqyXvMujq95BYc1Nd0ka8AXbu24f1BP6foVpKUnmU6cOxJ/wBKOGai84G4Vu23a3V5YyfKPaU4r3Pf5mxAHK3vozv0nLTtVsrtdIzzTk/tXzOc1jhfXNHiqmo2To0m8et7alD4ps9OUpR3Ta8j070e8BcM+kngm40nW7i5tb6jeSdG5t5qNRRcIezumpRyuTXk1uN6WY2+HzRw3xBqmk3tOpZXVaVOE0qva9pdnPJJ7LbJueIOJNV16j6utQpUo7tJRfa57de49yvfwXdS06nJaZr9pf8As/i3Wpui4vr7K7Sbe2/a6HE6t6EvSFpUpVXo8rqCe7o4m35KDkwjyfRHa0tRmtQ05X0HH2KTm4ZefBr4ZO2o8W17GwVto/DlOzSeVGNJyj45xLd+Jp9e4U4g0+5autIu7aonmLlDDTXcnv8AI0d27+ylioq1JdntYls0veBk39K1q16laOi3NBzk5dlSlGEc9yxsjf6NxTpWkypVI8H6V+Kh2Z/zqvGpVbxu5OWU/wBnBpLalr9WlCrQp3soTXai1Rlhrv5F12HFty/Uwsrypnf2qeFt4tAdK/SVPs1oW+jW9OM/yfavZzdPwy08+/JZp+kLVez/AKJYS8ZVH92DQ/5M8YyW2nVV/wDEgvvK48J8ZPnZ1F514fvA33+XmtS+rbaXH3zf3kPjXiGXJaZH+jN/eaX/ACM4vk96PZ87mP7y7DgXiqXOVFedwBs3xbxHP/8AEafHypP95H+U3Ej/APWNnHyoGvXo+4le7uLNedaX7ir/AM3nEMtnd2H/AMyb/ugZb4i4jfPWLdeVvH9xS9e1+XPX4LyowX3FiHo41ptdu/so+Tm/uNnaeip1buFKvxRaUKUo5dX6LUkovua+9ZA18tY1iT9riSS8oxX3FD1TUH9biev7ppHQ2fodp3FCrjjSxjcUpfk5WzjGpHvhKU0m/B4M6j6DbWoqdxDjaE7OSxOpDT8zoy7pw9blLxWQONep3P53FF57qzRanf8Aa+vxHfS/+PI9Bj6AGvxNTin+cy3oRVqlSuY/qVO3jteDRNT0CWvrPXUuJLqdrTeLqDs0q9u/1o9rePisgeb1bqznFxq6xeVE+adWTyWql1p3519eT/8AiSPS/wDzDwV3GlPVe1Rq721dXCUK/gn2MKXg2veair6LdGo1Z0q1zqMZwbjKPrabw08NZUWBwVW70tdbqfnUZttO4psNJ02lY0bKtKKTllNfnPP3nTL0Z8PLZ3Gov/4sf/CW/wDzX6C+d7qb/wDiQ/8ACBw2s69Z6lWjP6HUhV5Rm5br4GXw1xRcaDcUL+3pqq6KdOUG+ko4bXds38TptT4A4b0qxq3tW51Cp6uL7EZVI+1Lotop8zzSlBypyj212nvgD7U9HGraZq+gW8rW49ROdNSjPt8s75T+5nU0qF4qLhUvKV3jOHFpNrue58sehPXdQoUp2MVOp9Gkn2U8vsvk/uPZP5fdek6sburQg4Ya9VnsyXUDtaVte0tTWoXdKlRUIOFPE87eO2F8T5k9PWh0OHuLamo6fNQ0/U1Kq4UFtRrfnxSzybxJebxyPaNP4l0jVLBO21S0uoRwpOnUTXaxn3HH+kixtda0G5tXUhOXZc6Usp9ia5P7vJgW/wAEvi7WLrjOpo1zrOpXdjKzqVpU7u5bhBxcVHsJt778l036Humt0Y32sxlNRcE8Qj3+J8N6HdahY3VvqWmznC5o1Izpyg+TT+zY+0+D9T/ygoaTrFFqNG4oqUoNbqbays9MNSTX7gPM/Sv6RdF4e4/ueFbrh2N9RhQhOtVoXLt60ZyXawnHn7Li+aOZ1Hjfgq1rW17p9zxtZ+tpufZnXp1IRaeMJTT7Syms9rmmc7+EJaV4+lPU9TudOuLeNavtXqwklUjGKpx7O2GsRz70cbKEZRiucWtvIDrOIvSBxZe17W403U7+wjCOPZdOUJtvZySS6YT2a8iLL0j+knVLKrp09TjawfszuVbxhXS6qLXLzxk4qV3f05L6LP1dutorGe0vHPIvR+n1ZutZ3DjOSSlT7XZ5LG3RgbS706xtLec6snKpNuU6tSWZzk+bb5tnNXk6EG3COI97K9UuLxXU43MamU8Q9YnnHeaus5Sy5PIF+FxQlyqLPcdBwlpN7qdG8rW9rVrxo9lz9Wm+ynnovI4ymv51BfrL7Tv+Ba11QjcztasqM1KDU4bST35PmgK5aTcKh6/6NXVLl2+w+z8THdnJfnP3nZvXtZVf16rxjXf1qkaUYuf7SSxP+kmXHxBdfm2FjGm/r0VSbpT84NtL+j2QOGdtVT2aPSKkHbWVvbt5dKlCHwSRr3qelLEY8PUPVS/KUpzz5uE0lOPvckY2ratN1qjl2IRXewMipXjF7ly3u7aalSrpSpzi4yT6pmh+levi5wkpxy12oPKMSvXnT3ywMPX7CemXjh2u3Qn7VGp0lH966mrdacXmE5RfemdVpNK413t2MJ0YqC7alWn2Y9rpFN7Ze+E8ZwXZcHao6sqP0CPr0m3RbiptLrFP639HIHLT17WfU+o/lW9lRxj1cq8nH4N4MWOrV4R7Dt9OqLr27Ci3/W7OfmdE+GLupGpKOm3EvV/lFGDzD9pLde811XQU6frfV1owb2kuXxYGqhfWWH67QtPqt/ndqtF+5Rml8imL0Oc26um3kc8lRvFFL+tCT+Zm1NEjj2a80/GOTHno1dfUqxl5rAFqVpoNSWY3WpUF+i6EKvz7cfsIekWE1mjrlCC7rihUi/8AkUl8yZaZeR/NjLykUO0u4rLoy27twH8gV5LtW9/p1Zf+8qn8qnZYWga5UzCha1LpLmraca39hso7NWP1qc15xZPbAxbvS9Rs8xu9MuKD6qrbOL+aMKVOm9nT7L702b60vrq0n27W5rUJfpU5uL+Rmw126qNx1GFvqlOTXbje0/WSaXRVPykV+zJAcdUouKcovtR+aLR1Or6Tazspapozqyt6ePpNvVkpVbZtpKTaS7VNtpKWFhvstfVcubrwUX2lyfyAtAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbjhPhnXuKtUjpugaZXvrh7yVOPswX6UpPaK8W0BpzsvR/6N+JeMoyu7ShTsdJpvFbU7yXq7eHgpP678I5e6zg9N4Y9GnCXCXYuuJatHifWYbqyoyasaEv15c6rXcsR5p5M7irjKtddilVqwnGhHsULajFQoUIpYUYxWyWML3AXuHNH4O4CpqroVvHVdWgva1nUKSxTffRpPKh5yyzScQcV17u5qVvXVLmvN+1XqtvPkc7qGo3F5PNWeY9Ir6q9xhttvL3AvXFxVr1HUqzlKT6tlkBtJZbwgJKKtWFKDnUmoxXNs1uoazQoZp0F62p4ckaK5ua91Pt16jl3R6IDbXutSnmFnHC/2kvuRrMuU3UqSc5vnKTLSZWmBeiy5GRYiyuLAvxkXIyMdSLkWBkRkXIyMaLLkZAZEWXIyMeLK4sDJjIrizHjIuRkBVXn2aUn4Gi4MtnqPHOnW+M+tvIL3do2ep1OxZ1H4Mveg+h670jWFWSzG37deX9GLYHf8Z3CuOKL+ouSrOK8lsalFV7Vda8rVW8uc238SiLAuIqTLaZUmBdTJTLaZUmBcTJTKEyrIFQIyMgSwQAABAEnr34Nlxbz1PUtMuZRfrI06tKPaw8rKk18YnlFpaV7mWKccR6yfJHRaLQelXVK8ta1SF3SfahVi8OL8O4lkymqywzywymWN1Y+sbqhXUk6N5Uh4SipIUJXMW1WnTmujjFpngP8A5xuMuyk9XbSXN0Kef7JRL0hcXv8A9by91Gn/AOEwnHMbuf7bs+pzzx7ctX9pv/35fQVdUa0HTr0YVIPnGSTT9zOX1z0f8B61n+UeGdMnJppyjRVNvPPeODyJ8e8Wy56xU91KH/hLcuNOKZfW1iv/AFYr7jY53qM/Rxo1vZ07XSIwoW9KChTpNuSUVyW+TiONOCKul2VS/goU4095LtLDXh4mgnxbxHPaWr3PuaX2Iw77WdWvqXqbzU7uvTznsVKzcfhkDA27hlBkYAlsjIZABtEZD8iAHIZ3IYyAyTsnlcynmAN1pnEN3bRjSuErugnnsVXlryf78nQUru31e6heabq1xZajFYjGpNvK7t+a8vgcnaadVu6Pbtq1GrVXOh2sT9yfP3GLVhUpTdOpCVOaeGpLDQG91p8QWdKvQvJ1Po1efan6tJUpPv2WE/hk0TNvpXEt/Zr1VaSurfk4Vd3jwf78mf8ARNB1tZsqv0C6f+ql9VvwX7vgBzOehpeJOI7DRLWVWtPtT5Rit8s6HiLRNTs7SvSlBxc6cowrQeYptYTz0PF7qvqkJzsNU9ZUl6x5pTy1LG+Mx5oDE1nVtZ4nm6s5u0sO04Kc3iPil3vlsvDPeYa0qUK7dFqrSUFFPtLOfLzNjrN9VvacZz7EIQ7MKVOnHswhFJ7JdDE0pdrU7Vf72P2oDa8PalW0u4m7ap9FuXH1bUZpNp42bzsbC5401TRqdxbO6m5XNKSdOpUdR5ksKSeNnnrlcuph8VaFVu7iF1Y006jTU03hcthwdpFfReILbXLm0deNnP1ip4XZl7OE23yw3nO/IBwXx3R4T02Wl19I+l9uq6sq8LhLOUljHZfJJdTZ8TekGtdWNSNnpcLajWouGLmOZtvKeMPljGDWcV6xQ4i4nq3ukaPOrfzil2KadXs9nqklu8d3cctf0runKVW/p16dRdKtGSfkljYDb8JcP8UatSitE02vVpv/AFsmoQ25+1JpfA+i/R1YazwfwJKz12/pQup1Z1KUadTLpxkllZXXOXt3nmH4O/ETqK80Cs2pUv5zbpvLcW8VI/2Wl4s9m4isFf6XRnbaZ/KanVh66hKcV24rfDc3js57OV3Z2fIDVW3Gl9acPanq97qlLUNLsqblONxGNWM20nGKbW7eYrn1PnjiPjR8TX9KrLQNM0y5lJRcrBSp02s8+w20n5YOu9O+oS06UeEaFan7dWN/fwpP2YVJU4xVNdcZUp9Prx7jyzT4f5xtv+JH7QNnKz1C3j7EXNRX1qcs59zLNLUalObU4+0ueNmvNHSo5PUUp6rcZ3TqAb201ujVgqNzCnXg/wAypEqr6NpN/Fysrl2lV8oT9qHx5r5nNSpZ5P4ldGtcUX7E2l3PdAV6hw/qmnXMKta2c6HbX46m+3Dn1a5e/B3Po1029vqF9O0tK9woSgperg5YypdxzemcR3VpNKcpQ+xna8GcarSLipVs406Drteu9XFdmeM81yzu9+YGyurCvbz7Nxb1KM3+bODi/mWPUR/Ryd1a8a0tXtPV3P8AJdaL3dOtSmvhjtb+41er6VF20tQsqdvSoLacVdqWH3JSSlnweWByzoLGTl+ILa4vtRqVZXsVp1u49ihGL7ecYUmksyy1nKzsejabqtGtUqadPQre9qwp59bGm6SptrbtSjhS6bc/E5DXOEbmVd1PptP2p+snmnHKk+eHjKXhnAGk4TlcRtrjTYahUqeuXrKc4weKST3575ed+i8zb/yVUccVb2tUfjGP7hTpWHDWmXN5KMq9dqKlOT/WS9y3Nzwvc2Gt14W1xWqaZXqQ7dOVxD8VNdMTT5+7oBqrbT4W0XGn2n2nlt9TLlVvXRjRd1X9VFpxg6jwn4LodtW4D1bsduhVtK8cZj2Kj9r4rBzd/pt1Y3Do3dvUoTXSSxny7wMWpq2rt05zv6zqUvqVHj1kfDtfWx4ZL74i1ZT9ZB0IVn+UqQpJes/bS9mXvRTQsa9fteot6tXs8+xBvBaqW8oScZxcZLmmsAX5a7BQ9nR7KLl+Uh2c0p9/sPPZf7LRQ73QVFtaG/a+vRlPKXjCosSXk1Isugmm9i26GeSAyHR4U7Ll2b2VOXODfZq034PeM159ko/kXh9py/lmTpS5TVPEoftU3hv+i5eRY+j5yW5UN+QGQuFaLjlavZOD3jVi+1Tx+s17UP6UUvEs1ODNQTjGorWLqP8AFOdRKFXu7E/qvPdnPgWpW+/ImMKsIyjTqTjGSxJRbSa8QMSXB2oyrzofyVL18Fl0sJVGu9R5teK2Ne+Gas41JRsLrs0n2ajjCWIPufcdC73Ufo8LZ3laVGn+ThKbkoeMc8n5GW9d1iXq5TunKrTWIV+ylVS7u2vaa8G2gOQsLCnaVqs49qfrLerRcZbr26coZ812srxSNfbcHVrytGjTu1HtvGZU+XzPTKPEN805O3to15flKsIdn1v7cPqS88Z8Tb6Pf6apeuWiWtOvL66jFOlLyi94f0ZLyA8uq+iXVpT7NnqVpWeMpShKL+SZRdehvjejRVaFrY16T5TheQUX5Sk0vdk98stU0uhTU6FrOlPn2fzo/s1V7S8mmbCnrNi068bmcZy+s1FQq/0lj1dT34YHy1d+jnjm1l2anDGpSlz7NKl6x478RyzR3uj6tZQc7zS762inhyq28oLPduj7DpXNtKM+zeW0rSpj8XCmsRx30pPKfjHJdtatzCk3O4pVrd5UK7qyaiu7txXajvj62EEfFIPsurp1tddu11PRrC4hPelKrSpe233Npwn06Z3NNPgXg27nWs7nhKxhVy3mjQcJcubiuy0v2ZY8Ar5NB9NP0Q8CXVpUqR0+8tMt+3C7nKMPB57Th/Si/M09x6CNAlBulrepW8p/kvWernGXlJYjL4xfgB8+g9T1v0OXVlUlTtNft7ionjsVLeVNx8JbtJ+WTnL70ccUW2XTtqF0v9zWX2SwBx4NhqGiaxp+fpumXdBfpTpPs/Hka8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABm2WlanetK00+6r5/QpNr4gYQOt0/0d8VXeG7CNvF9a1RR+SyzoLL0SXWFPUdYoUYr6ypwcvm8AeZA9cocD8DWMuze6tUvKn+zjVTfwgsm903RuHqXZ/krg65u5L6tSpbYX9aq0B4Za2t1dT7FrbVq8u6nByfyN7YcEcU3qUqWj14RfWrin/a3PdKVtxF6vFDTNM06CX+tqubXugkvmcLxfxbxDpN46FO/tJSXPsW2Evi2BpLD0Ua3VqQ+m3lpbU2123BupJLrhbJv3o3svQ9p35vEt8//APmQ/wD85y79I3FUJ/6bRlvydCP7jteDOL9W1LsfT4UOy2k2o4wBrdf9CWo0NAu9W4c1mnrbs49uvafRnRuOx1lGPakpJeeTyU+tdOvLnTb2le2dXsVIbxkuTXc+9HCem30cWur6fc8ecH2qpSh+M1jTaa/JPrWpr9F7trpz7wPBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAu2tvXu7mnbWtCpXr1ZKFOnTi5SnJ8kkt2zveAPRRr3EtpDV9QqQ0HQG8vULyLXrV3UofWqPny22e56zo8uGeBbKVLhCy+j13Hs1tYvMSu6vf2elOL7o78uoHFcI+hqnY0qep+kS9np8JJTp6RayUryqunbf1aS5c8vnsmdpqfElnpWk/yNo1nbaHpK+rZWa9qp41J/WqPxkzlNY4kq1qk/USk5SeZVZvMpPvOdq1Z1ZOU5Nt823uwNnqmt3F1mEH6qn+jF7vzZqXJvmQgAJLNzcUbem51pxhHxOf1LXqlXNO0ThD9J82Bur/UbazjipPtT6QjzOdv9UubzMc+rp/ox+817k5ScpNtvm2MgVplcWWkypMC6mVpllMrTAuplaZZiytMC8mXIyLCZWmBfiy5GRjxkVqQGRGRcjIx4yK4yAyYyK4yMeMi4pAYevVezZSWeex0XoJh6u/1rUGv9H02aT8ZNR+85HiKp+LjHvZ2/oph9G4F4gvMYderRt4v+tJ/YgMvtZeSpMsplSYF5MrTLKZWpAXUypMtJlSYF1MlMtplSYFxMnJbTJTArDZCy+W7NnZaVVqYncZpx/R/Of7gMCjSqVpqFKDnJ9Ebix0mEMTun2pforl/ibC3o0qEOxShGK8Opdx1YFMYxjBKPZSXJJcir4jHgMd4DOSQsdC5b+p9dH6QqjpZ9r1bSljwyBTThOpUjThBynJ4SSy2Z9vp9N1ZW99cTsLj8yNai1F+cs7fA3NHQNL1O2U9F1GTrxWZUq+M+9Jbee6LE/wCWrSLtNT02pe20fzZwbUf2ZrkBqdQ068sWnXpfi5fUqRfahLya2MQyq11WpxqUKE7ihbze9F1G17+X2GJkA+YBAB7EMkhJyliMXJ9yAgCWU2msPuZDYEvuIGSMgMAZIbAqTaaabTW6aZsqWryqU1R1KhG9pJYUpPFSPlLn7nk1eSG/gBt5aZRu129JuVXfN0KuI1V90vcaupCpSm4VIShKL3UlhooUmmmsp9Gjpa7cuGZy1vCuM/zRz/KtePXGe/8AcBiaVxJf2UVSqtXVDk4VN3jwf/U1d76OOF+Lr241a1v9Qs7ycnKpSUopLPTl9XGyx0LNGnVr1FSo05VJy5RistnZcPafDQ6VS/1KvClOcez2c8lz9726BHmmrcBaZZunp9e3qUHQjhOE/wAov0m2t/l8jzy+nYadrEZ2NOU6MKyhGc6nu7TPUPS1xNGrB/R12JVYeppZ59lZ7Un8ce88dpW1fV9ZttJtcesq1FBN8k3zb8Et2FdBxbO5tpUnY3NWr2o5coPKflg56F/q3blbXtW6pp03NwnXdLtR6PfmufI900vhmvp+nUbOxpwqUKEeypp9lS73vjm9zg/TRobnp9pezjCnXp1vVdrtJ5i03hteMdvNgcRwZxhqvCOr1r/RqlJVKlN05xqJyjJNp74afNd53tp6e9ccOzqGkWd139ifYX/MpHkrsK9JueYSS54fQx502syS2A9ssfTHw9O4jc3nCFG1uI8q9tTpyqLv9rEWdLpPpw4UtaVWU6GprEcxpOjHMn3JqTXxPmxZb64Ko5zzA3fEGq3Oua3e6xeSzWu60qslnKjl7RWeiWEvBIsaU1LVLeP6+fka+hPfsP3Gx0WKerUH3N/YwOsRyl1vqdd/72X2nVpnM1bHU53Fa5o6Zd1qLqzaqQpScXiT6pAWUm28Fy3tp17iFKn7U5yUVvjLfcbChoGs19Pq3lGyniCyoT9mU/JPdnMqrcO4U22pxfLlgo9FocEqdpmvdOlXfKMY9qK8H3+75mj1LhvU9PbqOlJwX+so+1H3rmjoOEuKFUp07XUZYkliNV/f+86p1J1PyGMfpvl7u/7CDyy1v722akm5xX50GdnwjqV7rU5WnZdSkl7cpNrsrzW69xsL7QNNvE3UpONZvLqweJZ+z5GZp8IaZa07O2pqNPG80t5vq34gbu3VnpVkrSyhGEcttrq3uzUX9xKo22zIkpShlmuunjIHPcYV3Q0S4qrsOSccKSTT9pdGY/B+ja5DXqGqavSpKhRUuzQlLn2k8YUHst+8xPSJXcdKpUoyw51lld6Sf34N1wBq077TI2lZzlWoQWJNPePLd965eKx4gd3bapRt6Kp0HqFDm/xN52YL3dl/aZ9C8utWirH6fe3DmtqU7SnVS8e03t57GvttJjRoxu9Wqu0oSWYU0s1av7Mei8WbVwqOyXrMaLpMvzE817hfa/sCNRU0m8t9Vjb6XdRurpb5tpPNN+MsJL4mwo2tnZV5zuHDV9UXtVO1P8RRffOT+s/+hN7fRtbZ2lKlLT7V86FN/j63jOX5qf8ACaNBe3dWrT9XCChQi8xo09l5vvfiyjY8VX1rfulUo2NlTn/rKtCbbm+7DSa82veaLsrdNIrx2l7UcLqmFBKOIpRxySWCKo7PNroUukmi8o77CWOuceAGO6Uc4W3uKXSWeRk4zyXxIfLljoBjOksbLIVJZwkveZSiny22JcfkBapUlnfCfcbOyglhow7dxnlxkpYe+GbG3S7sAZsM4wV5LUXlFYFeepMakoyzGTi+9PBQh1AyKV5dUqTo07irGnLnBSfZfmuRfpatfwoRo/SJTpxeYqaUnHyb3XuwYHUZA29PiC+ik6nYqVI/VrPKqJd3aTWV55KLnWruvQlTSp0u2mqrppx9Yv1kts+KSNYM7Yxv3gTnqMkZJAk1Op8N6DqcWrzSrWpJvLmodif9aOGbTJOQPOda9FdlVjKekX9W3nhtU667cG+iysNL4nAcQ8K63oWZX1nL1KeFXpvtU371y9+D6FyROMZwcJxUotYaaymB8wA9i4v9HFjf9u60Vwsrnm6WPxU//D7tvA8p1bTL/SbyVpqFtUt6q6SWzXenya8UBhgAAAAAAAAAAAAABOH3AQC5To1qjxTpVJvujFs2Frw9rt00qGj3088sUJY+wDVg6+w9HHFd012rGFtF9a1VLHuWWb619FFSlD1mq65b28VzVOOce9tAeZA9esuEOAbWajUvbjU6q/NpSlU38qaOk0zTLC3x/I3BFXPSpWpQpfOb7XyA8LsdH1W+aVnp11X7XJwpNr48jotP9G/Fd3vOzp2q769VL5LLPbadlxRWWFHStOh5yrSXw7KLseGrytvfcQX0++NvGNFfJZ+YHl1n6JJQjGpqet0qcfzlSp/3pP7jYWvB/AFlLs172rqFVfmxqub/AKtNHpFHhHQYTU6tl9Kmvz7mpKq/+Zs29taW1tBQt7elSiukIJfYBwGm6dpVDs/yNwXcVGuVWdvGmvjUafyN3RtOKKyxTtdL06HTt1JVpfCKivmdWAOahw5qNZp33EVy11ha0oUl8Xl/Mv0uEdDT7VxbVb2XfdVpVfk3g3wAxrSwsrOKha2lChFdKdNR+wyQVRpzl9WDfuA1HEN2rSxrVG8YifOXE167vU69ZvOZPB716QNJ1u902pTsbZyUlu8njmocBcQ07Wd1OlTUFz9oDjqEHWuYwXVnb295HT6VG2ptKezkcnpvq7K/n9L9mVPpjqXIXk6946sm/alkD3vhfUo3VlCnUls1s+5nSaPqN1o+pQuqDy1tOD+rUi+cX4M8p4Lv2qUYOR6HZXEa9KMJP2vzX9wHnXp89HNvpeONeFKGeHr2f84oQW9hWfOLXSDb26JvHJxz48fX+g31tbu403VaCutJv6boXtvNZUoPbPmj529MfAdzwHxVKzjKVxpV2nX0665qrSfRvl2o5SfufVAcSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAyNPsrzUb2lZafa17u6rS7NOjRpuc5vuSW7PZ+FPQ7p2iQp6h6Rrtu4+tDQrKqnVfd66osqmuW0cvD5rkB5jwRwZxHxlqDs9A06pcKG9avL2KNCP6U5vaK+fcmez8McD8GcExjcXSocV67Df1lWP8AMLeX6sH+Va75bcmkjYa9xTTt9NhpVnRt9L0ul+R06yj2Iecsbyfe5ddzhNU1m4u24p+rp/oxf2sDqeK+Mbm/u5Vrq6ne3HJNv2ILuSWyXgji72+r3dTt1qjk+ncvJGM23zIAnmCDA1HVbWzi05dup0hEDPbUYuUmklzbNNqeu0qOadqlUn+l0RpNQ1O6vXiUuxT6QiYWAL1zc17mo6leo5vx6FoACSSABUipMoROQK0ypMtplSYF1MqTLKZWmBeTKlIsplaYF5SK4ssJlxMC/GW5XGRjxZcTAyIyKlLYsJlXawgNTrk+1XjHnhHpHC0Xaei20i9nd39Sp5qMYxX3nl+oS7d2z1e6X0XhHhyyaw42bqyXjOTf2YAxFIqUjHUitSAyEytMx1IriwL8WVqRYUitSAvKRKZaUjIs7avdT7FCnKb6tcl5gRkzrDTri79pLsU/05fd3m007RqFHE67Vap3fmr95tksbJAYtlYW1rFOEe1U6ykt/wDAyduiwS3jkU83vzAlhcgkum5PQCMsLI6YNlo2jX2qVMW1LFNPEqs9ox9/X3Aa5c8HR6HwreX3ZrXebWg9917cl4Lp5s6XSeH9L0Wl9JryhUqx3derhKPkun2mr13jFLtUNKjl8nXmv7K/f8AjdOWjcN2ePYo5XLnUqfe/sRy+pcZahVuM2UYW9Jck4qTl5t/cWNL0LVNbrO6upzp0pvLrVd5S8l1+w3mo8O8O2Foql3XrUkljtes9qb8Fj7ANRHjLUnHs1KFpUXXMHv8AMzaF/K5oK4qcI060ZfnwprfxXsmitr3S7O/lWo6dO5pL6iuKiyvHCWPtO0sdYq39g5UvoNrXlj1cJ3Kn72ktvIDQXFzw8pL6doF3ayfdmK+1Fp0+EK7zGveW3hhv7mWOIrC8hdRq6pqtrOc30lKTS71FLZG10WXCdnTw7qncVWsSnWpS+SawgMB6Pw/V/wBH16MP+Kl/gRLhZVFm11ezrZ5b4+zJevtY4clcOEdGjUpL/WQioNvwW3zM+x0XRtToOtT0y7tIP6spyx2vJZf2AaSpwnq0I+yreov1an78FpcMap2O1OEKcv0XmX9lNG71LRdL0uh616reWq/NSqZcn4JJNnOQ1fVIVnC1v7uacsQU32pPu233ClXQ9Sp86UceMuz/AGsGLLTb6Lwracv2F2vsO44fp69Ls1tTu1GnzVL1ce0/NpbfxyL+uarp1hBq67FWo17NJJOT/cEec1aFaltVo1YftRaLZsNW1avf1Go06dvR6U6awve+pg0qdStUVOlTlUnLZRisthVHN45GXYaddXspeqUYUofXrTeIRXizM+hWWmJT1SfrrjpaUpcv25dPJFqVbU9bqRtqFLs0I/VpU12adNeP+IFyV5p+lezp8Vd3S53NSPsxf6kfvZVp+j6jq9b6XeVJ06ct3UqfWkvBfwjNp2Ok6DSjX1GpG4ucZjBLPwX3s0utcQ3moZp59RQ/2cXz831CN3X1bS9Dpu20qlGvcYxKo3lZ8X18lsc1qOoXF3UlcXldy7KzlvCivBdDzm79IttQvq1GOnVKlGE3GNRTx2kuuMG54c16w4oqu1uKkdLsmsVrmvmS6eylFZfXuXeFcvxDe3WsapOpa29a4nJ+rt6VODlJrosL4noXoh9HGq6PXqazrdGhC6qQ7NKlL25Uk+beNlLkuu2T0zg6jwDpdpB6Zrmk1qslh1p3MFOXgk3svBHYW1C1uYKpb1qVaL5SpzUl8UBw9bRL2tW/GS9dT6L1jh8sfeecel/h3j3V6Vtp2laBKtptFqtJ0qkHJ1MNJY7WcJN9Op9FUrGGeSMmnZwj0QHwNrGm65o1b1OsaXe2MnslXoyhnyytzcej7hRcVXFeNbULextLXsyrTqPM2nnHZj15H2Nrmr0P5L9VV4WuNboz7Ua1CnToScGnjEoVJrPLpk874f4N4W1niapf6VwPr3DV3CnNSjdW8qNrVzh4azJLdL6rXkwOX0rg/hq2sZULDRrT6Mtqt/qFKNWpPyysR8orJF1pXC9OhK1s9A06aksTr1baHbf7Kx7K+Z0esaXfTu50bhTToycOxTnGUYY6JPs4NRd2P0aS9bOdJP8AOqU9vjFyA8y484KpK3jf6DauFSn+VoU8vtLviu/w/h83o+lalRqUb+5tKtCiqjpp1I9luTjLkn5HtDpKTxTuKFTyl2f7SRznHtKpTsbNz7GPpX5tRS/1c+5gcumdVwc3/k5b5xjt1X/9SRyaZ1XCsZQ4etFJNKSlJZXNOTYG2k0+XXuOU4u4Vo6lOV5ZRjTvcZlFbKr59z8TqnCajHtRlFSWU2ua8AkksLkB4xUhVt60qdSMqdSDw4tYaZ1PC3E0rZxtL1uVHlGXNw/w8DpeJuH7fWKTnHs0rqK9ipjn4S719nyPN72yubG6lbXNKVOrHp3+K70VHrtGca0IypThOE1mMk9mj1PRfRJVqxjPVtVhCnJZ7Fqsv+tJY+R808M65dabPs1ITq23OUe7xXievcL8ZaxY29Gvp19OtbTScadX2oY7t915JoirfEmmz0bVLvTaj7Tt6jgpY+tHo/esM5e9nuzYek7j20ramrm6t+xeVKUe3TpPKeFjtb8s92/I1no903iL0halO20PTKdOjT/K3NzX7FOn3ck2/JJlHp/oe4A4S4h4eqaxxRpsL2qrqULaNSpNRjBKOX2YtKWXnnnkdfxbpnCmi6d9F4c4c0201OvH8RWo2VOPYxKOW54ztk7HhnQbfhzhyw0em41PotJRnUxjtze8pe9ts4/0yapZUtKoWPr1OpOp2pUqU0m4pP63cs46b4Oju4Jj7S2/rNf6a9cm/M089xbWVWVz21qV+3+Mu6u9Gk/1c/Wa/dyNXqGoP6V6+FapcXTWJV54fZ/Y2281y6FKne3sex7bowx2nGHswXe8Ll+4VrXTKVRweoVqzS3lRt8x8k5ST+Rzs2umnJtybbby2w14GVcq1WPo1WtPv9ZTUfsbLLjnv+JFWuznkSolyUHglQ8ALLj7x2O/JecHkdnf/ACw4BQy9vtMqnRnN4hCcn4RMiOlalPDhp93Lyoyf3Aa1xaeEvcKMZNPtxSeeSecm5paBrNR4jpd1/SpOP2mVHhTX3v/ACdJedSC+8DSUqe/MyaS25G1jw1qUZL17tKH/EuYfc2ZNPh6SWZaxo8f/wAzn7EBqo5KjbLSbOm8VtcsV3+rU5/Yi4tP0NL2tfz4RtJgayztLm7qOna0KlaUY9qShFvEVzbx0N6uEbz2O1e2ajVX4ir236qq/wBFTxtLweDGtno9jdQuLbWNVVSDzGVvDsb++RsaHEFnQq1qkbvVqqrr8bSq0qDpVPGUcc/HmBrnwxrPZrKNspVqG9SgpfjEu9R/OXismDU0zUadrC6nZ1VQm8Koo9qOe5tcn4G3XFV1DNLsKvQhLNs55jVt/wBmaeceDyi7DjS6UHOenUpXUlitUjUxTuF3VKfZw34ppgaa20fUq919GVrUpTx2n61dhJd+5VUtLCg+zW1JVJLmqFJzXxbR0MuMdNVKjB6bfSpJ5dBSi/oz/So1e0pL9lrHkZVXXeGq17SnXuacq0l7F3Kybf7NeHZ7Mu7tRfwA5B/yYtl9Ml4+zH94xpsn9e7p/wBGMvvR3tnfaMruSt62mRrtZnbO4j6up+tRqP6r/Vl8uZiVeHtIvvpNwqtxUhnM6tKK9bbS/wB5SS9qP60OfzA4xWtCpWUaN/QjBr/XwlB593aXxZN7pt5aQjUrUs0pfVqwalCXk1sdNX4Poq1iqd6oV54dGpNqVC48ITWMS/Vkvf1M2ejXFpw/X0zTsXFSUk7yNbKqUE8e0qcc5XLeLf7g4Dtx7zPsNMuL2LnSnQjFL8+rFN+7n8jLu+GtVoV6MadGlcUK/wCSuqVaPqZP9HtSaw/BpFmroesU7qpay0259fTWZU4w7Tx34XNeK2AmrotzS7Sq3FlTlH82VdJ+45/WdJ0/V7V2mpWtO4p52zzi+9Nbp+KNjKhWjD1k6VRQTx2nF4z3ZLeEB4zxf6OtQ0xTutKc760W7hj8bBeS+svFfA4V7PDPqHbvPOfStwhb1rCtr2nUY0riiu3cwgtqsessd65vvWQPJAAAAAAAro0qtaoqdGlOpN8owi237kAowlVqwpx+tOSivNnsHDNtdaHZK2s+H9IqSf5StO6bnUfi3Hl4LY890ThbiG4vqE4aRdxhCpGUpVIerWE/1sHrdrpd8t5unHwct/kBVC71ypjs8NaO8/8AtcV/dLsp8SR/J8O6JF/+85+yJm29o6bTdXPki/cXtC0p9q5uaNCH6VSaiviwNdTlxnPalZ6FbeLlOWPhgyaem8V1Vm64j0+1T5+otM498ma6+4z4btZdmrrNCT/3Wan9lM0V76TtDp9pW9te3ElyfZjGL97efkB2S0Oi/wDtDifVrp9Y0qipRf8AUS+0yLTR+E6NRTemRrVF/rLhOrL4ybPKr30pXs44s9Kt6L76tRz+zsmqn6Q+Jqlbtq4oU4r8yNFdn55fzA+iLa502lBQoOlSj3Rh2V9hkxuKE/q1qb8pI+eaHpK12DXrLexqLr+Lkn8pGxo+lGeEq2jQk+rhcY+XZA95TT5PIPGLb0naThOrZ39KX6nZkvjlG2s/SRoc45eqXNB/o1Kc39iaA9SBwllx3pdbDp67ZvwqyUP7WDcW3FcriGLS5sKr76bUvsYHSxp1JfVhL4EypdhdqrUp013ylg4TV7jie9clR4hdrD9GFsvtychqvDPE15NuprquU/8Aa1Jx+WGB6vqHEHDunpu81m2i1+bGWX8jnNR9KHC1rlW8Lm8kuqjhfM83o8BavVr9mpc2dOP6cpya+Syb+x9GNh6tTv8AiOmsc40KLb+f7gLupemK7eY6bpNCiukqj7TOX1P0hcWahmL1OVCL6Uo9k6i40T0d6LvdrU72ce+Ekn8EjDlxfwlZ5jpPCdCUlylW9r97Ap4b1r0h3Fiv5MudRuI8u16pTj8WsGi4iu+P1OrUvoXiTb7X4pdn4JYM3XPSFxXdWyt9Noqxt0sKNGlj7f3HN3XEfF1e1lSr3ty6WN1tgDnrqdxKMpVZe037SccMmzi8Ka6c0W51atSTjOUpNvL7XeZFD1ttUhLs9qEuneB2HC9z2Kkdz0fS7uKpdqUkopZb7jynSqkYzhKm8xb28PA9G4d02vqFGnO5UqdlzkuTrY/NX6ve/cuuA7KlVhcWFtcxcs1Yt7rDazhP3lr8I/SK1P0CaTOsm61tqFOs+0t4QnGoseH1onbcE8LyuKlLVdTp+rtY4lQotYdXHLbpH7fIxfwnE6/oV1ypNLMJ20l4fj4L7wPi0AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA6LgfgriPjPUHaaDp8q0Yb17ib7FChH9Kc3sl8+5MDnT0jgD0Sa1xBaU9a1uvHh7QHury6g/WV13UafObffsvF4weicLcGcGcCqNxONDirX4LP0itD+Y20v93B/lGv0pbbJpIx+KeL7i9u517q6ne3T27Un7MF3JLZLwQG60664e4K06pZ8G2P8nRlHs19TuGpXlwuvtfmL9WODjdW4inUlKNs5LtPepL6zNJe31e7qOdao5Pp3LyRigXKtWdSTlKTbfNt8ygESkoxcpNJLm2BJZurqha0/WVqiivmzU6nr1Ok3TtEqk/0nyRz1xWrXNT1leo5y8QNpqWu1q+adqnTh+l1ZqMNvMm2+rZKROAIwMFWBgCMDBUkVJAUYHZLiROALXZYwy+okqCAxwZPqkyulZzrz7FKLlJ9yAxEV7rmmjo7TSbewpqteTi6nSJFxdWlwnSq0Iun3pYa8gOfTKkybuj6mvKEZduCfsyXVFpSAvplaZYUitSAvxZWmY6kXIyAvpkzliDZZUiK88UpPwA1aTq3qiucp4+Z6xxbJU9SpWqfs21tSorw7MEn8zzThS3+mcTWFu1lVLiCflk7jXrtXOtXdZPaVWTXlkCmMitSMONQrjUAzIyK4yMRVCuM/EDLUi5SUpzUIRcpN4SSy2ZekaLeX3ZqSXqaL/Pkt35Lqdbp2nWthHFGmu21hzlvJgafTNAqS7NS9fYi9/Vrm/N9DoqNGhQh6uhHsRXJKJUMAMdWM+YWRswGOowRvnBOQG/mTGOZKKwm3jd4ITytxsB2HDXDmnz7Ne+vba4nzVGlVTivNrmddcOpbWXZsLWnUlFYhTUlCKPIc9CujXrUZdqjWqU5d8JNfYEdJrlhxTf1PWXlvOcF9WnTnFxj5JP5mn/k/VbSpGr/J9zFweU5UW0visClrWrU3mGo3WV0lUcl8GZdPirXItZu4zS6SpR/cFbLReJ7uFw1q13UjTj+bG3WX5vobe84p0N28pZdxJcqbpPL+KwaGPGepdns1bezqLrmEt/mVf5SWFSPaueHbWTfOSS++IRH+UGkVW/pHD1uvGDWfsRr9TvtKnUhLTtLjRcXlupJvPh2c4wZ/8ocJ1nmtpFelJ/oSePlJfYHR4PrvMbq7tvDD/cwrK0XiWVapGjqErChQisfkZZfgsbI3M/8AJ+6t6lenbW10ofW9TRU5fBLJzctF0Os/5tr9OHhVSz9qNjaWGsW1CNHS9dtJ01yj2I4+xhGg1G8hR1BVrPSoWapv2VVg234tPZGVHi/VPVuLjbttYUuw8r54MvWNL4ovYRhcVaVeC/NpzUU/FrCyZWh8KULdRrai416i3VNfUj59/wBgGhsNM1XXrh3NWc+xJ+1Xqcvd3+462x07S9Ct3Xk4xkl7Veo9/d3eSMfW+JbPT4uhaqNevFYUYv2Ieb+5HEanqN3qNf1t1WlN/mx5Rj4JAdBrnFlSr2qGmZpw5OrJe0/JdDlpzlOTnOUpSk8tt5bL1hZXN9W9VbUnOXV9I+LfQ2Ha0zSfqqGoXi6v8jTf95hWPZ6XUqUVd3lWNpaf7SfOX7K5srqapGhF2ui0ZUVL2XVe9Wp7+nkibez1XX7j19WUnDOHUntFLuiv3G0nV0fh2DhSj9KvsYbzuvN/m+S3CMTTeHpdj6Xq1X1FJe04OWG/N9PtGo8R0baj9E0alGnBbescfsX3s0urapeajU7VxV9lP2YR2jH3Fq2tbtKncRow7Dfs+t7OJeSlzCsavWqVakqlWcpzlu5N5bMS7qQhbVZVJOMIwbk0uSwb2dvaXFWTnaVqL54orK+1r5GNdaZZVYSpyqzpQknGSqPtNrySA4n1PDFz9XUbbf8A2lLBXwppul1a6pVXTlQy6ksP2ZPoscjD1L0e07ftO21tS6wjUoY+yX3FjTqNTTKUbWpUjKdH2ZuL5b5T+f2CD0y60fhy/svo9a3s2seyls15PvOXuJaZosY29fTLS6p0szVzJzjUlF8lKUZLdfHY1cLmv6zMK8kvM1PHN7VeltKT7SW7XvZldCm79IGq09UnDQ9R1Wxodv8AF0KN9UxFdE23v8Df6B6W/SBpd5Fy1id3TTTlb31ONSEl3dtJSXxPOuFNNp3spzd46FWDXZTWVL5+Xf1N5cqFlczp3lzQrTUV2V2t+z0f29TEfXHoq460/j3Sala3puzv7VqF5Z1HmVJvk0+sXh4Z2r7FJbc+8+MvQzxLV0j0q6JWsaknQurlafcQ7X1oVMKKeOfZeH/RR9a6jqMKUW3JJLxA8r9LMatnxo7qyuKdGVehCc4uTXaabWXtjkkai21ipKm6d7b0quVj8VVhLPubRjekDWI6xxJVr0ZKVKlFUoST54y2/i2aFNgb2/ttPqxdWjC5tZc8OjKUflnBx/Hl3cVNOtbWrU9ZCFx203Hf6klze+NzbJtcs5NBx5cy+g20a9abh6xtKUns8Ac2nuj0Ph3h+8fCemXdvP1kaltCo4xSysrPLrzPMnXt1ByVXks4yeicIXlWhw1p8YQUf5vDMlUnF8l3SRR0FlexqRVrqV3cRw8ZnCNSK84tZRtocP0rmmqtvXs68XyfYcV/yyOZqXNOth3EbmcukvX5x/WTfzNFqXE1TSKzp6Le1fpWcT7cU4Q88fWfgEdXqFg7KbV1p1WEW9p063sv4pnIcfU9Njb2lRScK3rXFKo4/Vw29+u/Z6HIVOLtfqatLVqmrVKlzCLcYTj2odiW2HH6qW6eEmjQ6rUub66lc391KtVl1k+XglyS8FsB0N06VKyrVINNqDxg9A9H+lJ8M6eq97bW9N0nUlKdTMkm3L6vNvD5Hjmn29xXrQs7OFSc67UFFbuTb5H0zw3wjdVrS0esVpKjQowpRpP6zhGOEn+iv42A8pXo/wBV4m4turv11J6eqr/G9p4jTTwstpLkuniemaJa2GgW6paFCNSrSXZlfVI4p0vCC6v4tm31/W9JtLT+TdPtaFwobYx+Ki/L85nHTqVKrfaqdlN5wliK8ktkBsr7W9QnKVKF5e1ZvaVWtUbm/BL81eW/j0KbfRnTnTepV4WzqbxovLqNd7XT34LGl39TTbh1qVGhVnjClUi32fFcsPxMmlqtL1yq19PoVMNt7JuWe9zUm/iBtoQjQ01RtmqlvN+zCrOMY1H1wm8NmM6VOqsz0erNd9O3Sj8YCevqpdUqqVahGlFqnBe0o57sOL92ceBj3eteri6em+ug3Hszuass1pruT/Mj4L4gWLz+SIRxG2bqZw405Ti15uWfsMGtK2nBKjaunL9J1O0/sRVaW1xe3EaFrRnVqy5KKz8Ts9O4b07RbZajxBWpzmvq0ucU+7HOT+QHPaHwzqWqwVWnTVGg+VWpsn5LmzPr6ZwxpMnC8va9/cR+tTopKOfF9PiTr3FV5qUvoenQnb28vZSh+Un4bcvJGPT0e002irnXqsozks07Ok16yX7T/NQFE9b0ui3Gy4es0ukq7dRv4k0te1ur7Nja0aXcqFov3MxKms1YS7On2ltY01lRdOmpT985Zln4GPVv7+4WK97c1F3Sqyl9rA28r3jGouV9Hyodn7EixOpxRL8re3VNfr3Sp/a0aZ7vmyANnOhqlX8vqlF/t38Jf3mWXZL/AFuqWa7/AG5S/sxZhJLljfwJSXRfIKz42dmvratQf7NKo/tii7GhpsV7V/Wl+xbfvkjXxjjoXYJJJAXJqCnJU5SlDPsuUcNryyxghFWGQRgkkARgYJ2ADBGCQBHZXchCCjLtR9mXfHZlQAu0bu9o21S1o3tzChU3nTVV9iT72uRlrXtdVOhBapW/m/5GbjFzh4KbXax4Zwa8AdBDi/Vo1XVdOxbmsXEHRfYuV17cc9nPikmZVHjL8bClX0tu0hvTUbl+tt5d9OeMpfqttHKjzA7e14xsqleq69K4oVpLH0hUoyhcL9GtTTw/2lv5FVPWeGrq0nRqUqSp59qzrQaS/Wo1cez+zLC8jh2Ry5gdvUjwfeW0I/S7Lsxfs1qclTrU/wBWtRTTa/WiaD0jaNpFtwpdTi6lrd3lpUpW1CNVVaFzOUXFOnU3aw2m1J8n7zgeMuMdL4ehKjlXV9j2aEHy8ZPovmeUVuM9eq6v/KMrmHbTzGn2E4QXck/MCLjgjimh9bSak0+tOcZfYzLsvR3xNcNest6FtF9atZf3cs2FH0n65HCq2en1F19iSb/5iLz0na3UUo29pZUE+T7MpSXxePkBmWXorrtp3ur0od6o0nL5tr7Dc2vo14eto9u6r3dfHPt1FCPySfzPPr3jPia7j2amr14R/wBylT+cUmaa6u7q7n27q5rV5fpVKjk/mB7B9G9HekJqS0ntR5qc1Wkvc8sprekHhWxp9izhWqpbKNCh2F/zYPGwB6be+lRbqy0h+Eq1b7kvvNJe+kjiOvn1LtbZf7ull/8AM2caANvecT8QXmVX1i8afOMajgvgsI1U5znLtTnKTfVvJSAAAAExIJiBXkZKQBVkkpGQKskp4ezKSQMu21G/tnm3vbmi/wBSrKP2M2NvxXxFReYavdPH6cu39uTR5JyB1dvx/wAR02u3cUK37dFfdg2Nv6S9Tjj1+n2lT9hyj97ODKkB6Xb+kuhLa40mpHxhVUvtSMulx3w1cbXNpXh/xKMZL5NnlSKkB65S4g4LufZVahTb6Soyh88EytOD73Mad9apy6Quln4NnkaJTA7niLhfh6xpTuKWoV4tLLlOUZQXwW5z1tSo1qHYp1I1qUntKPOD6ZXNGHY0rq/qU9No0al06kvxdGOW+13ru+zvPYfRn6N4abe06zpyvtWqP8XBLMaC7l0bXWT27sc2Gr4D4EqxqwvNXpScpNeqs0suT6OS+yPx7n9B8JcH07SML7WKUZ1sZpWrXsw7nL9xs+FeG7XQ4q5ruNxqDW9TGY0vCPe/E3j575z1AicnNtuSb8ziPTzau89DfEtFLeNqqv8AUqRn/dO3eceBzfpRip+jDimLX/qe6fwpSf3AfBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAXrK1ur67pWdlb1rm5rSUKVKlBznOT5JJbtmXw1omqcR65a6Lo1pO7vrqfYpU49erbfJJLLbeySPpPhzQ9G9Femz0/SpUb7impHs6hqvZyrfvpUc8sdXzfXoohxXCfocsNFp0tS9I91KNZpTpaFZ1E60u711RbU1y2WW0+aaaOp1/iqlb6bDSrOjb6XpdL8jp1lHsw85dZPvbOU4n4nVCpOMKrqV5tuVSTy2+rOVV5O5m5zk233sDc6prNxdtwT9XT/Ri/tZq223uyAABaubijbU3UrTUY+Jzupa7Vr5pWidOH6XVgbjUtUtrJYlLt1OkI/ec1qGpXV82py7FPpFGLhuTlJtyfVk4ApUUirBOCUgKcFSQwVJAU4JSKkiUgISJwSkSkASKkgkT0AJFS7upkWNlcXk8UYez1m+SNtTpWOmLO1e473yQGLYaTVqx9dcv1NHx5szJ3tvZ03SsqaT6zfNmFeXta4lmUnjojEyBdrVqlWTlOTbZbBKQEEOlCXOC+BWkVpAYztYPlmJS7Sa+rJPzM6MS5GAVqnRqx5wb8tyMtPDymbyNJPmiv6JTmsOIGiUi3dzxRkb+WkU5/V2MO70C4nFqnIC36PUlxHTuHyoU51fhF4+Znzr9qpKTfN5LehadcabG8nU7LnUouFNLOctrPyMWarU37dOUfNAbCNbxLiq+JrrZVrivGjQpzq1JPEYwWW/cd5w5wROXZuNYn2VzVCD3/pP7kEaTSLG91Ot6q0ouePrSe0Y+bO60Phq0skqt0/pNfpleyn4L72bq1t6FrRjRt6UKVOPKMVhIuvYCFhdCPeS9wkBG45FWCAIHuGPENd4BMjIwSBGRkEgUonkMDG3IAgBkAGAAIyAA8TIsLK6vq6o2lGVSfXHJeLfQx+hm2GrahYU/V2ly6cM5x2U180B1NnZ2nDtBXWp39SpWxmNKE3jPguvm9jRa7xLeah2qVHNvbPbsxftSXi/uRqLmtWuK0q1epKpOT3lJ5ZlabpdzfKVSPZpW8Pr16jxCPv6gYKjKTSist7JG1paVRtKcbjWazoJrMLeG9Wfu/NXmXJX9lpidPSIetr8pXdSO/wDQXQr0vh+/1Sf0u9nOlSl7TqVN5z8s/awMWvfXeoY0/Tbd0KD2VGkt5eMn1NnZ6DZ6bQV5rVeGVypZ2z3frPwRdutX0zRaLtNHowq1eUqnNe9/nfYcve3dxeV3Wuasqs31fJeC7gNtrPEle4i7ewj9Gt0sZW0mvuXkc7JvnnmVMokBesbmhbVXUrWquGvqpywk+/GNzIpaqo3LuKle57UtmuzGTx3J7YXkayZbljrJ/ADcrUdO3UKMYpt7VW8b9MRi9jEr6nC3zGzk5zxh1WsRiu6MfveWa+hb1rqsqNClKpN9EjobPQraxo/S9TnGTj+b+av3gaay0+4vFK4qSVKgsynWqPCx1ficTxErOetXFa1t61WjlRp1+zH28RSbTbW2cnpF1qVDUtQsbD1OdOleUI3UpeynT9ZFPP6MeWW+mT1O1teHq06sKd3oN9XU6lKFj6iMfVdluLk8LtPv36blTb5WVapF4VV01jG9JyfxjlGPWoUbmjVo3F7NQqQcc1mopPo90nz7jc6lRt7biC9t4V4xp069SEJUo5hs+WH0Oc1a/uaOowtaHq4ubft1MKPzIrmbi3r2ledtV/F1abw1nGfEtzlWlvUm5JdW85/edZNWuo03TvFSqXFPrTlvjwa5eX/V4lfRaEKULqaqRovPYnOTknjnjkVG+9COj1briq11WpGULHTKn02vVxtFQTcfi8e7J6NxTxlqGuXE4U5yoWecRpp4cl3y/d9pkcDU7e1/B3uLulCMK95dSpuWN2lVSx/Vi/mclCGH3EVmUpt8y/BrqzCg8dS9CeAMyOMdTkvSd/2daf8AFf2HTRqrvOX9Jnblo1CrCLfYrrtY6JplHBtvD3PWOH7ijS4asqlWrCFONvDtSk8JYWOZ5A6uVjr4mTcVrq4t6VCdzP1FP6lNt9mOXnOPeVHYcU8aUFTdtpLlUbeKlZeysd0Xzz4/A5ynqUa1tVfq5U5dhqOeTb22MKlQS+qsvva+4yLW2qXFeNG3pVLitJ9lRgs5fcBb9VntyTi3OnGKjHd7dnd/Bm54V4R1TX7j1dnQbgnidWW0Ieb7/Dn4HecGei+pNU77iKap017X0WLxt+tLp5HRa3xdp+k260zhyhRk6a7KqRilSp/spc38vMC7wxw/w5wFCF9fVIV75rEJNZk33Qj95l6xxPe6wvVUc0LaTwqUecvN9fI4XTrXUNd1R1q9ftTk/wAZXryxGPm+nkdlpVvqmlTqqyt6d1PHZjcU6frMeMe7zwBiXFvWtanq7mjUpTxnszi4vHvKOZcuI3XrZVLmnXVRvMnUi8t+8t47yKY6EpZIfMv2VtcXlxG3taM6tSXKMUBY8jf8OcL3urONaadva/7SS3l+yuvnyN3pvDmm6LbrUderU5zW6pveKfdj85/I1mv8VXmpS+h6fGpb28vZSj9ep4PHLyQRt73V9H4ZoSsdIowr3XKcs5Sf60uvkvkc3Rt9X4kvJ3Nar+Lhn1leq+zSpL7F5Iv0NHtNMpRutfqOMms07Om/xk/2v0UYer6zc6hCNCKhbWdP8nb0liMfPvZRmz1LT9Fi6Gix9dc4xUvqkd/HsR6ef2mhr1J1q06tSpOpOTy5Te7ZFOM5zjCnFznJ4UUsts7Lh/gqdaCuNVnOimsxowx2v6T6eQHF435jc9D1Lh/hy3qKlcXNrb1ZLKVSTi8d+0kvkYn+SOmXKzaXtOpnk4XCx8MP7QOG6b4JxsdjW4Hrxf4uvN+UYtfHtL7DV6lwvq1ql6u3qXGefq4NtEGiXLOckxWeexfr2V5bv8faV6T/AF6bRZSaCq48ipeBHs9lYzkqWwFaKlyKEVICSU9inqSBIReo2t1WpOrRtq1SnF4coQbSfmi1JOLakmmuafMgDqR1AE9COe45gASRzJyAG5GxquJOIdN4ftfX31Zdt/k6Ud5zfgvvA2detSoUZ1q1SNOnBZlKTwku9s8w419I8qnbseHpOMeU7trd/sLp5v8AxOW4u4s1PiKs41pOhZp5hbwe3nJ9Wc6BNSc6lSVSpOU5yeZSk8tvvbKOpJHUCSGSQAAAAAAAQAJBAAkgAATEgrp05zy4QlLHPCyBBJU6dRc6cl7ilprmmgBJAAkkpJAkZIAFSKkUIqQFSZUihEoCtGy4e0bUNd1CNnp9FzlznN7Qpx/Sk+i+3pk2HBfCV/xJcdqCdCxhLFW4ktvKPfL7Op9GejjgKl9ChQs6LstMpvNSu1mdaXXf86XjyXyA0Poy9H8bV/RtMpeuuXFfSb2osKK/urwW78T27QdHstEtHQs4udWS/G3El7U39y8DJsbW2sLSNnZUo0aEei3cn3t9WXsLlzAjksp7h5z+9NDyeRy6YAhvbOMmm45t5XfAvEFrCLc62l3NNLvbpSX3m5578/mPVxrRnRnvGcXFp9U9gPznBVOLjOUXzTwUgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB9KegbSqXCXogr8c0KUZazrVxOzt7h7u3oxbTx3NyhLf9nuNbrlScbGrU7fTMm+p3fC1k6X4K3CzxjEpVn/Sr1l/eOD1yPb0m5jz/ABbfw3A8p1VVY3s51G5OTzll2yq8tzKvaSuaGfz4r4mldwrVvt58gOkhNOGW0kubNTqeuUqOadtirU7+iNRcXd3ers5dOj3LqW4UIwWy97AtXFS4uqjqXFRyfd3FPZS5G3p6HqtWjGtT027nTksxlGjJpr4GLcWdxby7NehVpS7pwcX8wMLAwX/Vh0wLOCcFzsEdgCnASK+yMARglInBKQEJFSRDajzNjYaTXuIqrXf0eh3y+swMGlCdWoqdKDqTfRG4tdJo28VW1Kom+apRf2l13VrY03RsaaT6zfNmurVp1ZOU5Nt94Gdd6lKUPVW8VSpLZJGulJt5byQABKCRUkFQkVJblSiXIxApjEuQgVwgXoUwLcIeBfp0/Au06XgZ1np93X3oWtaqu+EGwjEp0jIp0jdWvDWq1MN2ypp9ZzS+XM2ttwjX/wBfdUor9SLl9uAOYhS5bF+FLwOxocMWMF+Mq1aj80kUXFfhPTHKNzd6fTnHnGpWTl/Vbz8gOXjTOj4Z4C1PiS3rXNraQdKjHtOLaVScd8uEecltjYv2d3wpfVYRt7nT605v2acaqjKXhjKZ6Pp/Gt3a6HHTVbuMqKjG2q0qnYlThH6sXs+0l3MDkNP4Pno2nK/ttKmqM32ZXEYqbznlJrLj5PBl3Vhf0KkI3Nnc0pTXsKdKS7S8O86+HG9HKvFp8qV5UXZvKKw7e6jybaf1ZeOH7zIhxRoDULJq7emVJdqNOcfxtlP9KnJN5j4cwOAalFuLTTXR9BjJ6RHWdG1C/jSranaq/ox/m2ozpONOvH/Z1oySx5/DxqtqGg39zWubOw0+vcR2vdOjOMu0v06El9i+TwB5rgdD0X/J3h/6HKripU0ytJqnewlL1lnP9GrDOMct8efRmvu+D7Sk6di6t1T1Oq82zWJ293HvjJR9l473tt37BxXXkQ15HY1+DoTt3Xs7uvJ0Jdm9t50fx9v49lP2l4r9+Ma54PuqdW3lT1CynZXS/E3jlJU3L9GWz7L8wOY2XMjBvpcJa6tQrWCtFK5pR7fY9ZFOcf0o5a7S8jAek6n9GqXP0C4dKlJxqzUG1Brmpd3vAwceBBfr21zQhCda2q0ozWYOcGlJd6zzLPvApJ6DYkCMAnuKoxlOWIxcn3JZAttExi3sk35GwtLHVqc1VtrG77S5ONGT+42FO14rqvswp6jDwy6a+4DTU7K8njsWlxP9mm2Xf5K1FvexuI+MoOK+ZtamgcTVE3WjUx1dS5j/AOIs/wCT1dPFfUtLoy7ql0s/IDA/ku8S9pW8P2rmmvtkR/J7X5S9sof/ABlL+zk2EtEtqf5XXdOS6+rk5/YiFp2hL8pxA34QtJga92VBfW1S18oxqP8Auj1Gmx+vf154/Qts/bJGe6HDdN+1e6hW/YpRj9rKlW4ZpfVstQrvuqVYxT/qgYEbjTKDzRsJ15rk7mpmP9WKX2syY0tc15wpwpSdCL9lRj2KUF9n2svx1uzt97LQrOlJcpVG6rXxwWb3iHV7pOMrt0ofo0l2Ps3A3FDTtF4dpqvqNaNzeJZUEs4fhH72aXWdd1DV6joUoyp0ZcqNPLcvN839hq1CpWqqMYzqVJvCS3cmbm10R0su5qzhcRSfq6UkvU9VKpPlDy5gc/OMotxknFrmmuRS0+47WrpN1cUI3EL+N45rMU6MKkpJdznhs1/8nL6dTtLihaurNZ7Co1FNr+hsBzDRNS2rwoRryo1FSltGbi+y/eb69WladcuMLWNzXg8dlyzST9zecdxiQpaprt4lCM60ksd0Ka+xIDRzRuNG4aur5RrV+1b273Ta9qXkvvOmtND0vQqUbvVK1OpW6OS9lPuiurNdd6rqeu15Wmj0J0qPKdR7PHi+nktwiL2+0nQKLtrOnGpX6xi98/rS+77DSXUK93OF3rt07ai3mnRivaa8I9F4sy69Ohw3eU1UsJ3dbnKtUXZh/Q7/ADZuFPReIqXZaj65LaL9mpHy7/mgMStpWn6hpH0awrqnQkt+xup+E1zZyPEd7xfw9pdaytb5U7arHsSrwoRdTs9zqpKTX7Tyb6+4f1XSqrudMrTqwX6H1seK6/xsYEeK6Mm7bVqXq3ydSMcr3roUeZ6Vpta+fqqetUKdw3+TuKGIyfhJNmbqHAXEdxQlG5pWNzFPtxdCs41E8dO1HG+23gjtL7hDTNUj9L0yrToylunTfapy93T3fAxKNfXuG6kaV3B1rbOIqXtRf7Munl8gPM7bRdQ0q4qQq6fXU1s1JrPwLlew17UafZt9JvatOD5UqMpJfBHtdleaHxFSVCvCCrdKdTaSf6r6+74GHe8Kajp9X6Xo1ec+zv2U+zNL7H/GwFrRqmraT6JNJ0Spp9xRqxualzXk20oJyl2YtJ565efAtaZK3vn6q4vLe3m3yqWsMP8AptN/E2mj8V1aNT6LrVvKMovsyqKOJL9qP7vgba74c0fXKDutPqwpTlv26X1c/rR6P4MDU1+GbhRbp2dhNYymqlTtPy3SOeu6VvaVnRu7K7o1F31FFe72Xn4m6nDiPheXtxdazT5/Wp/vj8vebmx1fQtfpK0v6cKNWXKnW5N/qy/6MDiqFG2rp+pq3EmuajR7ePN5X2GDrNlSrWVa2r14qE44alF9pdzwk+R2+qcCXNvU+k6LcSco7qnKWJLyl+8x7HXatjW+h6/p8qnZ2cnDFRe57P5AfPOpabf0ridN21SUYP68YtxfjkrsI16kMK3qSafZ7SW2fFntfH1xw/qE6VtodhUq3s2kqlGLgt/zezj2n7jJ4X9HkvYveIZpU4JyVt2+S/WfTyXxKPOOEeCNW4hrrsw7FunidR7U4+/85+CPV7DSuGOA7CNWq4zu3HaXZTqT8Ir81fw2WeJOOLTT6P8AJ3DtKlJwXZ9co/i4fsrr58vM89ua9ze3M7i5rTrVZvMpzluwN1xLxVqOtydFP6NZ9KEH9b9p9fsM3SOHKdC3pXup0qtxGeHTtrZOWfGpNZUF5ZZzNKG5nW6aacW4vvTwB3dgrWFWjZ2tnRqTbku24yVKEufZhGpPfzeM9xtKVWq6cad1ShXqt9mFCncLLl+j2Kccp+bx4nD2l9fwSUL25S7vWPHwNtbatqEITjCv2PWR7MpRpxUmu7tJZA3uq3NfTaSpUtS7FeTzK2pU4/ivBzT5/E1P8pag5ZV5Xi31jNpliytri8uIULajKrVm9ox3Z2+l8N6dotstR1+tTlJbxpveKfd+s/D7QNJw/wANX+rzVeq5ULaTzKrNZc/JdfM6C81bSOGbeVlpNGFe7xiUuaz3yl18l8jUa/xXdai/omnRlb279lKP15+G3LyRYt9HttPoxvOIKkqaku1TtIP8bU8/0V/GwFilR1fiS8nXq1HKMfr1qj7NOkvsXkjInqOn6LB0dFiri7xid9Ujy2/Mj08/tMLV9auL+EbenGFrZw+pb0torz72a2MXKSjGLcm8JJZbAmtVq1qsqtapKpUk8ylKWWzN0TRb/V6/YtqWKaft1ZbRj7/uOi4b4MqVexc6spUqfNUE8Sl+13eXPyOp1PU9L4fsown2KaSxToU17UvJfewLWg6Bp2iUHV2nWS9uvUwsLrj9FGj4m40UO1baQ1KXJ3DWy/ZXXzZz+ua9qOu11QjGUaUniFvSy8vpnrJlOlWKsrl1tZ0jUalGO6iqTim/HPQDXVKd9cp3c6VxW7by6ri5ZfmUW1tcXNeNvb0Z1KstlGK3PU9J1eN9p06um6XcqNNYpxmo04Sfcnk5PXLTjC6u3dVbatT7O0I29ReyvDsvIGx0nhDUKdknU1m4ta737FGTcY+e6yzG1v8Al7QoxjR1u7u5PfDt3JJeMpZRRoOq6ppTnU1i21qtFbRUlLsLz7XP4nZaLqUdUtfpNO1uaFN/VdaKj2vLDewHEaNxNxRdXCtqFKleVX0nSxjzawkbire661nUeFLe6/Yw/wDxHV1JRpU5Vey3hZfZjlv3LdnB8V65rd1GdGjYXllacnKVKSlNeL6LwQFD1rhqtKUL3hx0JJ4fqsJp+7slS/yIudl9Ms373/4jntEpabUvF/Kl3O3oR3ahBty8NuR3tPWOEaGnuNH6POlSW1P1Lcn/AFlu/EDS/wAicM11/M+IPVN/7ZL7+yS+DnVX8x1izuH3cvsyaPWb6lqN5m00+ha084hTpQSlLzxzZ0XDXB1Sq43WrdqnDnGgniT/AGn08ufkBhVuDdbpLMIUK37FT9+DM0jg2pKnOvrNZ2tKKfsxnHK8W90l/Gx1mq6pp2h2cVVcY4WKdGH1n5Lu8TzvX+IL7V6jjUl6q3T9mjF7e/vZBm6hrlawkrPRdTuJ20I9ntTpwx/R2z7zn5TlOo6lWUpyk8ybe7fmUIkKDIJAAAgZGdt3hGJqmpWel2c7u+uIUKEebk+fgl1fgeVcYcb3+tudrYess7B7N5xUqrxfReBR0/GfH9tpzqWOkdi7vFtKpnNOk/vfh/0PKr+6ur+7nd3tedevN5lOb+S7l4BQUVhLBHYCMeSLckZUoFmpBoKtFPUqZT1IJIJ6EAQSQAAAAALdlagBQCvslShnoBaJSbMinQlJ8jZ2GndqSckBr7HTrm8qKFOOM9WekcJ0Kuh6c7e3rPtVJdupLHN4x8DA023hRisRSNrTlsBtlqt5ydXPuK1qdzJe1KL80auLLiYGc7pz+vRoS86SZS42k37en2MvO3i/uMaMi5GQFz6Ppre+kaY/O0h+4fQtJlz0XS//ANJD9xEZFakBblpeiT+tomne6gl9hanoPD0+ei2v9HK+xmWmVJga58NcNPno1L3Vai/vEf5L8Mv/ANTx91er/wCI2eSuClKSjFNt7JLmwNRLhThl8tLcfK4qfvM6w9HnD9fs16tnXp085UfXy9v9yOp0XRZzrU1Ok69xNpU6MVnf734HrfCnCdHTOxe6ooV73nGm94UX498vkgNJwTwPSp2tGvqFvG1sqcV6izguy5Lx/RXzf2+grCjGEIRhTgsRhFYUV3JEyk5PtN5b6kfDHgBHnEnPTCIYzzxl94DG/PKH8Me9D3gQ290lz6k0vyiIa8hT+uue4H578S2rseI9TsmsO3u6tJ/0Ztfca86b0rU1R9J/FNNLCWr3WP8A5sjmQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+3rCzdH8FfhpYxjTLep/Wq5/vHld3Dt29SH6UWvij36jp8V+DXw/QlH6nDNCb840IT+1Hgk+TA8odScW0uaZz2qXlGV24xpZUX7T72dHqqVC+uKf6NSS+Zp7OjYxuKte9g55eYJvb3gYMrvKxRpPPfL9xsdA1CdhdfSK2m0Lya3j67OI+SW2fFpkV9Qs1L8TbxivBFiV/nlBID0C09IMEl9I0ice9wrJ/JpG1t+PtCqLFald0u/tU018meTSupMp9dJgexfy3wRfvNeVjOT/ANtbYfxcSf5H4G1F/ilp0m+lG47PyTR4522Mtgev1/R7w9XXaoSuaS/UqqS+aZr7n0ZW0v8AR9Uqw8KlJS+xo81oXFxQl2qFarSl3wm0/kbG14j1+2adLV7zylVcl8HkDqLn0Z6jFfze/tKn7SlH7ma244A4ipfUtqVbH6FaP34Ldvx3xPRknK9hWXdUox+5Jmyt/SXrEJL19jZVY9VBSg/jl/YBz9xwtr9DPb0m6wv0abl9hgw0vUZ3kLOFnWVebwlODWPF5PQKHpRpOSVxo1SMerhXUn8HFGyoeknh+o0qlK+o55uVJNL4NgcD9BtNHniv2bi7XPqosw72+rXD9qWF0SPUHxHwLfTzWrWM5S5utatfNxC0vgO/l+LemTlLpSuuz8lIDyRvLB6zW9H3D9Z9ulO6pRfL1dVNfNM19x6NLdyf0fVasI91SipP4poDzdIqSO/p+jWr236zVoKPRxoNt/M2tp6PdGopSuK91Xa55koxfuSz8wry2MS9Ro1KklGnTlN90VlnqcrXgbSMuq9MhKPNVaqqS+DbZj1+PuE7CHYtJzqr9G2t+yvnhAcVZ8N63c49Xplwk+s49hfPBurPgTV6mHWnb0F1zPtP5L7y5eelSgpSjZaRUmukqtZR+ST+00d56SeIq8XGirS17nTpZf8AzNr5DY7C04Boxw7nUJy8KcFH5vJsFw1w5YQ7d044X51ev2f3I8lveKOIbx/j9Xu/KE+wvhHCNVUqTqTc6k5Tk+bk8tk2Paqmv8F6ZvC5sVJf7Cn238YpmuvfSbotHKtbS8uGuWUoR+1v5HkeSGxs09BvvSlqEm1Z6ZbUl31Zym/lg0F/x7xRcpx/lH1EX0o04xx78Z+ZzUmUsDIv9U1K+f8APdQurnHL1tWUvtZhMllLCtvpV7U07ULe+p7zoVFNLvwz3rh/VrPW9Np31jU7UJbSjneEusX4nz3L8k/Iy+G9e1HQL9XVhVxnCqU5bwqLua+8Jp9EsM0nCHFGncSWnbt5KncwX423k/aj4rvXidnQ1mVHs+r0vTE44xJ0HnPfnJUadJcwlHPJZNrqGsV7un6v6LY0Y4eXTt12n73lr3YNV6tbvHzA6DhnTJX1teX1xeXELe2g+3ToyfrJrGceRZu9fruzjYWdOVG0hJSpwnWlUcWuTWXhPySMDTry7sKyrWdedGfLMXzXc11Mm41Chd5+maZbyk+dWg3Rm337ez/ygX58U65K9oXqvn6+hHsxqOnHtOP6Mml7S8zf6bquodqpc1Lrh2NpeRzXtJ1HGE3+k479mXfjHI5VS0fspfQ75Y2z9Li//wBmSp6Rv/NL5+dzBf3AOzo6xp8bR2upajbVKdF9qyqUKspVrZ9ynjdeD+ZlW2sxuLC41u3qQ+nWsezXq06bVK7ilyqRaW+Oqzj5HCRvNPprNDSYSkutxWlP5R7KKL/Vb69pKhUqKFvF5jRpRUIL3L7wOw/yl4ejaKmozdhXea+m1INuhJ850ZLZb74yvDBk3lzoV47e2v8AUbSvHZ2OoyjFzg+kK8XzXi/fg84wMdQPRZaPpF7qMXT0+wjqdGOZ2cKmLe7h+lSaeE9v3lVahw7p8JajWsreWmuap1ac6EVcWdTuccZmvn5nnHX7BNdufbm3KeMZe7+IG5q8S3kLiorGjaUKEZNU5K2hGU1nZtb426ZZRU4m12osfTnFfq04r7EajkNwM2es6vN76nd+6q195j1bu8rP8bdV6n7VRv7S0kTjHMCl+JS2s4ysvvLqgnByc4rwfMolGMliUU/NAMEYK2QwKceAwVYAFGPINEgC5a169tV9bb1p0amMdqDwy6tRulGUJOjNSeZdujB5fi2smNu2oqMpNvCSTbbOis+GqtHT3qeo0K1SmlmNvQ3nL9p/mr4vyA01O+UFJOytn2ubXai/dh7FctQrO1+i2VvG2pyWKnqsudTwcuePDkdLw5regJxoXWmULOefZqdntr3t7r+ORsOJrLTaDV1V1i7s4t5lRp1nLtr9VZ2+wDQ6BwfcXXZuNS7VvRe6p/ny8+5fM2epa7pujUVp+j0IVqy27MPqxfi/zmaa+1u81PsaZYdu3tPqqPbcqk14t7vyXzN3oVDh7TqMvxtx9JacZV6lCcZRfhtiIRr7Hh6+1m6d5rlxJYxigpe0l3Nfmr5mPrC4g0S4VW3o0o2FPaMKMM00v1uufH5lF/p07S7d9o2t0riTeX+PSq+/L9r+Njc6PxDeqKpavptwly+kUqLa96+9fADE07X9J1el9F1GlCjOWzjU3py8n095h6zwQs+v0ir2Jc1TlLb3S/f8Td6pwrpOr0nc6fOFCpL86lvBvxj092DnlU4i4VqKFWLq2mdk/apvyfOPy8ijEttc1PSa30TWbadSK6y2ml355S/jcyr/AErh/iui50px9elvKHs1Y+a6/M31prGg8RW6tL2nTp1Zf6qt3/qy/wCjOd4h4FvLat9M0K5m5ReVTcuzNfsy6/L3gcvX4X4i4aryutOqzr0OcnTWdv1of9Tc6NxLYXsPo2r0YUJSXZlJrtU5eeeRc0XjO+0+4+ha/a1JOLw59js1I+a5P5e86G60Thzii2d3Z1Kcaz/1tHZp/rx/fh+IGi1TgO1vIfStIrRpSku1GOe1Tl5NcvsMC11LX+HK8bbVbedajyj6znj9WfX5mVOx4m4QqurbTdezzluK7VNr9aPOPn8zodJ4t0XWKCs9Wo07ec1hxq+1Sk/Pp7/iBbt/8neKqCpyUfXpbRl7NWPk+q+KNJf8J63old3miXFStBdIbVEu5x5SX8YNprnAtOT+l6Fceqn9aNKc3j+jLmvfnzNdYcV63oNwrLXbWpWgts1NqiXepcpL+MgXtF40pN/Rdbt/VS+q6sY7f0o817vgZeqcH6PrFH6XpNanQnLdSpYlTl5rp7jPrU+FuLraVVTgq8Y5lJPsVqfn3r4o4F1brRtblb6Bqc7pt9mMqMW1N93Z3UvmUbCOocRcI3MLS8ca9u/qQnLtRaX6L5r+Ni7cfyzxvcwVG3hbWNKW05LKj3+1zk/BbGx0XhC4ubmWqcT13UbXbdN1P7T6Jdy/wI4n43tbGi9P0CFOUors+uUcU4fsrr9nmBXUocOcE2nr60vWXk47SaUqtTwS/NXyPO+KuK9S12cqTl9Gs87UIPZ/tPr9hblb6jrV5OvUqSqSnL8ZXrT7MV/Se3uNlqfClDsUY6Nd/Taij+NlltOXclGLSXi5MDkowKoxNxX4d1e3/KW0V4KtBt+5PJZ/kbVk0v5Mvcv/AHEv3AYlKJnUI7FqrbXFrVVK5oVaE8Z7NSDi8d+GZunW1a7rwt7ajOrVm8RjFZbAu0V7zquF+GL7V3Gq07e0zvVkvrfsrr58jbaDwnY6PbLUuIq1PMMNUm/Yi+5/pPwW3mY/EHGFe9zaaWpWtr9XtLac1/dXgv8AADd3WqaLwtbSstLpRr3bWJvOd/15fcvkc3Tp6xxNfyqzm6ij9apN9mnSX2IuWWiULO3hf8QVZW9GW9O3j+Vq+7ov42LGra3XvKKtLalGzsIfUoUuT8ZPqwM532m6HmlpPZur7GJXlSOYw/YX3/aaG5q1bitKtXqyq1JPMpybbfxLfIy9LdgruMtR+kOgt3GilmXhltYAu6NpF7q1x6q0pZS+tUe0YebOzo2uh8IUI17qf0i+a22Tl/RX5q8TBv8Ai+1tdPhaaBZugsfWqQS7HksvL8X8zjq9etXryr1as6lWTy5yeW35gepPiO1VCFSva39vGpHMW6OWl34Wce9HNXencMX9xOtLX7mNaTzJ1+fzijmf5RvHU9ZOsqs/0qsI1H8ZJlVxq1/XjJVK3alKXalPspSl0w3zx4cgOn07h521b6Ro3Etq6mMJqMXt8WZ9e143VKcKeoWlZSWFKKSkvJ9lHGW+qQhS9XXsKFbKalL6sp9ybXTyw33ldhqdGlUfrbeVGnhv+bVJxk30WXLC88MDLqaFxTbVJVY0bntyeXOlWTbffs8lDveLbJ9qdTU4Y61IykvmsGRb67OlCMaGr38ZdnMpXFXMU+5LsSyvHK9xEeMdaoVGlc07mK61KKX2YA2/D/F1OnSf8s3tedV7dn6PFRj71uzfUeJ9Drw/F6lSi2tvWRccfHByMON69Xa/0uzuI9yTj9uStcQ8OXOVecOU6a76PZz8lEI2N+72HrdQo8ZWzgt+ykuyvBRTf2GlpcZ65Sl7VejXint26SWfhgyVU4GutnRvLPx9p/fIlaNwlcbWuvzpS/32EvmohUVeNbuvRlCvp1jUbWE3BtL3NvJptM0++1m8lC1opybzOSiowhny2S8DorHg6wq3Eca5RuaXOUaSXafv7TOuqVdK4f01Z9XbW8eUUt5Pw6tgYvDnDVlpEVWlivdYy6sl9X9ldPMwOJeMKNr27XS3GtX5Sq84Q8u9/LzOb4k4qu9VcqFv2ra05dlP2pr9Z/d9poE9wL9zcVrqvOvcVJVak3mUpPLZbTXeRklATuSQAqQQUV6tK3oyrV6sadOCzKUnhJeLILnI5ni/jHT9Bi6EMXV+17NCD+r4yfRHMcYekCpcuVjw+3CmvZndtYb/AGF95wij7UpylKc5PMpSeW34sukZms6nqGtXf0rU67qSX1Ka2hTXckYeCoYKIwFEqJSIKewiHTT5oulmpW9r1dJdufyXmBauKVKEMyeH08zHvqEaFaEYt4cO08+/9xfq0uzSlOo+3U7308inVv8ASIf8L72RWDkZIBAAAAAAVQ+sXYxLdJZmZVOGQKYwyZFKhlrYrpUzMoUwJtbdbbG4tKaikYtCKRnUgM2i8IyYSMOnIvwkBlwkXYsxYSL0ZAZEWVpliMitSAvplaZYTK0wL6ZUmWUzO0vT699P2fYpJ+1Ua2Xl3sCm1oVbmsqVGLlJ/BLvZ2XCnDdxd3UbeypeuuGvbqvaNNefRfN/I23BXCFbUI/ik7axi/xtxJZc33LvfyX2+p6dZWmmWcbOwoqnSW7b+tN98n1YGFw5oNloVH8RitdzWKldrfyj3I2mc75fxGP4wFz7wBC57JZJWUsP7R06fMCGE0uvLvCzjvCzlbAMdOgzv94a3TXPzHPL7PzAN46v4kRx6z/EPL38AvrdQPhz06W7tfS9xNTax2r6VT+ulL+8cUelfhN0lS9NeuYW01bz+NCmeagAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACUstLvA/Te20yFX0VaNpk1iH8lUbeXgnRUWfKtSLi3FrDTwz7L+j9jhulbrb1VvBf1Uv3HyZxvauy4v1W37HYUbqbiu6LfaXyaA8W4vp+q1y7j3z7XxWfvOenFzqyp88nrPEHD1nq79bNyo3GMetis5XiupytTgjUqN5GrSr29anFp83GT93L5gcNVt505tOOClRa6HoupcKXVbenQTk+6S/eaz/ACK1ZvalSj51EByCgu5orjTT/OXv2OwhwLqr51bSPnN/uMinwHef6y+t4/sxb/cBxKpSxnDwVKmd3DgF5zLU0v2aH/7xk0+BbRL8ZqFeXlBL7cgeeKBKgejrgzRo/Xurn/5kV9xcjw1w1S/KTU/27jH2YA817BHYPSnpXB9LeSt9u+vJ/ePW8F0Nv83PHfFS+0DzRwKXDwPSnrHBtJ5jSsm11hbx/cFxlw9RXZoKfgoUuyFeZqhUm8Qpzk+6MclX8i6tV3paXe1F+rQk/uPRavHunRz2bWvL4IxKvpBpJ/itPm/2pIHu4u10TienNfRtP1Kk+9QlD9xvrG09JNJJ0bi/ilyVS7i/lKRm1PSFc4xCwpJ9MzZh1ePtWk/xdK3j/RyDTZ59JlW3dOrdOPjGdKL+KNRf8NcaXj7V1Uq3L7ql2pfay3U4212af4ylDyply34q1uWe1ee7soDBnwXxLFNvT08d1aD+8wavD+u02+3pN7t+jScvsOqocT6umk7iL7+1BGdb8U6gpP1nqpruccE0PN6lOpSm4VacoSXNSWGiEz1ijxKq8XTubKM4vZrOz+Ji3tHg6vT9be2FK1S5yjHsL/lGleZZGTO1+rorv+xokLn1KbzOrLZ/srGcebMAipIYIZRDKGVMoZBSyllTKWVGb/qX+yYxkr8hn9X7jGIq/YXdzY3dO7s686Nam8xnF4aPZOA+PbXWYwsNUlTtb/lF8oVn4dz8PgeKEroVNPqLHINdx4rwp6QNZ0qELa7xqFpFYSm8VIrwl+89N4e4w0HW3GnQvIULiWPxFw/Vzb7lnaXuKljetEYL1WjUpy7NSnKD7pRwU4CLeNicFWBh8gKUo79rK22wUuJtJaLqka1OlO0nCdWOafbaj289zbw34cyzW02/pXLtqlnXjWSz2Ow28d4GCluMbl2rSqU6jhUhKElzi1uinHiBQ4jCK8EPYCjAwVYGAKfIeBU+ZS8Jc8AQHgo9fQ3XrqWf2kVJqSzGSfkBLLtra3N1KUba3qVpRWWoQcsL3E2VrXvbiNvbQ7dSXJZ+86C206lplXsQi7i/is+sk5U6VLv7PJzlz7kBzEsp7rDIydxBvULtq0VC8tovE6lTt1Zx2znE/ZwWKlnaVakIQs6FddntVZZp040V+u4JJf1sgcabDQ9Mr6vffRbdxjJRcpSlySX/AFLmty0yFZ0tPpwmlzqxU0vJKUnleOxhWlxcWlzCvaVJ060X7MovcDubrh280nT1LQ4U6101+NrT/K/0E9l9o4U4g1W4unYX9jUquntUrKPYdP8AbzhfY/Ms6FxvF9mjq9Psvl66C296/d8DXcY6jrF1HMuzHTZv2JUJdqnPzkub8HjyCMriTWtFt751tMsqFxfL/wDEOPsRfel+c/E5C8ua95cSuLmrKrVlzlJlojG/MKNdSZ1Kkp9tzk5fpOW5D8yMgXFdXMVhXFXHd23gRva0Hlerz/w45+OMllkNAZ1PW9Rp1IOF1XhGL3jGtPD+LZnR4svuw6dWpVnBppxl2Jp9yeY9xoMIomBn1bvS61SUqllSgn3Rks/CWF8DZ6VxXS0/s0fWyq0e12VGrXk1Bd6bhlLwyzlKz2MC43yEd7q/E3DOp0adLVNPqTc17M6coSlD3qWUaCWnq0uVe6DqleEFvGVShUi14OUYuLXwOW6lSZR6joPFVwlGhrH0Oae3r6NzTz/Sg3n4fAyta4T0fWKbuLOULatNZU6SzGXnH92Dyynf3tOn6und14w/RVRpfArt7+pTks06E1zxKjFt+/GQOndTing6piSdeyT65nSf3xfw95uZcYcOappU46rbSU4rLoSh2+0/1ZL7djj1rWp14K1s1WhKfsqNGtVbl4dntNfI6Lhjgacou71mC7WW4WqlhN/rNcl4L/Ao5/SNButdv6lTTbedrY9tr1lSTagu7P5z8DuqdpofBmm/SqkZVK0vZ9a45nN9y6JGhvtf4m4f1HF5YUYWW0adGnDFJRX6Ml18/gdNc8R6YtBje6lQnQhWi0ratBOdTyXVeLwB53xRxTqGtSdJy+j2mdqMHz/afX7DX6JDR3UnV1a4nGMFmFFRlio/1pJNpeSyYt9OlVu61WhS9TSnUlKEM57EW9l7jFnyA7Gw1e1leUO1XsKlGOyp5dKlbx/UhJx7Tf6T3L2oXNpcSuKdCV3eycvxTtqcayin4dp8uW55/OOWUqO4Ha1tQp6Iqn02p625ksU7KjJRdPxqTpqO/wCqtzkry8uby4lVrVJtvp2m0l3bsspFSiBNNbnsPossrKnw7C7oxjK4rSkqs+qw9o+Cxh+88ipx3N9w1rl/odw6lnNOE/ylKe8J/wCPiBncVX99e63cxvXKPqqsoQpN7QSfRff1MbTq7tbyhcqCm6VSM1F8nhp4OwjecNcWKMbyL07UMYU20s+Ha5P34fcanWuFNT0tOpGH0m2W/rafReK5r7ANlrtCHE1wtR0qsqldU0qlnOWKkcdY55ryOYqU6lKpKlVhKnOLxKMlhp9zRRByjJThJxlF5TT3TN5R1xXNONvrdsr+mliNXPZrwXhLr5MaGmQZu6mhRuaUq+h3KvqaWZUX7NaC8Y9fNGlnTlCThOLjJPDTWMMohME4Q5jQggqIYFLQKgwKQSPcBT1JJwMADM0rTb3U7j1FnRdSXOT5Riu9vobDStB7dstQ1Wt9BsFyk17dTwivvOgrujR0xfSVLSNIe8LaD/nF1+11w/46EHE3VCVtcToznTlKDw3CSkvc0R25Twpzk0uXXBXezt6l1OdrQdGi37MHPtNLzLSAqRUnyKUVICpFSKV4FSAkBSa/Rx1yjh+MuPbewc7HRlC5u17MqvOnTf3v5AdFxJxDpugWvrb2qvWSX4ujDec/JfeeScT8SalxFWf0mToWieYW0Ht5y72aq6r3F5dTu7yvO4uJvMqk3l+S7kUF0KopJYWyJKSUEVIlEIkKkSajFyk0kurLdWrGnjOXJ8ormyhUpVGp18Y6QXJAO1O4/J5hT6y6vyL1OnGnHswjhFSHIgs3X5F+77Sxq/8ApEP+EvtZeuvyT819pZ1Z5uI/8NfayDAABFAAAAAF22WZ+4z6UTCtPyj8jY0UBfpR8DKpLBZpoyKYGTTMmmzEgy/BgZdORfhIw4SL8JAZcJF2EjFjIuwkBlRkXYyMWMi7GQGQpFcWWqMZ1KkadOLnOTwopZbZ2PDXDdSdxSVSi7m6m8U6MV2kn497+SA1+i6LO4Ua90nCi94x5Sn+5HrPB3BfrKVK81Wk6NqseqtksSmu99y+b+3dcLcJUNLcLzUlC4vucYc4Un3+Mv48Tp5Nyk3KTb6gURUIU40qcI06cF2YQisRiu5E5WMoEff3AS9tufgHy5B57x0xn5ARnffCHgS34/FkLwAeC5Dw3Hvz7xjnt8gGenVDbrsRhcmvEnAEd3LccpxzzbI8ufLYPoB8f/hW0XT9MN3PH5a0t5r+p2f7p5Qe1/hi0FT9JWn1kvyuk0234qrVX7jxQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAZWkUHc6tZ2yWXVrwh8ZJGKb30eUfpHH/Dtu1lVdVtYfGrFAfqZOKVFxaTj2cYZw/Efo64P169nfX+kU/pc0lOvSnKnOWFhdpxazhJczua20H5GH12A+cPThwFpfAnBlxxFpl/eVp0qtOCt7jsyjLtSSwpJJrGc75PnqfpBvO00tPo7frs+o/wzazpeiyhTTf43UqUX7oTl9x8TyT7XLqB19Xj7U5fk7WhH4sx5cca5J+y6Ef8A4ZzGXjoOuUFdHPjLXZL/AEinHxUDGq8Ua7L/ANYVV5YRpm9+hOcvuINnLX9Ymva1Cvv4mPV1HUKm8r64f/xGYe3cTtnKCq3dXUtpV6zfjUZbm5t5lOT9+SMvfbbqPfzAh5Te+fcFuscyWueX7sjOM4eNgDyuuc9CE+y3y8iVh7vDI9lbt7/aFVNPntjzKZZIwsvyJjnPcQQuSZTnJXJpsjb+EUTBPHIy7bny8zGp7vdmTb47W7CNjRXXCexfpJZw+8x6KT26mTGL7a2KMugmltjn0MPituOkyTW3aSRnUUkk0sbd5rOMJY02C75Ajkaf115mQY0X7S8zJMVhkhsqjCU3iMXJ+CyXIWlzOagqM+03hJrAGOylnYR4D1OnpE9Uvrq1tKEI5w25SfuSx8zjqrSk1F5Wdn3lEMpYbZBBnx3t1j9H7jGwW6X5SPmi/UWJgUJEpEkhdLlH6yS+w2VrCEotVIKUW+5fea6hz7sdTa2fNZzt3dwHVcM61r2kxjS0jXNQtKMdo0JTVWh5+qqKUPkdlY8ea326UNR0Ph/U4cqlSnGrZ1pPvThKVNP/AOHjwOB099nCbz2X0NxRitpPKxstsIsYu5tuOOHqkZvUOHuI9NcXhfRKtC/jJd+7oteWGdNw7xz6NaNmrieuXuk3scrOp6PWamn+bKNONSKXTKnnn5HlMI4xn3fuKm+1T7aWJdzKmnvGlcU8F6nYtadxhwuoS+vZXWoQp0Kn7Cq9ipTf9HBubXTdU1G3qq1tqV3bQSlRpevVxOm+rhVg9l3e1n7D5c0ihQq6bCFWjTn7c89qCf58jYwpxpvNOPYljZw9l/FA0+i6dtqdWhN3+nXdWvQyqdScPU3MEumd+2vDDXmzX0Y0NRhUrXWnZrUIt/SKFBN7fm1abSi38H3YPGtO4j4j0/bT+IdYtUulK8qJfDJuLT0k8cWlT1q1+pXnybubelW+co5+YR6MrPSL9tzt6dnXpvCnTblQrecYy7Ufc2l3lueh6Xd49Q6llWU3D1dWb9TWa6QqtbPwaz4HIWHpW4jtq061TTdCupVPymbSVLt+L7Ekn8DY2HpZtqfbp3PCNCVCr+UoW99KNJvq1GcZJe4Dd3HC0Z1FRtK1ancredpcRUajXVweezP3NGJX4buZU5Tsq0blw/K0XBwrUvOD3+GSiHpH4Wr2s7evp+t2tCUfYpzjTulRl+lB9qEkYdfX+H7pO9o8YXsq9OD7H0uyrU58toqUXJLPmBruJ5LRKFKcq1vcTrp9iNKeWsc+0nhx96OKu726updutUlJc+yvqr3FWq6hcanfVby7rVKlSe2ZS7TxyS37jEfLHVY3yBLf6KKoOUcuMnF9Hk+nfwVLOzu/Rpeq7s7ev/nWqmqlJS/1dLvPL/wmdL03SfSRCjplhbWVGpYUqsqdCmoRc3Kabwts7IDzeF7d0/q3NZbbLtvGTNtOIdZt3+J1CtFLpF42NVzby0+9L7A02va2fgB0dHjLXKcu1OtSqvo5x7f25L9xxle3NCNvcU4qjFt9im4wjnvxFJNnL4T64zyDzhY2zyTA6OOvWjft060fHCx9pvdL4t0rT7Z/RLeMb57fSarcux+zHGF55ZwDntnO3iU7IDs5ahb15upK7pznJ5bc1lmdpup3Nk3K1rexL68HiUJruaezPPccnuypScfqtpruYHqHZ0nU/qOOl3b6SbdCb8+cPmjX6hY3dhUULqjKn2lmMucZLvTWz9xwsby6pr2LmsscsTNnZ8U61aW0rRXMa1CWzpVoKcV4pNbPxQG9yiGc7DXLtJduFKWXts195dhrzx7dt8Jf4AbtlLexrIa5bSz26dWOPBP7y7HVbKa/KuKf6UWBmNvBaqMuWrpXabpXdmsP/WXMIP4SaL89PljLvLFLwuYy/stgauszCrbm5q2VqvrataLwUKr/ALhi1bbTYv2tSnL/AIdu3/aaA07W5SzZyho0c/jr+p5UYR/vMpp/yXKpGnRsNQrzk8RiriKbfkoMqNdk2uhaFear2qsXC3s6f5W5qvs04Lz6vw+wu39GnpdSn9K0OFOckpRhWuXJ48Yxaa9+C/f8W3t7Ywsqthp0baDThTjRaUccsblHd8LaRoi0mp/Il+p3M49mV3FJ1Ivya9leHz6nPz0fi3h/UnWsK87ulUk5SmpZjLxnF8ntz+ZpdDu9aq1Z1tKt7e39Us1K1O3glTXjJrYz9SseJL+DVbUql1T5tP1kIfOKiBtuIuN6dKyjaW1O2u7xxXrakV2qEJfq53l/HM8/vLq4u68q1zVlVqSeXKX8cvAzZ6LqKl2YUqdaX6NGtCo/hFtmNd6ff2se1c2dzRj31KbiviwMN8i1NZ6l6XItyQGPNEJFyS3CiUUqJKiXFELnjD+AFVOO5k0lsUUpTinGMpJPmky/SQF6mnjkdHw/xRqek9mnGp6+2X+qqvKS8HzX2eBz0O4vQSyEd8qXDPFG9GX8m6hL83CSk/LlL3YZoNc4d1PSG516PrKC5Vqe8ff3e800fM6jhvijVraUbeUXf2+VFwqPeKe31ny9+wHOUalSjUjUpVJU5xeYyi8NeTNne6zUvrN0r+1oXFdbQuWuzUXm19b3nV6nw9pOqVMUqdbR9QlHteorU3BS8ey9mvGLOO1XSL7TZuNempQzhVab7UH71y8nuBgbBsl7c9iAAAQD7AwSgKcbAqwQBGDdcKKMr7sw0l6jc4zSUqmKcP1pLG/vZpzItby8tqdSnbXNalCosTjCTSl5gdTquqW+m3Lr3NWGqawuX+wtvCK6tfxg5W/vLm/uZXF3WlVqS5yl08F3Is8iAIwVJBIkCUSgiV8wqUY+p6hZ6bZzur6vChRit5SfPwS6s0vF/FunaBS9VlXF817NCEt/OT6I8k1zWNQ1u8+k6jWc8fk6a2hTXckXW023/F/G97rPbtNP7dnYPZvlUqrx7l4HKRSisJYRSTkujaonJRklMCtE5KMiU4xWZPCAuJluVWU5OFFZfWT5IoxOr9bMId3Vl6OIxUYpJLoRSjSjTzLeU3zk+Zc3KM4GSC4CjPxJAtXKzDOX0+1FnVf9I/8Ahr7S9cfk/evtLWpJyry7KbxTWcLkSjAABioAAABvuEuEtd4nnXlpdo/otrB1Lu8qvsW9tBJtyqTey2TeObxsmBqrFe1JmypGLSpRpScIT7cU3iWMZ8TLpgZFMvQLMC7FgX4MuxZjxZdiwMmEi9CRixZdhIDKjIvQkYkZF6MgMqMjO0yyub+4VG2h2n1k+UV3tl3h3QrnU5KrLNK1T3qNfW8Irr9h7DwLwU7qhTkoOz01PeePbqvwzzfjyXyA0fBHCFa4r+osKXra2Px1zNYjBfcvDm/s9h0DRLHQrfsWq9ZXmsVLiaxKXgu5eBmWdva2VpG0sqMaNCHKK6vvb6su7pAF45z4Do2uYz4h8s8/mBDRC36L3k9cIjploB+dnOO4lZ/6kZx3jO37wJ93vyRv1+ZLzvsQ/iA25DbGMtsLbuD65AFK7lv7iceX7xgBl8lnzIl0xnGSeRS+X7gPmT8M6jjivQLjG07CcP6tRv8AvHgh9G/hp0duFbnvV1Bv/wCU19585AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADrPQ1R9f6XeD6WM9rXLP8A76ByZ6D+Dhb/AEr068H0sZxqdOp/UzL7gP0or/k3uYhl3D/FsxAPA/w2aqj6OdOp53eqQ28qVQ+Mvay+nifXX4ct12OHNAs8/lbqrUx+zBL++fIy5sLEe8Z35ZyHyIzjdMKqyQ3jBGXlp7kvDz0IC59PMPl3dBLPRZ2D2xnl4ARus7bfaEm5eQj4LbqSot8kAWyXwD7knnoN8523IT78PfvIqV7MuRS84fNhrD2ZL67v3lEdd87c1kLfDWyIlnGE+fiSs43yQRlslPr9pSnjG2xKzkorhju3Mui9sfYYsO/HmZdvs144REZlNNpLGxlU21Ld5XcY1JJ4WF8TLorfqvcUZtvjx3NPxtP+bUaa5dps3VBZS6M5/jaW9CHgwRzdPsqpHtLMcrK8D3C34W0CjDFLTaO/Wa7b+Msnhp9B6VVdfS7Ss/8AWUYS+MUxFrBq6DauHZpwjFdElgxKPDtNanbrsxeZZOjKbffVqK7othHH+mbV1RsrfRreWFjtVMHkjOx9K1f13E1VZ+qsHHhdIAwSRU0/ykfNGRW+u/Mxo/WXmZFf8owRCJ6FKKlyCrlFpSS95trPDfZaymaui/aSNpa92d/AFjf6dJZbUV2e7uNxSw+vLOF9xpdNz2t9+qwbm3W+NluugYVkxzhSk1jw8iU1hpNrPcTBRl1zzSIy2nD2WljdZ3KMLQdtMhzb7dTkuftyM/P6OxgaJh6fjZ/jKm6fP8ZIzoPaO232FBpJ4XkVPGU+TfNIt9pJt46dStcv3bBFWVnOW0uW4ljOezhvfmQ5Lm+Qll4aWMeIB9ruw+/PUZy298kN889MDLysY58giW8p7YfRkpZw28YKUnhc2t+gz7X72B9T/gj/AP3cX/8A/Nqn/dUiz6cPRNr/ABtxVS1rSL7TacIWkKDp3M5xlmMpPKcYtP6y+Bd/BG29HWorKeNWqf8AdUjO9LPpgfAXFVHRZaAtQhVtIXHrVdera7UpxxjsPP1O/qB4xqPoM9Ito06Wm2t6uroXdP7JOJz2o+jjjyw9q44V1VqPWnburj+pk90s/wAI3hKcE7vRtaoT/OUIU5xXvc19h0Vj6b/Rvcwi565VtpvnCtZ1cx83GLXzA+RLmjXt6zoXNCrRqr60aicZLzTLbb7z7qq0OFON9FcpQ0zXNPqZiprs1Yp9cSW8ZLww0fJ/pt4GjwNxd9Etas56dd0/X2kpvMoxziUG+uH8mgOEfe2389hLdrmeiehn0X3fH9W4ua15Kw0q1koVKvY7U6k3v2ILktt23yytnk9gl+DrwY44/lTXk8f7al//AIwPlrdNeXeHs/Doz6Ov/wAGyxnUk7Liu5ow/NjWs41H72pR+w5rV/wd+KbWNatY61pV3SpxckpudKcsLOEuy1l+MgMf8HPgThvjShrr1+0q15WroKjKFaVPsdr1na+q1n6q55MD8ID0eaLwHc6RLR7m9qU9QVbtwuZxl2HDsYw0lt7fXPIsegu89I1nW1WpwHp1vfx7NL6ZTruCivrdh+1OLz9bk/PoR6ctc471e40u2420CjpVS1jVdu6NOSjV7XY7XtOUlLHZXJ7Z35gebdrpnzwUvfeW3uJ2ymllLuIT8G31QBN8seO5PtZXL4EJvPJtfYT2tmtn7wC3fgRsn7L38ic5bXwIb9rLxuyiv11eK2q1F4dpkq5uMYdRvbuTLOXnmwmu7dAX/pNXC+q/Fo2GkcQX+kzqVLGNKFWcez6x04ylHybTaNR132WMjfv5FGVWvZ1q0qtedSpUm8ynKXabfi3zMvSKunu8hK+VapbJ5nGi0pPw35Gp59dhy9pNprk8hHdVdV027r0aTuqttp1GXap2lOh6uKfRtrt9p+LT9xtdR1HTNQvbWori1qQ5XEpwpxqy7mpyhH7vM89tK/rMxe013dV3mQmB3Oq6hZ066vKt/cK1jFQo2Fvcpurj86bh7MU/ezktY1CeoXTrOhb28elOjTUIpe7m/FmIMAUSRaki80UNFFnAS8C44hIIhIqSyVKOSpLuAQRkU0W4rlsXoLfAF2K3L0UvMtQWWXYrkyivDw8c/E6vSOKrbS7SlTstIjTlKCheU5VnOjdLq5RksqXin7sHKpEpDQ798X6Io0rSdvqFTT204Qkl66yn0lSqJ5aXc+73GdDiPRbq/jGpq1FXfZ7NLUXbuMasf9ncQkkn5+/Y805M6Hge10u61WVPU3CT7P4mnOWIzl3P93Umldbb22h6jdV5Wtjpda67P85slOLp1V+nQqreD8MryWMmLDhfh76NXqU6N1Us28TqwnL6VYS7pweVKK78Z8+Zy2s6hrdC6lb3Cqaf2XmNGjH1cYrwxzXjubPQeMJ20uxqVD1ykuy69Laol49/yGhlVOCLSNuqD1KdO4rb2d3Jxla3HcspJwl4NvwyYdXgm9dCpC2uFVv6G9eyqUvV1FH9KG7U15GPqWs29GhWt9DvNRpW9d5q29xGM6bz1WW8FiHFOs/QqVtUuIVpW7zb15x/G0vCMk08eDyNUTd8K6lRtad5TnbXNpOXZlWpVHilLun2knH3oxrrh7Wra7p2lXTqyq1Y9ql2Wpqov1XFtPyR1nDXEGpavqsXC2pUblQ/nNxTp/i60e6rDOHnllNMorcXadaXs7OGnzq6cqz7dv241IQaf16Mn9Xv7PLxHuOIq2l1SrVKNW2r06lL8pCVNpx810LSi3HtJNxfXoei/wCWOkXmpRlXubq2msu21CNDs1aH6lSKypx/jC5l211LQr26ryp1NJp3klmtTrRStL2PVrtLNOfn8wPNFzJ3PR7PR+HLqlXnY2P0m3i2ri2pVMXVo/0oOLxUh8enPdGNV4R0Z2abvKlOjV/0bUqcu1Rk8/VrReXB9M8vLGAOB6ko7O54HlJ/Rbe6nR1GC7Tta6TVeK5ypTWFLyeH34MC64Qu1RV1ZXVG4toPs3LnGUKts+vrKaTaS71n4Ac2SjeXXCmr0LijT7FCpTuPyFxCtH1VV9yk8b+Dxk5ji7UbfhSVWlraqWtxT/1Lj7c+7srqvFbeJUZU5wp05VKk1CEVmUpPCSPO+MfSA36yx0CX6s7trl+wvv8A+pzfFnFuo8QVJU5Slb2WcxoRf1vGT6+RzxlMTaZSnOpKpUnKdSbzKcnlyfe2MkEF0irIyUjI0Ksk5KM+JC7U/qvEf0v3EFTnh9mK7Un0RMYpNSm+1L5LyIj2YrEV5vqyMjSruc8wms9C3kdrYgu9odotdrchy7gL/bHbMdzKZVMIgvVmnFeaOp4Ts6cqFxfWNy7i8jBxrW8NqkYd8V+cn1a5fM4uc29sk2txXtbmFxbVp0a0H2oTg8NPzMaq/f1KFzd1JRt1aSztFcveuj8tvtMScJQl2ZLDOrhfaTxPFUdYdPTtVxiF9GOKVZ91VLk/1l7zS6xpl/o1z9D1O3aXOEk8xkv0oS6oitYC5OlhdqD7cG8Z5Y810N/ourWXDqjc6dbUr3V+cLu4pqVK1ffTpy2lNbe3NYTziOUpEHU8JcC6Lo9lb8S+k68qadp00qtppFL/AE7UF0fZynTpv9J4z0xsy36RfSTe8UWlLQtJ0+hw/wAMWz/m+l2m0JNPPbqNJduWd+7O/PLfDahfXup31W+1G7r3d1Wl2qlatUc5zfe292U0yoyKUTJposUuRkQAuRLkS2itEVdiy5FlmLK4sDIiy5FmPFmZptpdX91G2tKMq1WXKMftb6LxYEwbbSW7fQ7jhbhGpUlTuNTpy9pr1dsk+1J9O13eXM3vAfA84XVOFGj9N1GW/aS9iku9Z5ftP3Ht/C/DVnocY1puNzfY9qs17MPCC+/n5AaThLgmnQhSu9ZpJKKTpWaSSS6dvuX6vx7jt228JYjFLCSWEl3Ih5ec9fMhpvPeAwx1648RnKfMe4CGsr5h8u/vx1Jff2coh+WVz5gPL95HT72TtyXMPnus+QEb+Ie/LbyJfiQua7vMCW8rHMh5W6b8gumch47mA88jltz8xtzbD7uviAWVstvNkNZ5JMlZeyeSHz35+YB9/T5B8tskPbw8AmnvnqB4L+GdS7XDfDlfn2LutDP7UIv+6fMR9W/hiUe36ONMrpfk9WjH40qn7j5SAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHqf4JtP1v4QnCkccq1aXwt6j+48sPYfwNqDrfhC8PyxlUqd1N//AKeovvA/Qi4+oYvUybn6nvMdAfLf4dk1nhiGeUbt/wDdHyxnoz6a/Dprxlq/D9v1hb1p7frSiv7rPmTm+/yCxV/CI3S+0POc4Ie65ZIqMrHJFWduZSlzwSt10yAb655cxnLIfaWy2y/LA78EE9OmchYeO0G3s8fMb5CmU1vuI4XUfm8l47BLO2AIk3v19wW7z0Il3Lb3EvO+NvACHjmGsbBYCy5PkttwGMpEeHwJeW93lkL62GFXYbY5N+Jl226W6MWk8P7DKoPHNfMIzaG2zz8TLotNJ9VtyMKm8qOFty5mZRSfNeQTTY0OS57HNcaP+dUo/q5OlobR54OW4ylnUkl0iguLRYPd+EqvruF9MnnP81pr4LH3HhSPafR3Lt8G2D7ozXwnIq2OhKbRr+Vc/o0yS1Sl2bq5qfo0QxeLcb1vX8RXU859tmiM/XanrdUuJ9839pgkZ6QCcDAXQuaL9f67LCW5fr/XBIoRUiESF0v0cJptGztHuknjxNXSNraR2XZwIWN9p/1lLb3G6o5UWpY8djT6fiTTX2dTc0m+ynLLWPPqGtdhNJJvOStt9lPONtvBFPYTinl4T5J8irHVPKfnh7/IqNfoaSsWlLOKtTdf8SRsHjDeeS6mv0X2bRxzsq9Xl19uRmvd9nd7gVS26LbmSsN9+2NiM7rKWUu/BKbcuv7iiNu1jOMvOBjtRSxt4dRzbe3mSm+vd3cgiNuW/cT0ylhNbExaTW3Lw6EZT3Saz0XMA+0n2cc9u8LGfrIZjjlhvx5fx9xU3ywtu5II+o/wRHn0ealvn/O0/wDuqR53+Fus+kyy2/8AVNJ//VqnV/gf6tQVnrmgyqRjWVWF3Tg3vKLXZk15dmGfNHQ+n/0W6jxvdWOsaHXt431tSdCrSrycVUppuUey+SacpbPZ557bh8prm85eQuxlZW3mdxqHoh9I1om6nC9xNLfNGpTqfDsybOeveF+KLNr6bw7rFvjpUs6kfm0B7X+B1c1PX8R2Tk3T7NColnZPM0/jt8DI/DEt4u14bu8pTjO4p+aapv7vmVfgj6Fqdk9d1O9sbi2oVo0aNCVWDj6xpycsZ54zH4lX4YtSH8k8O0m/blXryS8FGGftQHS/gr2StfRXG4U+07y+rVn4Y7NPH/Jn3nPen30r8ScLcW09A4edtbxp0IVa1WpSVSUpSzhLOySWOmdzq/wY3n0RWC7riuv/AKjPE/wotvS1cro7Si3/AFWBn6f+ENxtbpRurTSLxLm50Jwk/wCrJL5G+t/wkbmdvOle8J03KUWu1RvXFbrbZwf2ngGX1+Qy8PvA+hfwOHvxR/8Alf8A9qVfhk/k+F/O6/8A2Rb/AANtqvFC8LX/APal78MdfiuGH07V19lIDnPwbvRtpXFavNd1+i7mxtKqoUbftNRqVMKUnLDzhJx25PO/Lf37/wA3nAjSX+SGi4S//wBOGfsOH/BMio+jK5a2ctUqt/8Ay6aOO/Ci4w4h07jKz0bStYvdPtaVnCtONtWlSc5ylJZk4tNrCjt5geq6l6IPRzfvtVeGaFKXfQq1KXyjJI57UfwfOBbnLtq2rWT7qVxGS/54v7T580n0mcfadPt2/FeqS6Yr13WXwqdpHRaV6dvSLZ1e1cajZ6gukLi0gl/9NRfzA6ziT8HHUaFKpW4f1+jdySzGhdUvVSfh202m/NI8R1rS7/RtTr6ZqlpVtL23n2a1KosOL5rzTTTTWzTTWzPq/wBBfpRr8fK9stSsKFrqFnGNRu3b9XUg3jKTbcWtur5nD/hh6RRUtB12nQxWn6y1rVV+dFYlBPyzP4lHz/RpVa9WFGhSnVqS2jCEW5SfckuZu58FcZp9p8J69jv/AJPq/wDhPoD8Evh+xt+D7riKVtF393dToxrSSbjSgltF9E5drPfhdx6hxRxhwxwxUo09e1q0sKlZZp06ksyku/Cy8ePID4YvLK8tKkqd1aXFvOOzjVpuLXnlGL47+Z912PG/BeowStuJ9FrdpfUd3BS/qt5K6vDHBerZuKnD+g33rN3VdpSqdr34Gx8Jb57g8tSecn1Z6XvRfwLZcB67rVjw/RtL+3tZVaVSjVqRUZLH5il2fkfKXc3vksRkabfXenXkLyyrOjXpvMZLk/BrqjrbfiHTbuCup6VSpXb3m4U4zpVH+tTlyfPdNM4mXvfcbDTte13S6KoaZf29Oh2u26NxYULmDb5tesg2s+DRR1sK3DV2/pMrVWdf8+3n25UJ+MZR9qHzXgbGvwvYVrmCoQuLWrNZ+hVqii5rvpVGmpeT378HPafx7eW81WuuGeF7+so47UrGVBYfP2acoxz7jbWXpH0uNvK31DhBztp87e21F+qT74xqwk4+6SQE3HCUKldxsLmtJwWa1rVpqNzTXeo5SmvFP4mFdcKXfq3XsbileUIvFVwhJTov9enjtL4M3Fvx1wfc9mnqFHiGhQi3KnGdKnc1KL5rsVozhNLzT5Fen6vwa9SdaPGFSjGUXGNW8tqtG4p575QhJT97Xv2COYveHdStraF2oUri1m8KvRqKUF5v8334MW80nUbOEKlzZ1oU6n1J4zGXk1szvbNUKmtSr6VxboVWnWa7VajewhNrunTqzWeu+JszaOicRWt4qdC1p17Oq36121L6RaVl+tGO0Hy3WFtzYHl1SjVoy7NWlOnLulFp/MiKPStQpXmnVqenX2m9m1q5/mt1mVFvGypTe8Hz67ZWEStPtZ9qyq6dTqzeX9DuVGFzHxpVltUXg2/FlHnMVyLsF5HcU9B0mpTq0o2lS5jBe24N07y3/apt9ma5bpL3nN6lR0q27VC0r1LuezjXjLswx3ODjnPvwBgRRdiUIriii5Hs9nk8+ZKKYlXvAnBLIW5KXiBn0NZ1KlRVB3Lq0VsqVaKqRXulnAq31CvTlGvpdm8rDlT7dJ/8klgwtmEBmULjTqdOMFpMZJdZ3NSTfvbyXP5RoUl2qGj2Oe+bnP5OWPka9cxUcace1UahHvk8IaG3p3XEOq2tSjb06srWnhTp21NQgs9Go4zyNZc2txbS7NxQq0Zd1SDi/mYlPi2y0lt09ftrfD7TgriLTfjHOH70Y2v+lXRL6Cp3FxbTccYdC3k37m+XuL202zycbHE1/SNpMG/UWV7V7vZUV9pgXHpKrPa20ZJdHUr/AHJF7abeirZ5i2n4GRaX19aU6tO1vLijTrflIQqNRn5rr7zyGt6QdfqN+qoWNGP7EpNfM1t1xXxJcxlGpq9anF9KMI02vJpZ+Y7Km3vL4h1ShpSs6t//ADSk+1B1VF+qa6xm94+5mq1n00Q0+rRrQvbW9v6KUVXtqblUnH9Gck+xJeaPBbutXu6iq3lxXuqiWFOtUc382YtZpLCWEWcZ3PoCjx/r3GVG6utA0qjaQrwVC4pqUcbYy1289nkvq45I5jUeGJXuoyq8QqVe7UV7Lm/ZTz1zl57/AAMr8H9f5muH/v39iMX0o6zW0b0iQrwTnRdrTjWp/pRzL5roY699RVmfA2g1V7NKvSffCq/vycPxjw9V0G9j2HKpaVfyVR/OL8T12xuKF5a07q2qKpRqR7UZIo1fTbbVdPqWN3DtU6i5rnF9GvFCWweDgz9f0q60XU6llcreO8JpbTj0kjX5M0SUyaW4WW8R3fXuRKxB5TzL9L9xBHZ61P6v7yXJv7ilkP3k0KsjPiUpjIE5GSlspyRVbZDkUNkNkEtlLYZBioQECAb/AETiN0LP+StXt1qOlv8A1U37dL9anLo/DkaAAdHqvD7pWr1jQLl6hpv58kvxlD9WpH7+TNA5Qnul2X1XT3GTo2q3+j3sbvT7iVGotmucZrukuTXgzoXaaRxWnU05UdL1l7ytG8ULh/7tv6r/AFWRXKxL9MquLK5s687a8ozo14PEoTWGilRcXhoqMiDMiEjEgy7CQGZFlaMaEy7GWSKvIrTLUd+R3XCHA9W57F5rEZ0qLw4W/KdT9r9FfPyINNwzw9fa3V/FL1dvF4nXkvZXgu9+B7f6POA1Uo+rsqf0e0TXr7uosym10Xe/BbL7en4N4FhGhRr6pR+jWsUvU2cF2W14/or5+R6BBRjTjSp04QpwSUIQjiMV4IDF0nTrPSrP6LYUfVw5zk951H3tmTz5bhtvovcG29lt7wC8kMZWyEf+oyA9/MdeQWOmGMb9EBA8Oo7+vkg/f5gHjG7IXgT5ELfo2A5Y6B495L26kb+KALHQeexKefMhc9/tAdd0M48ENlyyk+oe/wC4CHjuHkyds8kR1AZ26kPv3+BLzjZ5IfuA8l/Cypet9Ebnh/itRoz+Uo/3j5CPs38Jyi6voX1eWPyVa3n/APVivvPjIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAe7fgM0FW9OlOpj8hplxU/sx/vHhJ9CfgDw7Xpnv5Y+poVd/wD1qC+8D7juvqrzLBfuehYA+Nvw37hy9IunUM7R0uDXvq1P3Hz5Jnu34alVT9K1BJ709NpR/wCeo/vPCZYfJsLB+LT8h4lOW8fwipv2Vt1IqFs85ySt3uQ9sLxJ27Oz3Ac29ufQbdlxI371yIT67hU9Fyz0IT3ec88oqaW3PCKJLfOcbcyCvKe+z2wRJqPLl48ylN5wkvcQ28soNp8yU99iPfyGOm/gBVh5yR7yN/HzDa8WRTrh7DGJbJh8hnYC/BcvEy6HLtGHTbwZtBJqLj0KjKp89nhmZRUu3ltLvMSjhpYW76mZb5ck858CDYUdjkOK23qsl3JI6+l39O84ziOXa1atno8BljGuR656KqrqcKRg3+SrzgvlL7zyM9R9EFVS0G6pdY3Ll7nGP7guU9nbmBeVfV2upVc8qbXyM5Pc0PENb1Wi6lPvWAweNXcu1cTl3yZaK6m82/EgjdpSMFRVCnOfJbd7BpRgvVuZfoW8E8yeWWKmGwulKRKIROAaX6K5bfM2lnt2W+a5rJq7fmt8eJtbXPN8gmUb/Tk+qffvv/ibekl2M4fa6/I1GmpNRys5w/A3FHffG66dPErVWRDliTbz4cymTl2V2XnZJ4fL+PuJg+ynvLGM4XNkOWzi2m2115BHHWUdWuNT1GlZ6m7eFvVbUJR7Sfabf3Ge1xXQi36+xufCUWm/kinhjL17Wmk/ykeXTeR0aSxlLfvBWgWo8R0Y5r6JTq460qq+zcmHEtWnHtXei6hRxvmMMr54N9LtJ5Txlht7Ry882+pUaShxbo1THratai+TU6T+7Jm2+u6PWTdPUbePT232H/zYMypTpVI/jKcJ9/aj/gYFXRdJqJJ2Fs23yUey/kUZ9G6tq+XQr0aqzjMJqX2F3aUl7ONuq5Ghr8K6RU3jQlSbfOFR7fEtf5MK32s9Xv7fbO09s+7AR0Scmmotb+JLzt7W3NnOLStetn+I1/tf8aln7cjPFtF4asbnD67P7gOy4f1jUeH9WoappF3UtbuhLMKkPmmuqa2w+Z7jw5+Ehcwp06ev8NxrNbVK9nW7Lfiqck/7SPlx6tr1H/SdDdRd9Gp+7JL4njSbV1pV/QfX2M4+OAj7V0r8IDgC8l2bmWp6e++4te0n/wDLcjotK9LHo61OTjb8VWNNrn9JUqHzqKJ8G0eKtFm/arzpv9ek/uyZlLWtJqr2NRt/DtTUX88AffMuPeB4wc3xfoOIrLxqFJv4KR8yfhC8e6fxtxJaR0eU6mmadSlGlVlBx9bOTTlJJ7pbRSzvszy6lVpVaaVOrCcX1hLJce3Lb7wPrn8Fx59Etqu66rr/AJs/eeN/hSJf+deu28L6HQ+xnGcK8c8W8K0pW+h63dWdCU3N0vZnTcsJN9maa6IwuK+ItW4o1aWra5dfSr2UIwc1CMPZitliKS+QGr5NYSw+SbIxs3jfmUvllv3hbS67gfQn4G+1xxRu/q2vPzql/wDDIX824Ya6TufspHIfg3cc8O8GX+sx4guKtrTvadH1dWNGU4pwc8pqKb/OXTozZ/hOcY8NcVWfD74f1WlfO3nXlVUYSi4JqGMqSWM4fwA9A/BMefRhX6/5zq/2KZ5Z+Fin/wCdGljrptL+1M9L/BMvLReju6tXc0VXWp1JOk5pSw4U8PHPDw/gebfhY9n/AM59HDz/AJspZ/r1APIU8vPLHgS+105sSzjnld5E3nm8+/mUe2fggSf+XOrR3S/k17f/ABYHcfhe/wD3e6Y+7VYf91VOF/BAb/y71RY2/kx/97A7v8L7/wC7zTds/wCdodP91VA6L8GtL/zN6NhLeVxn/wCfM8E/CdqVKnpf1GE5ylGnQoRgnyivVxePi2/ee2/gs3juvRNQoPH80u61Hx3anv8A1zyv8KThjV6PH1TiGFlXq6feUKWK9Om5RhOEey4ya5PCT8c+DA8ZeFjn4bFy2ubi2l6yhcVaUl+dTqOL+RaSx0WfAcsbblRuZ8VcT1NOradV4i1WrZ1YdmpQnd1JU5R7nFvGDT7b7PwIi0+a6EZWG8Iolt4zl57+4dGyG8tLpnkE+m+AiUs4323Da545Dtc11IeMFBNrZkfHO3QNvx5kbrC9yKInCEvrxUvOJTSpxpT7dFypSXJ05OLXwwVP+GHjGVgDaW/EnFFrTdO04o1u3g19WN7Nx+Em0U23HnHlxCvQrcR07ujb1uxCF5pltXb9mLT7ThlPfozW923TBhadhVb/AP8AeX/YgXQ6LVfSJxbd0adK8teH731a9mc7WrTmlj9KFRbY6cjVf5Y3kfyuhWTf+5uqkf7SZr7nOZLnlr3o19d5T33ZZE230uOasX/9nZvHdfr/APxlK4/a56DVXldxf905uonhY58jGmt+WDLtht10vSCkttFre+4j+4tS9ItTGVocvfcr9xx00W5LCwZdsTbr5+kW8f1NFpR7u1cZ+xGPW9IWsSX4qysoftOUv3HKPKKGXsibdJPjriSbeKtpT/Zo5+1mLV4r4lqv2tXnFd0KUI/caVIqRe2JtmV9W1iv+W1fUZp9PpEkvgmYVWCqvt1ZSqPvnJy+0qGPEuja36uCW0I/AjljYuNFLLoUPb/qFt4kvxKcjSbTkd5S3uMjSpbMeqy8yzUWxND2j0Af9h1/+O/sRznp234zku62p/edL6AV/mGq/wDfs5n05v8A9Nanhb0/sNUnxs74af0f8SPSbv6DeT/mNaXN/wCql3+T6/E9Yi8rKw0+p8+M9H9GnE3rYw0S/qe3FYtpyfNfoP7hnj90ldHxhw/R1/THSfZhc08yoVH0fc/Bni91a17a5qW91TlRqU5OMoyW+V3fvPoXBx3pG4XWq2z1Kxp/z6jH2or/AFsV0/aXT4dxjjfsyryZvbEViPcUsl7PD5kM2aYmSGCDHSjIJIZBBGQQzFQgEMgMgkGKoAQAAAgEptNNPDXJkADq9O4jtdRt6encUQnWhFdmjfwWa1H9r9OPzL11wvqMa1GFjRlqdC5/0etaxc41Phya6pmLwHwbqHFV23CStNOov+cXlRezBdy/Sl4fHB7HbVtN0DSFonDdF0LVfla0t6teX6Un/CXQg8quOBOI7d4na0E/0fpdJteeJGur6Dq9u/xlnL3Si/sZ6jWqtmrvZZTCvO3Z3kPrW817jJ0zS9Sv7uFraWlSrVlySXLxb5JeLO50fQLvV6vaivU2yeJ1pLbyS6s9X4D4JlXh6nT6P0e0TSr3VRZcn5/nPwWy8AOO9H/ACtrmlij/ACjqst4qKzCl4rP9p/Lr7twvwpaaP2bm67F1frdSx7FN/q55vx+BtdH0ux0a0dvYUuzlfjKsvr1H4v7uRlkEuTlvLZvfL6kc+/3oEvfnuwKW+a394z8OXInp/gRnxALPPmHs8c98DHht4k+SAh8937hh/AbkcgG/XKIZK+A5bL5AOXIjksvkTnuSDfVbPzAjx+4h8skjbuAZ7xu8YSHwDe2cgOvPfzI3a6bk/FkAM74wOS32GPAbYxgAsvkvgiH3LDJxtvy8R0/jYDhPT/Q9d6GuJKeOVvCf9WrCX3HxEfdnpfpeu9FPFEMb/wAmVpf1Yt/cfCYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+kf/ACftBy9KOuXONqeiyh/WrUn/AHT5uPqj/wAnlQUuIOLrrG9O0tqf9ac3/dA+vrnnEtPZbci7c/WRZfID4W/C9uHcemXUKbeVQo0Ka36erUv7x47zWM+R6j+FPVc/TZr7fJTor4UKaPLe/wCYWGMJtNvw7iE8t7f4E5y3lcu/mHtsiKJc3kjCaSyub3JaTb3RT5ZfuAb+DXQpxs8Mrm+mPcU5xlvoBUnjm992Uz2eOT6E57Xi+4htufluFFzeXh5GXz7+pD+3uI5JYIKl0Tx4kPCXMjuyVrDTzzKKXJYSKds4KpbPo/AhYIp0yFuHjZ4CYF+ku1JZZm0em+6MKil2k8YfeZ1N9lZx4AZdOMc4b+DMu337m38jDp47MU+hm22M7J4fIIz6S3Tf1TiNafa1Ks/1mdzTjiHl0OD1J9q8qPOfaZK2cc3ti4PRPQ7L2NThnbNJr/nPPDsPRVeK316rbTn2Y3NFqK75ReV8u0GWU9nqie5yXGlbsaBdLP154OrzhN+B59xrf2tW0dqriPac25Y3wVhjPd5/Lmy4rer2YylHsxkspy2yi+50ae1CGX+nPd+7uKHKUnmTbMW+Y7RGFOPTtPx5FeWylFSRNspiri2jEfMykY7W42vapRUhgqSLtO1eoRzszZ2uE9pJbbGut0+Rs7Nb4XyYjDKN5puyW+c9eRuab7UU0nleHQ1Gnp7ZXLmjb0MbYkllY5mUaayE0l2ZRXTb+OpDjFyTzlePQZbpr3vGCey+xJ5XPfL6hg5rhzbiPWk1GUu3HdvHVm/i89FFd+TntAXZ4l1fZP2ot93NnRL2UvZb6IFVTeO087Lm/wCPMip9ZZ6bhNx3e2NmQ8vOG30fIIlbtNLG3PGSX2Xlp8sZwY11dwovsJLt4ba6RXe+4s297a3HZhQ1C0nUbwoKS3+ZRndmTmpYSWPiQpOWXLsxe+ds5LdKpJTdKpF06mOT3z4rvRdjuui3+RQSz1SWcZZU22s7c8troQk45UuTEX7Kb5LuWwRV2dliT7OeXLBSljfOMd6J7eG0t/Jhtdp9lbeX8fwgMedtbVm/X29Gr2dszgmYlbQtIuG5SsKHLnCLhv7mbKKlLzfQiPs5TTywjQ1+FNIqtOjGtS/YqP78lD4brQxG01vUKEV0csr5NHRRljfOdnnYn6zw+oHPS07iajHFvrdKquvraa+3DKZ1OLqEcfR7G5fLMW0/tR0MUnyS+3/EZypZTbXd3gc/LW9at6f854fqv9J06mfuZK4stoJq6sL+jLrmmsL4m/ym1yx/HInaTzl5eNkgNNR4p0Wo03dum291KnL9xmUNW0uqswvrZrudVJ/Bl2dlZV8yqWtCp+3STMKtw/otVvtWEIrLw4tx28kBtKc4VIRlCcZ5XOMs4K5zlOeZzlN4S3edvM52XCmlyz6mVzQn07FX96Kf8nr2k8W2u3tNY2UsyX2gdE23h8vNE/mpR683g5x2PE9GS9Vq1vWXT1kEn9j+0K54spZcrGzrx74yxn/mKPVfRFx3P0fcRV9VWmx1GNe2dvOn671b+tGWVLD/AEe46j0y+lyz4/4Ts9Ko6LXsK9G9VxUlKvGpDChKOE0k2/b7uh4Gte1SnPs3HD1zt9Z025fd95EOLdPjJRrW93by69qCwvnkD2T0L+kq89H+q1IVqMrvR7tr6TQjL2otcqkM7drHTZPwwmvpPQPS36PtZ7EKHEdtbVZLLp3idBxfc3JKLfk2fCVLibRqjUfpijn9KElj5YMynqenVJJU7+2lLuVWP2ZGh9+yseD+Ic1pWehat0c3TpV/nuafU/RP6O9Rn27jhSxg/wD2ftUF8Kbij4mpzcH6yEnF9GmbfTuK+J9NX8w4j1e2S6UrypGPwzgukfQXpU9C3BOl8G6trmk0b6xr2VtKtTpwuXOnJro1PLx5NHzNFLff/E6y+9JvHV7o9zpV5xJc3NldU3TrU6sYTcovmu045XuZyLfNd5YEk3hpbEZWMrK7wl0aYfNYXzKiH1w+nMbY/wABLmk1v3FS5tciildd3u90MZTwufciXjLbS8skbPZc2Awt+7pghb75TT5Ep55t/AnGN8ZKiMNvHgYGndnt3zT/APxLyu/2IGe0vDYwNPWJ33T+cv8AsQLBRc9d9lua+qsJ9febG46prZ75wYFaLxjquRlIMSfLPN96LE+WeiMiSWc4ayWJLPNbeZnIxWJru95anu+4vTWyZamZyIssp68y41sUPJlIx2jBMUMbkoujaSSUQ8jRtDSKGXGihl0m1tlLZUymRdG1D2DZEnuUtjtVWy1UO54L07R4WFO9v7N3tSsnhSfsQSbXLq9uZ0f8m8LVsdrQYLP6La+xmq5SMpGx9AS/9HqjxzrSOV9OO/G9Zd1Cn/ZO14futO0Wj6nTLaraU3LLS9pZ97Zh8RaVoPEl/K9v7irC4lFRclLsbJY5NYNU+bbL7PFWhCU4TjOEnGUXmMk8NPvR6be+jK1lFuz1StB80qtNST96wczrHAnEGnxc6dCN7TSy3bvtNf0Xv8MmzulY6rueAeJIa5YeouJJX9CKVRcu2v0l95058/aZfXWl6jTvLWbp1qUs7/NPwPTKHHUnj12nR8XGr92DXlhrwylaj0m8LOjUlrWnUvxU3/OacV9WT/PXg+v8Y8+wz2elxpp0sKpb3MPHEWvtLs9b4Yu1i49RLPSrbt/cSWxdPEmiD2memcE3nOhpmX0hNU38mixV4B4YuF2qNOtTT5OlXb+3I7oaeOkM9SufRlp8n/N9Tuaf7cIz+zBrbr0Y3sX/ADbVLeov95TcPsyTcHnzIwdhd+jziOl+Tp21x/w6yX9rBrLvhHiS2/KaRcS/4SVT+zkg0LRBmXenX9n/AKXY3Nv/AMWlKP2oxWiaVQCrAwSwUAriszivE6nUdOtHwPC/9TGNzG7hTVRbPsuM20+/dIxWOTBVODg8PnhMhJt4RBB2nAvBUtUUdU1mU7XSovKS2qXHhHuXiTwXwxTqShqGq0+1TW9Og/z/ABl4He1q86rXaaxFYjFbKK7kgMy4vYK1pWFlQhZ2FFYpUKawku9978TClMoyXrS0ubyfZoU3JLnJ7JebIMecm+Rt9K4d9Y419SUlHOY0FtKX7XcvDn5HT8HcF317VVS1t/XSi/auJrFKn5Pv+fgescNcI2GjYuay+mXi39ZOPswf6q+9/IiuY4R4JdWnTutXpu2tYpeqtYrsymvHH1Y/PyPQIxhSpRoUacKVKmuzCEFiK9xXU3beUylp42AL+GOa3C3b3zgJ7eAEZT7husdCW3nGSF4YAdCN8/Ilp5IfPZAHzfMNbck/dkcuaa9w8eb7gHhyfiR4bfAqWz8Cnwx5dcgT13x8SnGX3kvx280Q34oCXuuiDe+dviOXcQ3y3QB8thulkfIICMb4y2Su7PPkGvh5gCHv/wBQ3h9CcZWyIW/Ln9gDpjH+I6c2QsNd68yXuwHjy3IaJefeQueNgNF6QaPr/R9xJRxvPSrpf/SkfAp+hWu0lX0DUqGPylpVhjzg0fnqAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPrT/wAnbD8bxvPujYr/APuD5LPsD/yeFBx0rjK5xtOvaQz+zGq/7wH1LcY7fuLcs9llyu/xnuLU3iD8gPz6/CUrKt6ZOIpJ5xcqOfKEV9x5snt4na+m6u6/pV4kqPn/ACnXis78qjX3HFb55hYnLzy6kpZTb9xCfR/JBZRFTtnb4DLTWFjcjLxsMrAB478ENYTa+OQ34LHeRLk+QVPwRGNuvxIy89loZTjn7iCOvN8irDfPuI26MjK2WWBO2NhGT5ZSXkHnCePkRt0xjxAdOfvHXcjy2J5YAh7dGSn7XcF7yVhyW3mGS9Q3kjYUXt0MCkllfBYM+jy657wMmKTZnWyTez3fMwqazF4e/wBhn2ywsPcIzEpepbjz6nB3ufpVTPPO538I5g+vecXxBayttRllezNdpMlbeO621yRkWVepaXNK5oy7NWnJSi/FFhFdOnKcsJEZyN/rXF2qajHsdv1FNreENl/ic/KUpvMpOT72bClZw7H4zr3dDFuLadF5+tH9IMsZPstIqSIRXFGNrfjjtKRUkTFF2MDG10Y8Nq3gsOO5m+rKqVlOa7ck1HoTuZZcFkYGAkZNxRVOWFnHQtJGW2jLjXbVdzNnaJJrv7zX28XnbdmztuSz+8ylc+eLeaf0fly6m2t87NtYXyNTY7xTy31NxR9qGJYTx73gyjnyi5Bbdlp9nZY7ympFTju93l4e7w9i4lsvZz978y3N7pr2lgrBzGhNf5U6slmKkk+Xj/idIpYxlrnzZzujf/a7VU00+wsfFHQ1JJbJuWzbXJBKlJKKystNLCJSck4x7S8F1/jcYbilnG3LHyI5JrCTbxgI5/W9SsrK1uber2p3dzTltFZxnKWX0NHQ0W3r2FH+c+pmpZ7bXPPQ23Hek1HSoaxbw7VNRVOul+a09n5dDQ0qkK8qKl2ZxgpPsy9wHc31RQs4zT7Uqe8Gnvnu962MqjJVKaccNNc13HK0Lxz0WlTXalNqMV4s6i1pyo2tODbbhBR+W5UXl+j2+fRkSa3SXXk+pEdprHLD5h74w1jPPuKJabXLP7yG3F7Ze/LOxEfZXJ+LJezxlvPKONgifZy84wttiJYc1ns9rw2wSsY5rHVyKduyt0s9McwiZSylhvd4xgLk3u8rcST7eO10zyZLbz9ZJJARupbcmnuQ21LPzEm03Hnnfn1D9nMspNLoAy4wzJtvvyVN7qTfv7iG8S5Ye/N9CMyb8cdwDCjD2Xl55p8yc7J+12ubeRLdYcsPq8hJNJt4cuuOQEvp9Z58Q1utviUyztFZx35Gcc1zWfIoqkoptx7sEYbhhp4edyFJqXjjngiWezzWfMCVns4eW/BkSw4/VTzvh9xKw1tsRvJ45LxKMSvpthVblUsbecn1dOP7jEr8OaNNycrOMW3j2ZSjj4M3Ek1vv/HiRlPLy8vvKjn58Kaep9q3r3dCW/1Z8vlkfyHqcJfzbiC6XcqmZfedA17O2M9/2ER3jjHLv6hHPzt+KaTXYvrW4Xc6aj9iJV7xRST7emW1VLrTnh/ab5b78yXiOZNRzuyjn/5fu6b/AJzod3Dvcfa+4qp8V6a32atO5ov9aHL4M30WsZWO03yZRKEJrEoRmt17STKNdR1/SKu6vacX+snH7TMo3llW/I3dvP8AZqJlqtpWnVsqpZW+/X1aT+KMStw3o8s4tuy+nZqMvsNuuS5b9UMrn4GgfDFvTbdre3lDynyEdH1ai/xGvVW+iqRyvm2XSN+S85bxhZOejT4ppZfr7O4XdJYz8EiYajxFTbdxpFOqs/6uePvZe0b/AH5vmYNljt33/vMv7EDXriKpTb+l6PeUYrql2vtSMrQLmne291c0lLszuJP2lh/VijLSLly8uTXRmvrJdp7Y7jPrPd+Lw0zDqL2028bFkKxKq57/AAMeSzkyqjbl0Map1S2ybJGNqxPdFuSLs1zeC3IzkYrMkUNF2SKMGcjG1SiUGVdTLSbEsE4CW5OB2m1LRRJF1opZdG1iSKZF2SLcuRe1NrUkW5FySKGh2rK7fhKXb0Kgv0Jzj88/edFQ3Ry/BT7WjzjlZjXfziv3HT2zWEjmznu2y+zNprJh6un9HbjnZ7GbTxjbYxtTj/N5czXpWu0rXrqwqKMpOdPPJ8jttJ1S01CmnSmlN84t/Z3nmNde2xbXNe1qqdGbi08i47Nu/wCKOEdJ16Mqlel6i6xtcUliX9L9JefyOB13hzU9Fk5VqfrrfO1emsx9/c/M7Ph7iqnV7NC92lyUuv8Aj9vmdXTlSr0cxcalOax3pow3cVeHRlsHJLqeg8UcC0buMrjRqkbS45ulL8nLy6x+a8jy7WKGqaXdu11G2qW9VdJLZrvT5NeRlNVPDNlVS6lv6V2XmMmn3pmoVSrN9S5GnVfMvabbmnr2p0VilqN1FLklVePtMulxlr1PZXvbX69OL+eDnlSa5sYSMbIu3X0fSBq0ElUt7Sp49lp/JmdQ9Im2K+l573Ct9zRwLaRDkjG4xdvT6HpA0ieFWt7uk/2YyX2l6fEXB+oJK6lbT8Li2b+1HlDkQ2TtNvVZ6bwJqCxGnpWXyVKoqb+TRarcAcMXMc0I1qWetKv2vtyeXZJp1alOSlTqThLvi8Mx0u3oNf0Y2DadtqdzBp/6yEZ/ZgytR4NuqnCy0i3vKMqiuI1lOcXFNKMljbP6RwVDXNZoNOlql4sdPXSa+DZvtF1vjq+nFWH0i7jnm7ePY98sJL4ksqytff8Ao+4kjU7VOlb1kopexWS5LxwZvD/Bd9a1VcahZzc4v2YLEl78Ho3C2lcealVjG5trBJr2oUaU6lRefZfZXnk9L0j0aazd4d2qVpB7S9ZPMsd6UfvaMfA8Zo06jn6uNOXaW3ZUd/gbWy0PULmUYxouLk9k+b9y3PoHRPRhoVjFO5dS6lzcY4pQk/Jb/M6zTdK03TodizsqFBYxmEUm/N82TavCOH/Rhq104zrWlSMX+fcfi4/1frM9K4f9H+k6dGDvIq8qx+rBx7NKPlFc/edxth7rPkWKmct7d6ILMadOlTjTpQjThBdmKisJLwwUVMeC95elyLM2vADGnz5Z95HJ55e4qqc3iK3KXz6/MA+mVgB+Q8ngA/EhrKxv38x0xle4YwuYBrbpgj3k92f3kPnttlZYDO/iFz5onlzxgh4XL3gOW+R8H4CS2b69/eH7wIz3DOxD54ZOcdwEbZ+wPljYn5EeK2QDpz2HLngct8/Yg2s9F8wI79+Y2fNBd2ScdMZAgcw34jbfcBjvGf8AEhtNZysE81vv35Ah+8bY32Q67k/Z3gW5w9ZSq0/04Sj8UfnbUi4TlCSw4tpn6LU/r+8/PfiSmqPEWpUUsKF3Vjjym0BrwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD7X/APJ9UFH0c8Q3WN6mr+rz+zRg/wC8fFB9yfgBRx6HtWl1lr9b/uKAH0DW/KMs1dqUn4F2t+UZZuH+Im/AD82/SrUdX0i8Q1M47WpXMvPNWRy+Nt9zdcb1lX4s1Ssn9e7qyz4ObZps9MkWI/d3jLa5cupDxnLRHvYVU89+5C6iO+2M7cgljfwIG62zuQ988l3ZY35ppEZ38e8qksp/4kpLnleIxt0CWFs1nwZBLfTYpb2zhBvOMt5yOa5e8oN+z0DT5kNc0x0fPuIC7sL9w5cupC55yS+ewVK6ErCe/eUrufMqi/a3QVkUlvy8jOorCX8ZMOjlmdRaST3yBfo9HnfPzNnbZxy5GBReOvvNla8uQGXH6vLmYmtactSsHTgv5xB9qk+99Y+8zI4wVR8soLHA0rSp23GpFxcXhp9GZlOEKSxFb9502rabC7Tq08QuOr6T8/HxOZrxqUakqdWEoTi8OLW6Ma3Y+6pyEpJpp7os9olMlbscVqra/nUt/wBUtRjuZbrqGy3ZZbc5uUubNdrt4sNqqcMmba2zqSSS5lmhHLO59E9lTu+PeH7erBTp1NSt4zT5NOpHKNHJnqPoPTuknJl8XhuuJvRFxTwhw5aa/qtpbVLa4UVUjTk5TtpS+qqixhZ5ZWVnY4O4jUnNx7OMH6F8UaPb67w7faRdRzSu6Eqb8G1s14p4fuPiu80CdK7lCrS7FSMnGcWt4yTaa+KNPFy23VTLHj5+G5yas/5Hn9zYylSba3W5rKtBwfLZno+qabClTUlFPHM5PVLT1akmt4SwzsleFn5ae3hvjvNlbLdJmJCGHnCM22+suRnHLyYtzp6e2dza0032W8d7+BrNPTeMNtG2pNrEc5S5e4zjiyit80sN+HeRU3p7vbGU30KsQjFtteO+/uLc+TWc7bLw8vcZNdczpHs8aalGL29Xth+MToUpde1yOc0pNcaailHL9T98dzpJLLzFRXPr0CZJ6tYwlzDTbw+01yXNsjtezlSwum7Jm8JJN4ksPO2QxXKNaVJyjHszhJYlCSypLxRqLvQNFr1XWpWPqpZ+qpvsfBGzi84UkvIRk1J+xLGebCMCw0u3oVY1Ek3BYhGKxGH8d5nty7W2Ft0DaXZbylncn6q3llvv6FBe01NqSS6YEkmueNupSlLtdrtYzty8vuEX7TzyxiXQonPaWcY367YIk1HfbuznzCaUFhRSS5DtbOK7/eEG12M428N9yNuzvHd77krsrnyXzJfLCS26hB5z2se/Ownzw5POdsPOSYuOzcYt/Aoh2uecrOWkBLbWyWNu7qMRS7XXGHtkS3SWMS+1cxnD8d0gJWMN5WPF8yc9pNZWcYfgURwnlJBOUk08JeAFSy8NYbfV80OS3fgs/IZzsnu984ISXYXZxjGAJ5xbxjp34KXLyXfuRFNJpLblu+RMlvvu+hRKeVtLPeRl5znbGeXMPPaeXnnjwQfPl179yhl4Wcry3GXnKf8AH8ZDe227+QXZcsLZ4CKVzysxf3lXfhJEx5tLHPOzI9ldf+hQ2WX0fLBS20ueXyJ3wnjD6eCGe7OcgTu8KOGkHu2sSw9/4+RT0wm9u4mK5LGPeZISWVhbhvwSxzx0Gc8iXhb42Abru33Jjs8rBS0+m+BJPktmUS8vO2PDBEl7O/Ideb35eRMUsb8kVEYajhfMY2ymkS+RG2Wmns9jKCzdxX0Wt9X6kua8DS8D/wDYdTCy/pEvsiby8x9Gq+MHle40vAz/AMx1f+PJ/KJnJ7JWdcOO/wBn8fxsYVb62U0Z9ZLLwua5GDWXNdPEykSsSfLl0LFTrl4MmoluovkY8ujee42SMKsyRZlzL8t15FuSNkjG1ZkUteJcaKGvA2TFhaowVYDTyThmfax2lLcnASKlsXtNqWilouFLQ7TazJFuS2L0l4FuaL2ptYkihovNFDQ7VldRwNL+aXcM8qkHjzT/AHHVUOXQ5HgZ/jLyn3wjL4PH3nX0t/zjl5Mfibsb7M2lyT5FN8lK3l0RVR3XNk3KzQe3Q06Z7cldRxUawYss55GbeLFSXmYkkZyIoivcbjRtdvNPqJRqSlDqnualFaT2yhcdm3p2i6/ZahBJzVOp3N7ZM3VdNsdUtXbahbU69N8lJbxfenzTPKaUpUpKcJOMl3HS6JxRc2qjSuPxtNbb9P3GnLjs8Mpk1fFHAV5YxlcaNF3lFbuj/rIrw/S92/gcDUrzjNwlBwknhprDTPoPTdTtL+ClRqLtfoN7/wCJreKOE9J16Mqlel6m6x7NxTWJf0v0l5/IxmWvK6eFOpJlPakzr9T9H3EFpUkrejTvKedp05pNrxT/AMTXvhLiJSa/ke428P8AEu4nu0G43OlocFcSVcdnSqkfGc4x+1m80b0V6/f1lBulBv8ANpRlVmvclj5ktiyPPkiuFOc5KMIuUnsklls+h+GPwfalVRqX9Oc119fV9WvdGHtfFnqXDnog0HSopzcVy7Ubekqaku5vdv4mu5xlp8i6VwVxFqDTjp86EG953HsY9z3+CPQOF/Qbq+odidy7ipBrP4qn6uDXhOfP4H1dpfDmi6Z2ZWenUYzjynOPakv6TyzacljPuWxhc6uniPDHoF0yxUKl1TtISWHlr108+ctk/JHoulcCcOafFN2juZR27Vd9pf1Vt8jqXnOevemE3hJbPz3ZjtVqjb0LeCp0aNOjTWyjGCSXuLm/Veewx9vPlkPd425kB45LDf7JGU+WCc8sYz3Y+YeW95bAQ8OOc58eZZqPfHUvt57mWKmOnvwBbn4tfvLM3u8/YXnjGN/iyzVW75+8DHnnJT47FU89PsKOW6QDPTn95OcPqM+PzDW3PAEN777P5kPZLJU8vk9mR0a6d6ANZ2wyEvEbrdJbeAeeWceQDdMPx3HLr8yHzzkB02Gz7gnt/iG9sgHu8ch0z9we/d8BzSSANeBAwvBeIz1+ADks7ojfuZOXyyRhYxs+/IBDG49w26psA8+ZLyvvIHTbIDLRDXfyJ8M/AjvXIAue3UdO8Zb5NeSHPy8AEPro+AON0o8Z65FclqNwv/qSPv8Ai26iyz8/uM5dvjDWp/pahXf/ANSQGpAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPuz8Auk6foUu5v/Wa3Xkv/AJVFfcfCZ+gH4Ett6j0BafUx/pF7c1PhUcf7oHslX8ozF1CahZVp90G/kZNT678zW8TVfU8P39b9C3qS+EWB+ZesTdXUa1XO8pOT+JiRe/iX795u6j65MfknyIqeuSlvEmu4dege4Ui2u/HkHJZ2bwR0z0yTjboBD5YEd+SI5dct9xPKWF9hFS1h89sh4S3RG+fILlnu7gD58g37K33Gctb594ey3x4d4Eb/AGbjpjGCE1vyW/eTy3CixvvhjDyHvHK2xyXeS8LbLbCoxhZ6ldPk879xbeNu8rprIGVQ365M2kllMw6Pazu8YMyk3y5LvAzbeSylvy5GxteSxg19DOV0eevU2VsuWwGXTTx34Lqj37J+BTBbYxj7yqO73WPIKmPLHuMXVNPoX8OzP2KkV7FVc14PvRmRSTRXh9rZkZz2cHfW1ayrOlcQcZLl3Nd6ELO5nR9dKDp0ukpbZOu1KVKF1ZKpRhVj65S7M1nZc/iajiG7qXVduWIx/Nitkkc/JnZl2x9L6b0PFy9Nl1HJfF1J/O2hlFJ4RMFuJcyuCJaxxxnd7Mm2juj0X0MLHpG4cf8A/Erf/vInntst0eiehxY9IfDv/wDMrf8A7yJzcvivqPS8Pa/pf9PurHsnyx6TLenpnGmr28t83tSon4TxP+8fVC5Hyt+EK3D0gai47e3Tz/8AKgc+PtlHlen4/id+H5b/APX/AOuJ1erSnTcU1ujh9aqLtVY82dB2pzmk3lHMcQ+zqFePTP3HoY32eXzcWs7GBT33xkzLZLtrfHgYtun2TMt1iS558jbHn8sbfT135Xkt0bWklGL7Mt/2cGtslssvY2dN+zhrxz3GyOHOLsFnd7pPfHd4lmWFTTTb6PDXgXuzJJ4w1F756FmssqaWe9eBY01zGn+zxtfpYw6G7z4xOik/ZSXPz5nP6dH/ANN772v/AMPnPvidA1JxzlJ+O5Uo23FZXPfvxuQ928ZePaw1jJV2Xh4eF1wyl7ZaWX4e8MENrplLGcFbaztnGft/jBTHC337+RO3ZzJ8+YExazzlFLoubIkvZ2WHjD33XcUtdqOGsdPa5YDx0bz1bKiW12Uuy11z/HvJfLP1kn5kLP5zy31JSy8PHLvKKn1Tedv48ijKxH2ljHQZ9nZ75wTGTi1hY8P3BETb7Lyn3L4FU85xhRxsssow+bTeF1JeWmse/wDwALeUd+W/Ldkx+rlNPP8AGAm2/rbNYwRuspZ39y5hFSSS3eds+RS4/m5zlc+qJf1suONt2/tI/O7KeMPlyAPLT5rwJW0e12cvpkjZwae+ekuYTS7nvuBVhyeHt1y+fmQ04PO/PqRh9p8svuEmu1jkvmUVJxznMW+7D3KXybj9v7iGluuucpdxC6LDe62QFUnu1l/DkRLOFt12JaaaWfiQuaz5FDPdiS55xsSm87/JkJ7LxJbe+cJ+CLERldpr+MkvaSbewzs0s8t9hjCfLCAh4xvJtZ3Je2WufJ74Ektt+XfyIXjtgqJ70m8hNvOfPGeZDbjnfGObCWX3vmUSsbZ5dzDwsbcsZIeHu102YwnyZSp/NSb58/Apb3y0Sk8bPbPJDbtPBUObfP3h4w/ERSXQPKXn3GUgPOEtmvMdHhIl77BfW6cmWItXaf0Srss+rl9jNLwOl/IdR7/l5Z/qxN3dtK1rc8diTTx4Gm4HX+Y6ixl+vl9kTZJ7Iz7h5eVjK2Xv/wChr6z3w452M+7aak0m1hGHV5vON1jJnjGNYc1lJluolnngvS2S3y8lqpuubybZGNrHqLdookXJLqUtZN2OLXasteG5S1tzLr2KMG3HFruSjHwCRXhDZGyYMLkhLBKWxVgnBl2J3KWtyGivHeOyXsTvWZItziX2iiUfAvYdzGcSlxL7iUuJjcGUybfgp41KtHH1qL+Uos7OjltM4rhJ9nWoL9KnNf8AK39x21DuOTmx+J0cd9mbQ+rlcy7VWaTXIt0No9xeksxxuc1jZK5K/jitJeJhTRs9VjivLzNfJd5njEtW8PJVFPsk4ytyYxfQy7U2qS27ti5FLv5FMEXIJNYJ2m1+3q1KMlKlNxfM9Q9E2n6txne1tOoXFtGpRoOt2qzaylKKxsn+kjyyKz1PYfwWbl0fSB6n/b21Wn8lP+4aObHWO2zC+7sqnop4qT9irpUvOvNf3CrS/RhrGZPWryytsP2Y2s5Vcrxcoxx8Ge4JZRrb1b82tzh7q26cXpvAfD1liU6E7uafOvLOPcsI6S0taFtBQtrelQp52jCKil7i78diHjuTfeQS3ybab6Bc/az9o5Y5fAPPPL8yKjZJrKxnv5iW3PbpnPMluSw+RGO/O3uAJp47K+JDxjn5k45bPOAn0wmBD5L2l4EZeOuV3Inm+b35jvWHjuALDW2F5hZxsn5DO+Fhe/kM74zuAe3dvv3ZLNVLrheSJuaro286saVStKMW1CGO1LwWcfcWqFaVezo3EqFS3lUgpulUx2qbaz2XhtZXXDYEPD35ItS8PsL0vHJaqLdtgYtTnh536so95cq558i3jfn5AS0iMrHiPDBPuAiXj+8nby8hv/0IfaxyAhvfP3Dn0fk0Tnw+RHi8fABzfiw9l2luN2Q/H7QHPbcJ7tYfuDa8vMbY5ZAh+a97G7e+PcB0AnP+JDXdn4BjADfrjce57hY7sse4BtnfcjvfuJefLwIxjuAhJ9zz5k/xyHTqMpdyXiBD5jwy8d/cH7x/HIATjfl/iRuR1zsmgEX7ez6n57a7W+ka3f3H+1uak/jJs/QO7q+ptq9bP5OnKXPuWT88JNyk2+beQIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP0R/A9gofg7cMvG8ndS/8A6qqfncfot+CKsfg7cK/sXP8A/dVQPTZP235mg9IdX1HAuuVu0l2NPryz5U5G/l9Z+ZyPpiq+p9F3E0+q0u4x5+rkB+cF0/5xN975llY57pFdZ/jppJvfvLbzlJLBFTnOe7oQ/wBF74Ecp+I25NhTpjC5kvONkvMJ43fwIe7+3AB5S23HNry6ES5hNY6ZIqJeHeSn4cw+XuIzhc2BVnb7Cl7sPON9ws5CmX3IlYe+5D54G7ewEtL95D8OY2zyGe8KnnyKqedijJcpYzjwCsuguXcZkM5w/cYlq8vCWxmUdtlyxzAzaEV1bab5Z5GztlyaWU+9mtt0nh+PQ2dDlhfIDKjnCzv5l1YaTz4ZKaays7vwRWu7GcIlWJhlPZ5ZXFYl9pTB7tFx5W/NLmGcarWW/p1pFvOFJ/I0GpSzUZutXk/5SpZ/NpN+Rob55mzl5PnfXdDe302T/Nv8MJ8y5TKOpcpkrn457sy2W6PQvQ/t6QeHn/8AxK3/AO8iefW3NHoHojeOPdAfdqND/vInPy+K+q9Lnw39K+7FyPlf8Ilf+nuo/wDEp/8Ac0z6oXI+WvwiV/6eag/16f8A3NM0TzHj+k/Uz/8AH+Y8sp/XXmc5xKv851/d9iOkj9ZHOcSr/OlbyX9lHdh4cXUz+pWvoZUTOto5nF4znu3MK3WYmwtsN4bS68jdHk80bex37Md3jbZdDZU4pY2T8Hy5M11nHEc5T8+bNkt0muq6J7GyPOzXJxfay3hd+cbGNNyUGuylzMjD682u/wAC1V7EoY32eefPqZNNczZ4hxxeLuttvjE6DZycVt/Hkc9at/5b3TS//DfLMToYpprljGdtyxMkSaxlSy1hdwSkoJd23vDy6j2WN/IpfZS7Me1jvxhBhVT2y02ltsuoaX1cdMDfsppNPxEV9ZYbfNvkEU7qXawnFd/LBPZimo5z/Hd7wko5bSzjOOQX18Yw8rC/ewh6vLT29nk1yfImMcY5Sec46B809tuudmS1JRSzs/AsEPu3aT5INOUm0sJfMht9jCajy5cictxbUvPy8SiJJtP2sLGEms7iSefZn2u/LJftRj7Eo+D6ESbTxlPGcdEwilN42aS5d4ee0t8LPcS0m0u0ot7PwyVNJNttv3bBFPKT2T6bsjk1u1nnvzKn2eymvdjHeUttRzj2V3ICY5Syuq6snfdJpZ3WURLlnGI8w0m0o9nOMrcCVJ4xBpPqQ8vftZeOeOQ27Lb3XciJ4bznbmUTjZLZLvWBlbbZRTjkt/hyKs5T359EUG01hYxy2Ci8trK278h801yS2Kc422xnqBOMT2xt4ZJyur3zz6ELtZW3XDDWHu+mxUTh83nP2E7PvSzgiOya3ISw8LCXTbmUTz3byvLoTnfC9xEpYz78opjnKabff3FRU+ecc39pVjGywsZRD3WE9vAJ4inu0nt8hBCjFrssPDfa5d+AlvjmR15bmSJzjfrgnOcNddynC7OMfAnBQy8cvfknGH08SOT5YCWd8v3GUgnHXp3FXXOGsrlkJdlpeBSnlvwMoxUXcc2dVYa/FyXPwNLwT/2HNcvx8vsRur14ta6w/qP7DTcFNrQ6nd66X2I2Yz4al8s245NSW+Hv15mJWS3WdsmfXjGWXj2nv8zBmljvXnzNmMY1iTXy2LVRY7i9NNbNbpst1I9xuxjXasSRQ0uhelEolH3HRji1ZVZ7JHZLvZfgFHvOjHBpyyWUiez7i6o536FUYd5vx42m5rSiSol5Q8CtQeDbOJrvIx+zsOyzJ7A9WZfgsfxWI4luUDMlTKJU/Al4mU5GHKJQ4mXKHgW5QNd422Zr/Dns65a/rT7PxTX3nd0Vvz6nCaO+xq9nN8lXh/aR3tOOJtd2xw9Rhqurhy9mTQW+xkPOMFqgn7jIisxb7jiyxdErmtZj+PZrGbnWl+MyalrmZYYpat4Kox7yrs5KlHwN0wYdxGJcjEmMdiuMB+GdyIR2PTPwcayoek7TFJ4U5VI/GlUX24PN1E7P0OVZUPSLobjLs5vqMfc5pP7TR1GH9OtmF+KPspbJGvvlhvf5GxfPxMC/XPc8Z1MCW7y9u7YPr17w+1l9Pm8jC5fcVDn4eOQ99lnbxG3Jod2635dfgFR15LYlddk/D+ENtm9iOb6EEvKfL5EZXe3gey/f8R4PbzQEY2zsEklht5Ja2ztgjLe6369wE8vZ3xz36D6z35d3QjK5ZDe2U/kAbTeH3Fupjxwy6003vt9paqZysZbYFiSw9n8y3Jb+JdksbSeC3PON0gMWsvEo5bbFyty7y30zuwIyH2ccufgPh7mM/wAYAN9cjn3Dd8x07/cAzjbbxeCny/wJ5dGFzzz94DbGfmFnrzHxIys82Ay1tjGOgaT3xljHuDznbPvAhpJJbY8xy6+7I/jlgdHv8eoBPbC5kjn0bKeWF9wD37+JLxvyC65Y7/uAjl/0C+A81sRnP+IDHjsTvnK294+OQ+byBD58yMbf4E+fJ+Ib8c5AjfGxL5EdOa+IfiBoePbv6DwPr94vZdDTbia81Tlg+Bz7Z9Pd8tP9EPENZvepbxoR35+snGH2NnxMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP0W/BGefwduFf2Ln/8Auqp+dJ+if4ILz+Dpwt4K6X/9VWA9PlzOG9PdX1Xoh4lkuthUj8Vj7zuWec/hI1XS9DPEMo83bqPxnFfeB+e1XarLuyUNJrdbE1H+Mlnnkp36Y9zIo+7mMZ2fTmO8jo+TYVLftL+MEbPbuJeM5SbISkln5ZIo+WOYSee10IfLIby+WwE9NuhTjZtbeAb257v5E5W6yFQ/gTHbfJG/MJbgSsNZS5DLWcEx5kyeW88wqjclY5jICoK4FG/Jlynnl4gZ1tjCaWTOopY6swrbl4+Jm0MptyXLqBm2uMpcups7fd/W3Nda749nyeOps7dbJPqBkx3Sx0RXHG/RlCTTbXIrg+uMkWK+ysc8Mqi+S7l0KUsJJ9OhXHHj45JWyNFq+2pvP5tH7zQXTzNm71eX+c7h4axCKNDcP2mcuXz19Xw3t6Hjn6/7WlzLtPmWlzLsOYrVxeWbbc0d56Kn2eOdDfdqFD/vInB23NHcejKXZ4y0aXdfUX/zo5+X5a+r9K99z8n3muR8u/hGxxxxeP8ASlTf/wBKB9RLkj5l/CVhjjKpL9KMH/yJfcc/3jxvSPq5/wDj/MeRR+ssnO8Tf9qVvFR/so6P89HO8TL/ADnV/Zj9iO7Dw5Opn9SsC2eEbC3SbWOXMwLZYhnbc2NotsN8t1jobo8jmbizW6zHPRvJsYx7T2Texr7BZgmuaxl4xg2MctNLG3Vm2PNzTNS9W0lsv3lmrL2s9pP3GRL2lzWeTeOfgY9RqMmuzu23nPJeRk01y9q8ccXWUkvo22OX5p0Cjz2w39hoKLUeOa31n/NfzljPI38sNbLC8+RYxyH9XL9lJLn9gxFe1s0n8P42JS7K7OUuQzGMk3LCW+3n/iGCMJJdp8uRLWYvEs9nbJDjmD35LCbEm3Lkoy6tLAQytur32fcS1FdpJyfV4KVLCcl2X4t/eTJJPPakkumE/wCOQRTlJpKXLd5Ku1jKee57dAsS8XnvKW4vEWvg+hRKee09455PwCfs47WXvglp4TWM+4htpYSWVvz2S/j7Ch2cLHXwYUVFYW+PDG5TmSW2Fnly7yW99o81zb38Ag5ey/ZTyvj3hYznsvPJ7FKTTxKXi8lS7XYw3v3oIiTzPtpy9/NkvaX1uRH1Yr2e/EeQjvNLZ79wE7JcliKzkNbJ96Iwm1l5ZK3SafmgCWzay/IlYa7WPkQtsNJLC32Hc+mEkURjKXtvn7yrC7PPZ9ClZbxnK70GsPO7w+4olbPKxv0fNIYw+2+yvMhpdrOH16krGW+bfNMBjknnPUpkn2sp+4qk4vfbHdkj2t08N+W5UI58eW4Wc5y9/ElNZy47BJP2VHbz5lgh5xs3jkQ29139CptdpJPp1XwIWVjxW+ehUTl4e/MjrtnvHazvhPyEm28vkWFTnZbPHTBGVltPnyCeMOPPvCXspY2RYiVz5/EZ69CFjK9l4x3kt7feZQOe22eeCpLb5EL2sv7wmm/fjzMolTjLbbC3/wAWIr2efvJxtuuZnEW7xZtKyW/sS+w03Ba/zHLH+3l9kTcXiTs66a/1cs/A1PBWFoU+ua0vsRtxnw1hb7s65eNnzXPBhT+sk8Z7l5mfcJPKa5PO5hVtpdrGTZjGNY01jbHhuWubyZE2njLT6FlpnRhi15Vaa2wUOKzkvYwvcUuODqwxaM6tOIUVyLjj3FcY5Ozj43LnmtKOS5Gn4F2EMovwpbcsHZhxOTPkY8aZcjS8DLhRy+Rejb+B14cG3NlzMD1PgHRx0NvCxqtZ9VP+qJWkkvag15o2zp2u8+mmlR8C1Kl4G5qWz7jHqUMdDDLp2eHPtqJ0yzOBs6tLwMWpTwcufC6sOXbDp+xXhNfmyT+Z6LKGK8/2mefzhg9BhibVX9OKl8Vk83quPWnodPnvbIoIvxT6os0lh5MiOO73Hl54u2Vo9dhv2uSwaaMcnQa9HFLKNLSjlmXFju6TO+ymNMuRpsyqdLKLnqkuh6ePB7OW8nuw409yvsGR6tLoR2DDLi0sz2sqJvOB7pWPFWm3n+wuaVV/0Zxl9xqUjI059i7g+qz9hx9Rh8Fb+PL3j7sfNowr/k/Iy4zVSEaieVJKSfmY18sxznofOPQavbngjruiWnl+0mwsNPPdjd5RUPPKzvuxl9797CWXy96CbynnbHeQOXd7upHjtn7xLddV38iUku/IVCz3N+4YeeTzjoh5rfwG2N1nBA5vOffkh74+8np4+A3y85a6ARt2eeM7YDbTy3JthrL22GF3sCNuuCiqu/JceefyyUVPf5AWJfxuW5vrsXJlua7/AHZAx6paa8Ny9U7ix7gGPD5Bv3h479/Fh9P3ZAPnkLzI+YfUB8/ePduQtly2JfuznmA5EeHeSFttkCEN85Q8e8PPfjyAjkPFL3hvPcF8Uu9gHut3/gMd/wAxs1l4+JK/joBC+A8ydiPkBGw357kvl95TusZ6eAE7+I67kfxkcunMBnG48B72h7gGfEpnhrvJ8XuUVGuzyyB45+Fpf/RvRlQtE8Su9Rpwa74xjOT+aifJ57/+GNqvb1PQNEjsqNCpdTWefbkoR/sS+J4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP0J/Azrqt+D1okM70a11Tf/AM+cv7x+ex94/gK13W9BtWm3+Q1a4prwzGnL+8B7j1PK/wAKqt6n0Ka30c/UxXvrQPVDx38L2r6v0N30M49ZcUY/86f3AfB1TLqN8iFnvwTJNyeG+fUp6f4hVST5Nkvrz7ynm3nmN87d2CKlt8s7MhfV3+wjKzvuSRTmsY+RHV4Dl1ZHfgCrbLaKXjfYZWcYKtm+XMKhL3DGX+4PvwI/AKdRt3MLy26hvfvCw2YaJW3mGFU745lymuSKc+JcpNbZIM22xsnzM+itt9mYNuzYW+5UZ1ot8YRsreKe+cLBgWsJYTXM2VKGMZX+AF+Cwt3v5FS5rcjs4yu8qimn0wRlExzyaTfkXOz1XzKVnn395VDaXPG3QlZxy+qyzf3T6qSXyNJWftG11GWbm5eedVmpqvc5r81fU+Om45+SlF2BaiXYErDi8s225o7L0fS7HFGlz7ruk/8AmRxttzR1XB0/V61ZT/RrwfzRz8vy19b6P75yP0Bj9VHzb+E1DHFMZd9KD+R9Iw+qj50/CgWOI6D77eD+cjn+8eF6T9az8r/DxX885/if/tOp+zH7Dfr668zRcUbajL9iP2Hdh4aOqn9Stfa/UNjar2totvpua+z3p8zY2cY7Nxy87b/abo8fm+7c2KxhPntsn/HQ2UU4dnMo+/b+OhgWaykn5LkbCCkksPuxtlvwNsebmttxy1mKw+XPJbe8G1h5fdzL1WKlDEVHHPGCxP2Vs843b6ZMmiuWjGS47mnBr+b7LPTCOgz2U1s8+Rzziocd42w6G3hsdCorkscuWBEyE9tsptYxn+O4Ntdrs79l79xDx2VFL2dubJcn2GsYzuisCOZfWw8+JEoZ79+S5ENey1h9+3UKeE32U8Bin81tyy+4iCk5clhLfu9w7Saw8tZ5rbBMU8rtNrPcgHaw29viRsm1h+CXMYknFteTGE90t0ueSxDG3LbAXPfO3dv8SN0sQxtnG3ILKw01l7+RRKzhd3LfYpUst4W3gV7qWFu8c+o3XNt8gimWE/afZyub6kSaTziCa5roSnj2F1WfIiKxJSf1umwRGYRbwnJ7brvK+vZ7Sjvs+5kdE029sErrFJtgEoLKw1hchLCa3Jy+0pLHaKW+nXnsiirnnbryZDxt7Kyl8Cl/o9OjKspt4WcvJRDlhdW+ax1C7TbSW/mQ9ot5cUF7T3+QCWOyk855FWG1t0ZS1jfOOfL5kNb5W2+BBViLSytu7r7iVvHMVzeX5lKbUuY9rGH15bFRMnJLbbm894zjGceJCwk1uk9k+4l5aXLPQqIXNc9+YTW3MZ6x+ISz+59SiX7ueweE9+Y5Nvq37w8NvG5UF7O6SyvEqiljpt47lLWXvtklNct/cZQTHGeTfj3ALDw/mxFYl4mUEx5pJvOOROMPDePHoRhJrZZeCVhdnfzXeZRilLLWFnxYSfPn4kN7rtLLb6lUc4zs8eJnEqzd4+hVc7fi3nv5Gp4JX+ZJP/fy5+SNxd/6HXysr1cuXN7Gp4Kj/mJ55eul9kTdjPhrC+WbXa7e7Xl3Pl95hVMNSeenwM+u+SXNbJmLNb4i9jbhGNYs+fJf4lvGz7i/Uw+mEW3ndHVhGnKrTjvsU4zy3LuGxg7ePFzcmShRK4wKoRL1OO56HFg4eTNFKnl8jLo0tyaEPA2enWc7itClFby69yPR4uOea4OTP7RGnafUuX7K7MFzkbu106FJJQp5fe1ubyz05U6EIQjhI3+haBUvKsYxpt5fcbpljJu+GGWNx9p5cpR06rNfVZc/kios5p5T6YPobhX0S1K1vCtqE1aRe6g49qbXl0OjuPRLos6LjSu68Z42coxaz5bHm8nrfSYZdu3Rh6f1Gc7tPkTUNJUcyjDHgaO8s3DOx9D+kP0d3eh+3OEatCX1asE+y/B9zPJNa031bknE9Pg5+PqMO7C7jh5ePPiy1lNV5/Xo4fIwa1PmdDfW/Yk9jVXFPmY8vE2cPLtqKkDurRfzS0ffb03/AMqOOrQ3O0tIuNjaprDVCC8sRR4/WYaxez0me6yKfsrOMl+OGsdS3TRfisLlueLyYvTxrW65HtUMrHPc0NDaeDpNUh2rZ7HOL2apjxe2UM/eNjRXsl3Cx0MejU2LjqbHt45ztefcbtEktyh4wTKaLbkaOTKNuESXLXa6p4z9ZFnLK6EsV4S7pJnDze8sdOHl9t8G15XXB+i3NR5nV0+3nJ55t04tmbe57BpPRhcRuvR3oNWLyo2UKXvh7D+cTe3X5PY+Yvl6Mahp5fPPiQmsbYb5+4qqZcmtuZS85wvsAY5bbh7rbIfXw8Ru+/OfgBGM8nnPQeC6dA2+eW+7cYzlLqRTfp7vEb9V4jZvfdkd2X8SCcPx/eR0Xjy8SfHHMjpz+fMBgc+73PmN+eX9g78f9QDSKamFHZ/Mq5Lnz+wpmu14vzAsSx0ZaljGU+ZeeV/gWpd75gWKu2yZY58ty/U8MIsN7/4gQ8rKb9ww/ALnncjHggC5bk7EYS6onr19wEPyeOoXN4TQ+ZH8cwCXPPwG3QDwwA692OrD/wAdh3c/iQvMBnYb7faGwn45QDnvncdE8fIdHhpEbc/mwJ6dwxjfbA8OncAIQXPkN2G3y3+IBvfDeCFyz3/MMcgHwHXC+DQXh7xz25gOe3XxLdR+1hsrfN/ey1J4km2ku99wHxz+ExqKv/S9qcI1FOnZ06NtHD5YgnJf1pSPNDacWao9b4o1XWJR7LvbyrcdnuUptpe7JqwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB9x/gC1O16GtWp53hr1b50KB8OH2V/wCT5vfWcHcV6bn8hfUa+P8AiU3H/wDZgfTHU8Q/DNq9j0TKOV7d9SXyk/uPbup4F+G5UcfRrYwWPb1KH/d1APinHafeQ8rmJPD5bkdpPoRklNtrpkde9EZWFsmSsLKYEeG+we3UNpojOPEipe2NyOjaYzsRnAFSwmS3sUZfMZ+AVXnzCw+rKM4GQqXzYyylsZCqs+8NlLZGQqtPBcpvfnjvLJXBkVs6Dyl3Gxs1ul0NTavx6G4sYrtJ9Co3lhS2Rs6NJYy0YunQ2zg21CnlLIYsZUtmljAjReei6Gb6rflsS6WFuSsowfVyUsIlwai/ZeM80Z3q8rPXl5kOmuy10w2Stked3kszqtvOakn8zW1OZnXEswbXVt/MwJ8zmnmvqOX248J+UIl6mWYl6BKx4vLNtuaOk4afZv6Eu6afzObtuaOm4apyneU8LlJM5+TxX1no91nH6CU/qLyPnj8KNf5/tX/7LD+1M+h6X5OPkfPf4Ua/z5aP/wBlh/bmc/8Ah4fpX/2L+leGraSfiaLin/tB+NOP2G9/O95ouKv9PX/DiduHhp6r6jAsvyZtrP62GlhvlnPmamwT7JubNPs7Sz5G/F4vP924sYxS5Y6/xgzZpR7S3WU2kvsMSzwo7uW2yRmPKwkuys79/wDG5tjzs1GWllvG/wACxUapuTjHZvGMrf7TIkmt+Sff0LE5J5l06Yw/sMmiuTqP/wBPKTcsdqi+S/VZ0TaUml7Xhjmc5df/AG7tXnDdF/2ZHQxeMtvk+YiZKnJtLPx5Z/6FNTDSae+SU0+rb5PfJEuXZXJLC3x5FYKZbtPtPPhuTjC6pc2hBZ57siLUcy7W76N8gxRLLaXXOWu/3kv2tsSbWMrG3mVU92orD6JJblGyTkl8vkBOHnxzjPX4lT5JvtJ7J7FMcqTjyeOm5KxjG/cyoYaTfYeWtu8jOWk9tsdCvkmmlhb8ymfZ7Wc803lblEZXa58uneM+y116YXkQlu2l5sQeZNZbYQUY9nEns+9ZGW2nlJ+K8iI45pt793MlJc1uEHz7OG3jZYIb7Ky3t1CbazlPlyGcvwXf1Ar2xjPLr95DS5eOO8iT6rm18BKTWY4bTW/gWCZNN4WX9pCSjDHaz5IPLlh7LqFhrHaXxKGcrCaePiQnlfb4kxSbzlPwDTw3tjPQA87pY55xgjo/Zys92Q98Zewxull+RRMfLfuGNvN74ZC3XMn6uEt19gRKePaaefkRlZ2bZHPbKeXsupK223yuZkiU2vZaWMhPrttyIeM9xPPHJ+CLBKbxlrOO8jOH2k3nuYfPbZ5xuQua8O8qK1sv3kNtvGVjbGwjsmTjwXkZQUvL2+ZUk22s7EN4fX4kx653fQyiVV4te9slc1yytuZSsJrxWfAYWct470ZRFT57YS7iWu9rwyQtu/uJSxs13tZ3NkY1Rd/6HX9rf1ct/czUcF/9iPZb1pdPBG2vov6FXi9s0pZ35bGq4Mz/ACI+71svsRvxnwsL5bCvvz55X8fx3mDUfsvx+0za0fal48tvmYlTv55wbsIwyq08Yyt0i2/LkXGn/wBPEpxnbr5nXx4tGVUJbeZKRVhYRUkehxYuTkqIxyX6cfAppx8DKoR35Hp8ODzeXJftqeyOr4StU6kqjXgc/aU89DvOCreMraTa37f3I7M/h465eP4uSN/Y2yqzhTiubwe6+iHhi3o2q1avSTcZdmgmtsrnL7l7zyfQrWP0yOFyWT6S4XoRt+HrClBYSoRfvay/mz5z1nqbhxTDH7vW6HimfLcsvs2KRIB8o9tha1p9DVNMr2FxFOFWDjy5Po/cz5R420r6Nd16TW8JtP3H10+R82elCinr2otL/wDET/tM+k/6e5spyZYfZ43q/HLjjl93hms2/Zk9jnbqnu9juNboOU5KKyzTw0OtWfbqfi4Pq+b8j7HkyxmO6+e4sMrnrGOasNPneXcaeMQTzJ9yOpnHDwl8jKpWlK1pKnRjhdX1ZaqLfPTwPB6rPvv5PoOmw7J7+VNPlkvxXToWo4W2C7H+PE8nlxejhVm/jmgzl7hYqM6y6j+KfkcverFVvBzyNlWo1MLmV+u25lhkZeDbOWxhcJWT6zLCk+8x03kuRfePxLTt0vJtEp+0mW1krT2NeV2ykfX/AKCpqXou0rf6sq65/wC+m/vOzuUvVM83/BsuZXHo7nTm8qhezhHPROEJfbJnpNbem+4+ezmsq754aertNrCx0RR2d+75lyusVHjYow9vkQRy35Y70HjxeNtmH3cveOnh5gQvrfW6beAztukTnHXJS3ul7uRAbfLKZK3TWGkM9zRHvRFMLLfMYxzb37tiV8PNkbYzhfEA303bHTGckdMfeMrvx7wJ/NxhZ8ymW63+RUt8pZfyIknnZAWJ464LcuWS7LOcYx95annfOwGPVXmWGu9mRV8jHfNrr3AU+S+JLz3/ABQ8B/HMCG+Sygku9jIa8twHVcufeR8GTld6I3zjZ+QE93QhsZ35obfwgIwg1nmT34WSHs98gS8+7zIfPf3DyXkRsuYE56ZHTPzHuwg93vzARxnuQecb4D7yHjk9l4gFjHN+8JrrjPUJjnya5gR12Q/jmOmyHdtkAmQ08YwhlYzsS/j7gIly5e85L0tarHRfRrxDqHrHTlGxqU6co81UqLsQ/wCaSOsnyyzxn8LLVJWXo1o6fCSUtQv4QnHq4QTm/wDmUAPlAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD6w/8nlWSuON7dveVOyml5Oun/aR8nn0h+ADeyp+k3XbDPsXGjSqNeMK1NL5TYH2kz56/Dhml6PdNjtl6jFp9fyc/3n0I/rHzt+HLn/ITS3hY+ne/8nID4xff3kc/cRnKIT8QyV594bfUpz3EZIKsoNlJDYFTewbbKc7DJFTkdSMhNBVWSMkZHQKnIyQAqRkjJIWJJi9yklcyMmdavOxvdMWZrJz9pnto6PSsuS5YyVjXT6ZDku821GGEzX6ZHeK79jdUaeXt8QxUKK5pblXZW2Es8vIyFS71uT2FncVlGOqbSwmyZx9h7JyaMhU3ndNMOk3l4znvIzjx+42jjub+0xJczMvY9irOHWM5L5swnzOaPpeTLeOP6T/SYl6BZiXqZKy4fLaaRbu5rqmuiy/I7TQZ0LatCnRSnUzvLojidNlKNTZtZWGdVw8v5xDzOfl8PpvS5vObffVB5owf6qPn38KT/tq1/wDdIf26h9AWjza0n3wX2HgH4Un/AGxbP/2Wn/bqHNPs8n0v26m/pXhf5yNHxWv59D/hL7zefnGk4s/02n/wl9rO3Dw1dX9Rr7BJw8Tb2UU8RWDU6e/Y5G5sVvvjPPY34vE6jzW5ssqOcvtdX1MtSaaSynn82JYtYrspPD+8v9pJLEt85w+RtjzclEksNNZz1azksVuy1FY3L7acGn2U2/P/AKGO5bJdjGz2yZNVcte+zx5aNpr8VLrnPsyOgT9jsvfp3mgvU1x3ZKWH+Jl4fmyOgksc+y2+7cRjkjOcxTwl0fQpxu+w92ue/IqlHtR2allYTz0EcqTk5PbvXUrBTjK7WV+5/wAMns4axLkuf2lLc8Zce0+q6e4qwsckmGJyeOb23zzWBHZ5y/HbmIvdYz7lz/jBCwt4rlv1AnDaw914c8kY5tybb225kdrDTayg95PfDfQqIcU8vLy+Ta/jxJTw87JN7tkRl7LbfhhIqnHMWsNrq0yiFhvHJPuEu/rnu5DCy8rnywRnON231TCJcl2XlPHcv48SndpNJ9/vIxmW2+eZLxssbrAQz+MXJZ+LI2UnLr4j2k1nGfzSEm/HHgBUst5T2z1fImXZ7K57iXLs9G90Rldn87C8PsKGXGLl8MvkEk4p4xhd/gTnMXlvf5Bp9pc/fu0UM43x5PJCT2wuXiQ1Jxba3XvZKeenMCd/gUvbdJc+8mMuklz8A2mtk8lQjl8uafJ9BJ52a5bbjGF9xC3jh57sZLBLafe/MRfTOwaxz2yhyT6fcVBZbWejJjtLbPmQs80s4EXtkpVSfhuFz3WSl/W6+8l7tvpnYsRMt8vOH3lTxnbO3RFK33lyyPZxz3S5GUEp5in4k9Vyz3ELltJrwwVc8YwzOJUfmpNvBVFyX8cinCz1+BK6tLO3eZxFXJezs+XPoSo8vIjGOf2FTT2TW2O4zjGrd7n6HWxtJwljzwang7P8hN/72ePhE2t2v5nWy9lTl9hrODlnQ8d9aWPgjoxnwtd8s+slu856PJiTym1tszLq85b4ZizS7fJ9nuN+Ea8llprYpwunMuyw3goaaff3nbxxz51CKooLwLkEejxYuLlqqnEyqEMlmmtl4mZbx3PU4Y83mrYWUD0DgmH81e35/wByOHsY8jv+C44tX+39yN3P9Np4PqO40GH86X7J9EaJ/wBj2S/9nh/ZR8+6Cv5z7j6D0X/siz/4EP7KPkPWr7Yvd9O+bJmAA+feshnz76QLb1+u37bxH1889/1mfQTPC+M0nqd//wC8z+09r0XK48mVjzPU8e7HGV5jfWNCnJtQXv3NNeQ593U6fVI7s528h7R9RcrlPd5vHjMfDTXC35GFVjvtyzsbKv13e5gVzk5I7MKsr4l+C+Za6PHMuw5bs87ljswqK6/FteBy+oLFRrbmdVNNrCW2DmtUilVZyye7c1s1tuUP4lxluXga6sEy9HzLC5l6DGJVyJWW0XEWkfTH4LdzGpwnqVsmu1TuYVGv2oJf3D16ok4NeB4V+CfVxDW6OfrU6EvhKov7yPdZfUfkeFzzXJXZh8sae6S9Z3lnuxjkZF23Ge+xZ8d/ea2SF1S+GCM9EsdNkS+XtNPPJcx5rLXjyAjd77/AhrnjZdSXzxnd7iXLrnwRBC9+e8PG2c48+YysZ6JdAue3/QKnPljz5EPdYyxvty8fAPPPHXbYgPfGW8DMs8mThvoilLfLS5gS91u/mU4WOXyKk03yfjuR08fMCxNLBamsfml6Xf8A4lmfPO2O8CzU5cjHwlt07jJqYzz8zGb3xlARjbZIb8kOu+5H8cwG6WOg5Z25/MeGMvzDzyW3kA69Q+fVN9zDxnDI3y1hgNsY5+8e9fELPX7Bz67d4B8+4jOFs8eRL2fMc1lP5ARlhd3R+I5d494BctlsM7Lce/3BvwfkBGElt8EFnns/EPxWBJ43+XIAu/mQ1s8vZ95O/OTI5vlkCNmS+/PzH2eIafTO4Bt9NgQ29+TJeMZXTvAt1ntu9j5k/DE1JVOIdC0eL/0a0qXEkuWaksL/ALv5n0xPdt958X/hDar/ACt6XNanGfap2s42lPw9XFKS/rdoDz8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD3D8CLUI2Xp2treTx9P0+5t15qKqf8A7M8PPS/wXr2Nh6fOEq83hTu5UPfUpzpr5yA/Rd8z56/Djhn0d2E8ctQiv/pzPoV/WPBPw2KTn6MaEkvqX9OT8uxNfeB8Ot7k5KXzHQKqyCEAoxkgEEsEABkAgKkZIBFVZBCJCpQIJQUK4lJMeZGUZtosSz3nRaRh9k52z7Lkso6bRUsxSLErsNLhmKN9b028NLPcajSIZjHC3OktKT7CDFaVHPPfrzJjReeWxsY2ye+FllX0bDSa3CxrHSa33wHSzubJ277h6jo/MjOV4XxJSdDWL2i1js15/bn7zUvmdb6S7N2vEtZ9nEa0Y1I/DD+aOTkjRZ7vexz7uPG/kRL0CzEvUzCunh8thYflonXaAvx8PM4+zfZqRl3M6vQq2a8FDv5nPyz2fUel/PH3zp7zYUH304/YeC/hSL/OdrL/ANlgv+eoe76Q+1pVpLvowf8Ayo8N/Ckj/OrOffQiv+aZyz7PJ9N9upv7vAvzjS8Wf6ZS/wCEvtZuX9Y0/Ff+lUf+EvtZ3cfhp6v6rA03kbuxjiXtb79TR6bySOgsG8LKeE0+b+w34vD6jzW4tWuykl4+Rekvawly8C3QWILvTL/a2Sf1VnfPM2x52SzL63ZUl3bsx6sX2X+b07n/AImTJOUXF+14YyY1VdmKe+75N7GTVXLXa/8ATmwWf9VPOV+rJ/eb5ybxt8DQ37T4409rK/Fy/syOgzya6vljmIxyRlZWMZ5bdwks5Tx7hjpnfp05MjZwTaT6eRWCGmlhvd9xTlNpYbXPD5lSeHzlz3YbjGbjlY5LrjcMUvfbtJbciO0+Wz7yG4xbaXgt/mSnz7SWFy35FEYWzit/PkS1heynh/Ipy8bJPPPz3JTSlFSW3n/H8MIh5XXcNLKUl16+4RzGLT6vmkE04ty5dCgsvls/AJZz2W0s7kRWVh9X8Bs+q+GGETPmm5RXV7ENvst8n4ErdYTx37ZKXjPfh4QQbXJ5eF8Astp74xtgiaby/ZaXLBOXj6zy1gCpYbWcYGF1e++/QRbTWGsLmQ2nLC6FCTa2fPHInHPG2XnmOaxhZfTzIzLrtkol+LTfMpk1jn8Qm3FrCwQ3jfPQCe7b/qSnu9k9iOW3f4kPdvOP3FRUmkv3Bb4fTHUh9+PngndrbPdksQbbxhJocuvwKYpN4e/vKua+RRC5vO4TSxt4jfff58yfNFSi2W+3gS+qxvyIznfdk5XL7zKBvGOOfgyWsdz9xD8Nicb4+BlBUtuq8ieb3KdtntklSws8jKMUy5vLy8hya6r3kp801z+0LLW3PweDOJUxfJ5x5lTbzvtuRjbvJWOzyx9xtjFbvP8AQq+yX4uXPyZrODcrRM7flZfYjaXr/mVf/hy+w1nBqX8hvL/1svuOjH5a13yza3LEcPHLfw5GNPHa8EZVbLynjHTwMWpnv3N/HGvJbzmTZDWfFFUk1zzkhbt9Du4o5s6R5LYrp7kLu+Rcit8npcMcPLV2mtzOtly2MSkt0Z1ut0enxR5vLW1sVnB3/Bsf5r/S+5HCWKWx33B6/mv9P9xn1P02HT/UdzoS/nC8j6A0f/sm0/4EP7KPAdCX84/onv8Ao/8A2Vaf8GH9lHx3rX9r3fTvmyZYYB4L1kM8M4v31PUf+O382e5vkeF8Wb6xqC/3kn/zHsej/Pk8z1Hxi4HVF7TOfvY4y+h0eqL2mc/dp74Z9TPDzsWnuV3mvrJtmyuFg19aK3NHJHTgsxw479xVD4MpjkuU+e/M8/ljrwqZYxvnfvOe1eOKj8zo2njY0Otx9vPQ4/u3/ZppFqRekWpLc15RlFKW5diti2luXY7kxi1VHJcRQiuPIysR7X+CpcSjxNf235tSynJrxjOnj+0z6MfJ/YfMH4MlwqXpAp0X/rqFWCX9Htf3D6fb6nidXNctdnF8rVXqxLK5mNvnZPbwMy/W7zlfeYbS2WV7kc7My0sfYg0tuiC8nnlnuIeOiwBL73n45Kff06ErHJYWPkHvnnuQQ335+JLTws8ue5GMNpYxzwEnz/wAPDxnHxGO7n4dB1658xsu5EVGNsk42zsiMYwxjfoAz/HIfxsS0vHv6kb7vGc8wLVV77Nssz792ZFV/rfPkWHt3b9eoFiosLlkx58+exkVFvnYx5rd74ApHiO7d7eAwscvgBHyD8cBLHd8A+vl8QDyRy2/wJ838wsdEBD8/kHnuQHLwAgZJa645d5C8NgHdzC36+/I6t94y285AbdHsF3bgd2MeYEeSDwS8fv2ITxjf4AR4jZZ3wS+fJ+5hvbGyAjYbY7hty5/IZx0AnfkUSfUnOM77fAtVHttn3ICxeXFO1tq11WfZpUKcqk33Rim38kfn/q97U1LVrzUKzbq3Vedab8ZSbf2n2Z6cta/kP0Wa5cqcYVa9D6LST5uVV9h48ey5P3HxSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA6H0ZXy0z0j8M6jKXZjbata1ZPuUasW/kc8VU5yp1Izg8Si00+5gfrDPmzxj8L239d6Ir2pz9VWpT/5uz/ePWtFvI6lothqMHmN1bU60fHtRUvvOA/CUsne+iLXKaWezQVT+pOMvuA/OyW0mQVVlirNdzZQBIyCAqQQSAAAUABBBKAIoSgCqlEoglEWVJMdmEipIMoyrSWGmdFpFVRnHc5ilJpm0sa7jJPIK9P0OtCUYnZadGM0sM8u0C99qKbPQ9DuO3j2kGFdLbW3aS2MlWKfTC8UU6fUclF/E3VBdqCeMsDTSsXjHZLU7PbePvOohRhJLMfgTKzhLdYCyvDfTXpDjZ2epwi8QbpVH3J7r5p/E8ma3PrTirhmhrmhXWm1opKtDEZfoyW6fuZ8rarY3GnajXsbqm6dehUdOpF9Gmas5q7et0nL3Ydv+GJFF6miiKL9KJqr0+Ksu1W6Om4fj+Ph5nP2kN0dPoMPxsPM5+Xw+m9Mz1lH3dw/LtaFYS77am/+VHjH4UsfYsZd9PH/ADM9h4Tl2+F9Kl32dJ/8iPJPwpI/zHT5eEl80ck8R53Re3W2fq+dJP20ajiz/SKD/wB197NtL6xqeLPy1v8A8L72d3H4aOs+q1+l4xlnQ2KxjOMHO6Ym03tzOk072sLZPl5nRi8LqfNbmhF9nfqvZ3K5J4fLffbnzIpR7NPGHnnuiuWYpbvPPfobI87JZm001tl45MsVXnbb2tsGRUaWcyy33mLWcWk20k9u5mTVXL6omuM9NbWM05r5SN9nxbeeRoNTi/8ALHTU9swly8pG+eWm0nksTITi1LbK+fIieZR7MNn1w/mSk1+bLfffcj6rT2DBSnl9G/gVt+KzjuKEljMXtyRK3fTHJJBil9YpZz37EPDwu0n05FLfPd5QTxvlLpjqUSn7OZNLw/j3h7Zba80Up5lvHD6Ex2xF+XiIiqOOSysc3ncoSaS9ndrGc/x/GCU3hpPKiRHHPf8AcUS1s+Wf3hNKC3TeN0iG/LG5Le2It+G+7CIeUlh7rdhcs4zndYGctJp/EZcVlvbvbAlbx6rHhsMPPTcpaXaT70TJteSYRMm8J4Ii32enXclY7L6eKEZbPZ5S/j7ShnnjfD6sjdPLxj7A28tLbbuHXD6FEpLOc+4jm842I35PG3eQ22vmBU4vK5/EdpLZ5WRN9cbd5Emmt+m7KiXhc/cMYl2kvmFv4vpuJbN7vYsB5ztJpvwHNx8iXjDzkiLfV48SoJ4XLqSl1eSMYay/kSsbFROOpKxjl4lK2e+SVl5wufMygPzKuv3lLbXLn9hL5mUEvdZ8Cds8upTn4k+HQzjFU0+1yJTaTKYt92Crm+uxnEqpYwuuxPJt+K5EbPue5Usd/LxNsYVbvnm0rcvycvsZrODl/mPfrVl9xs7xL6DcZ/2cvsNbwf8A9hNZ/wBbL7jow+VrvlnV8Z7JjT5PD8TIuF2X+98+ZjVUv46bHRxxryW28+KYxh+JKy4795Lylnod/FHNmhbZL0O/qWVzL1M9Lijg5ayKK3M62W6MKj4mfbc0elxPN5W4sVyO+4PX81/pfuODsFyO/wCD1/Nf6X7i9T9NOn+d2+hL+ce49+0n/su1X+5h/ZR4HoS/nHuPfNK/7Ntf+DD7EfG+s/2vd9O85MoAHhPWQ+R4VxU/893q/SlP7cnur5HhHFOP5euM9alRfaez6N8+X6PL9S8Y/q4nU1uznrzbPd1Oj1Pmznr1Yzln1GPh50ai53yjX1U85e5sbhczX10as3Rgx48y5BrG/MtrmVw2Z5/LHZxq+aa6s02tR2yzcpbbms1yPsJ96OK+XRPDnZdS1JF2fNluRhlFilcy5Dcojz5FyJMYWq0tytFMSpIzsSV6D6Aq3qfSZpMs/WquH9aE4/efW2D409FNw7bj7RaucJX1HPl6yKfybPsuR4nXTXI7OH5Wu1CO77jCa79u/c2GoLPa5Gu26/BI5G0eO9Z6EY7stfMl56fYQ99wG2UQ89Ps/jYnf/oNsrdZAYx4bkPPu5PwHwz5DHV5bIpy64X2BP8AV+AeeeyDwkk18yCFzzjkOmezhMc2+vgS8p77e8A8Z65z/HMh7rOE+uw69H7g33vO4FFQx5Zzj7jIqc+hYls+ePACzPGDGqc/42Mmba5Nox6vmBb+DD/jYlPu5Ee7IB+G/iHlLow9tgl4MAuq/wASPl5onrya+JD+HmA8xyfL3E8+rIXg/gwIHvzh9WSiM78wHLrnvC89x3MeYB/tYXgiMY2XzH2E+PuAYzjr7iPFDG/JkMCVuN/MhDpnPwAPz+exEtnhfEhvHV+RRKWdwDkn1RYqN5zleJVOXTPxRZk8vCYHgf4X+uqNnonDdOou1UnK+rwxukl2Kb9+anwPnQ7P018R/wCVHpI1W/p1HO1pVPo1r7WV6un7OV4Sfal/SOMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP0q9AWpR1b0KcH3il22tLpUJP8AWpL1b+cGbD0oWX8o8CazZ4TdayrQ97gzzH8B3V/5Q9Cj0+VTM9L1KtQUeqhNRqL5zl8D2rVqarafVptZTi8rwA/LPU4erv60OimzHydD6R9Nek8Z6nYYaVC5qUt/1ZOP3HOgSQABIACpBBIAAEUBIwAJSJSJSIqEslaS7iUAyMFSRCJTCpMi2niSyYxVB4ZGTp9KuOxJPPI9B4bve12Vk8osqzUksnY8O3zjUhHL8yxjY9n0ir24xec7HSWMs4zv5HC8N3Mpxi+mNjttOU5JPshi2tNLmXkWqUWl7SZfgvAESvI8X/CG4Q7UYcVWNLdYp3iiufSM/ufuPaUi1f2lC/sa1ldU1Uo1oOE4vqmSzcbeLkuGW4+MoRyZNGG5uuNuHK/DHE9xpNbLpp9u3m19em+T+73Gvo0/A57HvcPJv3jJs6e6On0KH42JobSHI6TRVipHzObl8Po/Tcvij7V4IeeDtGffYUP+7R5b+FIv80ac/Gp/dPUOBN+CtE/9wof92jzL8KRf5j09+NT+6ck8Rz9Jf++/e/y+a5fWNRxZtUt/+H95tpPEjU8V/Wtn+o/tO7j8NPW/Va/SX9p0+nc08/E5jSFl5ydTpmMJ5afmzoxeB1N925oxzFdO7mJL2u0u0881glJRg17vMhtLC8ehtjz8lqo203Lfw6mLU2ba5eGEZU1nKb5rv/jvMavHtfm4yscsFaq5XV/Z4u0tuTXsy3fvN9FyTz23v3nP62muKtKazl5WEb9OOyawuRYmSYJrL3TxsuhC9rMknnInHu5be8hZipY36YDBHZTzLbPPmThrO/Nch1xltcvLoIZm+XTkVFOZJJJteI3bct029skpN5bytm918yFhvZvCW24KmUX2cYxv8Rh5ym0S3jfrzZGcRbxstwiHjljK7pchUXaaxvnuCfPOW+uR2sbuOX122KIWVySbxu0JNqOXFE5WMJ5ed/ENNPdPzYRCl15kPuSefIjGUpZbT7+pV2ZZ3SWAh2sPCeMLHIjZSy1s+4bpPLe2yWSM427+oFaefjgjO2ct/cRyT7Lz7g1hZznfkUNn8ly3wHv3/wCBHl8e4fs8uW5ROU9std4zj7infC7WxPTLyAa8Mb7LH8YJx7WXjC8CHs8c/DqMtJ898YKirbtLO/uDbznD8e4iOduu4Sed8+ZUJJvk3jl4ktvGCFl7Y+4ndPC5+BROzw08PzIe3XcLKb7ujCbz/DMkS+uEsPPIqXLD3Iey5b5Jztv7ywU74WMfAqjlYy8kJ88PcLcziVWtuSHLotynDS8ScvKZlEVJN78veSn05LzKeezwSnmWyNkY1WtnjvKo7S5FMNtnzKo9OZsjGrV8v5lcc/yUuvga7hD/ALETz/rJfcbK+eLCu8b+rl9hreEMfyH1z6yWPkdOHyteXlm11lPCW/8ACMeos80/d8DKuVl/Wil3IxZNvdrpk6eNqyUS3xhf4kdAvP8AwJ64xz7jv4o5cyPLBfgWlzLsF7z0uFwcq/S5mwtemxgUTYWuG0elxvN5PLc2C5Hf8Hr+a/0jgbDod/wjtar9onVfTOn+d3Ohf6R/RPe9M/7Otv8AhR+xHgug/wCke49707/QLf8A4UfsR8b6z/a9/wBO85MgAHhPVQzwXix/59qeNdr4s96fI8C4wn2NYqVP0a7fzPa9F9+TJ5Xql1ji5HU9m9jn7vbmzodW/KS8znr3m+p9Nj4efGouDArde82Fwt2zX11uYZt+DG37RUvHoU8mVZODldfGuJdXy8TA1qK9TtuZyeNzF1RZt233HBl5dU8OWntJluWC7W2my1ImUIpjz3LsVuUJF2KLhC1UuRWuRTEuIzuLGVs+Fq/0XXrSvnHq6kZ58mn9x9wPHae3U+FdP2u4ebXyPt/SLj6VpVndZ3rW9Oo/fFP7zxfUcdZyuzp77VF+vZzz26ms6cvuNtfLMUamSxLD+089vRyXR+4j87L5+LJfP/AeIBrZ5/6kY6vnjPMl46b56ZHdus89yCJZ5OPzZLS323ZHaT+rnHVhc91gKYz3vyTJznol7iOfLcdHhLOSB1xJrzCXJvmEn0W3yHQBtz338xnHLmQsck9icbJLnjuAt1MdJYLDb73jr0MiryzlpeJjyxz5eYFqb3e2H9pjVNmZE3ksVFsBaa27mOS58vElop3AnfkQ9+e/gS/Hr0DeF/iAW3uIz3E/cQ8+AB8sv7QsL9+Rz7/DAfvAYXMPoR7uXgMY7wD8CP4eES2km29iFv0yBPL81jPgRtnboH16IAsrrhLuGXjb5rJHNhsBnzx4kSeXn/oRKTSLTlnK7wJlLxX2Fmcnh5YlJY7y1OWNgJlLfKXxOI9NHFP+Sno+1C/pVeze3Efoto091UmsdpfsrtS9x2EpJtYaZ8v/AIUnE/8AKnGVHh+2q9q20mGKmHs688OXwXZXg+0B4+AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD6j/wDJ/a96riDiXhibWLq1p3tLL60pdiSXmqkX/RPruvHt0pxfWLPzo/Bo4jhwx6buGtQrTUbetdfQ6zcuzFRrJ08t9yclL+ifo3JYbXcwPz6/Cq0ienelTUa3q1CncyjWh4qUVl/1lI8jPq38N/h78Zp2t06L3jOhUqdNn2oL/mn8D5SABEgACRgKDBKRKQEIYJwVKJFU4KkiUtiQIxsT0HUBUpjmQiftIsSgR1C5hlFQEU29i9Tpt7JBYqts5W51XDzaqR7be72NBa2+JJrLZ0+iUX6yPbTywWvU+DY7QUlmL8T1fQLKNSnHCTWOh5PwbGUHDsybS8Mnr2gTrU6MZqPZ22bXMrXW/o6L2o5SIraJUS2iYzlWrPNSpKXm+RGparZaBpdbU9U1KFhZ0Y5qValTspeHi30S3YRTW0ytDfsyXuMaVtVj0PCeOfwmtZeswp8H2Vr/ACdQk+1Uv6LlK59yacY+/L25cjL4f/ChpzjTp8S8HwlJv8ZX0+4xheFOaf8AbGlmTq/TVwe+IeHHd2tHOo2CdWk0t5x/Oh968V4nz7aSU45axJbSXcz6X0H02eivWkoz1ivpNWTwqd/byh/zR7UF72efek/0d07u8uuLOAb+w1uwqP1lza2FWNWVJvnJdlvKfPHQ1Z4/d6PR9RJ8GVedW3NHQaN+Uic3a1oOXZk3Ca2cZbNM6HRH268IQ9uT5KO7Zxcs9n1/pmWs4+0+AXngnRP/AHCj/YR5v+FF/wDZ6wf61T+4ei+juSnwLock8p2FHf8AoI87/Ci/+zlj+1U/uHJPEY9Lf++/e/y+ZJv2jVcVvKtn+rL7jZTftms4r/J2r8Jfcd3H4aOtv9VgaNnte9nW6XH2c7c/icjorfbfmdppEW0s74XQ6MXz/U34mzp7xbzltb78iKsI9cR7y7DCXZ7OeibWyLVVKKbTyltt0X8YNkcFqxUx2W9sPuwYtdyzJ58f8TJlmKbi2+e3JmPWccdnGcorCuS172eJdJk32W5b+G50K2lzXLr1Od4ib/yh0jGXir2d1t9ZHQuW+2X9qZUo228vmlts+ZHZ3afLHMLtrZ79dngLCeNsvbfkGFN8Yjt495OG8fq8sEP2srOV3rchzWFjKfJvvKiVL2d2xjvxutiGsrCT2HPlj3+8IhSaee1nfkt0S5dGljmUuOGm/kiGo5eefe+ZRPVrnt1eSVtvjGSEn3JY5eKC2Wc7/wCIEpPGfkUpNtZ9zXcO0mHjOW856YwEMrr1J2eE3yKHJ8+ysZ79ypPntuEQ3jCX1RjH5qefDoQlmLz3k5z3rpsBPZx7yHz2eyJlkYytviUUtSe3UlLrnkHjC6/cIrbGPkUTjfx8w0sd/gRvyw/f0Db5Y+YgmP3cgsJtkJ8t09iW9uW6XQqG78V4B83hbeRLfNpe7kRPDawsdMYLAWOf3hPuCWHhPHPYJfrJ5KhvjGGu/wAAsJY59zDWFjvIWMrOxkipYfPcnbu5EJ5x3Brfb3lgqbx3t+YWeb+BTtlOS27l1J3a59DOJVWVnfKJTae2zKUnnmviVcl3eJlESsYRMOfPcjqsdNiqLWUkbIxqqL323SJjz5LxKVy5P3oqgnndLwNsY1bvd7Ou3+hJ/I13CK/zGnhflJfcbK9z9Br8m/Vy5+RruEP+wlyTdSWHnyOjD5Wu+WbWeMbZ5mPLdrdvdZeDIrZcm8vn8P43MeaXbe+HjqdPG1ZLeF15krffPUlpNtpY6lKZ38TlzVxWxdgi3Hluy5D5HpcLg5WTSM+1XIwKXPuNha80elxvN5PLc2C3SO/4R/0X+kcBYc0d/wAI/wCir9pmPVfIdN87udB/0j3Hvmn/AOg0P+HH7DwTQf8ASPce92P+h0f+HH7D431nzi9/07zkvgA8N6qHyPn7jd/z64f+8l9p9AvkfPvG/wDptx+3L7T3PQ/qZPJ9V+XFzOtY9fU7LTTeU+85y85s6TUYx9VScd06MH7+ys/M0F5FZZ9JjdRwY+/u0lxnLXUwK3M2lzHnsa+tBdTXnXRhGDP6xVlNblU4biMN+SOHlrqwI7ox9QWaDWxlOntt9havI5ovHczgyvu6sXJXH5aRaZfulisyzgWIRRejEphHcvxjsb8MGvLJSlsVpEpFWNjO4pKqtfZuab7pL7T7P9H1f6TwJoVbOW7Cin5qKT+aPi+G00/E+vfQxW9d6MdGln6sakH7qs0eJ6pjrtrt6a+XU3a/F5NRJe0+qNxcr8WzT1frv4LLPIdSnlv80yOZLfh7u8l426/Eop2abe6J2zn7iNlsk13bErGO/wAcAJbtPfPmQ3zWPi8jD8M9xO+2c48CKjC5Zee4b9X8yebI2zsnv06kDx5Mhr3DOzePl/iTnb7gG+23kQ0muhP7L3fgQmsrO/TbYCmfLkWJrwL8s9nkk/EsVF7XT4bgWpppP7jGqGRPPcs4Mapy3SAt7LohnfYdeeSAJI+HxJ5vPLwCz/CAj7h05bD3fEeeEgDe/MPddQ3tzXZ8iPHmgHIYzyRPiyG+958mAe5DSzyyycvwY5pKWPICPjgLbkPe+XTYh+K+QE7Lcpk9uRTKWEsPL7y3OeFsvIBOXfj4lmUtu9eYlLbnhZ7i3KW334yAnJY6lmpNY39wqTxvj3FpZlhteQGr4v1yhwzwtqXEd3iULKi504S5VKj2hH3ycV7z4f1C7uL+/uL67qyrXFxVlVq1JPLlKTy2/ez3b8LLilSrafwba1Fiji8vcfptNU4vyTlJr9aJ4EAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABMW4tNNprdNH6a+iHimPGvoz0DiZzUq15Zx+ktRwlXhmFVY7u3GWPA/Mk+s/wC+N8w1fgC8rb5eoaepPyjVgv+WSX7TA9W/Cb4bjr3o6vWqfaq20fpEH4wTz/yuXyPz9u6ToXNSlLnGTR+pevWkbzS61GcVNOL9lrKe2+x+cvpj4arcNcaXtlKDVONVqDfWPOL+DQHE4JQwSkACJSyVKIVTgqXgVKKJ5BVOO8klpjBBAbJ6BkIhDqSQFETgciVHL3yFiFkrjTzu+Rdp0pNrbCMy3t91hZCselRb3awjY2tq5NYiZdnZOUl7OX9h0WkaTOpNJRz37A2wNL0yU2l2TveGOGp1pRzTePLmbbhXhaVVw/Fv4HrHD+g0LGlGUoJ1PLkVja1/C/DVGzpQqVqa7X5scHWUaHLCJqyt7S1qXd3WpUKFKLnUq1JKMYRXNtvZI+fvS56fsKto3Acsc4VdUlH4+qi/wC0/cuTDF6d6UPSfw3wBbOlczV9q8o5pafRku0ttnUf5kfPd9Ez5O9IfHnEXHOp/S9au26MG/UWlLMaNFfqx7/F5bOburivd3NW5uq1SvXqyc6lSpJylOT3bbe7b7y0VAAAC7bXFe1rRrW1erQqx3jOnNxkvei0ANzLijXpz7dxqE7qbWO3cwjWl/WmmzGvtZ1K8h2K90+x1jTjGnF+aikn7zXgnbPLZ+Lydvb3XX6v0X/Bvv4aj6EOFK8HtCwjQfnTbpv+yc/+FEv/AEZs33Sqf3TQfgOa+tQ9F13oc6idbSr6ajDqqVRdtP8Ards3f4VVRU+ELabfKU/7p43Jjrks/N9h6VnLnx5fl/D5hlL2zX8VfkLTyl9xXSuPWVMFvjReqt7HvlGT+w7MJ7Metz/qsLQcesfmd1o6TgnF7eZxvCVJ1ZOW56BptBqmtjoxjwuoy3kvdjGyXLm/EsVdop9rbvfkZdaOEmtu/ZIw7jtP60Vl9eRm47WNPOWljlvnkYlbmsZzHn4GXUa3XJbvBh1nGHsqLfdv7uZYxcnxM/8APukY/wBst/6SOiw4t5WV5czn+KsrXdJe7/nCX/MjoGsyfXfqwl8Izzy370wnzWO1lZIWNlmT/wCpEMLk15thgmMlLzwJZ71jvIXZ3xlN+Ai8vL3aXVgTFZzthdWyE2mk8+7qE4+PgQm8YS9nluVFS58sFLxlvtYS6vkHntvnnOM94aTw237kUSniOM5+1EPPful3kb898ruIXV46YAqf1cdA3h4bXLmyE8eHkMrOcJZ5hDm8tZeAuXXuIzh7vblsQ8ye7axvjIRV0TysBZ5rdjHN7YCeXzx4dwEtPOz3fQcsPOwb3TxnvIk9pdxRPgRLCWMrYLGcYRD88FE559EE08PJSkuuWvAnO778dRBKbccpJk4f+PQpw+4nOHnHLuKiZN42fMhdnDwu8LKSzld4T7l5dSiUlz5Y6hNJ+ZCbxnbPMPl+4sRPay/kT4FKST8cbk+HUyROFuuYx7RHiVPfDfJGUEdzZVHGPHuKXv5FUeRlEpFPuKsLGce8pWzXyKlnOTKIqjthPry35BPG7yRHnu1jqVLLazlfcbIxqqOVjkiVsls8bEYxyW32kxa93ibYxq3e7WNw8ZbpS+xmv4Q30Nc/ykjYXuFY3G/+qlt7mYHCG+hxS2bqSOjD5Wu+WbWXtuKSSaMWb3y+SMivtvnddCw481jL8zp42rJbe+d9+oXImSb3yFhLkd/E5c1UeZdg9y1F4LkOR6XC4OVlUehsLXoa+jzNhbdD0uN5vI3NhnKPQOEf9FX7RwFhzR3/AAj/AKKv2mY9V8h03zu50D/SPd9575Z/6JS/YX2Hgegf6R7j321/0al+wvsPjfWfOL3/AE7+5dAB4b1UPkfPfGz/AJ5cfty+0+hJcj5641/0yv8Atv7T3PQ/qZfs8n1X5cXM1JSlp9Ft8lKK/rM097yZsaFSLtKkM+1Co3juTX+DNZdy2fM+lymsq8/hu8Y1dztnkYFbBm3L8TX1ntsaM3Vgx5c+ZKeHjqUVH5FMZpvc4uV1YMqKzT2xkx7uOKEvIvUXyWdim7/JSb32ZwZeXTPDjb3as9upaismVqEPxufEs04nRjhWq5K6cS/GOwpQL0Y4R24cXs58s1tIYLrRTjcmWC45Lb5n1N+DvXlX9HUYN5VG7qwj4JqMv7zPltrc+kPwYrtVOEtRs1zpXcan9aCX9w8P1bH+nL+bv6W/FXq1dfi3k1Ff6/8AG5uKu9N7dDT3S9vHxPAdy03h4zv5kbrZ9CduayHusP4FEYeN17sE89tlgjGOWfHcJLPJgOuFv5IleTz/AB0IW3f8CZPPMgjD8PcOmOW3XoGk/Hy6B5b3WSKZXVr3hP8AjmMteHvJy8838QKXjfkmJeOH5slNpc2Uvn5gNsdPHYsTee/4F6Un1f2Fiq855AWp+TyY1Qvz3XLbzLNTrnAFp8yH125Ev5EPl9gBpbYwMPljcdeQfL/ACOv7w9t8dSdiHyALHPv2G+eTyPn5hY7twI+Qz0bRL3IwAXcsELv3J2zkhtt55vzAPddC3KTIlJLmWpyyBMpZ3x7sFqU+aWWRKTS26/Mtyk+f2IA3t097LU54znYicsLnj3Fv6zywG8t2Y+ralaaLpF5rF/LsWtlRlWqPvSWcLxfJeLMpLvPEvwruKvoej2PB1pUxWu8XV72Xypp/i4++Scv6K7wPAeJ9Yu+IOIb/AFu+l2ri8rSqz32jl7RXglhLwRrQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB0Po44pvOCuONJ4osU5VdPuI1JU849ZDlOD8JRcl7zngB+qXDusafxFoNlrmk11Xsb+hGvQn3xks4fc1ya6NNdD5r/DA4Idxbw1y1o5lBerqNR6c4P4tr3o1v4DvpNjRrVfRrrFw1GtKVxo85vZTxmpR8M4c4rv7fVo+mOO9Coa9w/c2Ven24zg4yXfHDyB+YUoSjJxezTw0FE7b0scJ3PDXElxb1abUVN+1jGe6XvW/wAV0ONSx0AiK28ScEpr3BhUdMjGWCQptgpZPfuG9yAR07ySOhBDCXUuQg34F+jRz0yFWqdJvfkjIo0cvCWX3mTRt3KSWGzY2tk29kUYlva5fLtM3Flp7nJey39xn6dpjnJezz8DsdD0GVSUU4e7ATbUaLosqso4hselcKcLOcofi/kbjhXhbPZlKniK5to762trXT7SdWc6dGjSi5VKk5KMYpbttvkhpLWPpGlUbKklGKc+/Bp/SJ6QOGuArD1usXXrb2cc0LGi061Txx+bH9Z7bPGXseXelj0/W9n67SOBexcV94VNTnHNOH/Ci/rP9Z7dyfM+ctSvr3U76tfahdVrq6rS7VSrVm5Sk+9tlR2HpQ9KHEnHly4XtX6HpcZZo2FCT9Wt9nN/ny8Xt3JZOGAAAAAAAAAAAAD0v8HT0iv0dceRvLmX+a76Ct71b4is5jPbL9l+ezezPb/wkuJlxLo9nHRZQvbepiSlbVY1Yxgt8txyk2+meh8in0J+CbmXD3FcU3mNa1a/q1f3Gjk4Mcsu77vR6L1Dk6fWM8OB061rxrRU6cotPdNGP6QKvbv7W2i8+qoLKXRt/uwfRN5badd0pwvNNoVqu/ZrJdma82vrfxucdH0faTU1KV/eVK1zVlLtdl7RT6LHPCEw0359Z35d2ThOCtOnG3TlDoumenkdrToulTTW+F0/ebevp9tZ05QoUoxil542MN03GnHCSy+ajsZyacWefddtfdPD7UcY588mFWliLxFry3M65efaf2GvrZbe+6WMMrWxriTUnhp7PfxMWo3l8lhfDYyJdlxynjpszFqp7pc/qtYxgrFy3FX/AG1o3RK4XPp7UTfyxlfV8s5NBxYn/LOkSWFm4W3j2kbxYct/jjkCqk008bL5kSSlDbv8sjL+tnZd6I8FnIYilLnnHmIvOeys+JEVnntknG+U/ZxjDKiM747OM7Iqk44x2kiEt8tZ8e4P67bAbtrMt+Y7TSTTTx4EJ5eXya2GeeWvgUT18PAp7lyxuVN81nciUV9bOMhEJ75znBOzT/O8CjGHv02JbeeW2AIw3nfHdglvLw+XzD5tYTx0GEnvjYIlvv5By9nm99uRS8Z3W/8AH+JLeUljPgBW9+WEUtNPDy/cUt4xjr0Ku17K26FCW370En2dl8SG0nt3ErtYw3lroUMYi0+fTBGEmunh3COz2TElhPml3sQS+eVsiMrKzh9V4B5Te2PeEunPvKip7vfl3hSXR7kLZbIhN7ckVE7ZXsvHzJe3yZCeXjrnmElu+8olJNbY5/EnHckQtlz5k/HBkhFLo+ROyec5RHf3E8sLvKJ5h4yIvvIWOWTOJVXawkyc9yyUrYqT35GURUs7ExW6KYp+BOXnBsjGq3jlnl3FS7tvEoT6P7SpPMcJ436GyMaovWvoFxs/yUvsNfwjn+Qo4/TkZ19J/QLh7N+rlz8ma/hJr+RIbfnzOnD5Wu+WfXSXe1jdt9TFnLm+7O5kVm8tdMZMaWMrbbn/AB8Dp42rIkl+c+RGU+7JDezWff3hcju4nNmqTysl2n0yWo5yi7A9Lhefysqj4GfbNZNfRM+16Hp8TzuVu7Dmj0HhH/RF+0zz7T+aPQOEf9EXmzHqvpnTfO7vQP8ASPd9575bf6PT/ZX2HgXD/wDpHu+9Hv1v+Rh+yj4z1nzj/wA/w9/03+5cAB4j1ES5Hzzxo/53X/bf2n0NLkfO/Gb/AJ1W/bZ7voX1Mv2eR6t8uLirepi6r0v04ZXmv8MmDeSw3uRdV/o9/TrN4jGXteXJ/It6jLsVJRf5raPq+XD328np8/axr7iWcmvry5l+5qLLMCvUW5ycmLvwq3UnuUdvHL4FqrUSLHrcdTi5I6sK2lColvnl4k3NReree410bhJ8xXuU4NHDlj7umX2YV7T7dOUu5mJSjuZkasZwnBtbmLRku1jJ6vHhMsZXBnnZlYyqUEXexiOehVQw8dS/2W1nojsnH7Obv92K1joUNF+ceZZkt3g5+THTfhltbfge9fgr1k7fXKGcNqhLHfh1F96PBX3nsX4LlRrii/pdraVhN472qlP97PC9Wx3w/u9HpL8b6Fl9V+Rp7tfjOT9xuHyexqr5Ylvh7nzD02Lh75W3iQ89M5ZOF4L3B5+8ohp45Pn3Dp1+GRtnoSsdzAhZXPPmupK69/QjlvhJDk+fx5kUeX9ZoLdrkQs52TyMrDxggntbcsIpzl4z78ch7mUuS6yznpkCpvu5EZeN38MlLllLfPmQ2sAS5NJeznzZYqvL6eLRcbwlktVd+/4AWpY7/kWp88v7C5LPiWp7cgLbIf8AG5PP/FEdQDaW7IyS8fx0G76MCFnuY6ctx55+AaxthoBv4j4rwxzCI5gS/wCMJojmsYD5FMpYeN/3gTJ4WHktVG8bfaRKXXBanPD5gTNvO2S1N+JTKXQolLwy/ICZvO/awWZS8d/Bic8LPQtLLe4E7yeWVJEY2wVLcC1d3drp1hc6lfVVStLSlKtWqP8ANhFZb+R8Qcc8Q3XFXFmo6/dpxqXdZyjDOfVwW0IePZiks9cHvv4VXFv8ncP2vCFpVxcahiveYe8aMX7MX+1JZ/oeJ80AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGRp15dadqFvqFjXnb3VtVjWo1YPEqc4vMZJ96aTP0b9AnpGtPSXwBbawnThqdDFDUreO3q6yX1kv0ZfWXvXRn5uHo34PnpKufRnx7Q1STqVNIusUNTt479uk39dL9KD9pd+6yssD6l/CZ9HlLWdIqalaUM1qcW9lvKPNrzXNe9dT4o1G0q2V3O3qxeYvn3roz9QZqx1zRqda2rUrmzu6MatCtBqUZxksxlF9U018T4u/CV9HVXRdXqX9pbuNGrJyikuUubj969/cFeDDHUlpoJAEOhL8QkFR7glkns5L0KUvICzGGS7Tpb5fwL9OlFeLMqlbN89vAgs0qKeNjNoWrlzWEZVtaSfJG40/TnNr2WyjCs7DOMRwu86DS9KlNr2Xj7Ta6Tozm0+zk7vh3htyccw38gjS6Bw/KUo/iz0zhnhmNOMalWGEvA2OnaVY6TYzvdQrUbehSj26lSrJRhCPe29kjxz0qfhA06Kq6TwFFTkvZnqdWHsr/hQfP8AalttyezCbercf8ecL8AacparcqV1KOaNjRalWqe781eLwvPkfLPpS9K3EvHdWVvcVPoGkKWadhQk+w99nUfOb2XPZdEjiNRvbzUb2re391WurmtLtVKtWblOT722Y4AAAAAAAAAAAAAAAAA+h/wP129K4whz9qzfyrnzwfRP4Gy7VlxlHwsn/wB+Fnl6ZWo/jG/F743KXSeNs+RsqlP8Y/MtuntlppeRi2bc7q9HaT6OPcaC4T7OVhSz06nW6zTTUs5+rzOVv1iDTWVz3+QNtReSfZ579/Rmur4Ty1y3T22M+9322be2UYFdS37LXdnl8QMaT69nZck9zEqt4eJNxazu9zIq75a5eZjVJc47ppt4f2II5bipuWp6Q8bfSVtn9aJve/KfVrKNDxW4/wAoaVhPP0lb+9G+xs21svEFQvZltHH3sNJ837skvnnOenuIimmksp9z6BCOW1tjvJxybeW87+JS2+fZ3fRdSY4cXJJxj5lQeM5ay3yx5h4T6LPMJpNJ49xD5vL3KJeMLOOZGG3l7e8hv2nsG09846AOf53w6hPbm/JMPC2fJEbNbbZCJwk8dXzGHl/vKcPn0xzJzy3TS5hE7Y3+whPk+ZG+M459Bn3PuAq2T67kPnkc3sUvGXuvLIFTedkunIPtOW7zzXkFy2XMb4xn4lCLae2Hjqgm3JvL3JfjhLPRlPJY3RRK2a3b/cTlpYTZC3XdhEppiA/ILGWObXMhrm0+RUTzy88+eBLOE+nQpjstir83f4liIUd1nG3TvJXPvyR7ycY5b+JRKeJefJE5xJZKYvfkSllrbJUVvHJbEeGceZQtn3ZKu10zyZlBMsrkVb7ZRR9ZvHMlfaZRFWUSsLkUx2RPdkziKlv4NbFUe7OC2s9xUm9tkjOMar65fUqUm2s8yhZwVp+O5tiVRftKxuG916qePDY1vCT/AMxx7u3LPibC+bdlX/4cvsNfwpn+QopL8+XU6MPlar5Z1fm32Y782Ysmkls9jIq9rLw8ruRjSfty3R0cda8jOPfvkiPzDWUuzzKe7bGTv4q5s11NYxguRZZi8Y8C5Bno8NcHKyqTwZ9s90a6k9zPtnuj1OJ5nK3tg90eg8JP+aLzZ53p73R6Fwg/5pHzY6r6adN87veHv9I933o9/o/kofso+f8Ah1/zh+X3n0BS/Jx8kfF+s+cf+f4fQem/3KwAeI9REuR86cZv+c1v2mfRU/qs+cuMn/Oav7TPe9B+pl+zyPVvlx/d5vrb3kYde69dZ0aucyjH1c/NcvisfMydbe8jm43Hq686UpYhVWPBSXJ/d7z7bs7sXzk5OzNVc3G73NdXuMPdkXdbGd8Grr13l45HJycbv4+VlVrnCe5izut8tmFWr88sxqlfd9Tg5MHdhyNo7x88lErzO2TTOu1uslMrjx+JwcmOnXhnttXctSynuU07j8a9+ZqpV090yFXxJPJlw8tx9mPLhL7uss7hY3fmbKnUyn0OTsbrlubi3uk4rO56vHyyxwZcdlbKpvnd/vMapz3DrJpbv4FNSon4GHJZYz49xRLyZ6Z+Dfdu39IdGgsJXNGrSf8AUc/7iPMHLpvsdz6CLqFr6StKnNpdqt6tZ75wlBfOSPD9Sm+DJ6fS+2cfWXuNXqC9vobXmzW6hHd55Hyb1mDu0106hvOzwFhvGGsPlyGevNddiodXjkRnxWeuOZDzz5ohywsuT7yCdlyeBn3eBGfD5kNvo2u7uCpb64I65XPk2iHyIbTe2/uyQJY5qK95Dxs+4jLaaa3Ib7vlsBLeVju7v42KXy6kZeN87d5TJoCcrbrgtT552+BLl0+ZRN56b/NAUTa8C230eCZSXfv48iiT36fECG+aCznI64KX5AT9gzvzWB7t/MjbYB8BtjbA5Pp4INe8B38yOS229xPJZbxjvKHN+HnzAiT7vsLcpY6CcnvhosTlzey94FU57Yyi1Uk+/CIlLfbfBanLOdt0BMmsbtIszmufTpsROphbR92ShZe75gOb8ytYIWPEqSTe7AR7yLq5trGzuL+9qKla2tKVatN8owist/BFxdx4x+FTxf8AyZw/b8H2dXF1qWK95jnGhF+zH+lJfCD7wPA+PuI7nizi7UdeucxdzVbpwbz6umtoR90UkaIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD6y/Al9K2y9Gmu3P6VTRqs3/AEp0M/GUf6S7kfQnpQ4UtuJ+Hbi3q08z7GMpbrua8U9z8z7K6ubK8o3lnXqW9zQqRqUqtOTjKE4vKkmuTTWT9BfwavS3a+k3hP1F/VpU+JNPgo39BYj66PJV4r9F9Uvqy7k45D4r9JfC91w5r1xQrUuylNp4W2e9eDOVzhH3F+Ed6NKWuadPULSh2qqT2j1XPs+fVf4nxdrOl19Mvp2taDTT9ltYygrXYbK4U5N8jIp0VjdIyKFLPTYCxTpdyMmlQy9/gZVG3bwkjPt7RLCw2/IDCoWv6ps7Wycmsoz7PT22m15I6LTNJk2vY92AbavT9Lc2sxb8EdZo2hylKPsZfd3G50TQpSa9jc7OnaaVoGlT1TW7yhYWVJZlVrSwvJdW/BbsIw+G+GnKUfY+Rd454/4S9HNq6V5WV7q3ZzTsaDTqeDk+UF4vfuTPJPSX6erq5pVdI4FpT060eYz1CaxXqLl7C/1a57/W5fVZ4dXrVbivOvXqzq1aknKc5ycpSb5tt82B2PpK9JfE3Hd0/wCUrn6Pp8ZZo2NBtUodzf6UvF+7HI4sAAAAAAAAAAAAAAAAAAAAB9F/gYb0OMY/qWf21j50Por8C1+1xfH/AHVo/wDmqgj2WrD8ZLzLThjkZVfKqyx34LTW76kZbaXW44g+jcTj9Q2g3jLzudtraXY67L3nE6ltFvsp+4ixo7tpReHzZra8vZ3k3npkzrzKX52/2+Pga6u+y2+/bGArHc87Lyxkx6jXZwklz5FybwublnzLEpNpbv4hHN8W4jd6W0tlcrr4o3UnnMcLBouMU/X6fJYWLhbe9G9b2bTT32yBMsPeO2fmHnPf0IzmWXFbvPIhv2cZ8AipZbzJpFLee/x2JTWMLHh1Jxjomv8AAonuS5EJtdMteBS8Lm3sTlNb5yUGsrO+A0km3nfbyIy+eXvzIT7879Ahye6X7wm8c/mS3vJPl5kZTeHjC+YEtrHPZcyHl7rGSE1lvbmEsZ677vIRPLDS3RDeza6ILn4jZPfC8gJ54y14Ect9vMPn1265zgLnzAqxnGWQkug32TSfmTJrtYyigvq8tgn7Oz69SM557IZ2658Shs/D7AlhZ6kJ7c9ydsbLkIg+fPzCxyWEvFhS6PDGPayUT0eOhK93LGxSuSy/HfqM9OZYiZNY35hZe2HghZT57E9f8SiUOXUjq/ILGeZUSnnr7wm1z5BpcuWCU+mDISm0svPuCRDe6RO7fLL8TKCrvXQlPfCI6J5eCU0ZRil88htY3+IW+ctd5G2MmyIrW6XMq2bwvkUb4azgrS5dDZGK1f7WFxl7+qly8mYPCTxocMp7Tn0M6/f8xuF09VL7GYHCWf5Dg8r8pI6Mfla75Zdd43e22OXNY/wMefPx8jIrPDwk8oxaje+Nmb+OteSE/ae/Ib5TfIp+xhS7lnuO7jrmzXIvljmy5B8iyiuL36Y6HpcNcPLGXSZnWz3NbTe5nW8t0enxV5nLG9sHuj0ThB/zSPmzzawluj0bg9/zOPmzPqfpsOm+d3/Dn+kPyX2n0FS+pHyPnvht/wA4+H2n0LT+qvI+M9a+bH9/4fQem/3f8/yqAB4b1FM/qs+b+Mpfzir+0z6Qn9V+R808ZS/H1fNn0HoE/qZfs8f1b5cf3ec63LeRx+pS9pnV63PeRx+oy9pn3HH7R8vyeWDcVnNZb3XPxNdXm+9l+vLDMKtI5+WuvilWKsmY1Sb7y5VZjVGeby16HHionUfRlqVVrkJsszZ5/JXZhFz1zCrGNKRT2jkyuq6J7ttaV8dTa21zy6I5i3rOM8ZNlQr8jZhz2JePbpKV14l1V+0ub+Bo6VZrrjYyadU2XqfYnC2bqp7ZTxzR0no1u4W/G+k15vEad3Rm2+nZqRf3M42NXK548jZcP1nHU6O7WW0t+rTwcXU8nfx5R0cWOso+8X9b3mBqSTTeEtupl21WNe1pV4tONWnGa8U1kx9RWzzufMPTarrhLL8UQ31Yk1uuvRkJvuKh4EPO/eG989fgUp4SSb27gJ38fALPPGSl803uJPK3fz5kBeGX5EPxWUR8yHhY+/JFJNYfJe8pb8NxJpvDREmnyQEN4KW3n/AOW3Pf3FEpd+fiAk8/9C3JvkJS8n7i3J9MfICJPDxjBTKXuKXLbn8CFnnnAFf2h5XeinJOV4e5gT16j3hb9SHugGfEZa7yG9sYKZtYwAb2yWpyw+pE5cunkWpSW+PkAnPOd014llvL6iUu7kWpPrsBMp78yxUn06ipNraLfxKNwHiypLqETFb5AmK22K4rASJxnYCzqF7aaZpt1qmoVVStLSlKtWk+kYrL958Qcc8RXfFfFeoa9eZU7qq5RhnPq4LaEF5RSR7t+Fdxh9D0y14Ksqv425xc37i+VNP2IPza7TX6se8+bgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAG04V4g1jhbXrXXdBvqtjqFrPtUqtN/FNcnFrZp7NGrAH6B+g701cM+lPRoaRqlS307iRwca9hOWI12lvOi39ZPd9n60cPml2nxHp59Dyl63UbGi5wm3LMFum+q8e9dfPn8bUalSjVhVo1JU6kGpRnF4cWuTT6M959Hf4UHGWiW1PTOLLWhxVpqSg5V32LlR5flEmpf0k2+8DzjUdCvNNuZUrmm1FPEZ42l/HcTb22/ZUcnu+pcZehjjiPrbXWJaBeVl7dpqlvKMM93rI5hjxbT8jn6nAFG4hKtod3bXtHpVs60bik/fFtr5gedW1k20sG70/TJZSUefVnSW/Cl5byxKnFtPDbzHHuaRuaGlUbGj6+9q0aFJfWqVJJRj5voBqNK0aTaxF572jtdB4ey17GPHBx2rek7gzQYunbVamrXK/Ntkuxnxm9seWTy7jn0qcT8TwqWka60zTZbfRbVtdtd05c5eWy8APZuN/SvwvwdTqWWkqnrWrxzHsU5fiKMv15rm/COeTTcT57404v4g4w1L6brt/Ou4t+qox9mlRXdCPJefN9WzQgAAAAAAAAAAAAAAAAAAAAAAAGz4b0PUNf1JWOn005KLnVqTeKdGC5zm+kV/gstpAYFCjUr1OxSg5Sxnbou99yPo38DmlQoT4rjCo6lX1VspSW0X7VTkufv8eR4Zqk7KhJ6dpUnO2g8TryjidxJfnPuj3R6Lxyz038GfW4aPxFqVtUmoxu7eC3fWLf7ybZ9upt9E3Evxsub3LMpbczClqFObk+2nnLLbvE2kmvAqLetPNN45YZw+qVMRcE8NPffkdfqlwnTw2mpb89kcXq1bp3sg0l3tLm2s8zX1ZSezi0+ixzMy6bUXun7W+3LwNdOXZjlPbk8EVbm0mkvfsY03s22l13e/vL7cZPKeJeCMatySSTeM9Abczxm1GVk1n8unhruwdBnDa+Bz/G7Tt7XH+32Zvt3tzZUT2ot9+wbxH2Vu18iH7PLk+8hbNPZPwAlSjs87lSfXCxy3LfZefFe8lv2sNPYoqWMb45B+zFpv5lD54fMRznw5thFUd3h7dBJ7ptPyKHJSeGxHbp/gBVlJ7Pr7iHlPfL8iPJp+8nKXKT36JBErZ8s+CHRNopz0QWFsscuqAqbytuvQpi8t5ST6bkNrG7wTvy328AKo8sJjfKXf3lPJZw92S3h8gKvLJEd2889ugTzuRldUsFFXf12IWM46Z5iOE99/eOmGsFBRw+vIPmkt/eG29sciG9nnuEDPyJytskOWPcJfV2e5USub32aJXTdYXNEdPId5UVJPOcL4kN8vEjrnZY7ycrpvsUHu9yVnOM/Apf1l4lW+NipVWOi5+fQJfnMiPVNJBPZpZLBLfPPMqbeV5FLa5v5B8tuhnBPOOc4x3lW2cIoT2xnYlZTyZRiq2fL4ZKk1vnoUp557hbr95lEqtN7MqWzf3FHXm/gTnozbGKi/a+hXC3/JSz8Ga/hPL0OK/wB4+viZ1/n6BcPp6qX2GDwo/wDMMWnynL7Tfj8rC+WZVeVleX8fMxauzfVl+u2pfWeWY1XK7XXPzRvwrXlFLYjlJvJTnbn12DeG/HmdvHXPnFSlguxfzMZPmXIy7z0OKuLljLpPcy7eW6NfTkZdCZ6nDk8zmje2Et0ej8HT/mMfNnl9lU3R6PwZUzYwx3s39R78bV031Ho/DUv5y/d9p9Ew+qvI+buGKmbr4fafSMPqryPjfW5rLH9/4e96b/d/z/KoAHhPUUVPqPyPmPjKf46p5s+mrmcYUKk5PEYxbb9x8scZV162pv1Z9H/09jvPL9njesX2x/f+Hn+uVN5HI308zZ0Ot1t5bnLXc92fbeMXzWt5MG4lzMGrIya8jBrSOHmyejw4rdWRi1GXKkjHqM83kyd/HioqMsTZXNlmTOHOurCIky22TJlts5cq34xLlvnqZlpXyvEwGxTqdieehpt02SOgpVPFGTSn0yam3q5xuZtKfUwubORsYS3W5m2FZUryhVfKNSMn5ZRq4S2TMinLK+01ZZ7ZyPu70dXKu+BNEqxl2sWkKbec7w9h/OJtr5ZpZ+ZxH4P9/G+9Gtp2Xl0Ks4vw7WJ/3zurpZo4PEymq7I0Mtsvl8Sl43W2xXX9mo1lfuKJPDT+CYEPKbWX5EZwl4kdP3kZw/L3EB7Lf4cw845rfu2IbXRopb2ayl3hUvDW+SG8dMEPYpb8iCXnHL3FEmG/f7ihvd7J/MA3lFDkk+n2ByLcpN9QE5rq/kWpS7kTKXRtFmUvIA22t8fAqi88sPzZaf8AGUSpPvyBdjl7tde8qys80UJ5W/wwVbPnkCrZLd4KHJLfDXuJeEsb+RQ33fIA5dfvLc5pp4fwDlh5WV7izUnvjLz4sBOa3f3FqcnjlhkSnzX3FmcvgBMp77vcszqZzheZTKTfJhfxgAkveSviFuVRQBIrRCXkVJNgTFZ3MHiXWrHhrh6+17UpONtZUnUklzm+UYLxbaS8zYxT2SPl/wDCU9IK4g1hcL6VWUtL06q3WnB7V662b8Yx3S73l9wHl/FWuX3EnEN9rmpT7d1eVXUnjOIrpFZ6JYS8EjWAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACqlUqUqkalKcoTi8xlF4afmUgDbvifiV8+IdXf/AOdqfvNbc3Fe6rSrXNepWqyeZTqScpPzbLQAAAAAAAAAAAAAAAAAAAAAAAAAAHScB8H6jxZqEqdBq2saGJXd5UXsUY/fJ9I9fBZaDG4O4Z1LijVFZWEFGEF2ri4ntToQ/Sk/sXNne8UX2k8M8NVuGeHs9mosXVy9ql1LGMvuju8Lkl4tt5nEuvaTwxon+TvDlL1NvH8pNv8AGVp9ZzfVvu6eCPLL67q3NWVSpJtsjPGKKT3TMzR9Vq6XqsLmlJx7Ozxsa6EuhRUWJZIz2954c40VzQi5VV2msPJ1VnrsZqKjVysc8nzRpmo1rOonGbS8zu9B4hlKMc1PmVhY9mur2NSlhOXZb67nN6lNyk8vfPNcvea7TtYjWjJOfR9M5/jBNzXjUf1ljPMIxq9RqOcSy/1fuMGbTz28tfAuVpdqfsvbp3eH2GPUmmstrdbZ22RAbWHlJ42WO4s1JZhl4znP8eZXU3a57e4sVWstY5dHzfcUczxxlWVHEs4rLPwZ0KznPa8/E53jhfzCk1j8qvsZv0+/bbdAVx7s53ZTvKXXC6ZKZdrs5znbfwGd1l+4ot3VX1dPKi5Se0Y97NNd619FvJULh1XUik5Kmto7ZNnrF5Cy06d4qfrJwaSXLdvmzl7G+hdahcXd1Tppzj7Sxt2UsfcB0NnqUa1BVqU3XpJ4kpLEl38ufTx8zYKUZLKaaa27nsajRJ2E7SdKzhKlmXaalnZvrv02MjSKvbhVp8ownt4Z3+3IRsmtnyyQ88lu1zEm2lzx3NlOds4yBVLOW85b7hNJQX2JkLfL7uS7hlNLb3BDkl18mMp7tY+5ELb2spd2CptY5toAnlZRDls92RjbK8mvHoSs5eYy5fACU1jwXTAz8O8iSefawl5bjCf3ICpPKyviPztsFOGot8/eT2XjOGUTlOSbxgiT9rbcp7lvnoSpNRXXO+EUVp7JLDKUujKcp9Msnbq0BU903lkLkE+/G4k8J56FRUtkvMZS5FK5rOw5rpgsRVLPJPkQmkkQnlPmTn9Ln1yUTjbK6BN9XgJ52zkhPwz5FhVTa55ZKwljkRnbLfxJeHjll9xYieQz2ufzCa54HTxRlBMUkOb95CbazjYZ3XhzMolVQe+V4ErHLZ7lKl3Pcl4csLOxnEV8+b2KotJlEcPD9xKe/j1wZysat6g/5hcbf6qXXwZgcKP/ADHTzsu3L7TO1Br+T7lprHqpcvIwOFN9Cg3jCnL7TfjfhYXyy6+cN7tNZXh/Gxi1MPKzzMmu8Y7C8UYk2srOyN2FYZRTJ4WOXeRJvC3KW88+fmRk6+OtGUSnuVxZZzjbqVJs7+LJx8kZEJckZNGe5gxljmX6Uz0+HN53Li3FrUw10O+4Gv4KhOjJ+1GWV5HmtvU65N1pV7O3rRnB+DXejv7ZyYdrimV4s+57bw5fRV43nmtvtPp3SLqF7pltd05KUatKMtn3rkfFOi6x2Zwmp8me4ei30j0bChHTtRblaN5hNbypt89uqPB9Z9Oz5cJlhN2PT6DrMMM7MvFe4g1NpxJoNzRjVpavZ9mXLtVVF+9PdGp4h490DSqM/V3cLuultTpPKb/a5HymHS82eXbMbv8AR7mXPx4zuuU0yPSJrNPR+G7iTlFVq8XSpx81u/cvuPlniq/U5zeTpvSBxnc61ezuK9RJcoQi9oLuR5breoObl7WT7f0j0+9Lx/F5r5n1Dq/x8/bxGq1ev2pPDNBdVMmVe1+03uau4qcz1uXPTh4cN3axXkYdWRdrTMSrI83lzepxYLdSRYnIrqSLEmefyZO3DFRNlqTKpstSZx5104xTJlEpEyZQzmyrdIhspbDKWzTlWyRlWlZp9hvyNpQqZXM0Ck08rp1NhZ1+0l39TTlWUjd0p5Wxk0pY5mvoTRl055NVrZI+qPwS9T9fw/qGnSeXTcKkfjKL+XYPbqq7UGvA+XPwS9U9RxjV09tJXNKccPy7X9z5n1JzyjzeWazrfj4c/fJqo171stixldHnxW2TM1KOKnLYwXnHL7zBTHVr3kJtdw27/DBTKWGk+0vvAqy095NlMns+We/OMENpvmvc8lL80gpnfm0u5FLa6488DPd18Cht4zyz4kFUpb8/3Fty71kSbbxktybw87AJPbfctyltz378iUkWpy8QE5Y2TexQ+7PzZHN7/MZYDmQ852b2J6ZfcFHbfC8QK4Syva+0rT28u4xFRpxvJXKnWUnBQxKrJw9yzjPiX008Z3Aqk1hrkWpvv39xMpLxT7mizUljkvkAnLCfL4licvGXkxOaw/aXvZjzqYW32gVTl1e5Yk5SfPJDbn9YnqA58nuSkvgFnPMqQCK8dir3BctipLqBKXeVpcyPceY+m70o23BtnLStKnTr69WhsucbWLW05frd0fe9uYaz8IP0n/5PWdThjQblfyvcQxcVoPe1ptck+k5Ll3Lfm0z5fLt5c3F5d1bu6rTr1603OpUm8ynJvLbfeWgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAegcCcC0a1pT4i4sqTstGx26NBPFa87uyucYfrdeneg1/o84GuuJ6sr27qOx0WhL+cXclvJ/oU885ePJc30T7Li3jCw0vTo8PcM20KFnQTUYwe2espPrJ9XzNXxtxpO8ow07T6cNP0yhHsUbah7KjH3f9X72efXNxKo8LZdyIykVXlzUr1XUqzc5vr0RjPcgEZbSS90QiQq29jJsrqdvNNSeCzJZ36lCeAjueH9XzOKcljxOqpXym1ukm99uZ5La3FS3mpweUdPpGtwniFSfZ7wldbWrdqTcn7XfjYonKDks7eKlzNfG8jNqUZ9rpv1LjqTltHdr5FRlxliWFJrfolv8CmUsrtNpPDw+ZZU25LMeUiuq32exz7Lym10A5vjZf5spvP+tXJ7cmb1PG/a5rds0PGb/wA2qP8AvV95u4TzGO6fgii5u/Z7TxzQbwv42LacsuTfPp1JUo5WfkwLd5Rhc2tS3mnKE44klz93lscPe2VxYTlGpFypvKjUivZl/Hcd229913ci3Wpwqdp+1Ftc4vGf3hHI6Rc1KVac1Ful2UpSfJY6eZ0egUJwtZVaqcZVZdvD7uhejY2yqdualUa3ipvODKTeNm+WyAlNvOdu7qM7pLPuKHnGXjK8eYl2X9aWXgIqjjKWNu7kSuX1XhYKU0t1zxlPOB2n2njOf3AVrnnYjLzjO3TfvIUlldl7LYZ8VECrOzw/gTnlvumUJ5aai8LqSsSzlvzfUCpvC22bI/Ne2zXy54KPab8SrOMJ4x0fMCeyk3zb+ZVvjntzTSKG32t+fyJk/Z6ZfJLuKJ5vu8hjOF4ckQ/qpvfmSvZeMc1uUNk8rf7yVtvn3lG+26a6BtqOOYRUt32c7eIf1sP7RJ4lssrO5PPbbzbKIXdn3kv2ejWeRGXhd+Mkt5T3Khnr4dBnfHJkdl74322HJZxgoqWNiM4TyMYWc8/AJJ7ZEFfazu9u4Zz8ShZ2Xeu8l93eZIrTSeW0Qn1xhFHljxK+1jxKJe72Ca8fIo2bzjbBV4vGTKIqfR5z4dxK5PYpy0vElPkupnEVp4fJcsMlYfkUbEpvnn5GURbv3mwuE2serl08DX8KpfyJDv7UvtM+/eNOuXv+Sn9hgcKZ/kKD3+vL7TdjfhYXyy7iSbTlzWTFqNLOy3Mms+1LyZhT25ckbsKwsUtYbyQ99+fgKnN95QpLu5nThWrKJzsE9ilvBHab2Ozjzc2eK7F9Ni7CRi5K4z8T0OLkcPJg2FGpjqZtvWw1uaenPDMmnV8T0OLl04eTi26WxvHFrc6HTtYlTxibXvODpV2uplUrtrqduPJL5cd47HptHiWpGOPWv4li74inNP238TgY30u/5kTvpP8AOLrDynxOi1HVXPOZczQXl25t7mHVunLO5iVa+c7ky5ZJ7LjxXK+6uvV8TBrVPEVauTFqVDi5eXbv4uLSKszGqSKqkzHnI8/k5Hdhgici1JkykWmzjzydOOKmTLcmVSZbkzmyyb8YpkyhslspbNGVbZESZTzJ8Slmq1nEMqo1PV1E+nUokyIxcnsasqyje2lVNLcz6Msmhs6nYag3t0NtQqZSZz5M49C9DmsPRuPdMvG8QjWi5pdUmm18E17z7lxiT6+R+d+j3Ltb+3ucZ9XUUn4rJ97cDajHVOENKvlUdR1LaKnN/nTj7Mn8Yv4nJzT3224rur087moeMbdOvM3+pwzRz1Rz9TKk4468zTGY+mCM43W/lsQ988/cUuTzvnDXmVEt58e7Yol1bXyJb57NlD36NeZFGylvpkSeXzyu5sttvrjJAk8Lb7MFuT/hCUtvHzLcnuAlPxLbeQ34EbPm2AfuyQ8755Ev+NyF8QJTSIT3J37ylcst5AmT6v44Lcqm/wDjkipPu5Z7zGqTx1Avurz3ZZnUW6yv3GJUrOLWOXUs1Kkp4Sba8QMirVy8R5ltttYciIrq2StmBK2C26kkx/jIExRUkR9pVFJgSlkrSwuREV3nnHpp9J1pwVp7sNPdO412vD8XTe6oRf8ArJr7F18gKPTZ6T7bguxlpumzp19erwzThs420X+fNd/cuvPlz+T767ub68rXl5XqXFxWm51atSWZTk+bbKtRvbvUb6tfX1xUuLmvNzq1ajzKUn1ZjgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArpU6larClShKpUnJRjGKy5N8kl1Zn8O6Hqev6grLTLZ1anOcntCnH9KUuSR6Xp9voHAdt26E6V/rDi4zvJL2abaw40l06rPN78lsBhcLcG2XD9CGs8WQpVLlJTt9Oluk+jqLr+zy7+qNVxlxVc6reTnVrOb5JZ2iu7wNZxFxBd6lWnKpUkk31e7OfnLL5kZSKq1WVSTbZaDIIoSQSUCUQSkADjkklIgtpygyVLG6eGXFHPQOj1TwUX7TUbi3e0njzNxZ6+mkqjfPL8TnnRqLpkoakuaA7i31mjVe8lnvbMmN5Tm8KUVl7bnARk4vKbTL9K6qxxibfgEdPrqp3NvQpVO0ozrxTfhuY0dKnT/0bUbmiuicsr7jTSu60oRUpZ7ElJeG5lQ1SssZA2Ko65Sead/TqJf7SC3K4Xeu03+Ms7eultmMsfeYVPV9sSi/MyaWq02st4xv3lF2Gt3EPy+k3EVHbNP2l9hXT4j0+SxU9bSfL2ofuIhqNFrCaafQvO5t6kcT7DXXYIu0tW06qu0rujv+lLsv4MyadaFSHap1Iyi3zi8mrqWem1ViVvRX7KS+wx5aNps/q9um11jJ/eEdBzeUmiJPfEcZ57M0D0yvBJW+qXUEukpZRW6eu037N9RqpbYnDGfggN4pY57vzKm5/VynvnKNE77W6bXrLGhVS5unLDfxZUtbrQX850u5glzaXaQG8fLZdM5YUo5W2c7LOxp6PEGnN4nVq0nndTg/uMqGp6fUfsXlLfvnj7QM5vp3bEp7c+Rap1YJJwnF422eck5b2Ta8QLmFjd7Pm/Ehb81t3d5Tlrv5ZKu0mscvACrdJPby8SMuK59erIzh55ZEm85bfgBW3hZzvzKW8P2f+ncM5aw+feUtvsrfbnuyivKab3ax3D62c8i3HCe/NPLa7u4lKMY457brvKK1hrbcndRx3rk0UyljdP5+ITxs845bfx4liKt8Yygufd59SnOcPu7yYyj2tt/ACrtJb536k7Z57d5RnPRZJxye+fHoVDOZb7NE7N45YKVyW7JUlusrvexRVFZ5PCS2ZLym+T8yhNvn8WVNtvZ4x8ioL5eRUv4fMpfPZ9Cc/IsE58w+XPl0IXLATy0mZCuO6y9gnh55e8pUnnD3XR9xL38u7vMolVx3wsfIKXiiO1jbPLmu4l4+OxnEWdRb+gXLzypS+xmBwo/8x091tOX2mdfrOn3PPalJ/JmBwq/8x01z9uW3vNuN+FhfLKrPnhPwZiVHu/D5mTcPdsxai2aybMaxq3KTKHyJnvuUN+LOjGtWUHLv5FPa3IcinOx04ZNOUVuRKlsWck9o68M3NngyIzLsKniYal4lan4nXhyubLjZ8KviXY1vE1qqFaqeJ0487TeFsfX+IdfxNf63xDq+Jn+Ow/BZkq3iWp1fExnUKJVDXlzNmPEuzqZLE5lEplqUznz5XRhx6TORalIiUi3KRzZZujHElIttiTKGzmyybpiSZbkyZMoZpyybZEN9xDD2KWzTazgyhvJLZVCGd3yNVrOKYQcvIu4wsIkg12qGxsaylHsvmuZriqhUdKqpr3mrJlHS28uh9gfgt65/KXAVTT51JTqWVRYT6RksYXvi3/SPjm0qKaUk8po90/BS4gWn8Zz0qrLFO+g4Ry9lJ4x78qK97Obkm424vqq5j26MuuxzN3FRrNJfI6h7vHxNBq9Nxqt45PBzxmwG+mfMpk8eK8CJvPTbkRn+H0Kg35oobx3By6ZKM/rY+ZFTKXiy3N42EpLr83gtylzyAk+5v3lqTbJk8lOxA6jyW7HUZf3YAheD+Y3Tx3d4eHz+YfiAynjl7iicsLPd8iZvHLmWKkk3l5yBTVk+57dTEqzeSutPxeDFqSz4AW6ksvmTRSe7KMtvD5F+CwkgKt0SuYXgVAFywVEIqS33XMCYx6sra6EJbHmXpq9KVtwbaS0vSp07jXq0PZXONrF/ny8e6Pve3MLnpo9KFpwXZS03TpU7nXq0PYp8426f58/Huj18j5R1K9u9Sv61/f3FS4uq83OrVqPMpN9WRf3dzf3ta9va9S4ua83OrVqS7UpyfNtlgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAG+4P4bra9dOVSo7bT6L/nFy1ns/qxXWT7vezW6Jp1xq+r2umWkc1rmooR2ylnm34JZb8Ed3xTf2unQjommezY2i9XBLZ1H+dOT6uT3+C5JAZuqcQWOkaZ/IvDtvG3tI/XknmVWX6U5fnP5d22xwmoX061SU5zcpP85lm8upVHzz9iMKUm3uyMkzk5MobDAEAkYAIlIlIqSwFQkSkVqJVGJBRGO5Wo77or7Ka8ytRz0KKIxwslyMM4aW5VGO/JfEuxhnGVuDamENt8MvRpRkvaim/IrhBeBk0KeXlr4hGPHTaFRbw371sRPQ4NexVw+5o2tOCSzgrXPdeJdDQVNErr6k4y95aqaXew5Qcl4G/g7l3jThTVBR2lndsyUuTY0jkJ0LmH1qT+Bb9uPOD+B2aSb3SIlb0Z/XpQefADj41Envle4rVbCyp/M6iWmWc+dFe54MapodnP6rlEG2jV3UTTU38S5HUK8fzl8DYVOHotv1df4oxp6Bdr6soy94ER1WqnmTyZFLV+/YwKmk31POabZjTt7mDxOlJe4DoYatR7OG+zlbl2Go0ptZa+Jyv4yO7g/gVRmk93j3BHWOtb1vykYT7u0sotTstNqy9u1pJfqrC+RzkKjxhVEveXoXFeOGpPzA3EtFsHJSpupSf6lTGPjkqjpd1TbdDVriCxtGft/uRqo6hcJ7y+Zehq047STeeayBsIQ16llRvLatHunHD+SENQ1qmvx2mUppL/AFc8fezHhrCfNY8C9DVISWO0luBchrzhTTuNNu6WOqjlfF4LlLiDTaixKtOnlYalBkQv6MlhzXmVOpa1licKc/2kmvmUZNLUbGpiUbqg/DtpfIyYzUo52kmsZRqKmn6ZWj/o9P8Aors/YWpaLYdp+qq1qLX6E/3gb6KSi28795C2WU2aL6BewWaGr18J7Ka7X3lSevUd417Wv4Sjj9xRvlJPOMc+jCbfPk+neaL+UNXo4VXTIVF/u6mfluVPXYU12bqwu6D73DP2hG6ba5rw2Kk0tlzXM1VPXtMqez9Icc/pRa+7Bk0b60qyxTuqM33KayUZvPv8+4PPaSfXmW41FjnnfzJ3e+SxFbefgtiE8dctlOd+WxV2uqZRLzyaT72mTFvGG30+BSpbPpsQ28rLee4sFbfdywE0vZ3xnYLnjxxzKWm/rPGSouKXsppDK7/Itxw1uVJx5r/oZCvfKWH4sn2nHG37iht9755/6BS7938SwXHJr3d5MW8Yf/QoWMbkxezws+fQyjFb1Br+Trl5f5GX2GBws/8AMVNLGe1L7TN1F40+45r8VJNLyZruFpJaJTW/15Z+Jtl+FL5ZVfKeG89TGnLffoZFeWU1j5GFUfe8bGeNYUm8756FqT5rkJSbyW5PJuxrCwyUthvcokzfjk1WKu10Ici3khs6Mc2rLFc7RUpllvqFI3Y8jVcF9TKlMx+0O0bJyMLgyVPxI7ZYch2jL8VOxecylzLTmUuRjeRZgrlItykUuRRKRrubZMUykUSkQ5FDZqubbMUyZTncpbIbNVybJBkNhspbNdrOQbKG8jmy5CON2arWUiIQ6srDINVqpZAIMLWUCAGYWsoz9HuuxWVGb2k/Z8zueDtRqaRxBY6lSm4So1oy7S6b8/vPNls8p4Or4dvFdUHCTXrYfW8V3mnNnH6IaPfUtT0u11Gk12Lmkqqw+WVuvc8oxtao9qHaS5nnf4NHEf8AKvBj0qvUzcafLCTe7g/3PH9Y9Ru4KrRa96OW+1bHItvLTxkply2W/wDHgXbyHqqrj9rMdvoAk3yXLzKJSeGs58BKRRJ5XeBDbxvnHcW23uu4lvxRRggnL89+uR8R9pGeXIBhpeBDRL+zYpAlvux55KZPqS889y1OWOuPcAqSSWFh+OTGqz2wV1J53z8zFqz2/wCoFutLfqY0nnnjBXUeSjm9twLlGOdy8l3FMI4isYK/AAuZVjL2IS3K4rqASwytIhLbqeQem/0t0uHIVtA4cqwq6y12a9dYlG08PGfh069wGb6bfSrbcI289G0adO416pHfk4Wia+tLvl3R972wpfLN7c3F7d1bu7r1K9xWm51KlSWZTk+bbKbitVuK9SvXqzq1aknOc5yblKTeW23zZbAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA7r0JWkrji65rweJ2mm3NaPn2Ox9k2afiiUldyW6faeTqfwaNRsrH0s2NvqPZVvqNCtZNyeEpTh7K98kl5s3Ppv4DutF1erc0Kbna1Jtwkl8n4geQNkFyrSlCTUk0W8BkDBKJSIISKkiqMS4oNoKojHwLigy5CmsLfOStRePeQWlDvK1B4LsIcsrzK4xXcUW1Hz8V0KlHPfsXVHPuK4RXIIoVN422LsI9Nn5IrhHbpyL1On3lEUqeeafiZcIpLbZFEI45fYXVnlyCJxiOOfmSl3+4jrncqW2dgKl4Ex+BSslfkBUl3YD+QS38CrPeEH/HiSt3lY8EQ/Iqx8gEUltyKnzSe3cUt5STKk8bdQJSXPCb6E9lYa2a8QuvcTFpNZ29wFmVtbzXtW9N+7cx56TYVJb0MZfRmd1zgqzjDW4Gmnw9ZT+pKcc97Ri1eGnzpV0vNHR4k1nGSvGFh7vuzkDkavD+oQTcZqS7+0YlXS9QpbujJrvwdyluk3uuobXRZXLzA89lRuYZ7VKS80UNyXOLR6E0pJpxUu4tVbK2qRzK3pPHVIDglUknzaLirzXKb+J2FbRLCeX6pw8mYlbhu1l9SpKOeXIo56N5WX57fvL8dSrJbs2FXhiax2K8X57GJW4fv6eeyozS7pIoqhqs1hSXIv09Xg1iWTW1dMv6azKjPbwMaVOvDaVNoDpKWp0n+dhF+F9Rz201k5LtSjnMWiqNVx5ZKjrakrStPtVKVKo3+nBP7SxPTNLqNylbw/oNpfJnORuaiW1R/EvQvaqxmefeXQ20dGtIVM0Lm5oyfJwnhIrjY6jSblQ1epJ91SPa+bZrYalVi+hfp6tPZYLpGdCWv0HlytLjrh5X2YKlqmp09q+k9vD/1c8/vMenq0Xs3gyaepUZPL2ALiCjDCr2l1RfjAv0tc02osfSIp8kpJrHyIje03spJZKZxsqyxUoUpeLgmyjNp3ltUyqVxRnt0mmy/ntR26s0tXS9LmvyCi++MmUrRqMf9Gu7qlLHSpt9hUb2WOyt8Z6IR7sv4mk+iarTivUarKfhUgvt3JlU1+lt2bW48VlP54LBvVJ4ynnJMfjv8jRfytqFKSVbSqj73Tn2vsK/8oLNTxXo3NB/r0zKSjcp7vp1K09lh4XmaunrWm1dldQz+snH7TLo3NCtn1ValP9maZUqrUcfQLnn+Sn79jXcLvGiQ/al9pnajLNhXWc/ipfYa/hl/5jht+dLf3myX2Ysq4x2uuWYU2sYT5GTX9xiTkstPkZ41jVMnjmi0+vIqm9ueS3J7Nm3GsbCTKJNYDeShs2ysLBvvKXISlsUN7cjbjk12Ku0Tktt7EZ72bJkwuK6pbDJb7WB2jLvTtXMjteJbyG9y96dqtshyKMkNk717UuRQ2Q34lLZjcmUg2UtkNkNmNyZyJbIyUtkZNdyZyJbI3b2IW5cikkYWrojFIqIzkZNdqpI5EA12soMEAwtUIAMLWQy9YXVSzu4XFLnF7rvXcWGy3UnjkY1X0h+D9xVR0nimzu1N/Qb38VVSfLO2+OeH078H1xz65XevtPzd9GuvrTNap2t3U7FpcTS7eceqn0l4Lv8Aj0Pvv0Z6rPU+GKEK7X0i3iqc1yysbPy6HPnNVsi/r1BwnnkvDqaST35eW/I7DU6Kr0HJLdJ7v+PI5G7hKlUcXHHmYKtSliPZwiiTSx+8OT8inyyA26v5keOwW/VJfAj3EEp8yH4DzI9wE5x4ZIz1yGUyeOvuAicvAsTk8bcvImpIsVJfs7AUVJ97MapPPg/ArqzfLLLEnkCmTedi5Rinuy3FZaWVuZUIpICUu/clc/ARWcFSX+CAJbdEVYHJZfLrk+fvTZ6Y+36/hzhC59jendajTf1ujhSfd0c+vTbcDa+nH0vR0pV+G+FblS1B5hd3sHlW/fCD/T73+b03+r83zlKc3OcnKUnltvLbIe7ywAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFdGrUo1oVqM5U6lOSlCcXhxa3TT7z6N4D9KWmcY6TR0Tid0qerKKp9qphQunyTT5Kb7u/l3L5vAHuvG3o3tK8519LkqMm8+rly9x5frHDGpabVcK9vPC5SSyvibDhj0ka9pNONreSWp2kVhQryfrIr9WfP458MHb2PHHDmsU1Cdx9DqywnSukorPXEvq482vIDyR2c08OIVpPkkz16+0fS7umqqoU3GX1alN7Pya2NLdcN2yUvVVZQb5ZX7iaXbz+FtUzjsvzwVqjjGUzrZ6JcU9k4z95j1tOuIrPqG/vGjbnPV78sFUYPLZu6lnJfXoSjjw5Fj6NSe/Yayu4aNtd2W3nqVKOFnBnfRoZzGTI+j9FNNAYqhuvtK4U9i+qE03smirsSXR+4oohH3l2Ky87EKL6IuJbrmBMcblWPIpRVFd4E4wVdSmWIpt4S5lNrXpXNL1lGTlHOM4wEXfIriUr4lSfgBXu+hVyXcUqXImO+yAqXeg85C27hnGQJSXUmLKd+z1Ko9F9oFTaTx9xK+sRnw+BPPf7QHTDKktnjcpW/g8FcXh7MCd10KkllbYfkRhNc3nzCxywBPVrG2ORRhvd8ipKS5ePIlJYTy/ACmCz7PXwJTTkotKXu2CeZZi9wm+qSSAlL3+ZKeHlN/EhZlHHXw2Jkk+oCSyl+8pWXvy94ku0ueN9ipbPbZd5QWVHaWNyidOnNYnCLXislSXawVqKwnhvv36lGJU06yqP2qEN9uWDFq6DYVHiEJRfg8m1xJ5yk33B47Oefi3yCOfq8M0Hl060o/tIw63DNdP8XWpy2zzOpbTfLbr4E9Unn3llHF1dC1Cm3iHaS7mY8rO9ovE6Evemd7L2VtHvWSMcuTWC7R5+3OLxOjJeQdSHdJeaO9qW9Gp9aEJZ6OO5jVNNsame1bQTfLGxdjjlOPaTVRZ88FSuJR+rP5nSVtBsZt9n1kPDJjVeGqX5lw89ziWVNNRG+rRX18l6Gp1Y89y/V4cuov8XOMveYlTRr+G6pOS8NzL2GXDVZLDa26mRT1WD59e80lS0uqX1qUl5ot5qR+tBmWoOmhqMJdS7SuaTz9Xc5elOGfxjcTIhOGV2a0c+LwXSbb+dCxrbzoUZPv7K3LUtJ0yo9qPZfhJo1kPW84VFLyZcjXuI7ZbLBmLR4qElQvrmnFrddrKx5GRocVQsHRby4VJxz37mvjfVls45KLS+lSjOLT3m29+8y92LaV5rfGzxgxJzSbaMed52uecstSrZ6/EyiVkufiWpyefAtes8SHJfwzZKxqtspbyU58WRnGxnKxsS3uUtkPuIfMzmTHQ3uM9SnkyG0uZnMmOlYz5lCazjI7SMu40rz3jPcW+2ueSPWodyaXWyG9i06ryUuo2O5e1W31KWyhtkb95O5ZFTZS5FIyS5MpEtshMhsjJhay0rTK4vYtIrizC00rDIBhaugAZ6mFqhDDBjayCGw3gs1KnRGKpqVMci0syYSbeW8Lq30KKk85jH6v2mrPPTKRM6iXswfnL9x9W/gh+kyF/Tp8JapWS1O0p/wAzlJ4+lUFzp+M4rl3peDPk0ytJ1C90rU7bUtOuKltd2tSNWjVg8ShJPKaNFtrN+pScZxTT7UZLn355HOcQ2fYn6yKznqkc16APSVZekXg+ldr1VHU7dqlf2yf5Orj6yX6E92vHtLLaPQ7y3VehKnjfvIOCa33G3cjJ1G2lb1nHGPuMVd3vAN9W8kc1nC94fMiW4E5I6E526e8pb+ICTwWZyxyZNSWc7vyRZnJ9MfECmcs8l8UWKkvEmcl3de8x6kl0ApqPLz9rLWRJ7YKqcW2BcoRzLtF9ERWFhe8ritwCWCmvVpUKM69erClSpxc5znJRjCK3bb6Is6pqFlpenVtQ1K5p2trQj2qlWo8KK/jofLnpl9K15xhVlpOkupaaFCW8eU7prlKfdFdI+95eMBtvTZ6X6uuutw/wxWnR0reFxdR9md13pdVD5vy2fjIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAX7W7u7STla3Nag3zdObjn4G6s+Mtdt49idxTuYLpWpp/NYfzOeAHYUuOKrj/ONNpyl30qrgvg0/tM6lxnpbgu3b3VOT54jGSXvz9xwIA9Mp8R6FUcYq+p5l0lCUce9rHzMyFXT7msqdK4tK85fm06sZt/Bnk4A9Ynp9upYnSUJeRjy0q2lvGU0+7uPN7W9vLWXatbuvQffTqOP2Gdb8R61RbavpTzz9bGM38ZJgdnPSpLPYqe10LU9Ou4pYcZeT5mgo8ZalGPZrULWt4uLT+Tx8jPocbUnhV9Omn1lCqn8sL7QMudrdU1iVJ+7dFpxkt503y5pF+hxbpFR+1KvQed3Kl/4WzOp6xotzF9jULbfpOXY/tJAaiWE12oSWSqLhjLaj3ZZvqULWvD1lJUa0Vt2oNNe7BRUsLaTxKPZk/HxA03ZT/OjnqmVQpdlYhGKiu42EtJpPeE3F+JZlpVSP1K3kgMbDXR57iYr3e4uuxvIL2Wn5MtuN3FJOnJy67ASuSaZUl7y16+a2nRw/IqjWjnZYAub8ycbblMalJveXZ80VxnB8pJ+GQJxlcgsY6E4bJSaxkCEttyV3dPINc+XkEml0e4E4e+NipJrwI3TfTvyS89/wAAKkumwaws9e4o3z4FTk1HpjuwBVmXaS5vuJa2SzjuRCWd2sy55IznZ4aAnl0+BS2214d5Ki3JPOWg1h56gTFNZw/DcJNJ55P5h7Lu36DLXVgIttY6jtNsSfd9oztjr3AXFjZ9ccimPsyz17imOO/LLk+58yiG3jDzz5kPlglPL9rBOcLboVFCjFrljqM7PbrjmTJtrC5EYzu85238QEW0vrLbvDfLbDe/kRiK5Y26lSeyfXoURyf7hLdrHJfAN7cye7b5BFKxjdpvHxDe/LP3kpJ755k8k21goiKb5ReO/BOUm9446h4msvOOpMtt+pRS0m3lbt7bFmpb2817dGEtufZLs+ff0KU2nl5yywYU9LsajeaCWd/ZZjVNAtJLMXOPzNxvF9p7oiT6rLXXcstRz9Th3H5Ot8UWJaNqFN4pVMrwkdP18TGnVu/p8aUKEPozjmVTO+fIymVTTm5Wup0+cHLH6ufsLD+lUk1O3xlvLw0dok0tnjBZrKMm8pSk+r6mUyNOOdae+YSRSqvfFnS16UJPMoR96MSrRpcvVw+BsmTGxpnW8Aq5sJ0KaT9iOeZalRpZ3hHyMplGOmL6/vDrIvypU1+YiiVOP6KyZyppa9cUuqy44rol8CmSMpU0tuciMyKmiGZ7RGWAC7TQAQNmk5RDZGSGxtdJbIbIz3Bsm10NkNkNkEtXScgjIMbV0qT2KkyjJKMLTS6n3ApjyJMbV0qKQE8rPIxtVJEmkRKSRYnNt7GKpqTbeEUJL60nhdWNortS5dF1ZanJyeXsui7jVnnplImpPtbJYiuSKADnt2zAAB1von461P0fcY22vae3UpL8Xd2+cRr0W94+D6p9Gl5H6JcHcRaZxXw1Z6/o9zG4s7uCnGS2a7010aw010aaPzAPafwXPS1LgLiRaLrNd/5O6jUSqOUtrWq8JVF3Reyl5J9MMPtbXrFV6LqJe2luclUg4TcXtg9BThUpqcWpwnHKaezWDmOIdOVKo6sFtJ9wGj8Ql4eZHJ8hnfOUAbxtktTZVNssSe3iBFSX8cyxUe7WcvuKpv8A6FipLv3ApqS23T+JYnImb8i23nYAk302MihHCy+ZZpQy9u8zEklhYALc1fFXEWk8LaLV1bWbpULenslznUl0hFdZPu828JNmt9IfG+jcE6Q7zUqnbr1Mq2tYP8ZWku7uS6yfLzwn8m8e8Y6zxnrL1HVq/sxyqFvB4p0I90V9r5sDa+lT0jatxzqHZqdq00qjLNvZxlt+1N/nS+S6dc8QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFs8ozKWq6nSSjT1C6jGPKKqyx8MmGAN5S4q1uEouVzCol+bKjHfzwkzOt+NbtVe1cWNvOHWNNyg/i2zlQB3FtxnZSbVxaXFKPTsSVT7eyZtvxPotZNyuZUX0VSnLf4Jo86AHqVvqGl145heWk9sJesipfB7l9WlCpHtqlGUX1UdmeTFdKrVpS7VKpOm++MmmB6hLTbaWGlKP3liWlRUcxqSWeWThaGu6xRx2NRuHjkpy7f25M6hxfq9NYqfR6/wC3Tx/ZwB00tNuIbwqxeCJUL+nyy/maijxpLC9dp0ZPr2Krivg0/tM6lxhpk2lUpXVPPP2U0vn9wF7113HadPfxiT9MX59FYL9LiDRas+zG+ppyf58ZR+bSRmUqlhcz7FvXta8+6lOM38mBr4XdFv6rjjoXI1qD2VRZ7mjMq2NDte1Rw/hgsz02g8OLakvHkBEJwawpQa8xGOJ53eV5luemNrMau/fyLcrK7jHKl2vIDJxHOefmxFRzt713MxXTvKa+rLzW5T9KrqPtUvkBnc2sYx4opby9omMr2PKVNr7iuN5Qk/rNPPUC99ZLC8dhJrOeZCrUpL8rHwyVdmMnmMk0+iYBJ9lpk4Wc7+RLi8rEZY5ZRDWHhrAD2d2uaJTjukslG7eU1jqTKCSfd1wUVPGd2kvAhZTblmPgEstS5Ln5kZlNNKLl4lFc2u1lpYLfNtIlLbk2u8qxjbDzjmgiiOHnKfLuKtsYxgjC3y8rGcBpqeNwDw+WMFL7t85TySsrmEuXj0KIiuzv0ZcUu0ubXuKFlvuKo7Pl8GER+du9vMNPHPKxzwS03z3z8iM42WGvIoSTk0k8bEPCwnj4k8+y0+XMplHn3rmUUyfLrjxJ6KOG99iXJc8J+aKVlvzKKsNtt93Up3XPdvkTJvC2/wCo2znllbbhEvfCS7K29prBaqLK3bTzv4l147Lajs9t1yLMmub5Y3wZbGPUWG2YtVZXJLBmVfBLDfUxprZ9PEylRiVE8vJYkmssyakea2LE10+0zlRjy3RbknkvST8cluS3MpWOlt8+hQy40UNbmyVLFDTIcfArJxk3Y+7XfZYaIwXpRwUyjsZdqbWnzIZVJFLMKziGyMhkMx2uggMgm10EBkMx2qQQTzMdqlFSKUSibNKkyopROTG1dJyUyngickkWW3JmOzSZSbZDagk5LL6R/eJyVPbCc+7uLLbby3lmnPk+0ZyEpOUnKTyyADSyAAAAAAAAfWv4HnpgVzRo+jniS6/H0440e4qS+vFf6ht9Uvq+G3RZ+mryjG4oOnNZ2yflpbV61tcU7i3qzpVqU1OnUhLEoyTymn0aZ9xfg0+m634+safDvENalQ4nt6ezeIxv4Jbzj+uksyj/AEltlRDtNUtZ29zNNYWWYLaXcdzrOnwuqEmortLqcPeUqlCr2Kiw/FgWJS3fcWpybX7yZz2ysY7yzUe2cNrvAoqS8X7ixUl4oqm9i1PPiBTJtkJZeMjG/LKZcoxzJYAvUY4jnfc4H0telDS+CrednbKne63OOadsn7NLPKVRrkuvZ5vwW5zHph9MlDSPX6DwnVp179ZhWvFiVOg+qh0lPx5LxfL5yua9e6uKlxc1qlatVk51KlSTlKUnzbb5sDN4i1vVOIdWq6prF5Uurqq95y5JdIpckl3I1wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAZFrfXto27W7r0M8/V1HH7DOtuI9ZoJqN7KSf8AtIRm/i1k1IA6SjxhqEY9mrb2tXvl2ZJ/J4+RnUONKTWK9hUi+rhVT+TS+040Aeg0eK9GqY7Uq9F9e3S/dkzaOsaPX+pqNth/pvsf2sHmIA9XVO3rrNP1VWL6xaa+RalYW7e8Ozt0PLk2nlNp+BmUdV1Ojj1eoXSS5L1ra+HID0CppdKSzGcky1PTa0V7FTPdvg5OhxVrNOeZXEKq/RnSjh/DD+ZnW3Gl1Gb+kWVGcO6lJwfxfaA3v0a/p57E+XTJHbvYYzTeOm2xg23GVjOTdzaXFFdPVtVPt7JmW/EujVYuUrt0muUalOWX8E18wK1eTjl1KKTz3YRVG/pbxnGUH13L9G/024h24XdpJvknVjn4N5L87ShOClKlmMt89CjFV1QafZbWX5Faqwk8qazz3ZVPT7eTyljPPBaqabFv2KrWO8oupbYXZw99u4bY2542fcY0rCvDeFVeO5S6d/BbJyz45CMrDzlt58SXmOcGGri6i/bp7eKJWob+1Rw14gZPOXVr7ycpcua6FhX1BvDTj4F1V6DWY1XnosYKKm2+XkR2m+XMLsyyoTi8rvJUNsuPzCCk3s02u4hPLw8xXXwJlH2cpNeGCMJrKeM9QJz2Y9np3lLeHsufyDisc91uH79+ZRTj2N+fkF54x3FefZ9p8+/kUxxn5lEOWOT3Iby89O8oo1qFZ1PV1oz7LxLsvdFal4vEeXeUV59jZLOcFmWG91y688lyP1uuVsUyWM75QRYaTW+Xgtyi8bF+TTSaST7upaabMtjEqxxslgx6i8PmZlVdp5S9xj1I9TKVGLNdSzNZZkzXJFiS28fIylRZkiiW3iXWu8tyTT2RnKijoVIpb8BFm7jya8oraKJIuIiS2Oue8aPFY80WpGRNcyzJGnONuNWmQVSRSzTa2RBAZBjtkEEkGOzQSuZAMdqrSKimO5WTYIpnJJcyJzxsi2k5PPv8jG1dG8mROahtB+11l3eRTUqbdmHLq+8tmjPPftGUgADWyAAAAAAAAAAAL9hd3Vhe0L2yuKttdUKkalGtSk4zpzTypJrdNMsAD7U9Af4RelcTUrXh7jWtS03XMKnTvJYjQvHyWelOb7uTfLGUj2zXdGhe0pTppKpjJ+X56j6NPTv6QeBo07W21JarpkFhWWo5qwiu6EsqUduSTx4AfWupWlxZ1XGrBp9+DXzkm+/3HA6R+FfwtqMJU+KuC9QtMR9mdjXhcdp+U/V4+LLdz6e/RVXk3RocT0M9KlnSa+VVgd5OXRdC293jB5vV9Ofo9jFyT16o/wBGNjD76iOS4h/CBTpVqOgcO9mecUrm9rZ273Siuf8ATYHtep39jpVhVv8AUbqja2tJZnVqyxFf4+HU+fvSt6ZbrWqNbReF3VstNlmFW6fs1q66pfoRfxfXG6PN+J+KNf4muVca3qde7cXmEJPFOH7MFiK9yNMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALlKtWpPNKrOm++Mmi2ANlR17WKWOzqNeWOXbl2/7WTOo8XatD8oret39qnjP9XBz4A66hxrJbVtOi+906uPk0zNpcXaXNrt07mk3zbgml8Hn5HCAD0mhr2jVX2IX9JN/pxlFfFpIzKNSyu/ZoVreu+qpzU38jyoF2PVZ2VvJ+1DD8Hgsz02i1t2o46t8jzm1v761TjbXlxQT5qnUcc/AzrfiTWqEHCN45J8/WU4zfxayNo7N6Y8exWeOmUUuzuoxWJ5x1yc3S4wv4wUaltbT75JSTfzx8jOpcZUHj1thUg+rjVUvlhfaXZptM30NpRckR9IrZ9uDWNuRao8V6POW9StSz/tKXL+q2Z1DV9IrP8Xf22/6U+z/AGsF2Mf6XDte3Dm9sFyNzRcs7p9MozYUrevHNONOomucGn9hROwt5LPYw+uCox1VpzisTiVLGdmpZ6JkVNOpP6s5Ity06oo+xVz8gIo2dG2U/UUVTc5dp4XNl5yeMxi+vPYx3a3sHiMk/eQql7Typx6dxRfw8tLd49/MVHhNd/NYxgx1cuMn2qSXkvsJ+lwb9qMs9WBKw5PoJJ4b5Y5oOpSlym1LxD3WYzS70VFjsZ7kWakM9EjLn7bctty1Ujs8Y2LsYNSKMepHfoZ1SODGnFrO5lKMWUVjJblEyJp5Lc+X+BlKmmNJFO5cmuexbe77jZjWNiqLK9ty1FtdSuLydnFm0Z4okixNGSyxNbFzTFYkUMuyLbObJuihkMlkM17ZoZBJDJaBKKSUjDa6XIkTmUOWNgsJdqTwvtJbpdCWcybwlzbLdSp2vZjtH7SJzcvBLkik0ZZ7ZSAAMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAExk4vMW011TM2hq+qUWvV6hcpLo6ja+D2MEAbuhxTrFOSc61Ksv0Z0lj5YZnW/GVxF5uLGjU7vVycPtycsC7Hb0eMbGcX6+1uaUunYxP5tozbfiTRqkFOV46Uv0KlOWfkmvmedgbTT1Glf6bXisXlpNSwvyke0/dnJfdpQlHMqSimsrpnxPJy5Rr1qMu1RrVKbXWMmi9xp6fPTaEoNpyjt0LEtL2zGvJd2UcNR1/WaTTjqFeWP9o+3/ayZ1Hi/VYflI21ZfrU8f2Wi9xp1MrC5X1Zx+Janb3sHvBS65RqaHGr2VfT897hVx8mjOo8XaXU2qQuqT73BNfJ5+Re6Jpcl6yLxOhnv2LU/Ut+1T7PmjOoa7oteSUL+kn+unBfFoyqE7O6y6FShX7/AFcoyx8C7NNJO3oS5SawyzU0/P1KnkmdDKyt+zvTw2i1PTaW/ZlJPzMu5NOarWFZbxw0YdW3qweXBnVVrGcN41YvwaMSvSnCfZmoNY5pmUzSxzUozT3jgqizd1Y0k8Sjj3Fp07Xq0n4o2Y8umNx21faWC3UaNv6ize7qwIlQsVlutD4mz/5O/sx/CaSW5acX3G8lDTVzrJ+4s1JabHlOUvJGu8u2UxadwkPVyNlOvYpbU5yZYnd0l9S3Xg2zHuZaYfq2VeokuawV1Luo+WIrwRYqOcl2pNtd7JclTLsxfPPkW5Sb5ENwXOWfBIh1X2ezBKK6tc37zXc5F0l4ik5bv9H95ROTk8v/AKFINVytZAAIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAybe/vreDhb3txSi+cYVZRXyZnUeJNZp01TV524r9OnGT+LWTUADpI8YX7go1La2l3tKSb+ePkVVOJ6VVJTsZR72qif3HMgvdTTo5a1YzWWrhS/YWPtLMtStGmlOSz3xZogXuqabqV5ayX5ZLzjL9xbldW2Nq8X/Rl+41IL300z6lzS6ScvJFP0ig47yqJ+EF+8wgO+mmT6+GfqSfvwW5Vm3skkWgTupqK5VZySTawu5JFABjtQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAH//Z"
        export_html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>BRINC DFR Proposal — {prop_city}, {prop_state}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=IBM+Plex+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --black:#000000;--white:#ffffff;--cyan:#00D2FF;--gold:#FFD700;
  --ink:#0a0a0f;--surface:#f7f8fa;--border:#e4e6ea;
  --text:#111118;--muted:#6b7280;--light:#f0f2f6;
  --resp:#00D2FF;--guard:#FFD700;--green:#22c55e;--amber:#f59e0b;
}}
html{{scroll-behavior:smooth}}
body{{font-family:'Inter',sans-serif;background:var(--surface);color:var(--text);line-height:1.6;display:flex;min-height:100vh}}

/* ── LEFT SIDEBAR INDEX ──────────────────────────────────────── */
.doc-sidebar{{
  position:fixed;top:0;left:0;width:220px;height:100vh;
  background:var(--ink);overflow-y:auto;z-index:100;
  display:flex;flex-direction:column;
  border-right:1px solid #1a1a2a;
}}
.sidebar-logo{{
  padding:28px 24px 20px;
  border-bottom:1px solid #1a1a2a;
}}
.sidebar-logo img{{height:28px;filter:brightness(0) invert(1);display:block}}
.sidebar-logo .brand{{font-size:22px;font-weight:900;letter-spacing:3px;color:#fff}}
.sidebar-city{{
  padding:16px 24px;
  border-bottom:1px solid #1a1a2a;
}}
.sidebar-city .city-name{{font-size:13px;font-weight:700;color:#fff;letter-spacing:0.3px}}
.sidebar-city .city-sub{{font-size:11px;color:#666;margin-top:2px}}
.sidebar-nav{{padding:20px 0;flex:1}}
.sidebar-nav a{{
  display:flex;align-items:center;gap:10px;
  padding:9px 24px;font-size:12px;font-weight:500;
  color:#888;text-decoration:none;
  border-left:2px solid transparent;
  transition:all 0.15s;
}}
.sidebar-nav a:hover,.sidebar-nav a.active{{
  color:#fff;background:rgba(0,210,255,0.06);
  border-left-color:var(--cyan);
}}
.sidebar-nav .nav-num{{
  font-size:10px;font-weight:700;color:#333;
  width:18px;text-align:center;flex-shrink:0;
}}
.sidebar-nav a:hover .nav-num,.sidebar-nav a.active .nav-num{{color:var(--cyan)}}
.sidebar-footer{{
  padding:20px 24px;border-top:1px solid #1a1a2a;
  font-size:10px;color:#444;line-height:1.6;
}}
.sidebar-footer a{{color:#555;text-decoration:none}}

/* ── MAIN CONTENT ────────────────────────────────────────────── */
.doc-main{{margin-left:220px;flex:1;min-width:0}}

/* ── PAGE SECTIONS (each prints as independent page) ─────────── */
.doc-section{{
  background:#fff;
  border-bottom:6px solid var(--surface);
  padding:52px 60px;
  position:relative;
}}
.doc-section:first-child{{padding-top:64px}}

/* Section header bar */
.section-eyebrow{{
  display:flex;align-items:center;gap:10px;
  margin-bottom:32px;
}}
.section-eyebrow .pg-num{{
  font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;
  color:var(--cyan);border:1px solid rgba(0,210,255,0.3);
  padding:3px 10px;border-radius:100px;font-family:'IBM Plex Mono',monospace;
}}
.section-eyebrow .pg-title{{
  font-size:10px;font-weight:600;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted);
}}

/* ── COVER PAGE ──────────────────────────────────────────────── */
.cover-page{{
  background:var(--ink);color:#fff;min-height:100vh;
  display:flex;flex-direction:column;justify-content:space-between;
  padding:60px;position:relative;overflow:hidden;
}}
.cover-page::before{{
  content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 80% 60% at 70% 30%,rgba(0,210,255,0.08) 0%,transparent 70%);
  pointer-events:none;
}}
.cover-logo{{margin-bottom:auto}}
.cover-logo img{{height:36px;filter:brightness(0) invert(1)}}
.cover-logo .brand{{font-size:28px;font-weight:900;letter-spacing:4px;color:#fff}}
.cover-headline{{margin:60px 0 40px}}
.cover-tag{{
  font-size:11px;font-weight:700;letter-spacing:3px;text-transform:uppercase;
  color:var(--cyan);margin-bottom:20px;
}}
.cover-headline h1{{
  font-size:52px;font-weight:900;line-height:1.05;letter-spacing:-1px;
  color:#fff;margin-bottom:16px;
}}
.cover-headline h1 span{{color:var(--cyan)}}
.cover-headline p{{font-size:16px;color:#888;max-width:480px;line-height:1.7}}
.cover-meta{{
  display:grid;grid-template-columns:1fr 1fr 1fr;gap:1px;
  background:#1a1a2a;border:1px solid #1a1a2a;border-radius:10px;overflow:hidden;
}}
.cover-meta-cell{{
  background:var(--ink);padding:20px 24px;
}}
.cover-meta-cell .label{{font-size:10px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:#555;margin-bottom:6px}}
.cover-meta-cell .value{{font-size:15px;font-weight:700;color:#fff}}
.cover-meta-cell .value.accent{{color:var(--cyan)}}
.cover-meta-cell .value.gold{{color:var(--gold)}}
.cover-bottom{{margin-top:40px;font-size:12px;color:#444;border-top:1px solid #1a1a2a;padding-top:24px;display:flex;justify-content:space-between}}

/* ── METRICS SECTION ─────────────────────────────────────────── */
.metrics-hero{{
  display:grid;grid-template-columns:repeat(3,1fr);gap:1px;
  background:var(--border);border-radius:12px;overflow:hidden;
  margin-bottom:40px;box-shadow:0 1px 3px rgba(0,0,0,0.04);
}}
.metric-cell{{
  background:#fff;padding:28px 24px;text-align:center;
}}
.metric-cell .m-label{{font-size:10px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:10px}}
.metric-cell .m-value{{font-size:36px;font-weight:900;font-family:'IBM Plex Mono',monospace;line-height:1;color:var(--text)}}
.metric-cell .m-value.cyan{{color:var(--cyan)}}
.metric-cell .m-value.gold{{color:var(--gold)}}
.metric-cell .m-value.green{{color:var(--green)}}
.metric-cell .m-sub{{font-size:11px;color:var(--muted);margin-top:6px}}

/* Fleet split row */
.fleet-split{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:32px}}
.fleet-card{{border-radius:10px;padding:24px;position:relative;overflow:hidden}}
.fleet-card.guardian{{background:#0a0800;border:1px solid #2a2400;color:#fff}}
.fleet-card.responder{{background:#00111a;border:1px solid #002a3a;color:#fff}}
.fleet-card .fc-icon{{font-size:28px;margin-bottom:12px}}
.fleet-card .fc-type{{font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;margin-bottom:8px}}
.fleet-card.guardian .fc-type{{color:var(--gold)}}
.fleet-card.responder .fc-type{{color:var(--resp)}}
.fleet-card .fc-val{{font-size:32px;font-weight:900;font-family:'IBM Plex Mono',monospace;color:#fff}}
.fleet-card .fc-sub{{font-size:12px;color:#888;margin-top:4px}}
.fleet-card .fc-row{{display:flex;justify-content:space-between;font-size:12px;padding:6px 0;border-bottom:1px solid rgba(255,255,255,0.06)}}
.fleet-card .fc-row:last-child{{border-bottom:none}}
.fleet-card .fc-row .k{{color:#666}}.fleet-card .fc-row .v{{font-weight:600;color:#ccc}}

/* ── SECTION HEADINGS ────────────────────────────────────────── */
.sh{{
  font-size:22px;font-weight:800;color:var(--text);
  margin-bottom:20px;padding-bottom:12px;
  border-bottom:2px solid var(--border);
  display:flex;align-items:center;gap:10px;
}}
.sh .sh-accent{{color:var(--cyan)}}

/* ── TABLES ──────────────────────────────────────────────────── */
table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:28px}}
thead tr{{background:var(--ink);color:#fff}}
thead th{{padding:12px 16px;text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase}}
tbody tr:nth-child(even){{background:#fafbfc}}
tbody tr:hover{{background:#f0f8ff}}
td{{padding:12px 16px;border-bottom:1px solid var(--border);color:var(--text)}}
.tag-resp{{background:rgba(0,210,255,0.1);color:#0066aa;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}}
.tag-guard{{background:rgba(255,215,0,0.15);color:#8a6a00;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}}

/* ── MAP ─────────────────────────────────────────────────────── */
.map-wrap{{border-radius:12px;overflow:hidden;border:1px solid var(--border);box-shadow:0 4px 20px rgba(0,0,0,0.06);margin-bottom:32px}}

/* ── DISCLAIMER ──────────────────────────────────────────────── */
.disc{{background:#fffbeb;border-left:4px solid var(--amber);padding:16px 20px;border-radius:0 8px 8px 0;font-size:12px;color:#7a5a00;margin-bottom:24px}}

/* ── FOOTER ──────────────────────────────────────────────────── */
.doc-footer{{background:var(--ink);color:#555;padding:36px 60px;font-size:11px;display:flex;justify-content:space-between;align-items:center;gap:20px}}
.doc-footer a{{color:#666;text-decoration:none}}
.doc-footer .brand-mark{{font-size:16px;font-weight:900;letter-spacing:3px;color:#fff;flex-shrink:0}}

/* ── PRODUCT IMAGE ───────────────────────────────────────────── */
.cover-body{{
  display:flex;align-items:flex-start;gap:40px;flex:1;margin:40px 0;
}}
.cover-left{{flex:1;min-width:0}}
.cover-right{{
  flex-shrink:0;width:420px;display:flex;align-items:center;
  justify-content:flex-end;
}}
.product-img{{
  width:100%;max-width:420px;
  filter:drop-shadow(0 20px 60px rgba(0,210,255,0.18));
  pointer-events:none;
}}
@media (max-width:900px){{.cover-right{{display:none}}}}

/* ── TWO-COL INFRA TABLE ─────────────────────────────────────── */
.infra-grid{{
  display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:28px;
}}
.infra-item{{
  background:#fafbfc;border:1px solid var(--border);border-radius:6px;
  padding:10px 12px;font-size:11px;display:flex;justify-content:space-between;
  align-items:center;gap:8px;
}}
.infra-item .i-name{{font-weight:600;color:var(--text);flex:1;min-width:0;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.infra-item .i-type{{font-size:10px;font-weight:600;padding:2px 6px;border-radius:4px;flex-shrink:0}}
.infra-item .i-link{{color:var(--cyan);text-decoration:none;font-weight:700;font-size:11px;flex-shrink:0}}

/* ── GRANT STATS SIDEBAR ─────────────────────────────────────── */
.grant-layout{{display:grid;grid-template-columns:1fr 280px;gap:40px;align-items:start}}
.grant-sidebar{{display:flex;flex-direction:column;gap:12px}}
.grant-stat{{
  background:#fafbfc;border:1px solid var(--border);border-radius:8px;
  padding:16px;text-align:center;
}}
.grant-stat .gs-label{{font-size:10px;font-weight:600;letter-spacing:1px;
  text-transform:uppercase;color:var(--muted);margin-bottom:6px}}
.grant-stat .gs-val{{font-size:24px;font-weight:900;font-family:'IBM Plex Mono',monospace;color:var(--cyan)}}
.grant-stat .gs-sub{{font-size:11px;color:var(--muted);margin-top:4px}}
.grant-stat.gold .gs-val{{color:var(--gold)}}
.grant-stat.green .gs-val{{color:var(--green)}}

/* ── CRIME STATS BOX ─────────────────────────────────────────── */
.crime-box{{
  background:linear-gradient(135deg,#0a0800 0%,#120e00 100%);
  border:1px solid #2a2000;border-radius:10px;
  padding:24px;margin-bottom:24px;
}}
.crime-box h4{{color:var(--gold);font-size:12px;font-weight:700;
  letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px}}
.crime-stat-row{{
  display:flex;justify-content:space-between;align-items:baseline;
  padding:8px 0;border-bottom:1px solid #1a1600;font-size:13px;
}}
.crime-stat-row:last-child{{border-bottom:none}}
.crime-stat-row .csk{{color:#888}}
.crime-stat-row .csv{{font-weight:700;color:#fff;font-family:'IBM Plex Mono',monospace}}
.crime-stat-row .csv.accent{{color:var(--gold)}}

/* ── PRINT ───────────────────────────────────────────────────── */
@media print{{
  .doc-sidebar{{display:none}}
  .doc-main{{margin-left:0}}
  .doc-section,.cover-page{{
    page-break-after:always;
    page-break-inside:avoid;
    break-after:page;
    border-bottom:none;
    min-height:0;
  }}
  .doc-section:last-child,.cover-page:last-child{{page-break-after:auto;break-after:auto}}
  body{{background:#fff}}
}}
</style>
</head>
<body>

<!-- ── SIDEBAR INDEX ─────────────────────────────────────────── -->
<nav class="doc-sidebar">
  <div class="sidebar-logo">
    {"<img src='data:image/png;base64," + logo_b64 + "' alt='BRINC'>" if logo_b64 else '<div class="brand">BRINC</div>'}
  </div>
  <div class="sidebar-city">
    <div class="city-name">{prop_city}, {prop_state}</div>
    <div class="city-sub">DFR Deployment Proposal</div>
  </div>
  <nav class="sidebar-nav">
    <a href="#cover"><span class="nav-num">00</span>Cover</a>
    <a href="#executive"><span class="nav-num">01</span>Executive Summary</a>
    <a href="#fleet"><span class="nav-num">02</span>Fleet &amp; Coverage</a>
    <a href="#map"><span class="nav-num">03</span>Coverage Map</a>
    <a href="#incident-data"><span class="nav-num">04</span>Incident Analysis</a>
    <a href="#deployment"><span class="nav-num">05</span>Deployment Locations</a>
    <a href="#grant"><span class="nav-num">06</span>Grant Narrative</a>
    <a href="#infrastructure"><span class="nav-num">07</span>Infrastructure Directory</a>
    <a href="#community"><span class="nav-num">08</span>Community Partnership</a>
    <a href="#analytics"><span class="nav-num">09</span>Analytics Dashboard</a>
  </nav>
  <div class="sidebar-footer">
    Prepared {datetime.datetime.now().strftime("%b %d, %Y")}<br>
    {prop_name}<br>
    <a href="mailto:{prop_email}">{prop_email}</a>
  </div>
</nav>

<main class="doc-main">

<!-- ── 00: COVER ─────────────────────────────────────────────── -->
<section class="cover-page" id="cover">
  <div class="cover-logo">
    {"<img src='data:image/png;base64," + logo_b64 + "' alt='BRINC'>" if logo_b64 else '<div class="brand">BRINC</div>'}
  </div>
  <div class="cover-body">
    <div class="cover-left">
      <div class="cover-headline">
        <div class="cover-tag">Drone as a First Responder · Deployment Proposal</div>
        <h1>Protecting <span>{prop_city}</span>,<br>{prop_state}</h1>
        <p>A data-driven deployment plan for {actual_k_responder + actual_k_guardian} BRINC aerial units covering {calls_covered_perc:.1f}% of incidents across {prop_city}.</p>
      </div>
      <div class="cover-meta">
        <div class="cover-meta-cell"><div class="label">Fleet CapEx</div><div class="value accent">${fleet_capex:,.0f}</div></div>
        <div class="cover-meta-cell"><div class="label">Annual Savings</div><div class="value gold">${annual_savings:,.0f}</div></div>
        <div class="cover-meta-cell"><div class="label">Break-Even</div><div class="value">{break_even_text}</div></div>
        <div class="cover-meta-cell"><div class="label">Call Coverage</div><div class="value accent">{calls_covered_perc:.1f}%</div></div>
        <div class="cover-meta-cell"><div class="label">Avg Response</div><div class="value">{avg_resp_time:.1f} min</div></div>
        <div class="cover-meta-cell"><div class="label">Time Saved</div><div class="value gold">{avg_time_saved:.1f} min</div></div>
      </div>
    </div>
    <div class="cover-right">
      <img class="product-img" src="data:image/jpeg;base64,{_guardian_img_b64}" alt="BRINC Guardian Station">
    </div>
  </div>
  <div class="cover-bottom">
    <span>Prepared for <strong style="color:#fff">{prepared_for_city}</strong> by <strong style="color:#fff">{prepared_by_name}</strong></span>
    <span>{datetime.datetime.now().strftime("%B %d, %Y")}</span>
  </div>
</section>

<!-- ── 01: EXECUTIVE SUMMARY ──────────────────────────────────── -->
<section class="doc-section" id="executive">
  <div class="section-eyebrow"><span class="pg-num">01</span><span class="pg-title">Executive Summary</span></div>
  <div class="metrics-hero">
    <div class="metric-cell"><div class="m-label">Fleet Capital Expenditure</div><div class="m-value cyan">${fleet_capex:,.0f}</div><div class="m-sub">{actual_k_responder} Responder · {actual_k_guardian} Guardian</div></div>
    <div class="metric-cell"><div class="m-label">Annual Savings Capacity</div><div class="m-value gold">${annual_savings:,.0f}</div><div class="m-sub">At {int(dfr_dispatch_rate*100)}% dispatch · {int(deflection_rate*100)}% resolution</div></div>
    <div class="metric-cell"><div class="m-label">Program Break-Even</div><div class="m-value">{break_even_text}</div><div class="m-sub">Full cost recovery timeline</div></div>
    <div class="metric-cell"><div class="m-label">911 Call Coverage</div><div class="m-value cyan">{calls_covered_perc:.1f}%</div><div class="m-sub">of {st.session_state.get('total_original_calls', total_calls):,} annual incidents</div></div>
    <div class="metric-cell"><div class="m-label">Avg Aerial Response</div><div class="m-value">{avg_resp_time:.1f} min</div><div class="m-sub">vs. ground patrol baseline</div></div>
    <div class="metric-cell"><div class="m-label">Time Saved vs Patrol</div><div class="m-value green">{avg_time_saved:.1f} min</div><div class="m-sub">per incident, on average</div></div>
  </div>
  <p style="font-size:15px;color:#444;line-height:1.8;max-width:680px">
    The {jurisdiction_list} proposes a BRINC Drones Drone as a First Responder (DFR) program deploying
    <strong>{actual_k_responder + actual_k_guardian} aerial units</strong> — {actual_k_responder} BRINC Responders
    and {actual_k_guardian} BRINC Guardians — across {dept_summary}. The system is projected to cover
    <strong>{calls_covered_perc:.1f}% of historical incidents</strong>, reach scenes
    <strong>{avg_time_saved:.1f} minutes faster</strong> than ground patrol, and deliver
    <strong>${annual_savings:,.0f} in annual operational savings</strong> with a break-even horizon of {break_even_text.lower()}.
  </p>
</section>

<!-- ── 02: FLEET & COVERAGE ───────────────────────────────────── -->
<section class="doc-section" id="fleet">
  <div class="section-eyebrow"><span class="pg-num">02</span><span class="pg-title">Fleet &amp; Coverage</span></div>
  <div class="fleet-split">
    <div class="fleet-card guardian">
      <div class="fc-icon">🦅</div>
      <div class="fc-type">BRINC Guardian</div>
      <div class="fc-val">{actual_k_guardian} Unit{"s" if actual_k_guardian != 1 else ""}</div>
      <div class="fc-sub">{guard_radius_mi}-mile operational radius · {guard_strategy_raw}</div>
      <div style="margin-top:16px">
        <div class="fc-row"><span class="k">Unit CapEx</span><span class="v">${CONFIG['GUARDIAN_COST']:,}</span></div>
        <div class="fc-row"><span class="k">Call Coverage</span><span class="v">{guard_calls_perc:.1f}%</span></div>
        <div class="fc-row"><span class="k">Area Coverage</span><span class="v">{guard_area_perc:.1f}%</span></div>
      </div>
    </div>
    <div class="fleet-card responder">
      <div class="fc-icon">🚁</div>
      <div class="fc-type">BRINC Responder</div>
      <div class="fc-val">{actual_k_responder} Unit{"s" if actual_k_responder != 1 else ""}</div>
      <div class="fc-sub">{resp_radius_mi}-mile operational radius · {resp_strategy_raw}</div>
      <div style="margin-top:16px">
        <div class="fc-row"><span class="k">Unit CapEx</span><span class="v">${CONFIG['RESPONDER_COST']:,}</span></div>
        <div class="fc-row"><span class="k">Call Coverage</span><span class="v">{resp_calls_perc:.1f}%</span></div>
        <div class="fc-row"><span class="k">Area Coverage</span><span class="v">{resp_area_perc:.1f}%</span></div>
      </div>
    </div>
  </div>
</section>

<!-- ── 03: COVERAGE MAP ───────────────────────────────────────── -->
<section class="doc-section" id="map">
  <div class="section-eyebrow"><span class="pg-num">03</span><span class="pg-title">Coverage Map</span></div>
  <div class="map-wrap">{map_html_str}</div>
</section>

<!-- ── 04: INCIDENT ANALYSIS ─────────────────────────────────── -->
<section class="doc-section" id="incident-data">
  <div class="section-eyebrow"><span class="pg-num">04</span><span class="pg-title">Incident Data Analysis</span></div>
  {cad_charts_html_export}
</section>

<!-- ── 05: DEPLOYMENT LOCATIONS ──────────────────────────────── -->
<section class="doc-section" id="deployment">
  <div class="section-eyebrow"><span class="pg-num">05</span><span class="pg-title">Deployment Locations</span></div>
  <table>
    <thead><tr><th>Station</th><th>Type</th><th>Avg Response</th><th>FAA Ceiling</th><th>CapEx</th></tr></thead>
    <tbody>{station_rows}</tbody>
  </table>
</section>

<!-- ── 06: GRANT NARRATIVE ───────────────────────────────────── -->
<section class="doc-section" id="grant">
  <div class="section-eyebrow"><span class="pg-num">06</span><span class="pg-title">Grant Narrative</span></div>
  <div class="disc"><strong>DISCLAIMER:</strong> AI-generated draft. Must be reviewed, localized, and fact-checked by your grants administrator before submission.</div>

  <div class="grant-layout">
  <div class="grant-body">
    <p><strong>Project Title:</strong> BRINC Drones Drone as a First Responder (DFR) Program — {prepared_for_city}, {prop_state}</p>

    <p><strong>Executive Summary:</strong> The {jurisdiction_list} respectfully requests funding to establish a Drone as a First Responder (DFR) program deploying {actual_k_responder + actual_k_guardian} purpose-built BRINC aerial units — {actual_k_responder} BRINC Responder and {actual_k_guardian} BRINC Guardian — across {dept_summary} serving {pop_metric:,} residents. Modeled against {st.session_state.get('total_original_calls', total_calls):,} historical incidents from {_exp_date_range}, the program is projected to cover <strong>{calls_covered_perc:.1f}%</strong> of calls for service, arrive an average of <strong>{avg_time_saved:.1f} minutes faster</strong> than ground patrol, and generate <strong>${annual_savings:,.0f} in annual operational savings</strong>, reaching full cost recovery in <strong>{break_even_text.lower()}</strong>.</p>

    <p><strong>Statement of Need:</strong> {jurisdiction_list} currently responds to an estimated {st.session_state.get('total_original_calls', total_calls):,} calls for service annually — approximately {max(1,int(st.session_state.get('total_original_calls',total_calls)/365)):,} calls per day. Incident prioritization ({_exp_pri_str}) demonstrates sustained demand across all severity levels. Ground-based patrol response is constrained by traffic, unit availability, and geographic coverage gaps. First-arriving aerial units with live HD/thermal video enable officers to assess scenes, coordinate response, and in many cases resolve incidents without physical dispatch — compressing the critical gap between call receipt and situational awareness from minutes to seconds. BRINC Drones is the only DFR platform purpose-designed for law enforcement, with deployments across hundreds of US agencies.</p>

    <p><strong>Geographic Scope &amp; Participating Agencies:</strong> The proposed network covers <strong>{jurisdiction_list}</strong> ({prop_state}), hosted at {dept_summary} — including facilities operated by <em>{police_names_str}</em>. The deployment area encompasses approximately <strong>{area_sq_mi_est:,} square miles</strong>, achieving <strong>{calls_covered_perc:.1f}%</strong> historical incident coverage and <strong>{area_covered_perc:.1f}%</strong> geographic area coverage. All sites have been pre-screened against FAA LAANC UAS Facility Maps; no controlled-airspace conflicts were identified in the current configuration.</p>

    <p><strong>Fleet Architecture &amp; Program Design:</strong> The fleet consists of <strong>{actual_k_responder} BRINC Responder</strong> units ({resp_radius_mi}-mile operational radius, {CONFIG["RESPONDER_SPEED"]:.0f} mph, 2-minute average response, optimized for <em>{resp_strategy_raw.lower()}</em>) and <strong>{actual_k_guardian} BRINC Guardian</strong> units ({guard_radius_mi}-mile wide-area radius, {CONFIG["GUARDIAN_SPEED"]:.0f} mph, optimized for <em>{guard_strategy_raw.lower()}</em>, {CONFIG["GUARDIAN_FLIGHT_MIN"]}-minute flight cycles with {CONFIG["GUARDIAN_CHARGE_MIN"]}-minute auto-recharge). Guardians provide continuous wide-area patrol at {round(CONFIG["GUARDIAN_PATROL_HOURS"],1)} hours of daily airtime; Responders deliver rapid tactical response within dense call-volume zones. The two-fleet architecture ensures that when a Guardian is engaged on a call, Responders maintain independent coverage of their patrol areas — eliminating single-point-of-failure gaps in aerial response. Deployment mode: <strong>{deployment_mode}</strong>.</p>

    <p><strong>Technology Platform:</strong> BRINC Drones provides fully automated launch-on-dispatch, live-streaming HD and thermal video to dispatch and responding officers, FAA-compliant Beyond Visual Line of Sight (BVLOS) operations, chain-of-custody flight logging, and integrated data analytics. All hardware is manufactured in the United States. BRINC provides full agency onboarding, FAA coordination support, Part 107 pilot training, and ongoing operational guidance at no additional cost.</p>

    <p><strong>Fiscal Impact &amp; Return on Investment:</strong> Total program capital expenditure is <strong>${fleet_capex:,.0f}</strong> ({actual_k_responder} Responder × ${CONFIG["RESPONDER_COST"]:,} + {actual_k_guardian} Guardian × ${CONFIG["GUARDIAN_COST"]:,}). At a <strong>{int(dfr_dispatch_rate*100)}% DFR dispatch rate</strong> and <strong>{int(deflection_rate*100)}% call resolution rate</strong> (no officer dispatch required), the program is projected to generate <strong>${annual_savings:,.0f} per year</strong> in operational savings, reaching break-even in <strong>{break_even_text.lower()}</strong>. Cost per drone response is ${CONFIG["DRONE_COST_PER_CALL"]} versus ${CONFIG["OFFICER_COST_PER_CALL"]} for a ground patrol dispatch — a <strong>{int((1-CONFIG["DRONE_COST_PER_CALL"]/CONFIG["OFFICER_COST_PER_CALL"])*100)}% cost reduction</strong> per incident. The program also reduces officer exposure to unknown-risk calls, decreasing liability and improving officer retention outcomes.</p>

    <p><strong>Evaluation Plan:</strong> Program outcomes will be tracked quarterly across four dimensions: (1) response time comparison vs. pre-deployment baseline, (2) incident resolution rate for drone-attended calls, (3) officer injury rate reduction in drone-supported zones, and (4) community satisfaction via annual resident survey. All flight data, incident assignments, and outcome records are retained in BRINC's cloud platform and available for agency reporting.</p>

    <p><strong>10-Year Program Value Model:</strong> The following projections assume consistent call volume and constant dispatch/resolution rates. Actual results may vary.</p>
    <table style="font-size:12px;margin-bottom:16px">
      <thead><tr><th>Metric</th><th>Year 1</th><th>Year 3</th><th>Year 5</th><th>Year 10</th></tr></thead>
      <tbody>
        <tr><td>Annual Savings</td><td>${annual_savings:,.0f}</td><td>${annual_savings*1.05:,.0f}</td><td>${annual_savings*1.1:,.0f}</td><td>${annual_savings*1.22:,.0f}</td></tr>
        <tr><td>Cumulative Savings</td><td>${annual_savings:,.0f}</td><td>${annual_savings*3.15:,.0f}</td><td>${annual_savings*5.53:,.0f}</td><td>${annual_savings*12.58:,.0f}</td></tr>
        <tr><td>Drone-Attended Calls</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365*1.03):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365*1.06):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365*1.13):,}</td></tr>
        <tr><td>Calls Resolved w/o Dispatch</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365*1.03):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365*1.06):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365*1.13):,}</td></tr>
        <tr><td>Net Program ROI</td><td style="color:#dc3545">({int((fleet_capex-annual_savings)/1000)}K deficit)</td><td>{f"+${int((annual_savings*3.15-fleet_capex)/1000)}K" if annual_savings*3.15>fleet_capex else f"({int((fleet_capex-annual_savings*3.15)/1000)}K)"}</td><td style="color:#22c55e">+${int((annual_savings*5.53-fleet_capex)/1000):,}K</td><td style="color:#22c55e">+${int((annual_savings*12.58-fleet_capex)/1000):,}K</td></tr>
      </tbody>
    </table>

    <p><strong>Officer Safety &amp; Liability Reduction:</strong> Beyond direct operational savings, DFR programs measurably reduce officer exposure to unknown-risk call scenarios. First-arriving drones perform scene reconnaissance before ground units arrive, enabling officers to approach with full situational awareness. Documented outcomes across peer agencies include: reduced officer injuries in drone-supported zones (avg. 18% reduction, per DOJ data), faster suspect identification improving apprehension rates, and reduced use-of-force incidents through earlier de-escalation intelligence. These outcomes reduce agency liability costs and workers' compensation claims — benefits not captured in the direct cost model above.</p>

    <p><strong>Community &amp; Economic Impact:</strong> Response time improvements of <strong>{avg_time_saved:.1f} minutes</strong> translate directly to better outcomes in time-sensitive incidents: cardiac events, structure fires, crimes in progress, and missing persons cases. Studies by the International Association of Chiefs of Police (IACP) document measurable improvements in case clearance rates, property crime deterrence (15–30% reduction in areas with visible DFR patrols), and community trust metrics in agencies with active drone programs. For {prop_city}'s business community, faster emergency response reduces property damage, shortens insurance claim cycles, and improves the commercial district safety perception that drives foot traffic and investment.</p>

    <p style="background:#f8f9fa;padding:15px;border-radius:8px;border:1px solid #eee;font-size:13px">
      <strong>Applicable Grant Funding Sources:</strong><br>
      <a href="https://bja.ojp.gov/program/jag/overview">DOJ Byrne JAG</a> — Technology and equipment procurement<br>
      <a href="https://www.fema.gov/grants/preparedness/homeland-security">FEMA HSGP</a> — Homeland security CapEx offset<br>
      <a href="https://cops.usdoj.gov/grants">DOJ COPS Office</a> — Law enforcement technology<br>
      <a href="https://www.transportation.gov/grants">DOT RAISE</a> — Regional infrastructure and safety<br>
      <a href="https://bja.ojp.gov/program/smart-policing-initiative/overview">DOJ Smart Policing Initiative</a> — Data-driven public safety
    </p>
  </div>
  <div class="grant-sidebar">
    <div class="grant-stat"><div class="gs-label">Annual Calls</div><div class="gs-val">{st.session_state.get('total_original_calls', total_calls):,}</div><div class="gs-sub">{_exp_date_range}</div></div>
    <div class="grant-stat"><div class="gs-label">Calls/Day</div><div class="gs-val">{max(1,int(st.session_state.get('total_original_calls',total_calls)/365)):,}</div><div class="gs-sub">citywide avg</div></div>
    <div class="grant-stat gold"><div class="gs-label">Call Coverage</div><div class="gs-val">{calls_covered_perc:.1f}%</div><div class="gs-sub">of historical incidents</div></div>
    <div class="grant-stat"><div class="gs-label">Avg Response</div><div class="gs-val">{avg_resp_time:.1f}m</div><div class="gs-sub">{avg_time_saved:.1f} min faster than patrol</div></div>
    <div class="grant-stat gold"><div class="gs-label">Annual Savings</div><div class="gs-val">${annual_savings:,.0f}</div><div class="gs-sub">break-even {break_even_text.lower()}</div></div>
    <div class="grant-stat green"><div class="gs-label">Cost Reduction</div><div class="gs-val">{int((1-CONFIG["DRONE_COST_PER_CALL"]/CONFIG["OFFICER_COST_PER_CALL"])*100)}%</div><div class="gs-sub">per incident vs. patrol dispatch</div></div>
  </div>
  </div>
</section>

<!-- ── 07: INFRASTRUCTURE DIRECTORY ──────────────────────────── -->
<section class="doc-section" id="infrastructure">
  <div class="section-eyebrow"><span class="pg-num">07</span><span class="pg-title">Infrastructure Directory</span></div>
  <p style="color:var(--muted);font-size:12px;margin-bottom:16px">Public facilities evaluated as candidate deployment locations during optimization. All coordinates verified against FAA LAANC facility maps.</p>
  <div class="infra-grid">{all_bldgs_rows}</div>
</section>

<!-- ── 08: COMMUNITY PARTNERSHIP ─────────────────────────────── -->
<section class="doc-section" id="community">
  <div class="section-eyebrow"><span class="pg-num">08</span><span class="pg-title">Community Business Partnership</span></div>

  <!-- Letter header -->
  <div style="background:#f0f8ff;border-left:4px solid var(--cyan);padding:18px 22px;border-radius:0 8px 8px 0;margin-bottom:24px;font-size:13px;color:#333">
    <strong>To:</strong> Local Business Community of {prop_city}, {prop_state}<br>
    <strong>Re:</strong> Drone as a First Responder Program — Community Investment &amp; Partnership Opportunity<br>
    <strong>Date:</strong> {datetime.datetime.now().strftime("%B %d, %Y")}
  </div>

  <p>Dear {prop_city} Business Owner,</p>

  <p>The {prop_city} Police Department is proposing a transformational public safety initiative that will directly protect your business, your employees, and your customers. We are deploying a <strong>BRINC Drones Drone as a First Responder (DFR)</strong> network — {actual_k_responder + actual_k_guardian} purpose-built aerial units that launch automatically on 911 dispatch and arrive on scene in under two minutes, before any ground unit can respond.</p>

  <p>This is not a surveillance program. It is a rapid-response tool. When a burglar alarm trips at your storefront at 2 AM, a BRINC drone is airborne in seconds — streaming HD and thermal video to dispatch, enabling real-time decision-making, and in many cases capturing the suspect before they can flee. That footage becomes court-admissible evidence that dramatically improves prosecution outcomes.</p>

  <!-- Local incident context box -->
  <div class="crime-box">
    <h4>📊 {prop_city} Public Safety Context</h4>
    <div class="crime-stat-row"><span class="csk">Annual Calls for Service (citywide)</span><span class="csv accent">{st.session_state.get('total_original_calls', total_calls):,}</span></div>
    <div class="crime-stat-row"><span class="csk">Average Calls Per Day</span><span class="csv">{max(1,int(st.session_state.get('total_original_calls',total_calls)/365)):,}</span></div>
    <div class="crime-stat-row"><span class="csk">Incidents Covered by DFR Network</span><span class="csv accent">{calls_covered_perc:.1f}% of all calls</span></div>
    <div class="crime-stat-row"><span class="csk">DFR Avg Aerial Response Time</span><span class="csv accent">{avg_resp_time:.1f} minutes</span></div>
    <div class="crime-stat-row"><span class="csk">Time Saved vs. Ground Patrol</span><span class="csv accent">~{avg_time_saved:.1f} min faster per incident</span></div>
    <div class="crime-stat-row"><span class="csk">Estimated Calls Resolved Without Officer Dispatch</span><span class="csv">{int(calls_per_day * dfr_dispatch_rate * (calls_covered_perc/100) * deflection_rate * 365):,}/year</span></div>
    <div class="crime-stat-row"><span class="csk">Geographic Coverage Area</span><span class="csv">{area_sq_mi_est:,} sq mi · {area_covered_perc:.1f}% of city</span></div>
  </div>

  <h3 style="color:var(--text);font-size:16px;margin:24px 0 12px">Why This Matters to Your Business</h3>

  <p>Property crimes cost American businesses over <strong>$50 billion annually</strong> in direct losses — before accounting for insurance premium increases, business interruption, and the intangible cost of a less safe commercial environment. The businesses of {prop_city} are not immune. Consider what faster response means in practice:</p>

  <table style="margin-bottom:20px;font-size:13px">
    <thead><tr><th>Crime Type</th><th>Avg Cost per Incident</th><th>DFR Impact</th></tr></thead>
    <tbody>
      <tr><td><strong>Retail Theft / Shoplifting</strong></td><td>$559 per incident <em>(NRF 2023)</em></td><td>Real-time apprehension · evidence capture · deterrence</td></tr>
      <tr><td><strong>Commercial Burglary</strong></td><td>$3,000–$8,500 per incident</td><td>Scene arrival before suspect can exit · HD identification footage</td></tr>
      <tr><td><strong>Vandalism / Property Damage</strong></td><td>$1,000–$5,000 per incident</td><td>Suspect identification · reduced repeat incidents in monitored zones</td></tr>
      <tr><td><strong>Robbery / Armed Incidents</strong></td><td>$8,000+ in costs &amp; liability</td><td>Officer-safety pre-arrival intel · coordinated ground response</td></tr>
      <tr><td><strong>Trespass / Loitering</strong></td><td>$200–$1,500 in staff/facility cost</td><td>Non-confrontational aerial deterrence</td></tr>
    </tbody>
  </table>

  <p>Peer agencies with active DFR programs — including Chula Vista, CA; El Cajon, CA; and Westerville, OH — have documented <strong>15–30% reductions in property crime rates</strong> within covered zones, and average response times of under 90 seconds from dispatch. {prop_city}'s proposed deployment covers <strong>{calls_covered_perc:.1f}% of incidents</strong> and arrives an average of <strong>{avg_time_saved:.1f} minutes faster</strong> than current patrol response.</p>

  <h3 style="color:var(--text);font-size:16px;margin:24px 0 12px">The Investment We're Requesting</h3>

  <p>Total program CapEx is <strong>${fleet_capex:,.0f}</strong>. The {prop_city} Police Department is seeking community partnership contributions to offset a portion of this cost and accelerate deployment. Every dollar contributed directly funds equipment that protects your street, your block, your customers.</p>

  <table style="margin-bottom:20px">
    <thead><tr><th>Sponsorship Tier</th><th>Contribution</th><th>Recognition &amp; Benefits</th></tr></thead>
    <tbody>
      <tr style="background:rgba(255,215,0,0.04)"><td><strong>🥇 Founding Sponsor</strong></td><td><strong>$25,000+</strong></td><td>Named drone unit bearing your business name · press release · plaque at station · annual program briefing with command staff · tax-deductible receipt</td></tr>
      <tr><td><strong>🥈 Community Champion</strong></td><td><strong>$10,000–$24,999</strong></td><td>Logo on all program materials &amp; vehicle decals · certificate of recognition · annual impact report · public acknowledgment at City Council</td></tr>
      <tr><td><strong>🥉 Neighborhood Partner</strong></td><td><strong>$2,500–$9,999</strong></td><td>Listed on program website and department newsletter · certificate · public acknowledgment</td></tr>
      <tr><td><strong>🤝 Business Supporter</strong></td><td><strong>$500–$2,499</strong></td><td>Donor roll listing · formal thank-you letter on department letterhead</td></tr>
      <tr><td><strong>💙 Community Contributor</strong></td><td><strong>Any amount</strong></td><td>Recognized in annual program transparency report</td></tr>
    </tbody>
  </table>

  <p>For every <strong>$10,000</strong> contributed, the program is projected to generate approximately <strong>${int(annual_savings/max(fleet_capex,1)*10000):,}</strong> in annual operational savings and property crime cost avoidance for the {prop_city} business community — a {round(annual_savings/max(fleet_capex,1),1):.1f}x return on community investment.</p>

  <p>To make a contribution or learn more, please contact:</p>

  <div style="background:#f8f9fa;border:1px solid #eee;border-radius:8px;padding:20px;margin:20px 0;font-size:13px">
    <strong>{pd_chief}</strong><br>{pd_dept}<br>{pd_email_html}{pd_phone_html}
  </div>

  <p>Together, we can make {prop_city} safer, faster, and more resilient. Thank you for your commitment to this community.</p>
  <p>Respectfully,</p>
  <p><strong>{pd_chief}</strong><br>{pd_dept}</p>
</section>

<!-- ── 09: ANALYTICS DASHBOARD ───────────────────────────────── -->
<section class="doc-section" id="analytics">
  <div class="section-eyebrow"><span class="pg-num">09</span><span class="pg-title">Analytics Dashboard</span></div>
  [ANALYTICS_HTML_EXPORT]
</section>

<!-- ── DISCLAIMER ─────────────────────────────────────────────── -->
<div style="background:#fffbeb;border:1px solid #f59e0b;border-radius:8px;padding:20px 60px;margin:0;font-size:11px;color:#7a5a00;line-height:1.7">
  <strong>&#9888; SIMULATION TOOL DISCLAIMER</strong> — All figures are model estimates based on user inputs and publicly available data. Not a legal recommendation, binding proposal, contract, or guarantee. Deployments require FAA authorization and formal procurement.
</div>

<!-- ── FOOTER ─────────────────────────────────────────────────── -->
<footer class="doc-footer">
  <span class="brand-mark">BRINC</span>
  <span>{"<img src='data:image/png;base64," + logo_b64 + "' style='height:20px;filter:brightness(0) invert(0.5)'>" if logo_b64 else ""} BRINC Drones, Inc. · <a href="https://brincdrones.com">brincdrones.com</a> · <a href="mailto:sales@brincdrones.com">sales@brincdrones.com</a> · +1 (855) 950-0226</span>
  <span>Prepared by {prop_name} · <a href="mailto:{prop_email}">{prop_email}</a></span>
</footer>

</main>
<script>
// Highlight active nav link on scroll
const sections=document.querySelectorAll('section[id],div[id="analytics"]');
const links=document.querySelectorAll('.sidebar-nav a');
const obs=new IntersectionObserver(entries=>{{
  entries.forEach(e=>{{
    if(e.isIntersecting){{
      links.forEach(l=>l.classList.remove('active'));
      const a=document.querySelector('.sidebar-nav a[href="#'+e.target.id+'"]');
      if(a)a.classList.add('active');
    }}
  }})
}},{{threshold:0.3}});
sections.forEach(s=>obs.observe(s));
</script>
</body></html>"""

        export_html = export_html.replace("[ANALYTICS_HTML_EXPORT]", analytics_html_export)

    # ── Download buttons — always rendered so they're visible in the sidebar ──
    _safe_city   = _safe_city_base
    _ts          = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # 1. Save Deployment Plan (.brinc) — always available
    _brinc_data = json.dumps(export_dict) if fleet_capex > 0 else json.dumps({
        "city": st.session_state.get('active_city', ''), "state": st.session_state.get('active_state', ''),
        "_disclaimer": "No drones deployed yet.", "k_resp": 0, "k_guard": 0,
        "calls_data": _safe_df_to_records(
            st.session_state.get('df_calls_full') if st.session_state.get('df_calls_full') is not None
            else st.session_state.get('df_calls')
        ),
        "stations_data": _safe_df_to_records(st.session_state.get('df_stations')),
    })
    if st.sidebar.download_button("💾 Save Deployment Plan", data=_brinc_data,
                                  file_name=f"Brinc_{_safe_city}_{_ts}.brinc",
                                  mime="application/json", use_container_width=True):
        if fleet_capex > 0:
            _notify_email(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                          "BRINC", k_responder, k_guardian, calls_covered_perc,
                          prop_name, prop_email, details=export_details)
            _log_to_sheets(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                           "BRINC", k_responder, k_guardian, calls_covered_perc,
                           prop_name, prop_email, details=export_details)

    # 2. Executive Summary HTML — only meaningful when drones are deployed
    if fleet_capex > 0:
        if st.sidebar.download_button("📄 Executive Summary (HTML)", data=export_html,
                                      file_name=f"Brinc_{_safe_city}_Proposal_{_ts}.html",
                                      mime="text/html", use_container_width=True):
            _notify_email(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                          "HTML", k_responder, k_guardian, calls_covered_perc,
                          prop_name, prop_email, details=export_details)
            _log_to_sheets(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                           "HTML", k_responder, k_guardian, calls_covered_perc,
                           prop_name, prop_email, details=export_details)
    else:
        st.sidebar.button("📄 Executive Summary (HTML)", disabled=True,
                          use_container_width=True,
                          help="Deploy at least one drone to generate the proposal document.")

    # 2b. BRINC Tank — sidebar-free full mirror of the proposal page
    _tank_disclaimer = (
        "BRINC Tank is a presentation-mode HTML mirror of this deployment. "
        "All data is simulation-based. Not a guarantee or binding proposal."
    )
    if fleet_capex > 0:
        # Build a sidebar-free version: inject hide-sidebar CSS into export_html
        _tank_html = export_html.replace(
            "</style>",
            ".doc-sidebar{display:none!important;}.doc-main{margin-left:0!important;}</style>",
            1
        ).replace(
            "<title>",
            "<title>BRINC Tank · "
        )
        if st.sidebar.download_button(
            "⚡ BRINC Tank (Presentation Mode)",
            data=_tank_html,
            file_name=f"BrincTank_{_safe_city}_{_ts}.html",
            mime="text/html",
            use_container_width=True,
            help="Full-screen presentation version — sidebar hidden, optimised for screen share and client demos."
        ):
            pass
    else:
        st.sidebar.button("⚡ BRINC Tank (Presentation Mode)", disabled=True,
                          use_container_width=True,
                          help="Deploy drones first to generate the BRINC Tank presentation file.")

    # 3. Google Earth KML — only when drones are placed
    if active_drones:
        if st.sidebar.download_button("🌏 Google Earth Briefing File",
                                      data=generate_kml(active_gdf, active_drones, calls_in_city),
                                      file_name="drone_deployment.kml",
                                      mime="application/vnd.google-earth.kml+xml",
                                      use_container_width=True):
            _notify_email(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                          "KML", k_responder, k_guardian, calls_covered_perc,
                          prop_name, prop_email, details=export_details)
            _log_to_sheets(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                           "KML", k_responder, k_guardian, calls_covered_perc,
                           prop_name, prop_email, details=export_details)
    else:
        st.sidebar.button("🌏 Google Earth Briefing File", disabled=True,
                          use_container_width=True,
                          help="Deploy at least one drone to generate the KML file.")
