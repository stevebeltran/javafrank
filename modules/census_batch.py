"""
Helpers for preparing and merging Census batch geocoding workflows.
"""

from __future__ import annotations

import io
import csv
import json
import re
import time
import zipfile
from pathlib import Path
import http.client
import urllib.error
import urllib.request
import uuid

import pandas as pd

from modules.config import STATE_FIPS, US_STATES_ABBR
from modules.efficient_merge import merge_census_results_fast
from modules.data_validation import validate_census_results, validate_merged_data
from modules.cad_parser import _normalize_jacksonville_cfs_report, _normalize_loxley_priority_calls_report
from modules.numbers_adapter import load_numbers_dataframe


_STATE_ABBRS = set(STATE_FIPS.keys())
_STATE_NAMES = {name.upper(): abbr for name, abbr in US_STATES_ABBR.items()}
_EXCEL_EXTS = ('.xlsx', '.xls', '.xlsb', '.xlsm')
_ADDRESS_HINTS = [
    'street', 'street_address', 'address', 'location', 'incident_location',
    'addr', 'block_address', 'call_location', 'full_address',
]
_CITY_HINTS = ['city', 'city_name', 'municipality', 'town', 'village']
_STATE_HINTS = ['state', 'state_name', 'province']
_ZIP_HINTS = ['zip', 'zipcode', 'zip_code', 'postal', 'postal_code']
_URBANIZATION_HINTS = ['urbanization', 'urb', 'urbanizacion']
_NOISE_HINTS = [
    'date', 'time', 'priority', 'call', 'nature', 'type', 'desc',
    'agency', 'incident', 'case', 'event', 'status', 'zone', 'unit',
]
_INTERSECTION_HINTS = (
    ' / ',
    ' & ',
    ' AT ',
)


def _clean_text(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    if not text or text.lower() in {'nan', 'none', 'nat'}:
        return ''
    return re.sub(r'\s+', ' ', text)


def _clean_zip(value) -> str:
    text = _clean_text(value)
    if not text:
        return ''
    m = re.search(r'(\d{5})(?:-\d{4})?(?:\.0+)?', text)
    return m.group(1) if m else ''


def _clean_state(value) -> str:
    text = _clean_text(value).upper().replace('.', '')
    if not text:
        return ''
    if text in _STATE_ABBRS:
        return text
    return _STATE_NAMES.get(text, '')


def _dedupe_tokens(tokens):
    out = []
    for token in tokens:
        token = _clean_text(token)
        if not token:
            continue
        if out and out[-1].upper() == token.upper():
            continue
        out.append(token)
    return out


def _find_named_column(df: pd.DataFrame, hints) -> str | None:
    cols = [str(c).strip().lower() for c in df.columns]
    for hint in hints:
        for idx, col in enumerate(cols):
            if col == hint:
                return str(df.columns[idx])
    for idx, col in enumerate(cols):
        if any(hint in col for hint in hints):
            return str(df.columns[idx])
    return None


def _get_col_series(df: pd.DataFrame, col) -> pd.Series:
    data = df.loc[:, col]
    if isinstance(data, pd.DataFrame):
        return data.iloc[:, 0]
    return data


def _guess_delimiter(text: str) -> str:
    first_line = text.splitlines()[0] if text else ''
    return ',' if first_line.count(',') >= first_line.count('\t') else '\t'


def _looks_like_headerless_excel_columns(columns) -> bool:
    try:
        cols = [str(c).strip() for c in columns]
        if len(cols) < 6:
            return False
        lowered = [c.lower() for c in cols]
        if any(c in {
            'date', 'time', 'datetime', 'latitude', 'longitude', 'lat', 'lon',
            'priority', 'location', 'address', 'city', 'state', 'zip', 'zipcode'
        } for c in lowered):
            return False
        date_like = sum(pd.to_datetime([c], errors='coerce').notna()[0] for c in cols[:2])
        time_like = sum(bool(re.fullmatch(r'\d{1,2}:\d{2}:\d{2}', c)) for c in cols[:3])
        state_like = sum(bool(re.fullmatch(r'[A-Za-z]{2}', c)) for c in cols)
        zip_like = sum(bool(re.fullmatch(r'\d{5}(?:\.0+)?', c)) for c in cols)
        number_like = sum(bool(re.fullmatch(r'\d+(?:\.0+)?', c)) for c in cols[:8])
        return (date_like + time_like) >= 2 and state_like >= 1 and zip_like >= 1 and number_like >= 1
    except Exception:
        return False


def _sheet_score(ws) -> int:
    score = 0
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if not rows:
        return -1
    header = rows[0] or []
    header_norm = [str(h).strip().lower() for h in header if h is not None]
    if not header_norm:
        return -1
    hints = ['latitude', 'longitude', 'lat', 'lon', 'priority', 'location', 'date', 'time', 'address']
    score += sum(10 for h in header_norm if any(k == h or k in h for k in hints))
    score += sum(1 for h in header_norm if h and not re.match(r'^column\d+$', h))
    if len(rows) > 1 and rows[1] and any(v is not None and str(v).strip() for v in rows[1]):
        score += 25
    if len(header_norm) == 1 and header_norm[0].startswith('externaldata_'):
        score -= 100
    return score


def _looks_like_headerless_cad_export_frame(df: pd.DataFrame) -> bool:
    try:
        if df is None or df.empty or len(df.columns) < 4:
            return False
        known_header_hints = {
            'date', 'time', 'datetime', 'latitude', 'longitude', 'lat', 'lon',
            'priority', 'call type', 'call_type', 'call_type_desc', 'nature',
            'description', 'event_type', 'location', 'address', 'city', 'state', 'zip'
        }
        col_names = [str(c).strip().lower() for c in df.columns]
        if any(c in known_header_hints for c in col_names):
            return False

        sample = pd.concat([pd.DataFrame([df.columns], columns=df.columns), df.head(12)], ignore_index=True)
        if sample.empty or len(sample.columns) < 4:
            return False

        first_col = pd.to_numeric(sample.iloc[:, 0], errors='coerce')
        second_col = pd.to_numeric(sample.iloc[:, 1], errors='coerce')
        first_numeric_rate = first_col.notna().mean()
        second_numeric_rate = second_col.notna().mean()
        first_textish = sample.iloc[:, 2].astype(str).str.contains(r'[A-Za-z]', regex=True, na=False).mean()
        later_textish = 0.0
        for idx in range(2, min(len(sample.columns), 12)):
            later_textish = max(
                later_textish,
                sample.iloc[:, idx].astype(str).str.contains(r'[A-Za-z]', regex=True, na=False).mean()
            )
        state_zip_hint = 0.0
        for idx in range(2, min(len(sample.columns), 30)):
            col = sample.iloc[:, idx].astype(str)
            state_zip_hint = max(
                state_zip_hint,
                col.str.fullmatch(r'[A-Z]{2}', na=False).mean(),
                col.str.contains(r'\b\d{5}(?:-\d{4})?\b', regex=True, na=False).mean(),
            )
        return (
            first_numeric_rate >= 0.8 and
            second_numeric_rate >= 0.7 and
            first_textish >= 0.5 and
            later_textish >= 0.5 and
            state_zip_hint >= 0.3
        )
    except Exception:
        return False


def _normalize_headerless_excel_frame(df: pd.DataFrame) -> pd.DataFrame:
    rows = [list(df.columns)] + df.astype(str).fillna('').values.tolist()
    width = max(len(r) for r in rows) if rows else len(df.columns)
    padded = [r + [''] * (width - len(r)) for r in rows]
    norm = pd.DataFrame(padded, columns=[f'col_{i+1}' for i in range(width)])
    return norm


def _deduplicate_columns(df):
    """Rename duplicate column names by appending _2, _3, etc."""
    seen = {}
    new_cols = []
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            new_cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 1
            new_cols.append(c)
    df.columns = new_cols
    return df


def load_raw_call_table(uploaded_file) -> pd.DataFrame:
    fname = str(uploaded_file.name or '').lower()
    if fname.endswith('.numbers'):
        raw_df = load_numbers_dataframe(uploaded_file.getvalue(), uploaded_file.name)
        raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
        return _deduplicate_columns(raw_df).reset_index(drop=True)
    if fname.endswith(_EXCEL_EXTS):
        raw_bytes = uploaded_file.getvalue()
        engine = 'openpyxl'
        if fname.endswith('.xls'):
            engine = 'xlrd'
        elif fname.endswith('.xlsb'):
            engine = 'pyxlsb'
        try:
            import openpyxl as _oxl

            wb = _oxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            sheet_name = max(wb.sheetnames, key=lambda sn: _sheet_score(wb[sn]))
            ws = wb[sheet_name]
            row_iter = ws.iter_rows(values_only=True)
            headers_raw = next(row_iter)
            if headers_raw is None:
                raise ValueError("Selected Excel sheet has no header row.")
            real_idx = [
                i for i, h in enumerate(headers_raw)
                if h is not None and not (str(h).startswith('Column') and str(h)[6:].isdigit())
            ]
            if not real_idx:
                real_idx = [i for i, h in enumerate(headers_raw) if h is not None]
            real_headers = [str(headers_raw[i]).lower().strip() for i in real_idx]
            rows_data = []
            for row in row_iter:
                if row is None:
                    continue
                trimmed = [row[i] if i < len(row) else None for i in real_idx]
                if any(v is not None and str(v).strip() for v in trimmed):
                    rows_data.append(trimmed)
            wb.close()
            raw_df = pd.DataFrame(rows_data, columns=real_headers)
            raw_df = raw_df.dropna(how='all')
            raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
            raw_df = _deduplicate_columns(raw_df)
            _jacksonville_df = _normalize_jacksonville_cfs_report(raw_bytes, filename=uploaded_file.name)
            if _jacksonville_df is not None and not _jacksonville_df.empty:
                raw_df = _jacksonville_df
            else:
                _priority_calls_df = _normalize_loxley_priority_calls_report(raw_bytes, filename=uploaded_file.name)
                if _priority_calls_df is not None and not _priority_calls_df.empty:
                    raw_df = _priority_calls_df
            if _looks_like_headerless_excel_columns(headers_raw) or _looks_like_headerless_cad_export_frame(raw_df):
                raw_df = _normalize_headerless_excel_frame(raw_df)
            return raw_df.reset_index(drop=True)
        except Exception:
            all_sheets = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, sheet_name=None)
            best_score = -10**9
            best_df = None
            for _, sheet_df in all_sheets.items():
                sheet_df.columns = [str(c).lower().strip() for c in sheet_df.columns]
                sheet_df = _deduplicate_columns(sheet_df)
                score = 0
                for col in sheet_df.columns:
                    if col in ('latitude', 'longitude', 'priority', 'location', 'address'):
                        score += 20
                    elif any(k in col for k in ['lat', 'lon', 'priority', 'location', 'date', 'time', 'address']):
                        score += 5
                score += min(len(sheet_df), 100)
                if len(sheet_df.columns) == 1 and str(sheet_df.columns[0]).startswith('externaldata_'):
                    score -= 100
                if score > best_score:
                    best_score = score
                    best_df = sheet_df
            if best_df is None:
                best_df = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, dtype=str)
            best_df.columns = [str(c).lower().strip() for c in best_df.columns]
            best_df = _deduplicate_columns(best_df)
            _jacksonville_df = _normalize_jacksonville_cfs_report(raw_bytes, filename=uploaded_file.name)
            if _jacksonville_df is not None and not _jacksonville_df.empty:
                best_df = _jacksonville_df
            else:
                _priority_calls_df = _normalize_loxley_priority_calls_report(raw_bytes, filename=uploaded_file.name)
                if _priority_calls_df is not None and not _priority_calls_df.empty:
                    best_df = _priority_calls_df
            if _looks_like_headerless_cad_export_frame(best_df):
                best_df = _normalize_headerless_excel_frame(best_df)
            return best_df.reset_index(drop=True)

    content = uploaded_file.getvalue().decode('utf-8', errors='ignore')
    delim = _guess_delimiter(content)
    raw_df = pd.read_csv(io.StringIO(content), sep=delim, dtype=str)
    raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
    return raw_df.reset_index(drop=True)


def _series_non_empty(series: pd.Series) -> pd.Series:
    return series.fillna('').astype(str).map(_clean_text)


def _find_state_column(df: pd.DataFrame) -> str | None:
    named = _find_named_column(df, _STATE_HINTS)
    if named:
        return named
    best = (0.0, None)
    for col in df.columns:
        s = _series_non_empty(_get_col_series(df, col))
        non_empty = s[s != '']
        if non_empty.empty:
            continue
        rate = non_empty.map(lambda v: _clean_state(v) in _STATE_ABBRS).mean()
        if rate > best[0]:
            best = (rate, col)
    return str(best[1]) if best[0] >= 0.65 else None


def _find_zip_column(df: pd.DataFrame) -> str | None:
    named = _find_named_column(df, _ZIP_HINTS)
    if named:
        return named
    best = (0.0, None)
    for col in df.columns:
        s = _series_non_empty(_get_col_series(df, col))
        non_empty = s[s != '']
        if non_empty.empty:
            continue
        rate = non_empty.str.contains(r'^\d{5}(?:-\d{4})?(?:\.0+)?$', regex=True).mean()
        if rate > best[0]:
            best = (rate, col)
    return str(best[1]) if best[0] >= 0.55 else None


def _find_city_column(df: pd.DataFrame, state_col: str | None, zip_col: str | None) -> str | None:
    named = _find_named_column(df, _CITY_HINTS)
    if named:
        return named
    columns = list(df.columns)
    best_score = -1.0
    best_col = None
    state_idx = columns.index(state_col) if state_col in columns else None
    zip_idx = columns.index(zip_col) if zip_col in columns else None
    for idx, col in enumerate(columns):
        if col in {state_col, zip_col}:
            continue
        s = _series_non_empty(_get_col_series(df, col))
        non_empty = s[s != '']
        if non_empty.empty:
            continue
        alpha_rate = non_empty.str.contains(r'[A-Za-z]').mean()
        short_rate = non_empty.str.len().between(2, 24).mean()
        digit_rate = non_empty.str.contains(r'\d').mean()
        if alpha_rate < 0.7 or short_rate < 0.6 or digit_rate > 0.35:
            continue
        score = alpha_rate + short_rate - digit_rate
        if state_idx is not None:
            if idx == state_idx - 1:
                score += 1.5
            elif idx < state_idx:
                score += max(0.0, 1.0 - abs((state_idx - 1) - idx) * 0.2)
        if zip_idx is not None and idx == zip_idx - 2:
            score += 0.7
        if score > best_score:
            best_score = score
            best_col = col
    return str(best_col) if best_col is not None else None


def _find_street_columns(df: pd.DataFrame, city_col: str | None, state_col: str | None, zip_col: str | None) -> list[str]:
    address_col = _find_named_column(df, _ADDRESS_HINTS)
    if address_col:
        return [address_col]

    columns = list(df.columns)
    if city_col not in columns:
        return []

    city_idx = columns.index(city_col)
    start_idx = max(0, city_idx - 8)
    chosen = []
    for idx in range(city_idx - 1, start_idx - 1, -1):
        col = columns[idx]
        if col in {state_col, zip_col}:
            continue
        col_l = str(col).lower()
        if any(hint in col_l for hint in _NOISE_HINTS):
            continue
        s = _series_non_empty(_get_col_series(df, col))
        non_empty = s[s != '']
        if non_empty.empty:
            continue
        sample_vals = non_empty.head(25).astype(str)
        date_like = sample_vals.str.fullmatch(r'\d{4}-\d{2}-\d{2}(?: \d{2}:\d{2}:\d{2})?').mean()
        time_like = sample_vals.str.fullmatch(r'\d{1,2}:\d{2}:\d{2}').mean()
        if date_like >= 0.6 or time_like >= 0.6:
            if chosen:
                break
            continue
        fill_rate = non_empty.shape[0] / max(len(df), 1)
        alpha_rate = non_empty.str.contains(r'[A-Za-z]').mean()
        digit_rate = non_empty.str.contains(r'\d').mean()
        if chosen:
            chosen_has_house = any(
                _series_non_empty(_get_col_series(df, c)).head(25).str.contains(r'\d').mean() >= 0.5
                for c in chosen
            )
            if chosen_has_house and digit_rate < 0.1:
                break
        if fill_rate < 0.1 or (alpha_rate < 0.15 and digit_rate < 0.15):
            if chosen:
                break
            continue
        chosen.append(col)
        if len(chosen) >= 4:
            break

    return [str(c) for c in reversed(chosen)]


def inspect_census_readiness(raw_df: pd.DataFrame) -> dict:
    state_col = _find_state_column(raw_df)
    zip_col = _find_zip_column(raw_df)
    city_col = _find_city_column(raw_df, state_col, zip_col)
    street_cols = _find_street_columns(raw_df, city_col, state_col, zip_col)
    urbanization_col = _find_named_column(raw_df, _URBANIZATION_HINTS)
    return {
        'street_cols': street_cols,
        'city_col': city_col,
        'state_col': state_col,
        'zip_col': zip_col,
        'urbanization_col': urbanization_col,
    }


def _build_street_series(raw_df: pd.DataFrame, street_cols: list[str]) -> pd.Series:
    if not street_cols:
        return pd.Series([''] * len(raw_df), index=raw_df.index, dtype='object')
    if len(street_cols) == 1:
        return _series_non_empty(_get_col_series(raw_df, street_cols[0]))
    frame = pd.DataFrame(index=raw_df.index)
    for idx, col in enumerate(street_cols):
        frame[f'street_{idx}'] = _series_non_empty(_get_col_series(raw_df, col))
    return frame.apply(lambda row: ' '.join(_dedupe_tokens(row.tolist())), axis=1)


def _looks_like_intersection(text: str) -> bool:
    value = _clean_text(text).upper()
    if not value:
        return False
    if any(hint in value for hint in _INTERSECTION_HINTS):
        return True
    return bool(re.search(r'\b(?:AND|CROSS(?:ING)?|INTERSECTION OF)\b', value))


def _split_census_address_kinds(stage_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if stage_df is None or stage_df.empty:
        empty = pd.DataFrame(columns=stage_df.columns if stage_df is not None else [])
        return empty.copy(), empty.copy()

    work = stage_df.copy()
    if 'is_intersection' not in work.columns:
        work['is_intersection'] = False
    if 'has_required_address' not in work.columns:
        work['has_required_address'] = False
    work['is_intersection'] = work['is_intersection'].fillna(False).astype(bool)
    work['has_required_address'] = work['has_required_address'].fillna(False).astype(bool)
    street_df = work[work['has_required_address'] & ~work['is_intersection']].copy().reset_index(drop=True)
    intersection_df = work[work['has_required_address'] & work['is_intersection']].copy().reset_index(drop=True)
    return street_df, intersection_df


def build_intersection_fallback_rows(stage_df: pd.DataFrame) -> pd.DataFrame:
    _, intersection_df = _split_census_address_kinds(stage_df)
    if intersection_df.empty:
        return intersection_df

    fallback_df = intersection_df.copy()
    fallback_df['intersection_query'] = fallback_df.apply(
        lambda row: ', '.join(
            [v for v in [row.get('street', ''), row.get('city', ''), row.get('state', ''), row.get('zip', '')] if _clean_text(v)]
        ),
        axis=1,
    )
    return fallback_df


def build_census_staging(uploaded_files) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    stage_parts = []
    original_parts = []
    diagnostics = []
    for file_idx, uploaded_file in enumerate(uploaded_files):
        raw_df = load_raw_call_table(uploaded_file)
        raw_df = raw_df.reset_index(drop=True)
        source_ids = [f"{file_idx}:{uploaded_file.name}:{row_idx}" for row_idx in range(len(raw_df))]
        original_df = raw_df.copy()
        original_df['_source_row_id'] = source_ids
        original_df['_source_file'] = uploaded_file.name
        original_parts.append(original_df)

        layout = inspect_census_readiness(raw_df)
        street = _build_street_series(raw_df, layout['street_cols'])
        city = _series_non_empty(_get_col_series(raw_df, layout['city_col'])) if layout['city_col'] else pd.Series([''] * len(raw_df))
        state = _get_col_series(raw_df, layout['state_col']).map(_clean_state) if layout['state_col'] else pd.Series([''] * len(raw_df))
        zip_code = _get_col_series(raw_df, layout['zip_col']).map(_clean_zip) if layout['zip_col'] else pd.Series([''] * len(raw_df))
        urbanization = _series_non_empty(_get_col_series(raw_df, layout['urbanization_col'])) if layout['urbanization_col'] else pd.Series([''] * len(raw_df))

        part = pd.DataFrame({
            'source_id': source_ids,
            'street': street,
            'city': city,
            'state': state,
            'zip': zip_code,
            'urbanization': urbanization,
            'source_file': uploaded_file.name,
        })
        part['street'] = part['street'].map(_clean_text)
        part['city'] = part['city'].map(_clean_text)
        part['state'] = part['state'].map(_clean_state)
        part['zip'] = part['zip'].map(_clean_zip)
        part['urbanization'] = part['urbanization'].map(_clean_text)
        part['is_intersection'] = part['street'].map(_looks_like_intersection)
        part['address_kind'] = part['is_intersection'].map({True: 'intersection', False: 'street'})
        part['has_required_address'] = (
            part['street'].ne('') &
            part['city'].ne('') &
            part['state'].isin(_STATE_ABBRS) &
            part['zip'].str.match(r'^\d{5}$')
        )
        part['census_preview'] = part.apply(
            lambda row: ', '.join([v for v in [row['street'], row['city'], row['state'], row['zip']] if v]),
            axis=1,
        )
        diagnostics.append({
            'file': uploaded_file.name,
            'rows': int(len(part)),
            'ready_rows': int((part['has_required_address'] & ~part['is_intersection']).sum()),
            'intersection_rows': int((part['has_required_address'] & part['is_intersection']).sum()),
            'street_cols': list(layout['street_cols']),
            'city_col': layout['city_col'] or '',
            'state_col': layout['state_col'] or '',
            'zip_col': layout['zip_col'] or '',
            'urbanization_col': layout['urbanization_col'] or '',
        })
        stage_parts.append(part)

    stage_df = pd.concat(stage_parts, ignore_index=True) if stage_parts else pd.DataFrame()
    original_df = pd.concat(original_parts, ignore_index=True) if original_parts else pd.DataFrame()
    summary = {
        'files': diagnostics,
        'rows_total': int(len(stage_df)),
        'rows_ready': int((stage_df['has_required_address'] & ~stage_df['is_intersection']).sum()) if not stage_df.empty else 0,
        'rows_missing': int((~stage_df['has_required_address']).sum()) if not stage_df.empty else 0,
        'intersection_rows': int((stage_df['has_required_address'] & stage_df['is_intersection']).sum()) if 'is_intersection' in stage_df.columns else 0,
    }
    return stage_df, original_df, summary


def make_census_batch_chunks(stage_df: pd.DataFrame, chunk_size: int = 10000) -> list[dict]:
    ready_df, _intersection_df = _split_census_address_kinds(stage_df)
    chunks = []
    for start in range(0, len(ready_df), chunk_size):
        chunk = ready_df.iloc[start:start + chunk_size].copy().reset_index(drop=True)
        chunks.append(build_census_chunk_payload(chunk, chunk_index=len(chunks) + 1))
    return chunks


def build_census_chunk_payload(chunk_df: pd.DataFrame, chunk_index: int = 1, filename: str | None = None) -> dict:
    export_df = chunk_df[['source_id', 'street', 'city', 'state', 'zip', 'urbanization']].copy()
    export_df['urbanization'] = export_df['urbanization'].where(export_df['urbanization'].ne(''), None)
    csv_bytes = export_df.to_csv(index=False, header=False, lineterminator='\n')
    return {
        'index': int(chunk_index),
        'rows': int(len(chunk_df)),
        'filename': filename or f'census_batch_part_{int(chunk_index):04d}.csv',
        'csv_bytes': csv_bytes.encode('utf-8'),
        'source_ids': chunk_df['source_id'].tolist(),
        'frame': chunk_df.copy().reset_index(drop=True),
    }


def make_census_batch_zip(chunks: list[dict]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        manifest_rows = []
        for chunk in chunks:
            zf.writestr(chunk['filename'], chunk['csv_bytes'])
            manifest_rows.append({
                'filename': chunk['filename'],
                'rows': chunk['rows'],
            })
        if manifest_rows:
            zf.writestr('manifest.json', json.dumps(manifest_rows, indent=2))
    return buff.getvalue()


def make_sample_census_batch(stage_df: pd.DataFrame, sample_size: int = 250) -> dict | None:
    ready_df, _intersection_df = _split_census_address_kinds(stage_df)
    if ready_df.empty:
        return None
    sample_df = ready_df.head(sample_size).copy()
    export_df = sample_df[['source_id', 'street', 'city', 'state', 'zip', 'urbanization']].copy()
    export_df['urbanization'] = export_df['urbanization'].where(export_df['urbanization'].ne(''), None)
    return {
        'rows': int(len(sample_df)),
        'filename': 'census_sample_batch.csv',
        'csv_bytes': export_df.to_csv(index=False, header=False, lineterminator='\n').encode('utf-8'),
    }


def _normalize_census_result_frame(raw_df: pd.DataFrame) -> pd.DataFrame:
    width = len(raw_df.columns)
    cols = ['source_id', 'input_address', 'match_status', 'match_type', 'matched_address', 'lonlat', 'tiger_id', 'side']
    if width > len(cols):
        cols += [f'extra_{i}' for i in range(1, width - len(cols) + 1)]
    raw_df = raw_df.copy()
    raw_df.columns = cols[:width]
    if not raw_df.empty:
        first_row = [str(v).strip().lower() for v in raw_df.iloc[0].tolist()]
        if first_row[:6] == ['source_id', 'input_address', 'match_status', 'match_type', 'matched_address', 'lonlat']:
            raw_df = raw_df.iloc[1:].reset_index(drop=True)
    return raw_df


def parse_census_result_bytes(content: bytes, filename: str = 'census_result.csv') -> pd.DataFrame:
    text = content.decode('utf-8', errors='ignore')
    rows = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if not row:
            continue
        if len(row) < 8:
            row = row + [''] * (8 - len(row))
        rows.append(row[:8])
    raw_df = pd.DataFrame(rows)
    norm = _normalize_census_result_frame(raw_df)
    pair = norm['lonlat'].fillna('').astype(str).str.extract(r'^\s*(-?[\d.]+)\s*,\s*(-?[\d.]+)\s*$')
    norm['lon'] = pd.to_numeric(pair[0], errors='coerce')
    norm['lat'] = pd.to_numeric(pair[1], errors='coerce')
    norm['source_id'] = norm['source_id'].astype(str).str.strip()
    norm['match_status'] = norm['match_status'].fillna('').astype(str).str.strip()
    norm['source_file'] = filename
    norm = norm[norm['source_id'] != '']
    return norm


def parse_census_result_files(uploaded_files) -> pd.DataFrame:
    parts = []
    for uploaded_file in uploaded_files:
        suffix = Path(str(uploaded_file.name or '')).suffix.lower()
        if suffix not in {'.csv', '.txt'}:
            continue
        norm = parse_census_result_bytes(uploaded_file.getvalue(), uploaded_file.name)
        parts.append(norm)
    if not parts:
        return pd.DataFrame(columns=['source_id', 'lat', 'lon'])
    result_df = pd.concat(parts, ignore_index=True)
    result_df = result_df.drop_duplicates(subset=['source_id'], keep='first')
    return result_df


def merge_census_results(
    partial_calls_df: pd.DataFrame,
    result_df: pd.DataFrame,
    *,
    validate_outputs: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Merge CAD data with Census batch geocoding results.

    This function now uses the optimized efficient_merge module which:
    - Uses Polars for 5-10x speedup on large datasets (>50K rows)
    - Falls back to pandas for compatibility
    - Includes data validation at merge completion
    - Provides better memory efficiency

    Args:
        partial_calls_df: CAD DataFrame with _source_row_id
        result_df: Census batch results with source_id, lat, lon, match_status

    Returns:
        Tuple of (merged_df, ready_df, summary_dict)
    """
    # Use pandas merge — Polars' internal Rayon thread pool can deadlock
    # on Windows during Streamlit reruns, and pandas is fast enough for
    # the typical Census batch sizes (~10-25K rows).
    merged, ready_df, summary = merge_census_results_fast(
        partial_calls_df,
        result_df,
        use_polars=False
    )

    if validate_outputs:
        # Validate Census results. Parsed batch files keep source_id as text for
        # stable joins, so coerce only the validation copy to match the schema.
        if not result_df.empty:
            validation_result_df = result_df.copy()
            if 'source_id' in validation_result_df.columns:
                validation_result_df['source_id'] = pd.to_numeric(
                    validation_result_df['source_id'],
                    errors='coerce',
                ).astype('Int64')
            for coord_col in ('lat', 'lon'):
                if coord_col in validation_result_df.columns:
                    validation_result_df[coord_col] = pd.to_numeric(
                        validation_result_df[coord_col],
                        errors='coerce',
                    )
            validate_census_results(validation_result_df, raise_exceptions=False)

        # Validate merged data
        validate_merged_data(merged, raise_exceptions=False)

    # Maintain backward compatibility with old summary format
    summary['rows_with_census_match'] = summary.pop('rows_geocoded', 0)

    return merged, ready_df, summary


def _encode_multipart_formdata(fields: dict, file_field_name: str, filename: str, file_bytes: bytes, content_type: str = 'text/csv'):
    boundary = f'----FrankensteinBoundary{uuid.uuid4().hex}'
    body = io.BytesIO()
    for key, value in fields.items():
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode('utf-8'))
        body.write(str(value).encode('utf-8'))
        body.write(b'\r\n')
    body.write(f'--{boundary}\r\n'.encode('utf-8'))
    body.write(
        f'Content-Disposition: form-data; name="{file_field_name}"; filename="{filename}"\r\n'.encode('utf-8')
    )
    body.write(f'Content-Type: {content_type}\r\n\r\n'.encode('utf-8'))
    body.write(file_bytes)
    body.write(b'\r\n')
    body.write(f'--{boundary}--\r\n'.encode('utf-8'))
    return body.getvalue(), f'multipart/form-data; boundary={boundary}'


def submit_census_batch_chunk(
    csv_bytes: bytes,
    filename: str,
    *,
    benchmark: str = '4',
    timeout: int = 180,
    returntype: str = 'locations',
    retries: int = 3,
    attempt_logger=None,
) -> tuple[pd.DataFrame, bytes]:
    fields = {'benchmark': benchmark}
    url = f'https://geocoding.geo.census.gov/geocoder/{returntype}/addressbatch'
    body, content_type = _encode_multipart_formdata(fields, 'addressFile', filename, csv_bytes)
    last_error = None
    for attempt in range(1, max(1, retries) + 1):
        if attempt_logger is not None:
            attempt_logger(
                f"Census chunk request attempt {attempt}/{max(1, retries)} for {filename} "
                f"(timeout {timeout}s)."
            )
        req = urllib.request.Request(
            url,
            data=body,
            headers={
                'Content-Type': content_type,
                'Content-Length': str(len(body)),
                'User-Agent': 'Frankenstein-Census-Geocoder/1.0',
                'Connection': 'close',
            },
            method='POST',
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                resp_bytes = resp.read()
                status = getattr(resp, 'status', 200)
            break
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode('utf-8', errors='ignore')
            if attempt_logger is not None:
                attempt_logger(f"Census HTTP {exc.code} for {filename}: {detail[:160] or exc.reason}")
            raise RuntimeError(f'Census HTTP {exc.code}: {detail[:400] or exc.reason}') from exc
        except (urllib.error.URLError, http.client.RemoteDisconnected, TimeoutError, ConnectionResetError) as exc:
            last_error = exc
            if attempt >= max(1, retries):
                raise RuntimeError(f'Census connection failed after {attempt} attempt(s): {exc}') from exc
            if attempt_logger is not None:
                attempt_logger(
                    f"Census chunk attempt {attempt}/{max(1, retries)} for {filename} failed: {exc}. "
                    f"Retrying in {min(6, attempt * 2)}s."
                )
            time.sleep(min(6, attempt * 2))
            continue
        except Exception as exc:
            last_error = exc
            if attempt >= max(1, retries):
                raise RuntimeError(f'Census request failed after {attempt} attempt(s): {exc}') from exc
            if attempt_logger is not None:
                attempt_logger(
                    f"Census chunk attempt {attempt}/{max(1, retries)} for {filename} failed: {exc}. "
                    f"Retrying in {min(6, attempt * 2)}s."
                )
            time.sleep(min(6, attempt * 2))
            continue
    else:
        raise RuntimeError(f'Census request failed: {last_error}')

    if status >= 400:
        raise RuntimeError(f'Census returned HTTP status {status}.')
    if not resp_bytes.strip():
        raise RuntimeError('Census returned an empty batch response.')

    result_df = parse_census_result_bytes(resp_bytes, filename=filename)
    if result_df.empty:
        preview = resp_bytes[:400].decode('utf-8', errors='ignore')
        raise RuntimeError(f'Census returned no parseable batch rows. Response preview: {preview}')
    return result_df, resp_bytes


def build_corrected_export(original_df: pd.DataFrame, result_df: pd.DataFrame) -> pd.DataFrame:
    export_df = original_df.copy().reset_index(drop=True)
    result = pd.DataFrame() if result_df is None else result_df.copy()
    required = ['source_id', 'lat', 'lon', 'match_status', 'match_type', 'matched_address']
    for column in required:
        if column not in result.columns:
            result[column] = pd.NA
    if '_source_row_id' not in export_df.columns:
        export_df['_source_row_id'] = range(len(export_df))
    export_df['_census_merge_key'] = export_df['_source_row_id'].astype('string').str.strip().str.replace(r'\.0$', '', regex=True)
    matched = result[required].copy()
    matched['_census_merge_key'] = matched['source_id'].astype('string').str.strip().str.replace(r'\.0$', '', regex=True)
    matched = matched.drop(columns=['source_id'])
    matched = matched[
        matched['_census_merge_key'].notna()
        & (matched['_census_merge_key'].astype(str).str.strip() != '')
    ].drop_duplicates(subset=['_census_merge_key'], keep='first')
    export_df = export_df.merge(matched, on='_census_merge_key', how='left')
    export_df = export_df.drop(columns=['_census_merge_key'], errors='ignore')
    if 'lat' in export_df.columns:
        export_df['lat'] = pd.to_numeric(export_df['lat'], errors='coerce')
    if 'lon' in export_df.columns:
        export_df['lon'] = pd.to_numeric(export_df['lon'], errors='coerce')
    return export_df


def build_corrected_export_from_merged(merged_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the corrected Census export from an already merged dataframe.

    This avoids repeating the merge work when the app already has the merged
    output in memory.
    """
    export_df = pd.DataFrame() if merged_df is None else merged_df.copy().reset_index(drop=True)
    export_df = export_df.drop(columns=['_census_merge_key', '_census_filled'], errors='ignore')
    if 'lat' in export_df.columns:
        export_df['lat'] = pd.to_numeric(export_df['lat'], errors='coerce')
    if 'lon' in export_df.columns:
        export_df['lon'] = pd.to_numeric(export_df['lon'], errors='coerce')
    return export_df
