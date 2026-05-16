"""
CAD file parser and metadata extraction for BRINC app.
"""

import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import io
import json
import datetime
import math
import urllib.parse
import urllib.request
from pathlib import Path

import pyproj
from modules.config import STATE_FIPS, US_STATES_ABBR, KNOWN_POPULATIONS
from modules.numbers_adapter import load_numbers_dataframe


def _normalize_jacksonville_cfs_report(raw_bytes, filename=""):
    """Flatten the Jacksonville PD CFS report into one row per incident."""

    def _cell_text(value) -> str:
        if value is None:
            return ''
        if hasattr(value, 'strftime'):
            try:
                if isinstance(value, datetime.datetime):
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(value, datetime.time):
                    return value.strftime('%H:%M:%S')
                if isinstance(value, datetime.date):
                    return value.strftime('%Y-%m-%d')
            except Exception:
                pass
        text = str(value).strip()
        if text.lower() in {'nan', 'none', 'nat'}:
            return ''
        return re.sub(r'\s+', ' ', text)

    def _read_source_frame():
        text = raw_bytes.decode('utf-8', errors='ignore') if isinstance(raw_bytes, (bytes, bytearray)) else str(raw_bytes)
        first_line = next((line for line in text.splitlines() if line.strip()), '')
        delim = '\t' if first_line.count('\t') > first_line.count(',') else ','
        readers = (
            lambda: pd.read_excel(io.BytesIO(raw_bytes), header=None, dtype=object, engine='openpyxl'),
            lambda: pd.read_csv(io.StringIO(text), header=None, dtype=object, sep=delim),
        )
        for reader in readers:
            try:
                frame = reader()
                if frame is not None and not frame.empty:
                    return frame
            except Exception:
                continue
        return None

    def _first_nonempty(*values) -> str:
        for value in values:
            text = _cell_text(value)
            if text:
                return text
        return ''

    def _extract_city_hint(text: str) -> str:
        city = _cell_text(text)
        city = re.sub(r'\s+(?:police|fire|sheriff|ems|rescue)\s+department$', '', city, flags=re.I).strip()
        city = re.sub(r'\s+department$', '', city, flags=re.I).strip()
        city = re.sub(r'\s+cfs\s+report$', '', city, flags=re.I).strip()
        return re.sub(r'\s+', ' ', city).title()

    def _parse_header_state_zip(df) -> tuple[str, str]:
        header_text = ' '.join(
            _cell_text(df.iat[r, c])
            for r in range(min(df.shape[0], 6))
            for c in range(min(df.shape[1], 12))
        )
        m = re.search(r'\b([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b', header_text)
        if m:
            return m.group(1).upper(), m.group(2)
        return '', ''

    def _split_location(text: str, fallback_city: str) -> tuple[str, str]:
        loc = _cell_text(text)
        if not loc:
            return '', fallback_city
        parts = [p.strip() for p in re.split(r'\s*,\s*', loc) if p and p.strip()]
        def _clean_city_name(city_text: str) -> str:
            city = _cell_text(city_text)
            city = re.sub(r'\s+location\s*$', '', city, flags=re.I).strip()
            return re.sub(r'\s+', ' ', city).title()

        def _looks_like_street_piece(piece: str) -> bool:
            piece_u = _cell_text(piece).upper()
            if not piece_u:
                return False
            if re.search(r'\d', piece_u):
                return True
            if '/' in piece_u:
                return True
            return bool(re.search(
                r'\b(?:ST|STREET|RD|ROAD|AVE|AVENUE|BLVD|BOULEVARD|DR|DRIVE|LN|LANE|HWY|HIGHWAY|PKWY|PARKWAY|CIR|CIRCLE|CT|COURT|WAY|LOOP|TRAIL|TRL|PL|PLACE|TER|TERRACE|EXPY|EXPRESSWAY)\b',
                piece_u,
            ))

        def _clean_street_name(street_text: str) -> str:
            street = _cell_text(street_text)
            if '|' in street:
                pipe_parts = [p.strip() for p in street.split('|') if p and p.strip()]
                if pipe_parts:
                    street = next((p for p in reversed(pipe_parts) if _looks_like_street_piece(p)), pipe_parts[-1])
            street = re.sub(r',\s*[^,]+(?:\s+LOCATION)?\s*$', '', street, flags=re.I).strip()
            street = re.sub(r'\s+location\s*$', '', street, flags=re.I).strip()
            return re.sub(r'\s+', ' ', street)

        if len(parts) >= 2:
            tail = parts[-1].strip()
            if tail and not re.search(r'\d', tail) and len(tail) <= 40 and not re.search(r'\b(?:AL|AK|AZ|AR|CA|CO|CT|DC|DE|FL|GA|HI|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VA|VT|WA|WI|WV|WY)\b', tail, re.I):
                return _clean_street_name(', '.join(parts[:-1]).strip()), _clean_city_name(tail)

        return _clean_street_name(loc), _clean_city_name(fallback_city)

    def _priority_value(text: str):
        text_u = _cell_text(text).upper()
        if not text_u:
            return ''
        if 'HIGH' in text_u:
            return 1
        if 'MED' in text_u:
            return 2
        if 'LOW' in text_u:
            return 3
        m = re.match(r'^\s*(\d+)', text_u)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return ''
        return ''

    def _parse_call_detail(text: str) -> tuple[str, str, str]:
        parts = [_cell_text(part) for part in str(text or '').split('|')]
        parts = [part for part in parts if part]
        if not parts:
            return '', '', ''
        call_type = parts[0]
        priority = ''
        disposition = ''
        for part in parts[1:]:
            if not priority and (_priority_value(part) or re.fullmatch(r'(?:[1-9]|0?[1-9])(?:\s*-\s*.*)?', _cell_text(part), re.I)):
                priority = part
                continue
            disposition = f"{disposition} {part}".strip() if disposition else part
        return call_type, priority, disposition

    def _find_header_row(df):
        for row_idx in range(min(len(df), 40)):
            values = [_cell_text(v) for v in df.iloc[row_idx].tolist()]
            values = [v for v in values if v]
            values_l = [v.lower() for v in values]
            joined = ' '.join(values_l)
            if (
                any(v in {'cfs #', 'cfs#', 'cfs number'} for v in values_l)
                and any(v == 'location' for v in values_l)
                and ('calltype' in joined or 'call type' in joined)
                and 'priority' in joined
            ):
                cfs_col = None
                location_col = None
                detail_col = None
                for col_idx, cell in enumerate(df.iloc[row_idx].tolist()):
                    cell_text = _cell_text(cell).lower()
                    if cfs_col is None and (cell_text == 'cfs #' or 'cfs #' in cell_text or cell_text in {'cfs#', 'cfs number'}):
                        cfs_col = col_idx
                    if location_col is None and cell_text == 'location':
                        location_col = col_idx
                    if detail_col is None and ('calltype' in cell_text or 'call type' in cell_text):
                        detail_col = col_idx
                if cfs_col is not None and location_col is not None and detail_col is not None:
                    return row_idx, cfs_col, location_col, detail_col
        return None

    try:
        df = _read_source_frame()
        if df is None or df.empty:
            return None

        header_row = _find_header_row(df)
        if header_row is None:
            return None

        header_row_idx, cfs_col, location_col, detail_col = header_row
        filename_lower = str(filename or '').lower()
        department_name = _first_nonempty(
            df.iat[0, 3] if df.shape[0] > 0 and df.shape[1] > 3 else '',
            df.iat[1, 3] if df.shape[0] > 1 and df.shape[1] > 3 else '',
            df.iat[0, 4] if df.shape[0] > 0 and df.shape[1] > 4 else '',
            filename,
        )
        header_city = _extract_city_hint(department_name)
        if not header_city and 'jacksonville' in filename_lower:
            header_city = 'Jacksonville'
        header_state, header_zip = _parse_header_state_zip(df)
        if not header_state and 'jacksonville' in filename_lower:
            header_state = 'FL'

        rows = []
        current = None
        for row_idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[row_idx]
            cfs_value = _cell_text(row.iloc[cfs_col] if cfs_col < len(row) else '')
            location_value = _cell_text(row.iloc[location_col] if location_col < len(row) else '')
            detail_value = _cell_text(row.iloc[detail_col] if detail_col < len(row) else '')

            if not (cfs_value or location_value or detail_value):
                continue

            is_new = bool(re.fullmatch(r'\d{6,}(?:\.\d+)?', cfs_value))
            if is_new:
                if current is not None:
                    rows.append(current)
                call_type, priority_text, disposition = _parse_call_detail(detail_value)
                street, city = _split_location(location_value, header_city)
                current = {
                    'cfs_number': cfs_value,
                    'location': location_value,
                    'street': street,
                    'city': city or header_city,
                    'state': header_state,
                    'zip': header_zip,
                    'call_type_desc': call_type,
                    'priority': _priority_value(priority_text),
                    'disposition': disposition,
                    'agency': 'police',
                    'department_name': department_name,
                    '_csv_city': city or header_city,
                    '_csv_state': header_state,
                }
                continue

            if current is None:
                continue

            if location_value:
                # For Jacksonville format, only use the primary location (first row).
                # Continuation rows typically contain caller names, not address info.
                if not current.get('street'):  # Only set if we haven't parsed a street yet
                    street, city = _split_location(location_value, current.get('city') or header_city)
                    if street:
                        current['location'] = location_value
                        current['street'] = street
                    if city:
                        current['city'] = city
                        current['_csv_city'] = city

            if detail_value:
                call_type, priority_text, disposition = _parse_call_detail(detail_value)
                if call_type:
                    current['call_type_desc'] = _first_nonempty(current.get('call_type_desc', ''), call_type)
                if priority_text:
                    current['priority'] = _priority_value(priority_text)
                if disposition:
                    current['disposition'] = _first_nonempty(current.get('disposition', ''), disposition)

        if current is not None:
            rows.append(current)
        if not rows:
            return None

        out = pd.DataFrame(rows)
        out['_source_row_id'] = [f"{filename}:{i}" if filename else str(i) for i in range(len(out))]
        out['_source_file'] = filename
        out['street'] = out['street'].fillna('').astype(str).str.strip()
        out['city'] = out['city'].fillna('').astype(str).str.strip().str.title()
        out['state'] = out['state'].fillna('').astype(str).str.strip().str.upper()
        out['zip'] = out['zip'].fillna('').astype(str).str.strip()
        out['call_type_desc'] = out['call_type_desc'].fillna('').astype(str).str.strip()
        out['disposition'] = out['disposition'].fillna('').astype(str).str.strip()
        out['location'] = out['location'].fillna('').astype(str).str.strip()
        out['priority'] = pd.to_numeric(out['priority'], errors='coerce').fillna(3).astype(int)
        out['_special_layout'] = 'jacksonville_cfs_report'
        if '_csv_city' not in out.columns or not out['_csv_city'].fillna('').astype(str).str.strip().any():
            out['_csv_city'] = header_city or out['city']
        if '_csv_state' not in out.columns or not out['_csv_state'].fillna('').astype(str).str.strip().any():
            out['_csv_state'] = header_state or out['state']
        return out.reset_index(drop=True)
    except Exception:
        return None


def _normalize_loxley_priority_calls_report(raw_bytes, filename=""):
    """Flatten the Priority Calls workbook into one row per incident."""

    def _cell_text(value) -> str:
        if value is None:
            return ''
        if hasattr(value, 'strftime'):
            try:
                if isinstance(value, datetime.datetime):
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(value, datetime.time):
                    return value.strftime('%H:%M:%S')
                if isinstance(value, datetime.date):
                    return value.strftime('%Y-%m-%d')
            except Exception:
                pass
        text = str(value).strip()
        if text.lower() in {'nan', 'none', 'nat'}:
            return ''
        return re.sub(r'\s+', ' ', text)

    def _read_source_frame():
        text = raw_bytes.decode('utf-8', errors='ignore') if isinstance(raw_bytes, (bytes, bytearray)) else str(raw_bytes)
        first_line = next((line for line in text.splitlines() if line.strip()), '')
        delim = '\t' if first_line.count('\t') > first_line.count(',') else ','
        readers = (
            lambda: pd.read_excel(io.BytesIO(raw_bytes), header=None, dtype=object, engine='openpyxl'),
            lambda: pd.read_csv(io.StringIO(text), header=None, dtype=object, sep=delim),
        )
        for reader in readers:
            try:
                frame = reader()
                if frame is not None and not frame.empty:
                    return frame
            except Exception:
                continue
        return None

    def _first_nonempty(*values) -> str:
        for value in values:
            text = _cell_text(value)
            if text:
                return text
        return ''

    def _priority_value(text: str):
        text_u = _cell_text(text).upper()
        if not text_u:
            return ''
        if 'HIGH' in text_u:
            return 1
        if 'MED' in text_u:
            return 2
        if 'LOW' in text_u:
            return 3
        m = re.match(r'^\s*(\d+)', text_u)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return ''
        return ''

    def _looks_like_priority(text: str) -> bool:
        text_u = _cell_text(text).upper()
        if not text_u:
            return False
        return bool(
            re.fullmatch(r'(?:[1-9]|0?[1-9])(?:\s*-\s*.*)?', text_u)
            or any(k in text_u for k in ('LOW', 'MED', 'MEDIUM', 'HIGH', 'PRIORITY'))
        )

    def _parse_call_detail(text: str) -> tuple[str, str, str]:
        parts = [_cell_text(part) for part in str(text or '').split('|')]
        parts = [part for part in parts if part]
        if not parts:
            return '', '', ''
        call_type = parts[0]
        priority = ''
        disposition = ''
        for part in parts[1:]:
            if not priority and _looks_like_priority(part):
                priority = part
                continue
            disposition = f"{disposition} {part}".strip() if disposition else part
        return call_type, priority, disposition

    def _split_location(text: str, fallback_city: str) -> tuple[str, str]:
        loc = _cell_text(text)
        if not loc:
            return '', fallback_city
        parts = [p.strip() for p in re.split(r'\s*,\s*', loc) if p and p.strip()]
        if len(parts) >= 2:
            tail = parts[-1].strip()
            if tail and not re.search(r'\d', tail) and len(tail) <= 40 and not re.search(r'\b(?:AL|AK|AZ|AR|CA|CO|CT|DC|DE|FL|GA|HI|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VA|VT|WA|WI|WV|WY)\b', tail, re.I):
                return ', '.join(parts[:-1]).strip(), tail.title()
        return loc, fallback_city

    def _parse_header_state_zip(df) -> tuple[str, str]:
        header_text = ' '.join(
            _cell_text(df.iat[r, c])
            for r in range(min(df.shape[0], 4))
            for c in range(min(df.shape[1], 10))
        )
        m = re.search(r'\b([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b', header_text)
        if m:
            return m.group(1).upper(), m.group(2)
        return '', ''

    try:
        df = _read_source_frame()
        if df is None or df.empty:
            return None

        header_row_idx = None
        cfs_col = None
        location_col = None
        detail_col = None
        for row_idx in range(min(len(df), 30)):
            values = [_cell_text(v) for v in df.iloc[row_idx].tolist()]
            values = [v for v in values if v]
            values_l = [v.lower() for v in values]
            joined = ' '.join(values_l)
            if (
                any(v in {'cfs #', 'cfs#', 'cfs number'} for v in values_l)
                and any(v == 'location' for v in values_l)
                and ('calltype' in joined or 'call type' in joined)
                and 'priority' in joined
            ):
                header_row_idx = row_idx
                for col_idx, cell in enumerate(df.iloc[row_idx].tolist()):
                    cell_text = _cell_text(cell).lower()
                    if cfs_col is None and (cell_text == 'cfs #' or 'cfs #' in cell_text or cell_text in {'cfs#', 'cfs number'}):
                        cfs_col = col_idx
                    if location_col is None and cell_text == 'location':
                        location_col = col_idx
                    if detail_col is None and 'calltype' in cell_text:
                        detail_col = col_idx
                break

        if header_row_idx is None or cfs_col is None or location_col is None or detail_col is None:
            return None

        header_city = _first_nonempty(
            df.iat[0, 3] if df.shape[0] > 0 and df.shape[1] > 3 else '',
            df.iat[1, 3] if df.shape[0] > 1 and df.shape[1] > 3 else '',
            df.iat[0, 4] if df.shape[0] > 0 and df.shape[1] > 4 else '',
        )
        header_city = re.sub(r'\s+(?:police|fire|sheriff|ems|rescue)\s+department$', '', header_city, flags=re.I).strip()
        header_city = re.sub(r'\s+department$', '', header_city, flags=re.I).strip()
        header_city = re.sub(r'\s+', ' ', header_city).title()

        department_name = _first_nonempty(
            df.iat[0, 3] if df.shape[0] > 0 and df.shape[1] > 3 else '',
            df.iat[1, 3] if df.shape[0] > 1 and df.shape[1] > 3 else '',
            df.iat[0, 4] if df.shape[0] > 0 and df.shape[1] > 4 else '',
        )
        header_state, header_zip = _parse_header_state_zip(df)

        rows = []
        current = None
        for row_idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[row_idx]
            cfs_value = _cell_text(row.iloc[cfs_col] if cfs_col < len(row) else '')
            location_value = _cell_text(row.iloc[location_col] if location_col < len(row) else '')
            detail_value = _cell_text(row.iloc[detail_col] if detail_col < len(row) else '')

            if not (cfs_value or location_value or detail_value):
                continue

            is_new = bool(re.fullmatch(r'\d{4}-\d{6}(?:\.\d+)?', cfs_value))
            if is_new:
                if current is not None:
                    rows.append(current)
                call_type, priority_text, disposition = _parse_call_detail(detail_value)
                street, city = _split_location(location_value, header_city)
                current = {
                    'cfs_number': cfs_value,
                    'location': location_value,
                    'street': street,
                    'city': city or header_city,
                    'state': header_state,
                    'zip': header_zip,
                    'call_type_desc': call_type,
                    'priority': _priority_value(priority_text),
                    'disposition': disposition,
                    'agency': 'police',
                    'department_name': department_name,
                    '_csv_city': city or header_city,
                    '_csv_state': header_state,
                }
                continue

            if current is None:
                continue

            if location_value:
                combined_location = _cell_text(current.get('location', ''))
                combined_location = f"{combined_location} {location_value}".strip() if combined_location else location_value
                current['location'] = combined_location
                street, city = _split_location(combined_location, current.get('city') or header_city)
                if street:
                    current['street'] = street
                if city:
                    current['city'] = city
                    current['_csv_city'] = city

            if detail_value:
                call_type, priority_text, disposition = _parse_call_detail(detail_value)
                if call_type:
                    current['call_type_desc'] = _first_nonempty(current.get('call_type_desc', ''), call_type)
                if priority_text:
                    current['priority'] = _priority_value(priority_text)
                if disposition:
                    current['disposition'] = _first_nonempty(current.get('disposition', ''), disposition)

        if current is not None:
            rows.append(current)
        if not rows:
            return None

        out = pd.DataFrame(rows)
        out['_source_row_id'] = [f"{filename}:{i}" if filename else str(i) for i in range(len(out))]
        out['_source_file'] = filename
        out['priority'] = pd.to_numeric(out['priority'], errors='coerce').fillna(3).astype(int)
        out['location'] = out['location'].fillna('').astype(str).str.strip()
        out['call_type_desc'] = out['call_type_desc'].fillna('').astype(str).str.strip()
        out['disposition'] = out['disposition'].fillna('').astype(str).str.strip()
        out['_special_layout'] = 'priority_calls_report'
        return out.reset_index(drop=True)
    except Exception:
        return None
def _safe_notna_ratio(values) -> float:
    """Return a stable non-null ratio for possibly empty parse results."""
    try:
        if values is None:
            return 0.0
        size = int(getattr(values, 'size', len(values)))
        if size <= 0:
            return 0.0
        return float(pd.Series(values).notna().mean())
    except Exception:
        return 0.0

def _safe_notna_ratio(values) -> float:
    """Return a stable non-null ratio for possibly empty parse results."""
    try:
        if values is None:
            return 0.0
        size = int(getattr(values, 'size', len(values)))
        if size <= 0:
            return 0.0
        return float(pd.Series(values).notna().mean())
    except Exception:
        return 0.0

def _extract_file_meta(raw_df, res_df, filename=""):
    """
    Compute and return a dict of data-matrix statistics from a parsed CAD upload.
    Call this once per file inside aggressive_parse_calls() and store the result
    in st.session_state['file_meta'].  All values are JSON-safe scalars or strings.
    """
    meta = {}
    try:
        meta['uploaded_filename']   = str(filename)
        meta['file_row_count']      = int(len(raw_df))
        meta['file_col_count']      = int(len(raw_df.columns))
        meta['file_col_names']      = json.dumps(list(raw_df.columns))

        # ── City / state inferred from the file ──────────────────────────────
        meta['file_inferred_city']  = str(res_df['_csv_city'].iloc[0])  if '_csv_city'  in res_df.columns and len(res_df) > 0 else ''
        meta['file_inferred_state'] = str(res_df['_csv_state'].iloc[0]) if '_csv_state' in res_df.columns and len(res_df) > 0 else ''

        # ── Date range ───────────────────────────────────────────────────────
        if 'date' in res_df.columns:
            _dates = pd.to_datetime(res_df['date'], format='mixed', errors='coerce').dropna()
            if not _dates.empty:
                meta['file_date_range_start'] = _dates.min().strftime('%Y-%m-%d')
                meta['file_date_range_end']   = _dates.max().strftime('%Y-%m-%d')
                meta['file_date_span_days']   = int((_dates.max() - _dates.min()).days)
                meta['peak_month']            = int(_dates.dt.month.value_counts().idxmax())
                meta['peak_day_of_week']      = int(_dates.dt.dayofweek.value_counts().idxmax())
            else:
                meta['file_date_range_start'] = ''
                meta['file_date_range_end']   = ''
                meta['file_date_span_days']   = 0
                meta['peak_month']            = 0
                meta['peak_day_of_week']      = 0
        else:
            meta['file_date_range_start'] = ''
            meta['file_date_range_end']   = ''
            meta['file_date_span_days']   = 0
            meta['peak_month']            = 0
            meta['peak_day_of_week']      = 0

        # ── Peak hour ────────────────────────────────────────────────────────
        if 'time' in res_df.columns:
            _times = pd.to_datetime(res_df['time'], format='%H:%M:%S', errors='coerce').dropna()
            meta['peak_hour'] = int(_times.dt.hour.value_counts().idxmax()) if not _times.empty else -1
        else:
            meta['peak_hour'] = -1

        # ── Null rate across key CAD fields ──────────────────────────────────
        _key_fields = [c for c in ['lat', 'lon', 'date', 'time', 'priority', 'call_type_desc'] if c in res_df.columns]
        if _key_fields and len(res_df) > 0:
            _null_pct = res_df[_key_fields].isnull().values.mean()
            meta['file_null_rate_pct'] = round(float(_null_pct) * 100, 1)
        else:
            meta['file_null_rate_pct'] = 0.0

        # ── Coordinate detection ─────────────────────────────────────────────
        meta['file_has_lat_lon']  = bool('lat' in res_df.columns and 'lon' in res_df.columns and res_df[['lat','lon']].dropna().shape[0] > 0)
        meta['file_has_priority'] = bool('priority' in res_df.columns and res_df['priority'].dropna().shape[0] > 0)

        # ── Call-type breakdown (top 10) ─────────────────────────────────────
        _type_col = next((c for c in ['call_type_desc','agencyeventtypecodedesc','calldesc','description','nature'] if c in res_df.columns), None)
        if _type_col:
            _tc = res_df[_type_col].dropna().str.strip().value_counts().head(10)
            meta['call_type_breakdown'] = json.dumps({str(k): int(v) for k, v in _tc.items()})
        else:
            meta['call_type_breakdown'] = ''

        # ── Priority distribution ─────────────────────────────────────────────
        if 'priority' in res_df.columns:
            _pc = res_df['priority'].dropna().astype(str).value_counts().sort_index()
            meta['priority_distribution'] = json.dumps({str(k): int(v) for k, v in _pc.items()})
        else:
            meta['priority_distribution'] = ''

    except Exception:
        pass
    return meta
def _deduplicate_columns(df):
    """Rename duplicate column names by appending _2, _3, etc."""
    seen = {}
    new_cols = []
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            new_cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 1
            new_cols.append(c)
    df.columns = new_cols
    return df


def aggressive_parse_calls(uploaded_files, require_valid_coordinates=True):
    all_calls_list = []
    CV = {
        'date': ['received date','incident date','call date','call creation date','calldatetime','call datetime','calltime','timestamp','date','datetime','date time','dispatch date','time received','incdate','date_rept','date_occu','createdtime','created_time','receivedtime','received_time','eventtime','event_time','incidenttime','incident_time','reportedtime','reported_time','entrytime','entry_time','time_central','time_stamp','created'],
        'time': ['call creation time','call time','dispatch time','received time','time', 'hour', 'hour_rept','hour_occu'],
        'priority': ['call priority', 'priority level', 'priority', 'pri', 'urgency'],
        'lat': ['latitude','lat','y coord','ycoord','ycoor','addressy','geoy','y_coord','map_y',
                'point_y','gps_lat','gps_latitude','ylat','coord_y','northing','y_wgs','lat_wgs',
                'incident_lat','inc_lat','event_lat','y_coordinate','address_y','ylocation'],
        'lon': ['longitude','lon','long','x coord','xcoord','xcoor','addressx','geox','x_coord',
                'map_x','point_x','gps_lon','gps_long','gps_longitude','xlon','coord_x','easting',
                'x_wgs','lon_wgs','incident_lon','inc_lon','event_lon','x_coordinate','address_x','xlocation']
    }

    def _coord_column_matches(col_name, patterns):
        norm = str(col_name).strip().lower()
        normalized = norm.replace('-', ' ').replace('_', ' ')
        compact = re.sub(r'[^a-z0-9]+', '', norm)
        for pattern in patterns:
            p_norm = str(pattern).strip().lower()
            p_normalized = p_norm.replace('-', ' ').replace('_', ' ')
            p_compact = re.sub(r'[^a-z0-9]+', '', p_norm)
            if p_norm in norm or p_normalized in normalized or (p_compact and p_compact in compact):
                return True
        return False


    def _looks_like_headerless_geocoder_export(df):
        try:
            cols = [str(c).strip() for c in df.columns]
            coord_pat = re.compile(r'^-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?$')
            zip_pat = re.compile(r'\b[A-Z]{2}\b\s*,\s*\d{5}(?:-\d{4})?$', re.I)
            has_coord_col = any(coord_pat.match(c) for c in cols)
            has_address_col = any(',' in c and zip_pat.search(c) for c in cols)
            has_matchish = any(str(c).strip().lower() in {'match', 'no_match', 'exact', 'non_exact'} for c in cols)
            return has_coord_col and has_address_col and has_matchish
        except Exception:
            return False

    def _normalize_headerless_geocoder_export(df):
        rows = [list(df.columns)] + df.astype(str).fillna('').values.tolist()
        width = max(len(r) for r in rows)
        padded = [r + [''] * (width - len(r)) for r in rows]
        norm = pd.DataFrame(padded)
        base_cols = ['source_id', 'input_address', 'match_status', 'match_type', 'matched_address', 'lonlat', 'external_id', 'side']
        if width > len(base_cols):
            base_cols += [f'extra_{i}' for i in range(1, width - len(base_cols) + 1)]
        norm.columns = base_cols[:width]
        return norm

    def _looks_like_headerless_cad_export(df):
        try:
            if df is None or df.empty or len(df.columns) < 4:
                return False
            known_header_hints = {
                'date', 'time', 'datetime', 'latitude', 'longitude', 'lat', 'lon',
                'priority', 'call type', 'call_type', 'call_type_desc', 'nature',
                'description', 'event_type', 'location', 'address'
            }
            col_names = [str(c).strip().lower() for c in df.columns]
            if any(c in known_header_hints for c in col_names):
                return False

            sample = pd.concat([pd.DataFrame([df.columns], columns=df.columns), df.head(12)], ignore_index=True)
            if sample.empty or len(sample.columns) < 4:
                return False

            def _coord_pair_score(a, b):
                a_num = pd.to_numeric(sample.iloc[:, a], errors='coerce')
                b_num = pd.to_numeric(sample.iloc[:, b], errors='coerce')
                lon_lat_rate = (a_num.between(-180, 180) & b_num.between(-90, 90)).mean()
                lat_lon_rate = (a_num.between(-90, 90) & b_num.between(-180, 180)).mean()
                return max(lon_lat_rate, lat_lon_rate)

            date_rate = _safe_notna_ratio(pd.to_datetime(sample.iloc[:, 0], format='mixed', errors='coerce'))
            time_rate = _safe_notna_ratio(pd.to_datetime(sample.iloc[:, 1], format='%H:%M:%S', errors='coerce')) if len(sample.columns) > 1 else 0.0
            datetime_rate = _safe_notna_ratio(
                pd.to_datetime(
                    sample.iloc[:, 0].astype(str).str.strip() + ' ' + sample.iloc[:, 1].astype(str).str.strip(),
                    format='mixed',
                    errors='coerce',
                )
            ) if len(sample.columns) > 1 else 0.0
            textish_rate = sample.iloc[:, 2].astype(str).str.strip().ne('').mean() if len(sample.columns) > 2 else 0.0
            coord_score = 0.0
            max_scan = min(len(sample.columns) - 1, 7)
            for i in range(2, max_scan):
                coord_score = max(coord_score, _coord_pair_score(i, i + 1))

            return date_rate >= 0.8 and max(time_rate, datetime_rate) >= 0.8 and textish_rate >= 0.8 and coord_score >= 0.8
        except Exception:
            return False

    def _normalize_headerless_cad_export(content, sep):
        raw = pd.read_csv(io.StringIO(content), sep=sep, dtype=str, header=None)
        width = len(raw.columns)
        if width >= 6:
            c4 = pd.to_numeric(raw.iloc[:, 4], errors='coerce')
            c5 = pd.to_numeric(raw.iloc[:, 5], errors='coerce')
            lat_lon_score = (c4.between(-90, 90) & c5.between(-180, 180)).mean()
            lon_lat_score = (c4.between(-180, 180) & c5.between(-90, 90)).mean()
            if lat_lon_score >= lon_lat_score:
                base_cols = ['date', 'time', 'call_type_desc', 'priority', 'lat', 'lon']
            else:
                base_cols = ['date', 'time', 'call_type_desc', 'priority', 'lon', 'lat']
        elif width >= 4:
            c2 = pd.to_numeric(raw.iloc[:, 2], errors='coerce')
            c3 = pd.to_numeric(raw.iloc[:, 3], errors='coerce')
            lon_first_score = (c2.between(-180, 180) & c3.between(-90, 90)).mean()
            lat_first_score = (c2.between(-90, 90) & c3.between(-180, 180)).mean()
            if lon_first_score >= lat_first_score:
                base_cols = ['call_type_desc', 'date', 'lon', 'lat']
            else:
                base_cols = ['call_type_desc', 'date', 'lat', 'lon']
        else:
            base_cols = ['call_type_desc', 'date']
        if width > len(base_cols):
            base_cols += [f'extra_{i}' for i in range(1, width - len(base_cols) + 1)]
        raw.columns = base_cols[:width]
        return raw

    def _extract_lonlat_pair(series):
        s = series.astype(str).str.strip()
        pair = s.str.extract(r'^\s*[\(\[]?\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*[\)\]]?\s*$')
        first = pd.to_numeric(pair[0], errors='coerce')
        second = pd.to_numeric(pair[1], errors='coerce')
        first_first_valid = ((first.between(-180, 180)) & (second.between(-90, 90))).mean()
        second_first_valid = ((first.between(-90, 90)) & (second.between(-180, 180))).mean()
        if second_first_valid >= first_first_valid:
            lon = second
            lat = first
            valid = second_first_valid
        else:
            lon = first
            lat = second
            valid = first_first_valid
        return lon, lat, float(valid)

    def _coerce_coord_series(series, field):
        """
        Convert coordinate text into signed decimal degrees.

        Handles plain numeric values as well as strings like
        "37.10879898° N" and "113.59100342° W".
        """
        s = series.astype(str).str.strip()
        numeric = pd.to_numeric(s, errors='coerce')
        if numeric.notna().any():
            return numeric

        extracted = s.str.extract(
            r'^\s*([+-]?\d+(?:\.\d+)?)\s*(?:\u00b0|deg|d)?\s*([NSEW])?\s*$',
            expand=True,
        )
        values = pd.to_numeric(extracted[0], errors='coerce')
        hemi = extracted[1].fillna('').str.upper()

        parsed = values.copy()
        if field == 'lat':
            parsed = values.abs()
            parsed = parsed.where(~hemi.eq('S'), -parsed)
        else:
            parsed = values.abs()
            parsed = parsed.where(~hemi.eq('W'), -parsed)

        # If a row had no hemisphere marker, preserve the original sign.
        no_hemi = hemi.eq('')
        parsed = parsed.where(~no_hemi, values)
        return parsed

    def _infer_city_from_location_text(raw_df):
        text_cols = [
            c for c in raw_df.columns
            if c in [
                'location', 'address', 'incident_location', 'addr', 'street',
                'input_address', 'matched_address', 'department_name', 'department',
                'dept', 'agency_name', 'agency'
            ]
        ]
        if not text_cols:
            return None

        street_suffixes = {
            'ALY', 'ALLEY', 'AVE', 'AVENUE', 'BLVD', 'BOULEVARD', 'BRG', 'BRIDGE',
            'CIR', 'CIRCLE', 'CT', 'COURT', 'DR', 'DRIVE', 'EXPY', 'FWY', 'HWY',
            'LANE', 'LN', 'LOOP', 'PKWY', 'PARKWAY', 'PL', 'PLACE', 'RD', 'ROAD',
            'RTE', 'SQ', 'ST', 'STREET', 'TER', 'TERRACE', 'TRL', 'WAY', 'WY'
        }

        def _looks_like_street_or_intersection(text):
            text = str(text or '').strip().upper()
            if not text:
                return True
            if '/' in text or '&' in text or re.search(r'\b(?:AND|AT)\b', text):
                return True
            if any(ch.isdigit() for ch in text):
                return True
            tokens = [t for t in re.split(r'[^A-Z0-9]+', text) if t]
            return bool(tokens and tokens[-1] in street_suffixes)

        s = raw_df[text_cols[0]].dropna().astype(str).str.upper().str.strip()
        if s.empty:
            return None

        s = s.str.replace(r':.*$', '', regex=True)
        s = s.str.replace(r'\bCNTY\b', 'COUNTY', regex=True)
        s = s.str.replace(r'[^A-Z0-9 /,-]', ' ', regex=True)
        s = s.str.replace(r'\s+', ' ', regex=True).str.strip()

        candidates = []
        for val in s:
            parts = [p.strip() for p in val.split(',') if p and p.strip()]
            if len(parts) == 1:
                dept_match = re.match(
                    r'^(?:PD|FD|EMS|POLICE|FIRE|SHERIFF|DEPT|DEPARTMENT)\s+(.+?)\s*$',
                    parts[0],
                    flags=re.I,
                )
                if dept_match:
                    locality = dept_match.group(1).strip()
                    if locality and not _looks_like_street_or_intersection(locality):
                        candidates.append(locality.title())
                        continue
                dept_match = re.match(
                    r'^(.+?)\s+(?:PD|FD|EMS|POLICE|FIRE|SHERIFF|DEPT|DEPARTMENT)\s*$',
                    parts[0],
                    flags=re.I,
                )
                if dept_match:
                    locality = dept_match.group(1).strip()
                    if locality and not _looks_like_street_or_intersection(locality):
                        candidates.append(locality.title())
                        continue
            if len(parts) >= 2:
                locality = None
                if len(parts) >= 3 and re.match(r'^[A-Z]{2}$', parts[-2]) and re.match(r'^\d{5}(?:-\d{4})?$', parts[-1]):
                    locality = parts[-3]
                elif len(parts) == 2 and not re.match(r'^[A-Z]{2}(?:\s+\d{5}(?:-\d{4})?)?$', parts[-1]):
                    locality = parts[-1]
                elif re.match(r'^[A-Z]{2}(?:\s+\d{5}(?:-\d{4})?)?$', parts[-1]):
                    locality = parts[-2]
                if locality is None:
                    continue
                locality = locality.strip()
                if (
                    locality and
                    locality not in {'COUNTY', 'CITY', 'TOWN', 'VILLAGE', 'HAMLET'} and
                    not _looks_like_street_or_intersection(locality)
                ):
                    candidates.append(locality.title())
                    continue

        if not candidates:
            return None

        vc = pd.Series(candidates).value_counts()
        return vc.index[0] if not vc.empty else None

    def _normalize_chicagoland_cfs_report(raw_df):
        required = {
            'call', 'date', 'date time_call create', 'call type description',
            'lat', 'lon', 'address', 'department name'
        }
        cols = {str(c).strip().lower() for c in raw_df.columns}
        if not required.issubset(cols):
            return None

        res = raw_df.copy().reset_index(drop=True)
        res['lat'] = pd.to_numeric(res['lat'], errors='coerce')
        res['lon'] = pd.to_numeric(res['lon'], errors='coerce')
        res = res[
            res['lat'].between(17.5, 72) &
            res['lon'].between(-180, -64)
        ].copy()
        if res.empty:
            return None

        dt_source = next((
            c for c in res.columns
            if c == 'date time_call create' or 'date time' in c or 'datetime' in c or 'created' in c
        ), None)
        if dt_source is None and 'date' in res.columns:
            dt_source = 'date'
        if dt_source is not None:
            try:
                dt_series = pd.to_datetime(res[dt_source], format='mixed', errors='coerce')
                res['date'] = dt_series.dt.strftime('%Y-%m-%d')
                res['time'] = dt_series.dt.strftime('%H:%M:%S')
            except Exception:
                pass

        desc_col = next((c for c in res.columns if c == 'call type description' or c in ('call_type_desc', 'call type', 'nature')), None)
        if desc_col is not None:
            res['call_type_desc'] = res[desc_col].fillna('').astype(str).str.strip()

        dept_col = next((c for c in res.columns if c in ('department name', 'department', 'dept', 'agency', 'agency name')), None)
        if dept_col is not None:
            dept_vals = res[dept_col].fillna('').astype(str).str.strip()
            res['agency'] = dept_vals.str.contains(
                r'\b(?:fire|ems|medic|rescue|ambulance|engine|ladder|battalion)\b',
                regex=True,
                na=False,
            ).map({True: 'fire', False: 'police'})
        else:
            res['agency'] = 'police'

        top_city_name = _infer_city_from_location_text(res)
        if not top_city_name and dept_col is not None:
            dept_vals = res[dept_col].fillna('').astype(str).str.upper().str.strip()
            dept_vals = dept_vals.str.replace(r'[^A-Z0-9 /,-]', ' ', regex=True)
            dept_vals = dept_vals.str.replace(r'\s+', ' ', regex=True).str.strip()
            dept_cands = []
            for val in dept_vals:
                m = re.match(r'^(?:PD|FD|EMS|POLICE|FIRE|SHERIFF|DEPT|DEPARTMENT)\s+(.+?)\s*$', val, flags=re.I)
                if m:
                    locality = m.group(1).strip()
                    if locality and not _looks_like_street_or_intersection(locality):
                        dept_cands.append(locality.title())
                        continue
                m = re.match(r'^(.+?)\s+(?:PD|FD|EMS|POLICE|FIRE|SHERIFF|DEPT|DEPARTMENT)\s*$', val, flags=re.I)
                if m:
                    locality = m.group(1).strip()
                    if locality and not _looks_like_street_or_intersection(locality):
                        dept_cands.append(locality.title())
            if dept_cands:
                vc = pd.Series(dept_cands).value_counts()
                if not vc.empty:
                    top_city_name = vc.index[0]

        if top_city_name:
            res['_csv_city'] = top_city_name
        res['_csv_state'] = 'IL'
        res['priority'] = 3
        return res

    def _infer_state_from_text(raw_df, inferred_city=None):
        for col in ['state', 'state_name']:
            if col in raw_df.columns:
                top = raw_df[col].dropna().astype(str).str.strip().value_counts()
                if not top.empty:
                    state_val = top.index[0]
                    state_up = str(state_val).upper()
                    if state_up in STATE_FIPS:
                        return state_up
                    state_title = str(state_val).title()
                    if state_title in US_STATES_ABBR:
                        return US_STATES_ABBR[state_title]

        if inferred_city == 'Mobile':
            return 'AL'
        return None


    def _choose_priority_column(raw_df):
        exact_names = ['priority', 'call priority', 'priority level', 'pri']
        exact = [c for c in raw_df.columns if c.strip().lower() in exact_names]
        if exact:
            exact.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return exact[0]

        loose_names = ['priority', 'call priority', 'priority level', 'pri', 'urgency']
        loose = [c for c in raw_df.columns if any(k in c for k in loose_names)]
        if loose:
            loose.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return loose[0]
        return None

    def parse_priority(raw):
        s = str(raw).strip().upper()
        if not s or s == 'NAN': return None
        # Smart inference for PD offenses if priority column is missing
        if any(w in s for w in ['ROBBERY','BURGLARY','ASSAULT','SHOOTING','STABBING','CRITICAL','EMERG']): return 1
        if any(w in s for w in ['ACCIDENT','DISTURBANCE','THEFT','MED','ALARM']): return 2
        if any(w in s for w in ['NON REPORTABLE','FOUND PROPERTY','INFO','ROUTINE','MISC']): return 4
        
        m = re.search(r'^(\d+)', s)
        if m: return int(m.group(1))
        return 3

    _file_parse_quality = []
    for file_idx, cfile in enumerate(uploaded_files):
        _pq = {
            'file': cfile.name, 'status': 'ok',
            'input_rows': 0, 'output_rows': 0,
            'has_lat': False, 'has_lon': False,
            'has_date': False, 'has_priority_col': False,
            'error': '',
        }
        try:
            fname = cfile.name.lower()
            excel_exts = ('.xlsx', '.xls', '.xlsb', '.xlsm')

            if fname.endswith('.numbers'):
                raw_df = load_numbers_dataframe(cfile.getvalue(), cfile.name)
                raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
                raw_df = _deduplicate_columns(raw_df)
            elif fname.endswith(excel_exts):
                # ── Excel path ────────────────────────────────────────────────
                raw_bytes = cfile.getvalue()
                engine = 'openpyxl'
                if fname.endswith('.xls'):
                    engine = 'xlrd'
                elif fname.endswith('.xlsb'):
                    engine = 'pyxlsb'

                def _sheet_score(ws):
                    score = 0
                    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
                    if not rows:
                        return -1
                    header = rows[0] or []
                    header_norm = [str(h).strip().lower() for h in header if h is not None]
                    if not header_norm:
                        return -1
                    hints = ['latitude', 'longitude', 'lat', 'lon', 'priority', 'location', 'date', 'time']
                    score += sum(10 for h in header_norm if any(k == h or k in h for k in hints))
                    score += sum(1 for h in header_norm if h and not re.match(r'^column\d+$', h))
                    if len(rows) > 1 and rows[1] and any(v is not None and str(v).strip() != '' for v in rows[1]):
                        score += 25
                    # Penalize external-data placeholder sheets
                    if len(header_norm) == 1 and header_norm[0].startswith('externaldata_'):
                        score -= 100
                    return score

                try:
                    import openpyxl as _oxl
                    _wb = _oxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                    _sheet_name = max(_wb.sheetnames, key=lambda sn: _sheet_score(_wb[sn]))
                    _ws = _wb[_sheet_name]
                    _row_iter = _ws.iter_rows(values_only=True)
                    _headers_raw = next(_row_iter)
                    if _headers_raw is None:
                        raise ValueError("Selected Excel sheet has no header row.")
                    _real_idx = [
                        i for i, h in enumerate(_headers_raw)
                        if h is not None and not (str(h).startswith('Column') and str(h)[6:].isdigit())
                    ]
                    if not _real_idx:
                        _real_idx = [i for i, h in enumerate(_headers_raw) if h is not None]
                    _real_headers = [str(_headers_raw[i]).lower().strip() for i in _real_idx]
                    _rows_data = []
                    for _row in _row_iter:
                        if _row is None:
                            continue
                        _trimmed = [_row[i] if i < len(_row) else None for i in _real_idx]
                        if any(v is not None and str(v).strip() != '' for v in _trimmed):
                            _rows_data.append(_trimmed)
                    _wb.close()
                    raw_df = pd.DataFrame(_rows_data, columns=_real_headers)
                    raw_df = raw_df.dropna(how='all')
                    raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
                    raw_df = _deduplicate_columns(raw_df)
                except Exception as _xe:
                    raw_df = None
                    # Try all sheets with pandas and pick the one that looks most like CAD data
                    try:
                        _all = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, sheet_name=None)
                        best_score = -10**9
                        best_df = None
                        for _sn, _df in _all.items():
                            _df.columns = [str(c).lower().strip() for c in _df.columns]
                            _df = _deduplicate_columns(_df)
                            _score = 0
                            for _c in _df.columns:
                                if _c in ('latitude', 'longitude', 'priority', 'location'):
                                    _score += 20
                                elif any(k in _c for k in ['lat', 'lon', 'priority', 'location', 'date', 'time']):
                                    _score += 5
                            _score += min(len(_df), 100)
                            if len(_df.columns) == 1 and str(_df.columns[0]).startswith('externaldata_'):
                                _score -= 100
                            if _score > best_score:
                                best_score = _score
                                best_df = _df
                        if best_df is not None:
                            raw_df = best_df
                    except Exception:
                        pass
                    if raw_df is None:
                        raw_df = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, dtype=str)
                        raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
                        raw_df = _deduplicate_columns(raw_df)
                _jacksonville_df = _normalize_jacksonville_cfs_report(raw_bytes, filename=cfile.name)
                if _jacksonville_df is not None and not _jacksonville_df.empty:
                    raw_df = _jacksonville_df
                else:
                    _priority_calls_df = _normalize_loxley_priority_calls_report(raw_bytes, filename=cfile.name)
                    if _priority_calls_df is not None and not _priority_calls_df.empty:
                        raw_df = _priority_calls_df
            else:
                # ── CSV / TXT path ────────────────────────────────────────────
                content = cfile.getvalue().decode('utf-8', errors='ignore')
                _content_lines = content.splitlines()
                for _header_idx, _line in enumerate(_content_lines[:30]):
                    _line_l = _line.lower()
                    if all(token in _line_l for token in ('call', 'date', 'lat', 'lon', 'address', 'department name')):
                        content = '\n'.join(_content_lines[_header_idx:])
                        break
                first_line = content.split('\n')[0]
                delim = ',' if first_line.count(',') > first_line.count('\t') else '\t'
                raw_df = pd.read_csv(io.StringIO(content), sep=delim, dtype=str)
                _jacksonville_df = _normalize_jacksonville_cfs_report(cfile.getvalue(), filename=cfile.name)
                if _jacksonville_df is not None and not _jacksonville_df.empty:
                    raw_df = _jacksonville_df
                else:
                    _priority_calls_df = _normalize_loxley_priority_calls_report(cfile.getvalue(), filename=cfile.name)
                    if _priority_calls_df is not None and not _priority_calls_df.empty:
                        raw_df = _priority_calls_df
                if _looks_like_headerless_geocoder_export(raw_df):
                    raw_df = _normalize_headerless_geocoder_export(raw_df)
                elif _looks_like_headerless_cad_export(raw_df):
                    raw_df = _normalize_headerless_cad_export(content, delim)
                raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]

                _chicago_report = _normalize_chicagoland_cfs_report(raw_df)
                if _chicago_report is not None and not _chicago_report.empty:
                    res = _chicago_report
                    source_ids = pd.Series(
                        [f"{file_idx}:{cfile.name}:{row_idx}" for row_idx in range(len(raw_df))],
                        index=raw_df.index
                    )
                    res['_source_row_id'] = source_ids.reindex(res.index).values
                    res['_source_file'] = cfile.name
                    try:
                        _meta = _extract_file_meta(raw_df, res, filename=cfile.name)
                        _existing = st.session_state.get('file_meta', {})
                        _existing_names = _existing.get('uploaded_filename', '')
                        if _existing_names and _meta.get('uploaded_filename', '') and _meta['uploaded_filename'] not in _existing_names:
                            _meta['uploaded_filename'] = _existing_names + ' | ' + _meta['uploaded_filename']
                        st.session_state['file_meta'] = {**_existing, **_meta}
                    except Exception:
                        pass
                    _pq['input_rows'] = len(raw_df)
                    _pq['output_rows'] = len(res)
                    _pq['has_lat'] = True
                    _pq['has_lon'] = True
                    _pq['has_date'] = 'date' in res.columns and bool(res['date'].notna().any())
                    _pq['has_priority_col'] = True
                    _file_parse_quality.append(_pq)
                    all_calls_list.append(res.reset_index(drop=True))
                    continue

                # ── Census Geocoder: split combined 'lonlat' column ──────────
                # After normalization the file has a 'lonlat' column storing
                # "lon,lat" pairs (e.g. "-93.283,36.601").  The generic
                # coord-name scanner below matches 'lonlat' for BOTH lat and
                # lon but pd.to_numeric returns all-NaN on comma-pair strings.
                # Handle it here explicitly before the scanner runs, then drop
                # No_Match rows whose lonlat is empty.
                if 'lonlat' in raw_df.columns and 'lat' not in raw_df.columns:
                    _pair = raw_df['lonlat'].astype(str).str.strip().str.extract(
                        r'^\s*(-?[\d.]+)\s*,\s*(-?[\d.]+)\s*$'
                    )
                    _lon_cand = pd.to_numeric(_pair[0], errors='coerce')
                    _lat_cand = pd.to_numeric(_pair[1], errors='coerce')
                    # Census geocoder stores lon first, lat second — verify US range
                    if _lon_cand.between(-180, -50).mean() > 0.3 and _lat_cand.between(18, 72).mean() > 0.3:
                        raw_df['lon'] = _lon_cand
                        raw_df['lat'] = _lat_cand
                    # Drop rows with no geocoded location (No_Match rows)
                    if 'lat' in raw_df.columns:
                        raw_df = raw_df[raw_df['lat'].notna()].copy()

                # ── Census Geocoder: extract city & state from matched_address ──
                # matched_address format: "32 GOLFSHORES DR, BRANSON, MO, 65616"
                if 'matched_address' in raw_df.columns and '_csv_city' not in raw_df.columns:
                    try:
                        _ma = raw_df['matched_address'].dropna().astype(str)
                        _ma_parts = _ma.str.split(',')
                        _ma_cities = _ma_parts.apply(
                            lambda p: p[-3].strip().title() if len(p) >= 4 else (p[-2].strip().title() if len(p) >= 2 else None)
                        ).dropna()
                        _ma_states = _ma_parts.apply(
                            lambda p: p[-2].strip().upper() if len(p) >= 2 else None
                        ).dropna()
                        if not _ma_cities.empty:
                            raw_df['_csv_city'] = _ma_cities.value_counts().index[0]
                        if not _ma_states.empty:
                            _top_st = _ma_states.value_counts().index[0]
                            if _top_st in STATE_FIPS:
                                raw_df['_csv_state'] = _top_st
                    except Exception:
                        pass

            source_ids = pd.Series(
                [f"{file_idx}:{cfile.name}:{row_idx}" for row_idx in range(len(raw_df))],
                index=raw_df.index
            )
            _pq['input_rows'] = len(raw_df)

            if 'lat' in raw_df.columns and 'lon' in raw_df.columns:
                _direct_lat = pd.to_numeric(raw_df['lat'], errors='coerce')
                _direct_lon = pd.to_numeric(raw_df['lon'], errors='coerce')
                _direct_valid_rate = (
                    _direct_lat.between(17.5, 72) &
                    _direct_lon.between(-180, -64)
                ).mean()
                if _direct_valid_rate >= 0.85:
                    res = raw_df.copy().reset_index(drop=True)
                    res['lat'] = _direct_lat.reset_index(drop=True)
                    res['lon'] = _direct_lon.reset_index(drop=True)
                    res = res.dropna(subset=['lat', 'lon'])
                    res = res[
                        (res['lat'].between(17.5, 72)) &
                        (res['lon'].between(-180, -64))
                    ].copy()
                    if not res.empty:
                        dt_source = None
                        for _candidate in [
                            c for c in raw_df.columns
                            if c == 'date time_call create' or 'date time' in c or 'datetime' in c or 'created' in c
                        ]:
                            dt_source = _candidate
                            break
                        if dt_source is None and 'date' in raw_df.columns:
                            dt_source = 'date'
                        if dt_source is not None:
                            try:
                                dt_series = pd.to_datetime(raw_df[dt_source], format='mixed', errors='coerce')
                                res['date'] = dt_series.dt.strftime('%Y-%m-%d')
                                res['time'] = dt_series.dt.strftime('%H:%M:%S')
                            except Exception:
                                pass
                        else:
                            date_cols = [c for c in raw_df.columns if c == 'date']
                            time_cols = [c for c in raw_df.columns if 'time' in c]
                            if date_cols:
                                try:
                                    date_series = pd.to_datetime(raw_df[date_cols[0]], format='mixed', errors='coerce')
                                    res['date'] = date_series.dt.strftime('%Y-%m-%d')
                                except Exception:
                                    pass
                            if time_cols:
                                res['time'] = raw_df[time_cols[0]].fillna('').astype(str).str.strip()
                        desc_col = next((c for c in raw_df.columns if 'call type description' in c or c in ('call_type_desc', 'call type', 'nature')), None)
                        if desc_col is not None:
                            res['call_type_desc'] = raw_df[desc_col].fillna('').astype(str).str.strip()
                        dept_col = next((c for c in raw_df.columns if c in ('department name', 'department', 'dept', 'agency', 'agency name')), None)
                        if dept_col is not None:
                            dept_vals = raw_df[dept_col].fillna('').astype(str).str.strip()
                            res['agency'] = dept_vals.str.contains(
                                r'\b(?:fire|ems|medic|rescue|ambulance|engine|ladder|battalion)\b',
                                regex=True,
                                na=False,
                            ).map({True: 'fire', False: 'police'})
                        else:
                            res['agency'] = 'police'
                        top_city_name = _infer_city_from_location_text(raw_df)
                        if top_city_name:
                            res['_csv_city'] = top_city_name
                        inferred_state = _infer_state_from_text(raw_df, top_city_name)
                        if inferred_state:
                            res['_csv_state'] = inferred_state
                        res['priority'] = 3
                        res['_source_row_id'] = source_ids.reindex(raw_df.index).values
                        res['_source_file'] = cfile.name
                        try:
                            _meta = _extract_file_meta(raw_df, res, filename=cfile.name)
                            _existing = st.session_state.get('file_meta', {})
                            _existing_names = _existing.get('uploaded_filename', '')
                            if _existing_names and _meta.get('uploaded_filename', '') and _meta['uploaded_filename'] not in _existing_names:
                                _meta['uploaded_filename'] = _existing_names + ' | ' + _meta['uploaded_filename']
                            st.session_state['file_meta'] = {**_existing, **_meta}
                        except Exception:
                            pass
                        _pq['output_rows'] = len(res)
                        _pq['has_lat'] = True
                        _pq['has_lon'] = True
                        _pq['has_date'] = 'date' in res.columns and bool(res['date'].notna().any())
                        _pq['has_priority_col'] = True
                        _file_parse_quality.append(_pq)
                        all_calls_list.append(res.reset_index(drop=True))
                        continue

            if '_special_layout' in raw_df.columns and raw_df['_special_layout'].astype(str).eq('jacksonville_cfs').any():
                res = raw_df.copy().reset_index(drop=True)
                if '_special_layout' in res.columns:
                    res = res.drop(columns=['_special_layout'], errors='ignore')
                if 'priority' in res.columns:
                    res['priority'] = pd.to_numeric(res['priority'], errors='coerce').fillna(3).astype(int)
                else:
                    res['priority'] = 3
                if 'agency' not in res.columns:
                    res['agency'] = 'police'
                res['_source_row_id'] = source_ids.values
                res['_source_file'] = cfile.name
                try:
                    _meta = _extract_file_meta(raw_df, res, filename=cfile.name)
                    _existing = st.session_state.get('file_meta', {})
                    _existing_names = _existing.get('uploaded_filename', '')
                    if _existing_names and _meta.get('uploaded_filename', '') and _meta['uploaded_filename'] not in _existing_names:
                        _meta['uploaded_filename'] = _existing_names + ' | ' + _meta['uploaded_filename']
                    st.session_state['file_meta'] = {**_existing, **_meta}
                except Exception:
                    pass
                _pq['output_rows'] = len(res)
                _pq['has_lat'] = False
                _pq['has_lon'] = False
                _pq['has_date'] = 'date' in res.columns and bool(pd.Series(res['date']).notna().any())
                _pq['has_priority_col'] = True
                _file_parse_quality.append(_pq)
                all_calls_list.append(res)
                continue

            res = pd.DataFrame()
            exact_coord_names = {
                'lat': ['latitude', 'lat', 'gps_lat', 'gps_latitude', 'y'],
                'lon': ['longitude', 'lon', 'long', 'gps_lon', 'gps_longitude', 'x']
            }
            for field in ['lat', 'lon']:
                found_exact = [c for c in raw_df.columns if c.strip().lower() in exact_coord_names[field]]
                # Exclude bare 'lonlat' from the loose scan — it's a combined field,
                # not a plain numeric column, and will produce all-NaN via pd.to_numeric.
                found_loose = [c for c in raw_df.columns
                               if c != 'lonlat' and _coord_column_matches(c, CV[field])]
                found = found_exact or found_loose
                if found:
                    res[field] = _coerce_coord_series(raw_df[found[0]], field)

            if 'lat' not in res.columns or 'lon' not in res.columns:
                for c in raw_df.columns:
                    lon_series, lat_series, valid_rate = _extract_lonlat_pair(raw_df[c])
                    if valid_rate >= 0.50:
                        res['lon'] = lon_series
                        res['lat'] = lat_series
                        break

            # ── Fallback: no column name matched — scan numeric columns by value range ──
            # Lat: -90 to 90, Lon: -180 to 180. Pick best candidate for each.
            if 'lat' not in res.columns or 'lon' not in res.columns:
                numeric_cols = []
                for c in raw_df.columns:
                    series = pd.to_numeric(raw_df[c], errors='coerce').dropna()
                    if len(series) > 10:
                        numeric_cols.append((c, series))

                lat_candidates, lon_candidates = [], []
                for c, series in numeric_cols:
                    mn, mx = series.min(), series.max()
                    # Reject if already assigned
                    if c in (res.get('_lat_col',''), res.get('_lon_col','')):
                        continue
                    # Large integer coords (State Plane) — treat as potential coord pair
                    if mx > 1000:
                        lat_candidates.append((c, series))
                        lon_candidates.append((c, series))
                        continue
                    if -90 <= mn and mx <= 90:
                        lat_candidates.append((c, series))
                    if -180 <= mn and mx <= 180 and (mn < -90 or mx > 90):
                        lon_candidates.append((c, series))

                # Prefer candidate whose name hints at lat/lon
                def _score(name, hints):
                    return sum(1 for h in hints if h in name)

                if 'lat' not in res.columns and lat_candidates:
                    lat_candidates.sort(key=lambda x: -_score(x[0], ['lat','y','north']))
                    best_lat_col = lat_candidates[0][0]
                    res['lat'] = pd.to_numeric(raw_df[best_lat_col], errors='coerce')

                if 'lon' not in res.columns and lon_candidates:
                    # Don't reuse the lat column
                    used = res.get('lat', pd.Series()).name if 'lat' in res.columns else None
                    lon_candidates = [(c, s) for c, s in lon_candidates if c != used]
                    if lon_candidates:
                        lon_candidates.sort(key=lambda x: -_score(x[0], ['lon','long','x','east']))
                        best_lon_col = lon_candidates[0][0]
                        res['lon'] = pd.to_numeric(raw_df[best_lon_col], errors='coerce')
            
            _p_col = _choose_priority_column(raw_df)
            p_found = [_p_col] if _p_col else []
            if _p_col:
                parsed_priority = raw_df[_p_col].apply(parse_priority)
                parsed_priority = pd.to_numeric(parsed_priority, errors='coerce')
                parsed_priority = parsed_priority.where(parsed_priority.isin([1, 2, 3, 4, 5, 6, 7, 8, 9]))
                if parsed_priority.dropna().empty:
                    res['priority'] = 3
                else:
                    res['priority'] = parsed_priority.fillna(3).astype(int)
            else:
                # No trustworthy priority field — keep the app usable with a neutral default
                res['priority'] = 3
            
            # ── Event type description — carried through for CAD analytics charts ──
            _desc_hints = ['desc','type','nature','offense','calltype','call_type','event_type',
                           'eventtype','calldesc','incident_type','agencyeventtype','violation','call_nature','cfs_type']
            _desc_found = [c for c in raw_df.columns
                           if any(h in c for h in _desc_hints)
                           and c not in (p_found[:1] if p_found else [])]
            if _desc_found:
                # Pick the column with the most unique text values (most descriptive)
                _best_desc = max(_desc_found, key=lambda c: raw_df[c].dropna().nunique())
                if raw_df[_best_desc].dropna().nunique() > 2:
                    res['call_type_desc'] = raw_df[_best_desc].astype(str).str.strip()

            d_found = [c for c in raw_df.columns if any(s in c for s in CV['date'])]
            t_found = [c for c in raw_df.columns if any(s in c for s in CV['time'])]

            # Fallback: if no date column found by name hint, scan all string columns
            # for any that successfully parse as datetime (catches columns like
            # 'createdtime_central', 'call_ts', 'event_dttm', etc.)
            if not d_found:
                for _col in raw_df.columns:
                    if _col in (t_found or []):
                        continue
                    try:
                        _test = pd.to_datetime(raw_df[_col].dropna().head(50), format='mixed', errors='coerce')
                        _valid = _test.dropna()
                        if len(_valid) >= 10 and _valid.dt.year.between(2000, 2035).mean() > 0.8:
                            d_found = [_col]
                            break
                    except Exception:
                        continue

            if d_found:
                # ── Normalise date/time columns to plain strings ───────────────
                # openpyxl returns Python datetime.datetime / datetime.time objects
                # for date and time cells.  Concatenating them with strings via
                # `+ ' '` raises TypeError and is caught by the bare except,
                # silently discarding the entire file.  Convert to strings first.
                def _col_to_datestr(series):
                    """Convert a column that may contain datetime objects → 'YYYY-MM-DD' strings."""
                    try:
                        _p = pd.to_datetime(series, format='mixed', errors='coerce')
                        if _safe_notna_ratio(_p) > 0.6:
                            return _p.dt.strftime('%Y-%m-%d').where(_p.notna(), '')
                    except Exception:
                        pass
                    return series.fillna('').astype(str).str.strip()

                def _col_to_timestr(series):
                    """Convert a column that may contain datetime.time objects → 'HH:MM:SS' strings."""
                    try:
                        _first = series.dropna().iloc[0] if not series.dropna().empty else None
                        if _first is not None and hasattr(_first, 'strftime'):
                            # openpyxl datetime.time or datetime.datetime objects
                            return series.apply(
                                lambda v: v.strftime('%H:%M:%S') if hasattr(v, 'strftime') else ''
                            ).fillna('')
                    except Exception:
                        pass
                    return series.fillna('').astype(str).str.strip()

                # Build the raw string series to parse — combine date+time cols if separate
                if t_found and d_found[0] != t_found[0]:
                    _raw_dt_str = _col_to_datestr(raw_df[d_found[0]]) + ' ' + _col_to_timestr(raw_df[t_found[0]])
                else:
                    _raw_dt_str = raw_df[d_found[0]].fillna('').astype(str).str.strip()

                # Try explicit common formats first (orders of magnitude faster than
                # dateutil fallback on large files, and avoids NaT on ghost rows).
                # Format detection: sample the first non-null value.
                _sample_vals = _raw_dt_str.dropna().astype(str).str.strip()
                _sample_vals = _sample_vals[_sample_vals != ''].head(5)
                _fmt_candidates = [
                    '%m/%d/%Y %I:%M %p',   # 2/14/2025 6:03 PM  (Mobile AL)
                    '%m/%d/%Y %H:%M:%S',   # 2/14/2025 18:03:00
                    '%m/%d/%Y %H:%M',      # 2/14/2025 18:03
                    '%Y-%m-%d %H:%M:%S',   # 2025-02-14 18:03:00
                    '%Y-%m-%dT%H:%M:%S',   # ISO 8601
                    '%Y-%m-%d %H:%M',      # 2025-02-14 18:03
                    '%Y/%m/%d %H:%M:%S',
                    '%d/%m/%Y %H:%M:%S',
                    '%m-%d-%Y %H:%M:%S',
                ]
                dt_series = None
                if not _sample_vals.empty:
                    for _fmt in _fmt_candidates:
                        try:
                            _trial = pd.to_datetime(_sample_vals.iloc[0], format=_fmt, errors='raise')
                            # Format matched — apply to full series
                            dt_series = pd.to_datetime(_raw_dt_str, format=_fmt, errors='coerce')
                            break
                        except Exception:
                            continue
                if dt_series is None:
                    # Final fallback: let pandas infer (slow but handles edge cases)
                    dt_series = pd.to_datetime(_raw_dt_str, format='mixed', errors='coerce')

                res['date'] = dt_series.dt.strftime('%Y-%m-%d')
                res['time'] = dt_series.dt.strftime('%H:%M:%S')

            # --- COORDINATE CLEANUP: sentinel values & sign errors ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns:
                _preserve_direct_geo = False
                try:
                    _lat_geo = pd.to_numeric(res['lat'], errors='coerce')
                    _lon_geo = pd.to_numeric(res['lon'], errors='coerce')
                    _geo_rate = (
                        _lat_geo.between(17.5, 72) &
                        _lon_geo.between(-180, -64)
                    ).mean()
                    if _geo_rate >= 0.85:
                        res['lat'] = _lat_geo
                        res['lon'] = _lon_geo
                        _preserve_direct_geo = True
                except Exception:
                    _preserve_direct_geo = False
                _coord_scale_max = max(res['lat'].abs().max(), res['lon'].abs().max())
                if _coord_scale_max <= 1000:
                    # Drop obvious sentinel/null-coordinate rows before any further processing.
                    # Keep this limited to decimal-scale coordinates so integer microdegrees
                    # and projected coordinates are not discarded before conversion.
                    _sentinel_mask = (
                        (res['lat'] == 0) | (res['lon'] == 0) |
                        (res['lat'].abs() < 0.001) | (res['lon'].abs() < 0.001) |
                        (res['lon'] < -179.9)
                    )
                    if _sentinel_mask.any():
                        res = res[~_sentinel_mask].copy()

                    # Fix wrong-sign longitudes: some CAD exports omit the minus sign for
                    # western-hemisphere longitudes (e.g. 81.31 instead of -81.31).
                    if not res.empty and 'lon' in res.columns:
                        _neg_count = (res['lon'] < 0).sum()
                        _pos_count = (res['lon'] > 0).sum()
                        _total = len(res)
                        if _neg_count > 0 and _pos_count > 0 and (_neg_count / _total) > 0.90:
                            _median_neg = res.loc[res['lon'] < 0, 'lon'].median()
                            _pos_vals = res.loc[res['lon'] > 0, 'lon']
                            _would_match = ((-_pos_vals).between(_median_neg - 2, _median_neg + 2)).mean()
                            if _would_match > 0.5:
                                res.loc[res['lon'] > 0, 'lon'] = -res.loc[res['lon'] > 0, 'lon']

            # --- COORDINATE CONVERSION (MICRODEGREES / STATE PLANE / LARGE-INTEGER DETECTOR) ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns and not locals().get('_preserve_direct_geo', False):
                res = res[(res['lat'] != 0) & (res['lon'] != 0)].dropna(subset=['lat', 'lon'])
                if not res.empty:
                    max_val = max(res['lat'].abs().max(), res['lon'].abs().max())
                    if max_val > 1000:
                        converted = False
                        # Common CAD export pattern: integer microdegrees
                        # (-98281987, 30568167) -> (-98.281987, 30.568167).
                        _lon_micro = res['lon'] / 1_000_000.0
                        _lat_micro = res['lat'] / 1_000_000.0
                        _micro_valid = (
                            _lat_micro.between(18, 72) &
                            _lon_micro.between(-170, -60)
                        ).mean()
                        if _micro_valid > 0.80:
                            res['lon'] = _lon_micro
                            res['lat'] = _lat_micro
                            converted = True

                        # Strategy 1: Try common State Plane CRS at /100 and /1 scales
                        if not converted:
                            candidate_crs = [
                                "EPSG:2278",  # TX South Central (ftUS)
                                "EPSG:2277",  # TX Central (ftUS)
                                "EPSG:2276",  # TX North Central (ftUS)
                                "EPSG:2279",  # TX South (ftUS)
                                "EPSG:32140", # TX South Central (m)
                            ]
                            for scale in [100.0, 1.0]:
                                for crs in candidate_crs:
                                    try:
                                        transformer = pyproj.Transformer.from_crs(crs, "EPSG:4326", always_xy=True)
                                        test_lons, test_lats = transformer.transform(
                                            res['lon'].values[:20] / scale,
                                            res['lat'].values[:20] / scale
                                        )
                                        if (24 < float(test_lats.mean()) < 50 and
                                                -130 < float(test_lons.mean()) < -60 and
                                                float(test_lats.std()) < 5 and
                                                float(test_lons.std()) < 5):
                                            lons, lats = transformer.transform(
                                                res['lon'].values / scale, res['lat'].values / scale
                                            )
                                            res['lon'], res['lat'] = lons, lats
                                            converted = True
                                            break
                                    except Exception:
                                        continue
                                if converted:
                                    break

                        # Strategy 2: If CRS conversion failed, anchor to city column geocode
                        if not converted:
                            try:
                                city_name = None
                                for col in ['city', 'city_name', 'municipality', 'jurisdiction']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            city_name = top.index[0]
                                            break
                                state_name = None
                                for col in ['state', 'state_name']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            state_name = top.index[0]
                                            break
                                if city_name:
                                    query_str = f"{city_name}, {state_name}" if state_name else city_name
                                    geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(query_str)}&limit=1"
                                    req = urllib.request.Request(geo_url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"})
                                    with urllib.request.urlopen(req, timeout=10) as resp:
                                        geo_data = json.loads(resp.read().decode("utf-8"))
                                    if geo_data:
                                        anchor_lat = float(geo_data[0]["lat"])
                                        anchor_lon = float(geo_data[0]["lon"])
                                        raw_cx = res["lon"].median()
                                        raw_cy = res["lat"].median()
                                        city_radius_deg = 0.35
                                        raw_spread = max(res["lon"].std(), res["lat"].std(), 1)
                                        deg_per_unit = city_radius_deg / raw_spread
                                        res["lon"] = anchor_lon + (res["lon"] - raw_cx) * deg_per_unit
                                        res["lat"] = anchor_lat + (res["lat"] - raw_cy) * deg_per_unit
                                        converted = True
                            except Exception:
                                pass

                        if converted:
                            res = res[
                                (res["lat"] > 18) & (res["lat"] < 72) &
                                (res["lon"] > -170) & (res["lon"] < -60)
                            ]

            # ── Agency / source tagging (Fire vs Police) ─────────────────────
            # Look for a column named 'agency', 'department', or 'dept' and
            # carry it through as a lowercase 'agency' column so the map
            # renderer can colour fire calls red and police calls the default colour.
            _agency_col = next(
                (c for c in raw_df.columns if c.strip().lower() in ('agency', 'department', 'dept')),
                None
            )
            if _agency_col:
                res['agency'] = raw_df[_agency_col].astype(str).str.strip().str.lower()
            else:
                res['agency'] = 'police'   # safe default for single-agency files

            # Agency / source tagging (Fire vs Police)
            # Prefer a column that contains fire/police labels when duplicate agency
            # fields exist (for example Agency + Agency.1 after CSV import).
            _agency_candidates = [
                c for c in raw_df.columns
                if c.strip().lower() in ('agency', 'agency.1', 'department', 'department.1', 'dept', 'dept.1',
                                         'agencyname', 'agencyname.1', 'agency_name', 'agency_name.1')
            ]
            _agency_col = None
            for _cand in _agency_candidates:
                try:
                    _vals = raw_df[_cand].astype(str).str.strip().str.lower()
                    if _vals.str.contains(r'\b(?:fire|police|ems|sheriff)\b', regex=True, na=False).any():
                        _agency_col = _cand
                        break
                except Exception:
                    pass
            if _agency_col is None:
                _agency_col = _agency_candidates[0] if _agency_candidates else None

            if _agency_col:
                _raw_agency = raw_df[_agency_col].astype(str).str.strip().str.lower()
                # Normalize to canonical 'fire' / 'police' so the renderer colours correctly.
                # If a raw value already IS 'fire' or 'police', keep it; otherwise look for
                # fire-department keywords and fall back to 'police'.
                _fire_kw = r'fire|ems|medic|rescue|ambulance|engine|ladder|battalion'
                _is_fire = _raw_agency.str.contains(_fire_kw, regex=True, na=False)
                res['agency'] = _is_fire.map({True: 'fire', False: 'police'})
            else:
                # Fall back to filename-based agency detection when no agency column exists
                _fname_lower = str(cfile.name).lower()
                if any(k in _fname_lower for k in ('fire', 'ems', 'medic', 'rescue', 'ambulance', 'engine', 'ladder', 'battalion')):
                    res['agency'] = 'fire'
                else:
                    res['agency'] = 'police'   # safe default for single-agency files

            # City/state detection: store top values on rows for location detection
            top_city_name = None
            for col in ["city", "city_name", "municipality", "jurisdiction"]:
                if col in raw_df.columns:
                    top_city = raw_df[col].dropna().astype(str).str.strip().value_counts()
                    if not top_city.empty:
                        top_city_name = str(top_city.index[0]).title()
                        res["_csv_city"] = top_city_name
                        break

            if "_csv_city" not in res.columns:
                inferred_city = _infer_city_from_location_text(raw_df)
                if inferred_city:
                    top_city_name = inferred_city
                    res["_csv_city"] = inferred_city

            inferred_state = _infer_state_from_text(raw_df, top_city_name)
            if not inferred_state:
                for _addr_col in ['input_address', 'matched_address', 'address', 'location']:
                    if _addr_col in raw_df.columns:
                        _addr_series = raw_df[_addr_col].astype(str)
                        # Pattern 1: "..., ST, ZIPCODE" (comma-separated state and zip)
                        _states = _addr_series.str.extract(r',\s*([A-Z]{2})\s*,\s*\d{5}(?:-\d{4})?')[0].dropna()
                        if _states.empty:
                            # Pattern 2: "..., ST ZIPCODE" (state and zip in same segment, common format)
                            _states = _addr_series.str.extract(r',\s*([A-Z]{2})\s+\d{5}(?:-\d{4})?')[0].dropna()
                        if not _states.empty:
                            inferred_state = _states.value_counts().idxmax()
                            break
            if inferred_state:
                res["_csv_state"] = inferred_state

            res["_source_row_id"] = source_ids.reindex(res.index).values
            res["_source_file"] = cfile.name

            # ── Capture file data matrix for Sheets/email logging ────────────
            try:
                _meta = _extract_file_meta(raw_df, res, filename=cfile.name)
                # Merge into session-level file_meta (last file wins for per-field values;
                # accumulate filenames if multiple files are uploaded at once)
                _existing = st.session_state.get('file_meta', {})
                _existing_names = _existing.get('uploaded_filename', '')
                if _existing_names and _meta.get('uploaded_filename','') and _meta['uploaded_filename'] not in _existing_names:
                    _meta['uploaded_filename'] = _existing_names + ' | ' + _meta['uploaded_filename']
                st.session_state['file_meta'] = {**_existing, **_meta}
            except Exception:
                pass

            _pq['output_rows'] = len(res)
            _pq['has_lat'] = 'lat' in res.columns and bool(res['lat'].notna().any())
            _pq['has_lon'] = 'lon' in res.columns and bool(res['lon'].notna().any())
            _pq['has_date'] = 'date' in res.columns and bool(res['date'].notna().any())
            _pq['has_priority_col'] = bool(_p_col)
            _file_parse_quality.append(_pq)
            all_calls_list.append(res)
        except Exception as _pq_e:
            _pq['status'] = 'error'
            _pq['error'] = str(_pq_e)[:300]
            _file_parse_quality.append(_pq)
            continue

    try:
        st.session_state['parse_quality'] = _file_parse_quality
    except Exception:
        pass
    if not all_calls_list: return pd.DataFrame()
    if require_valid_coordinates:
        # Only keep frames that actually have lat/lon columns — Excel sheets
        # without coordinate data should not crash the concat
        valid = [df for df in all_calls_list if 'lat' in df.columns and 'lon' in df.columns]
        if not valid: return pd.DataFrame()
        combined = pd.concat(valid, ignore_index=True)
        # Safe dropna — columns guaranteed to exist now
        combined = combined.dropna(subset=['lat', 'lon'])
        combined['lat'] = pd.to_numeric(combined['lat'], errors='coerce')
        combined['lon'] = pd.to_numeric(combined['lon'], errors='coerce')
        # Filter to US territory bounds — eliminates corrupted placeholder coords (e.g. -1,-1)
        # that pass the zero-sentinel but would distort the bounding box and UTM zone calculation
        combined = combined[(combined['lat'].between(17.5, 72)) & (combined['lon'].between(-180, -64))]
    else:
        combined = pd.concat(all_calls_list, ignore_index=True)
        if 'lat' in combined.columns:
            combined['lat'] = pd.to_numeric(combined['lat'], errors='coerce')
        if 'lon' in combined.columns:
            combined['lon'] = pd.to_numeric(combined['lon'], errors='coerce')
    # IMPORTANT: keep the full parsed CAD dataset here.
    #
    # The optimizer is sampled later (after upload) for performance, but the
    # parsed dataframe itself must preserve every incident so:
    #   1) Total Incidents shows the true uploaded count
    #   2) the stations map can render a much denser full-history call cloud
    #   3) export/reporting math stays tied to the source file, not a k-means
    #      surrogate created during parsing
    return combined

def _get_annualized_calls(raw_count: int) -> int:
    """Return raw_count scaled to a full year using the uploaded file's date span.

    If the file covers less than a full year (and at least 14 days), the raw
    count is extrapolated to 365 days.  Falls back to raw_count when no date
    span is available (simulated data, unknown span, or span ≥ 330 days).
    """
    span_days = int(st.session_state.get('file_meta', {}).get('file_date_span_days', 0) or 0)
    if 14 <= span_days < 330:
        return round(raw_count * 365 / span_days)
    return raw_count
